"""BRSF Overall Star — month-wise star + cash reward report.

Pure report over the existing BRSF monthly finals: no star is recalculated and
no record is created here. Cash reward is derived from the FINAL net stars via
the single reward band table below.
"""
import csv
import io

from fastapi import Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from server import api_router, db, get_current_user
from brsf_stars import (
    MONTH_LABELS,
    _month_is_completed,
    _month_eligible_employees_raw,
    _require_star_admin,
)

# stars (inclusive lower bound) -> category / cash / action
REWARD_BANDS = [
    (25, "Research Premier", 11000, "Research Premier Performer"),
    (22, "Research Champion", 9000, "Champion Performer"),
    (19, "Research Outstanding", 7000, "Outstanding Performer"),
    (16, "Research Excellence", 5000, "Top Performer"),
    (9, "Research Elite", 3000, "Consistent Excellence"),
    (6, "Research Consistent", 2000, "Reliable Contributor"),
    (4, "Satisfactory", 1000, "Needs Improvement"),
    (1, "Needs Improvement", 0, "Performance Monitoring"),
]
UNSAFE_BAND = ("Unsafe Behavior", 0, "Disciplinary Review")


def cash_reward(total_stars) -> dict:
    """The ONE star -> reward mapping used by the report, export and UI."""
    stars = int(round(float(total_stars)))
    for floor, category, cash, action in REWARD_BANDS:
        if stars >= floor:
            return {"category": category, "cash": cash, "action": action}
    category, cash, action = UNSAFE_BAND
    return {"category": category, "cash": cash, "action": action}


def _month_keys(from_month: str, to_month: str) -> list:
    if from_month > to_month:
        raise HTTPException(status_code=400, detail="From Month cannot be later than To Month.")
    keys, y, m = [], int(from_month[:4]), int(from_month[5:7])
    end = (int(to_month[:4]), int(to_month[5:7]))
    while (y, m) <= end:
        keys.append(f"{y:04d}-{m:02d}")
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
        if len(keys) > 36:
            raise HTTPException(status_code=400, detail="Please select a range of 36 months or less.")
    return keys


async def _report(from_month: str, to_month: str, employee_id=None, team=None) -> dict:
    months = _month_keys(from_month, to_month)
    completed = [m for m in months if _month_is_completed(m)]
    skipped = [m for m in months if m not in completed]
    if not completed:
        return {"months": [], "rows": [], "skipped_months": skipped,
                "reward_bands": _bands_payload()}

    # --- one employee pass + one totals pass for the whole range (no N+1)
    emps = await _month_eligible_employees_raw()
    all_teams = sorted({e.get("team") or "" for e in emps if e.get("team")})
    if employee_id:
        emps = [e for e in emps if e["id"] == employee_id]
    if team:
        emps = [e for e in emps if (e.get("team") or "") == team]
    ids = [e["id"] for e in emps]
    y_lo, y_hi = int(completed[0][:4]), int(completed[-1][:4])
    finals = {}
    async for l in db.brsf_star_lines.find(
        {"employee_id": {"$in": ids}, "year": {"$gte": y_lo, "$lte": y_hi}},
        {"_id": 0, "employee_id": 1, "year": 1, "month": 1, "final_value": 1},
    ):
        key = (l["employee_id"], f"{l['year']:04d}-{l['month']:02d}")
        finals[key] = finals.get(key, 0) + (l.get("final_value") or 0)

    rows = []
    for e in emps:
        cells, cash_total, any_eligible = [], 0, False
        for m in completed:
            if not _is_eligible(e, m):
                cells.append({"month": m, "state": "not_eligible"})
                continue
            any_eligible = True
            key = (e["id"], m)
            if key not in finals:
                cells.append({"month": m, "state": "not_calculated"})
                continue
            stars = int(round(finals[key]))
            band = cash_reward(stars)
            cash_total += band["cash"]
            cells.append({"month": m, "state": "value", "stars": stars, **band})
        if not any_eligible:
            continue  # eligible for no month in the range
        rows.append({
            "id": e["id"], "full_name": e.get("full_name"),
            "custom_employee_id": e.get("custom_employee_id") or e.get("emp_id"),
            "team": e.get("team") or "", "date_of_joining": e.get("date_of_joining"),
            "confirmation_date": e.get("confirmation_date"),
            "inactive_date": e.get("inactive_date"),
            "employee_status": e.get("employee_status"),
            "cells": cells, "cash_total": cash_total,
        })
    rows.sort(key=lambda r: (r["full_name"] or ""))
    return {
        "months": [{"key": m, "label": f"{MONTH_LABELS[int(m[5:7]) - 1][:3]} {m[:4]}"} for m in completed],
        "rows": rows, "skipped_months": skipped, "reward_bands": _bands_payload(),
        "teams": all_teams,
    }


BAND_LABELS = ["25+", "22-24", "19-21", "16-18", "9-15", "6-8", "4-5", "1-3"]


def _bands_payload():
    return [{"stars": BAND_LABELS[i], "category": c, "cash": cash, "action": a}
            for i, (_f, c, cash, a) in enumerate(REWARD_BANDS)]


def _is_eligible(emp: dict, month: str) -> bool:
    from brsf_stars import _month_eligible
    return _month_eligible(emp, month)


@api_router.get("/brsf/overall")
async def brsf_overall(from_month: str, to_month: str, employee_id: str = None,
                       team: str = None, current_user: dict = Depends(get_current_user)):
    _require_star_admin(current_user)
    return await _report(from_month, to_month, employee_id, team)


THIN = Side(style="thin", color="FF999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@api_router.get("/brsf/overall/export")
async def brsf_overall_export(from_month: str, to_month: str, employee_id: str = None,
                              team: str = None, format: str = "xlsx",
                              current_user: dict = Depends(get_current_user)):
    """Exports exactly the filtered report with dynamic month columns."""
    _require_star_admin(current_user)
    if format not in ("xlsx", "csv"):
        raise HTTPException(status_code=400, detail="Export format must be xlsx or csv")
    rep = await _report(from_month, to_month, employee_id, team)
    months, rows = rep["months"], rep["rows"]
    stem = f"BRSF_Overall_Star_{from_month}_to_{to_month}"
    no_cache = {"Cache-Control": "no-store"}

    def cell_pair(c):
        if c["state"] == "not_eligible":
            return "-", "-"
        if c["state"] == "not_calculated":
            return "NC", "NC"
        return c["stars"], c["cash"]

    if format == "csv":
        out = io.StringIO()
        w = csv.writer(out)
        head = ["employee_name", "team_name", "joining_date", "confirmation_date"]
        for m in months:
            head += [f"{m['key']}_stars", f"{m['key']}_cash_reward"]
        head.append("cash_reward_total")
        w.writerow(head)
        for r in rows:
            line = [r["full_name"], r["team"], r["date_of_joining"], r["confirmation_date"]]
            for c in r["cells"]:
                line += list(cell_pair(c))
            line.append(r["cash_total"])
            w.writerow(line)
        return StreamingResponse(io.BytesIO(out.getvalue().encode("utf-8-sig")),
                                 media_type="text/csv",
                                 headers={"Content-Disposition": f'attachment; filename="{stem}.csv"',
                                          **no_cache})

    wb = Workbook()
    ws = wb.active
    ws.title = "Overall Star"
    header = ["Employee Name", "Team Name", "Date Of Joining", "Date Of Employee Confirmation"]
    for m in months:
        header += [m["label"], f"{m['label'][:3].upper()} Cash Rewards"]
    header.append("Cash Rewards Total")
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor="FFF2F2F2")
        c.border = BORDER
    ws.row_dimensions[1].height = 46
    ws.freeze_panes = "E2"
    for i, letter in enumerate(["A", "B", "C", "D"]):
        ws.column_dimensions[letter].width = [24, 26, 14, 18][i]
    for col in range(5, 5 + len(months) * 2 + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    for idx, r in enumerate(rows, start=2):
        line = [r["full_name"], r["team"], r["date_of_joining"], r["confirmation_date"]]
        for c in r["cells"]:
            line += list(cell_pair(c))
        ws.append(line)
        cash_cols = [get_column_letter(6 + 2 * i) + str(idx) for i in range(len(months))]
        total_col = get_column_letter(5 + len(months) * 2)
        ws[f"{total_col}{idx}"] = f"=SUM({','.join(cash_cols)})" if cash_cols else 0
        ws[f"{total_col}{idx}"].font = Font(bold=True)
        for c in ws[idx]:
            c.border = BORDER
            if c.column >= 5:
                c.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"', **no_cache})
