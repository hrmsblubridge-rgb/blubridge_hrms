from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query, UploadFile, File as FastAPIFile, Body, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
import re
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import hashlib
import cloudinary
import cloudinary.utils
import cloudinary.uploader
import resend
import time
import io
import csv
import openpyxl
import settings_module

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Indian Standard Time (IST) - UTC+5:30
IST = timezone(timedelta(hours=5, minutes=30))

def get_ist_now():
    """Get current datetime in Indian Standard Time"""
    return datetime.now(IST)

def get_ist_today():
    """Get today's date string in IST (format: DD-MM-YYYY)"""
    return get_ist_now().strftime("%d-%m-%Y")

# ============================================================
# Cross-midnight attendance handling (configurable)
# ------------------------------------------------------------
# Punches occurring at or before CROSS_MIDNIGHT_THRESHOLD_MINUTES
# (minutes since midnight, default 300 = 05:00 AM) are treated as
# belonging to the PREVIOUS working day (i.e., the OUT punch of an
# overnight shift). Punches after the threshold are treated as
# belonging to their own calendar date (a new working day).
# Configure via env var CROSS_MIDNIGHT_THRESHOLD_MINUTES.
# ============================================================
CROSS_MIDNIGHT_THRESHOLD_MINUTES = int(os.environ.get("CROSS_MIDNIGHT_THRESHOLD_MINUTES", "300"))

def get_effective_attendance_date(punch_dt: datetime) -> str:
    """Return DD-MM-YYYY attendance date for a punch, handling cross-midnight.

    If the punch time is at or before the configured threshold, the punch is
    attributed to the previous calendar day. Otherwise, its own calendar day.
    """
    minutes_since_midnight = punch_dt.hour * 60 + punch_dt.minute
    effective = punch_dt - timedelta(days=1) if minutes_since_midnight <= CROSS_MIDNIGHT_THRESHOLD_MINUTES else punch_dt
    return effective.strftime("%d-%m-%Y")

def attendance_shift_offset(time_24h: str) -> Optional[int]:
    """Compute shift-time offset in minutes for an HH:MM time string.

    Times at or before the cross-midnight threshold are treated as occurring
    in the NEXT calendar day relative to the effective attendance date, so
    their offset is minutes_since_midnight + 1440. This allows MIN/MAX
    comparisons to correctly identify IN (earliest in shift) and OUT
    (latest in shift) across cross-midnight shifts.
    """
    if not time_24h:
        return None
    try:
        hours, minutes = map(int, time_24h.split(":"))
    except (ValueError, AttributeError):
        return None
    mins = hours * 60 + minutes
    if mins <= CROSS_MIDNIGHT_THRESHOLD_MINUTES:
        return mins + 1440
    return mins

def is_sunday_ddmmyyyy(date_str: str) -> bool:
    """Return True if a DD-MM-YYYY date string falls on a Sunday."""
    if not date_str:
        return False
    try:
        d = datetime.strptime(date_str, "%d-%m-%Y")
        return d.weekday() == 6
    except (ValueError, TypeError):
        return False

# Temporary onboarding-skip window. New employees created within this
# window (inclusive of both endpoints, evaluated in IST) are auto-approved
# so they can log in and land directly on the employee dashboard without
# going through the onboarding flow. Outside this window the standard
# onboarding flow auto-resumes (no further code changes needed).
ONBOARDING_SKIP_WINDOW_START = datetime(2026, 4, 28, 0, 0, 0, tzinfo=IST)
ONBOARDING_SKIP_WINDOW_END = datetime(2026, 5, 3, 23, 59, 59, tzinfo=IST)

def should_skip_onboarding_now() -> bool:
    """Return True if the current IST time falls within the temporary
    skip-onboarding window."""
    now = get_ist_now()
    return ONBOARDING_SKIP_WINDOW_START <= now <= ONBOARDING_SKIP_WINDOW_END

def add_hours_to_24h(time_24h: str, hours: float) -> Optional[str]:
    """Return HH:MM string for time_24h + hours (wraps modulo 24)."""
    if not time_24h or hours is None:
        return None
    try:
        h, m = map(int, time_24h.split(":"))
    except (ValueError, AttributeError):
        return None
    total_minutes = (h * 60 + m + int(round(float(hours) * 60))) % (24 * 60)
    return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(
    mongo_url,
    serverSelectionTimeoutMS=10000,
    connectTimeoutMS=10000,
    retryWrites=True,
    retryReads=True,
    maxPoolSize=50,
    minPoolSize=5
)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'blubridge-hrms-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Cloudinary Configuration
cloudinary.config(
    cloud_name=os.environ.get("CLOUDINARY_CLOUD_NAME"),
    api_key=os.environ.get("CLOUDINARY_API_KEY"),
    api_secret=os.environ.get("CLOUDINARY_API_SECRET"),
    secure=True
)

# Resend Configuration
resend.api_key = os.environ.get("RESEND_API_KEY")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "onboarding@resend.dev")

# HRMS automated email system (centralized)
from email_service import ensure_email_indexes as _ensure_email_indexes, ensure_cron_settings_seed as _ensure_cron_settings_seed  # noqa: E402
from email_jobs import start_email_scheduler as _start_email_scheduler, get_job_handlers as _get_email_job_handlers, JOB_META as _CRON_JOB_META  # noqa: E402

# Create the main app
app = FastAPI(title="BluBridge HRMS API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============== ENUMS ==============

class UserRole:
    HR = "hr"
    SYSTEM_ADMIN = "system_admin"
    OFFICE_ADMIN = "office_admin"
    EMPLOYEE = "employee"

# Role permission groups for authorization checks
ADMIN_ROLES = ["hr"]  # Full write/approve/delete access
ALL_ADMIN_ROLES = ["hr", "system_admin", "office_admin"]  # View access to admin pages
SYSTEM_ROLES = ["hr", "system_admin"]  # System management (audit logs, roles, settings)

def require_admin(user):
    """Check if user has full admin (HR) access"""
    if user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Insufficient permissions. HR access required.")

def require_any_admin(user):
    """Check if user has any admin role (view access)"""
    if user["role"] not in ALL_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Insufficient permissions. Admin access required.")

def require_system_admin(user):
    """Check if user has system admin access"""
    if user["role"] not in SYSTEM_ROLES:
        raise HTTPException(status_code=403, detail="Insufficient permissions. System admin access required.")

class EmploymentType:
    FULL_TIME = "Full-time"
    PART_TIME = "Part-time"
    CONTRACT = "Contract"
    INTERN = "Intern"

class EmployeeStatus:
    ACTIVE = "Active"
    INACTIVE = "Inactive"
    RESIGNED = "Resigned"

class TierLevel:
    JUNIOR = "Junior"
    MID = "Mid"
    SENIOR = "Senior"
    LEAD = "Lead"

class WorkLocation:
    REMOTE = "Remote"
    OFFICE = "Office"
    HYBRID = "Hybrid"

# ============== SHIFT CONFIGURATION ==============

# Predefined shift timings (24-hour format for calculations)
SHIFT_DEFINITIONS = {
    "General": {
        "login_time": "10:00",  # 10:00 AM
        "logout_time": "21:00",  # 9:00 PM
        "total_hours": 11,
        "description": "Standard shift (10:00 AM - 9:00 PM)"
    },
    "Morning": {
        "login_time": "06:00",  # 6:00 AM
        "logout_time": "14:00",  # 2:00 PM
        "total_hours": 8,
        "description": "Morning shift (6:00 AM - 2:00 PM)"
    },
    "Evening": {
        "login_time": "14:00",  # 2:00 PM
        "logout_time": "22:00",  # 10:00 PM
        "total_hours": 8,
        "description": "Evening shift (2:00 PM - 10:00 PM)"
    },
    "Night": {
        "login_time": "22:00",  # 10:00 PM
        "logout_time": "06:00",  # 6:00 AM (next day)
        "total_hours": 8,
        "description": "Night shift (10:00 PM - 6:00 AM)"
    },
    "Flexible": {
        "login_time": None,  # No fixed time
        "logout_time": None,
        "total_hours": 8,
        "description": "Flexible shift (8 hours required, no fixed time)"
    },
    "Custom": {
        "login_time": None,  # Set by admin per employee
        "logout_time": None,
        "total_hours": None,  # Auto-calculated
        "description": "Custom shift (admin-defined timings)"
    }
}

class AttendanceStatus:
    PRESENT = "Present"
    LATE_LOGIN = "Late Login"
    EARLY_OUT = "Early Out"
    LOSS_OF_PAY = "Loss of Pay"
    LEAVE = "Leave"
    ABSENT = "Absent"
    NOT_LOGGED = "Not Logged"
    LOGIN = "Login"
    COMPLETED = "Completed"
    SUNDAY = "Sunday"
    WORKED_ON_SUNDAY = "Worked on Sunday"

# ============== ONBOARDING ENUMS ==============

class OnboardingStatus:
    PENDING = "pending"  # Employee created, not started onboarding
    IN_PROGRESS = "in_progress"  # Employee has started uploading docs
    UNDER_REVIEW = "under_review"  # All docs submitted, awaiting HR review
    APPROVED = "approved"  # HR approved, full HRMS access granted
    REJECTED = "rejected"  # HR rejected, needs re-upload

class DocumentStatus:
    NOT_UPLOADED = "not_uploaded"
    UPLOADED = "uploaded"
    VERIFIED = "verified"
    REJECTED = "rejected"

class DocumentType:
    AADHAAR_CARD = "aadhaar_card"
    PAN_CARD = "pan_card"
    PASSPORT = "passport"
    VOTER_ID = "voter_id"
    EDUCATION = "education"
    EXPERIENCE = "experience"
    OFFER_LETTER = "offer_letter"
    RELIEVING_LETTER = "relieving_letter"
    PHOTO = "photo"

REQUIRED_DOCUMENTS = [
    {"type": DocumentType.AADHAAR_CARD, "label": "Aadhaar Card", "required": True},
    {"type": DocumentType.PAN_CARD, "label": "PAN Card", "required": True},
    {"type": DocumentType.PASSPORT, "label": "Passport", "required": False},
    {"type": DocumentType.VOTER_ID, "label": "Voter ID", "required": False},
    {"type": DocumentType.EDUCATION, "label": "Education Certificates", "required": True},
    {"type": DocumentType.EXPERIENCE, "label": "Experience Certificates", "required": False},
    {"type": DocumentType.OFFER_LETTER, "label": "Offer / Appointment Letter", "required": False},
    {"type": DocumentType.RELIEVING_LETTER, "label": "Relieving Letter", "required": False},
    {"type": DocumentType.PHOTO, "label": "Passport-size Photograph", "required": True},
]

# ============== HOLIDAYS DATA ==============

COMPANY_HOLIDAYS_2026 = [
    {"id": "h1", "name": "New Year's Day", "date": "2026-01-01", "day": "Thursday", "type": "national"},
    {"id": "h2", "name": "Pongal", "date": "2026-01-15", "day": "Thursday", "type": "regional"},
    {"id": "h3", "name": "Thiruvalluvar Day", "date": "2026-01-16", "day": "Friday", "type": "regional"},
    {"id": "h4", "name": "Republic Day", "date": "2026-01-26", "day": "Monday", "type": "national"},
    {"id": "h5", "name": "Ramzan (Eid ul-Fitr)", "date": "2026-03-21", "day": "Saturday", "type": "religious", "note": "Subject to lunar calendar"},
    {"id": "h6", "name": "Good Friday", "date": "2026-04-03", "day": "Friday", "type": "religious"},
    {"id": "h7", "name": "Mahavir Jayanti", "date": "2026-04-03", "day": "Friday", "type": "religious"},
    {"id": "h8", "name": "Tamil New Year (Puthandu)", "date": "2026-04-14", "day": "Tuesday", "type": "regional"},
    {"id": "h9", "name": "May Day (Labour Day)", "date": "2026-05-01", "day": "Friday", "type": "national"},
    {"id": "h10", "name": "Bakrid (Eid ul-Adha)", "date": "2026-05-27", "day": "Wednesday", "type": "religious", "note": "Subject to lunar calendar"},
    {"id": "h11", "name": "Muharram", "date": "2026-07-06", "day": "Monday", "type": "religious", "note": "Subject to lunar calendar"},
    {"id": "h12", "name": "Independence Day", "date": "2026-08-15", "day": "Saturday", "type": "national"},
    {"id": "h13", "name": "Vinayaka Chaturthi", "date": "2026-08-22", "day": "Saturday", "type": "religious"},
    {"id": "h14", "name": "Gandhi Jayanti", "date": "2026-10-02", "day": "Friday", "type": "national"},
    {"id": "h15", "name": "Diwali", "date": "2026-11-09", "day": "Monday", "type": "religious"},
    {"id": "h16", "name": "Christmas Day", "date": "2026-12-25", "day": "Friday", "type": "religious"},
]

# ============== COMPANY POLICIES ==============

COMPANY_POLICIES = [
    {
        "id": "policy_leave",
        "name": "Notice \u2013 List of Holidays",
        "category": "HR",
        "version": "2026.1",
        "effective_date": "2026-01-01",
        "applicable_to": "All Employees & Interns",
        "content": {
            "overview": "This is to formally notify all employees and interns that the holidays listed below are company-declared holidays for the calendar year 2026.",
            "sections": [
                {
                    "title": "Operating Framework",
                    "text": "As a research-focused organization, we follow a flexible operating framework to ensure uninterrupted access to research infrastructure and learning environments. Accordingly, the declaration of holidays is for reference and compensation purposes only and does not imply a mandatory office closure.\n\nEmployees and interns may, based on individual preference and project requirements, choose to either avail leave or remain actively engaged in research and learning activities on these days. Those who choose to remain actively engaged shall be eligible for additional compensation, in accordance with the Holiday Work / Extra Pay Policy and the terms governing their respective engagement.\n\nEngagement in research and learning activities on declared holidays is purely voluntary, and no employee or intern is obligated to avail leave on the listed dates."
                },
                {
                    "title": "A. Standard Holidays",
                    "table": {
                        "headers": ["Sl. No", "Holiday", "Date", "Day"],
                        "rows": [
                            ["1", "New Year's Day", "01-Jan-2026", "Thursday"],
                            ["2", "Republic Day", "26-Jan-2026", "Monday"],
                            ["3", "Tamil New Year (Puthandu)", "14-Apr-2026", "Tuesday"],
                            ["4", "May Day (Labour Day)", "01-May-2026", "Friday"],
                            ["5", "Independence Day", "15-Aug-2026", "Saturday"],
                            ["6", "Gandhi Jayanthi", "02-Oct-2026", "Friday"]
                        ]
                    }
                },
                {
                    "title": "B. Optional Festival Holiday Bucket",
                    "text": "Employees and interns may choose any four (4) days only from the list below in a calendar year.",
                    "table": {
                        "headers": ["Sl. No", "Festival", "Date", "Day"],
                        "rows": [
                            ["1", "Pongal", "15-Jan-2026", "Thursday"],
                            ["2", "Thiruvallur Day", "16-Jan-2026", "Friday"],
                            ["3", "Ramzan*", "21-Mar-2026*", "Saturday*"],
                            ["4", "Mahavir Jayanti", "31-Mar-2026", "Tuesday"],
                            ["5", "Good Friday", "03-Apr-2026", "Friday"],
                            ["6", "Bakrid*", "27-May-2026*", "Wednesday*"],
                            ["7", "Muharram*", "06-Jul-2026*", "Monday*"],
                            ["8", "Vinayaka Chaturthi", "22-Aug-2026", "Saturday"],
                            ["9", "Diwali", "09-Nov-2026", "Monday"],
                            ["10", "Christmas Day", "25-Dec-2026", "Friday"]
                        ]
                    },
                    "footer": "* Dates marked are subject to change based on official or lunar calendar announcements."
                },
                {
                    "title": "Mid-Year Joiner Eligibility \u2014 Optional Festival Holiday Bucket",
                    "text": "For employees and interns joining during the course of the calendar year, eligibility under the Optional Festival Holiday Bucket shall apply on a pro-rated basis according to the date of joining, as set out below. Optional festival holidays may be chosen only from the declared Optional Festival Holiday Bucket, cannot be carried forward, shall lapse if not used within the same calendar year, and remain subject to prior intimation, approval, and organizational requirements. This clarification applies only to the Optional Festival Holiday Bucket and does not affect the Standard Holidays declared by the Company.",
                    "table": {
                        "headers": ["Joining Period", "Maximum Optional Festival Holidays Eligible"],
                        "rows": [
                            ["01 January to 31 March", "4 Days"],
                            ["01 April to 30 June", "3 Days"],
                            ["01 July to 30 September", "2 Days"],
                            ["01 October to 31 December", "1 Day"]
                        ]
                    }
                },
                {
                    "title": "Important Clarification",
                    "text": "The holiday framework consists of six (6) standard holidays and up to four (4) optional festival holidays in a calendar year, subject to the eligibility framework applicable to mid-year joiners. All holidays are non-mandatory, the office remains accessible for research and learning activities at all times, and this list is issued solely for internal reference and planning purposes, subject to organizational requirements."
                }
            ]
        }
    },
    {
        "id": "policy_it",
        "name": "BluBridge IT and Communication Policy",
        "category": "Company-wide",
        "version": "1.0",
        "effective_date": "2026-01-01",
        "applicable_to": "All employees, contractors, and third-party users who access company systems",
        "content": {
            "overview": "This IT and Communication Policy sets forth the guidelines for the proper use of BluBridge information technology (IT) resources, communication tools, and data to ensure the security, privacy, and integrity of company assets. It applies to all employees, contractors, and third-party users who access company systems.",
            "sections": [
                {
                    "title": "1. Purpose",
                    "items": [
                        "This IT and Communication Policy sets forth the guidelines for the proper use of BluBridge information technology (IT) resources, communication tools, and data to ensure the security, privacy, and integrity of company assets.",
                        "It applies to all employees, contractors, and third-party users who access company systems."
                    ]
                },
                {
                    "title": "2. Scope",
                    "items": [
                        "Company-provided devices and IT systems.",
                        "Network and internet usage.",
                        "Email and communication platforms.",
                        "Data security and confidentiality.",
                        "Social media and public communication.",
                        "Remote work."
                    ]
                },
                {
                    "title": "3. Acceptable Use of IT Resources",
                    "items": [
                        "Company Devices: All employees are responsible for the security, maintenance, and appropriate use of company-owned devices (laptops, mobile phones, tablets, etc.).",
                        "Network Access: Employees must use secure login credentials to access company systems and are prohibited from sharing passwords with unauthorized individuals.",
                        "Software and Applications: Only authorized software approved by BluBridge IT department may be installed on company devices. Use of pirated or unauthorized software is strictly prohibited."
                    ]
                },
                {
                    "title": "4. Internet and Email Usage",
                    "items": [
                        "Internet Access: Employees are granted access to the internet for business-related purposes. Personal use of the internet should be limited and must not interfere with job responsibilities.",
                        "Email Use: Company-provided email accounts must be used for work-related communications. Employees should not use personal email accounts for business purposes.",
                        "Prohibited Activities: The following activities are not permitted using company resources — accessing or distributing illegal, offensive, or inappropriate content; engaging in personal business activities or unauthorized file sharing; downloading large files that are not business-related, which could burden network resources."
                    ]
                },
                {
                    "title": "5. Data Security and Confidentiality",
                    "items": [
                        "Data Protection: Employees are responsible for protecting sensitive and confidential company data, including customer information, proprietary materials, and financial records.",
                        "Password Security: Employees must use strong, unique passwords and enable multi-factor authentication (MFA) where required. Passwords must not be shared or reused across multiple platforms.",
                        "Confidentiality: Confidential company information must not be disclosed to unauthorized individuals or entities, both inside and outside the organization.",
                        "Reporting Security Breaches: Any suspected data breach, loss, or theft of company devices or sensitive information must be reported to the IT department immediately."
                    ]
                },
                {
                    "title": "6. Remote Workers",
                    "items": [
                        "Remote Access: Remote workers must use secure methods, such as a Virtual Private Network (VPN), to access the company's network from offsite locations."
                    ]
                },
                {
                    "title": "7. Social Media and Public Communication",
                    "items": [
                        "Social Media Use: Employees are prohibited from posting confidential or proprietary company information on social media platforms. Personal opinions expressed on social media must not be presented as those of BluBridge.",
                        "Communication with External Parties: Employees must obtain approval before sharing company-related information with the media, clients, or third-party vendors. Any public statements or press releases must be vetted through the appropriate department."
                    ]
                },
                {
                    "title": "8. Monitoring and Privacy",
                    "items": [
                        "Company Rights to Monitor: BluBridge reserves the right to monitor its IT systems, including email, internet use, and company devices, to ensure compliance with this policy and to maintain security. Employees should have no expectation of privacy when using company systems.",
                        "Data Retention: Emails, files, and other communications may be retained for legal and business purposes. Employees are responsible for managing their email inboxes and other storage systems in accordance with retention policies."
                    ]
                },
                {
                    "title": "9. Violation of Policy",
                    "items": [
                        "Disciplinary Action: Violations of this policy may result in disciplinary action, up to and including termination of employment, as well as potential legal consequences.",
                        "Reporting Violations: Employees are encouraged to report any suspected violations of this policy to their supervisor or the IT department."
                    ]
                },
                {
                    "title": "10. Review and Amendments",
                    "items": [
                        "This policy will be reviewed annually by the IT and HR departments and updated as necessary to reflect changes in technology, legal requirements, and business needs."
                    ]
                }
            ]
        }
    },
    {
        "id": "policy_research",
        "name": "Research Publication Bonus Policy",
        "category": "Department",
        "version": "1.0",
        "effective_date": "2025-12-15",
        "applicable_to": "All interns and employees in the Research Unit",
        "content": {
            "overview": "This Policy outlines the structure and eligibility for the Research Publication Bonus awarded to individuals contributing to publications under the BluBridge affiliation. It encourages original research, fosters innovation, and rewards successful publication, peer review, and citation efforts that strengthen BluBridge's research excellence.",
            "sections": [
                {
                    "title": "1. Objective",
                    "text": "This Policy outlines the structure and eligibility for the Research Publication Bonus awarded to individuals contributing to publications under the BluBridge affiliation. It encourages original research, fosters innovation, and rewards successful publication, peer review, and citation efforts that strengthen BluBridge's research excellence."
                },
                {
                    "title": "2. Scope",
                    "text": "This Policy applies to all interns and employees in the Research Unit who author, co-author, or contribute to any research paper submitted, accepted, or published under the BluBridge Technologies Private Limited affiliation."
                },
                {
                    "title": "3. Bonus Structure",
                    "text": "The total eligible Research Publication Bonus per paper, per person is \u20b9 99,000 (Rupees Ninety-Nine Thousand only), disbursed in stages based on specific publication milestones and citation performance. Each stage's bonus will be released upon successful verification of the corresponding milestone.",
                    "table": {
                        "headers": ["Milestone", "Description", "Bonus (\u20b9)", "Condition of Payment"],
                        "rows": [
                            [
                                "A. At Publishing",
                                "Upon successful publication or acceptance of the paper on arXiv (authenticated pre-print).",
                                "10,000",
                                "Released upon arXiv posting and verification of BluBridge affiliation."
                            ],
                            [
                                "B. Peer Review (Conference / Journal)",
                                "Upon completion of peer review or acceptance in an approved international conference or journal.",
                                "45,000",
                                "Released after peer-review acknowledgement or presentation confirmation."
                            ],
                            [
                                "C. Citation-Based Bonus\n(Based on total citations achieved within the first year after publication)",
                                "\u2022 1 \u2013 9 citations\n\u2022 10 \u2013 19 citations\n\u2022 20 and above citations",
                                "5,000\n19,500\n19,500",
                                "Released upon reaching this citation range.\nReleased upon reaching this citation range.\nReleased upon reaching this citation range."
                            ],
                            [
                                "Total Potential Bonus (Per Paper, Per Author):",
                                "",
                                "\u20b999,000",
                                ""
                            ]
                        ]
                    },
                    "footer": "The list of recommended publication and presentation venues is provided in Appendix A of this Policy for reference and compliance."
                },
                {
                    "title": "4. Eligibility & Conditions",
                    "items": [
                        "Each BluBridge-affiliated author is independently eligible to receive the full \u20b9 99,000 Performance Bonus per published paper.",
                        "The publication must clearly mention BluBridge Technologies Pvt. Ltd. as an official author affiliation.",
                        "The employee or intern must be on active engagement or payroll at the time of bonus disbursement.",
                        "Proof of publication, arXiv link, acceptance communication, and verified citation data must be submitted to the Super Team Lead and Management.",
                        "Citation-based payouts occur once the verified citation count enters an eligible range within the first year after publication.",
                        "All approvals are subject to validation by the Super Team Lead and Management.",
                        "The Company reserves the right to revise, defer, or withdraw this Policy based on organizational or research priorities."
                    ]
                },
                {
                    "title": "5. Effective Date and Validity",
                    "text": "This Policy is effective from 15 December 2025 and will remain in force until amended, superseded, versioned, or withdrawn by BluBridge Technologies Private Limited. The Company reserves the right to modify, update, withdraw, or reissue this Policy in line with organizational, operational, or research-related requirements."
                },
                {
                    "title": "Appendix A \u2014 Recognized Publication and Presentation Venues",
                    "table": {
                        "headers": ["Category", "Name of Venue / Repository", "Type", "Remarks / Scope"],
                        "rows": [
                            ["Publishing (Pre-Print Repository)", "arXiv (relevant CS sections)", "Open Repository", "For authenticated pre-print publishing under BluBridge affiliation"],
                            ["Conferences & Journals", "NeurIPS \u2013 Neural Information Processing Systems", "International Conference", "Machine Learning and AI Research"],
                            ["", "ICML \u2013 International Conference on Machine Learning", "International Conference", "Core ML Research"],
                            ["", "ICLR \u2013 International Conference on Learning Representations", "International Conference", "Representation Learning / Deep Networks"],
                            ["", "AAAI \u2013 Association for the Advancement of AI", "International Conference", "General AI Applications"],
                            ["", "IJCAI \u2013 International Joint Conference on Artificial Intelligence", "International Conference", "Cross-domain AI Research"],
                            ["", "ACL \u2013 Association for Computational Linguistics", "International Conference", "NLP and Language Models"],
                            ["", "JMLR \u2013 Journal of Machine Learning Research", "Peer-Reviewed Journal", "ML Research and Theory"],
                            ["", "TPAMI \u2013 IEEE Transactions on Pattern Analysis and Machine Intelligence", "Peer-Reviewed Journal", "Vision, AI, and Pattern Recognition"],
                            ["", "AIJ \u2013 Artificial Intelligence Journal", "Peer-Reviewed Journal", "General AI and Methodologies"],
                            ["", "Nature Machine Intelligence", "Peer-Reviewed Journal", "Foundational and Applied AI Research"]
                        ]
                    }
                }
            ]
        }
    },
    {
        "id": "policy_admin_induction",
        "name": "Admin Induction Guidelines",
        "category": "Admin",
        "version": "1.0",
        "effective_date": "2026-01-01",
        "applicable_to": "All Employees",
        "content": {
            "overview": "Office layout, facilities, communication channels, and admin support guidelines for every BluBridge employee.",
            "sections": [
                {
                    "title": "1. Office Layout",
                    "items": [
                        "Employees are introduced to the complete office premises, including parking area, restrooms, cafeteria, mini pantry, and restricted zones.",
                        "Please ensure you use only the designated areas as permitted."
                    ]
                },
                {
                    "title": "2. Lunch & Snack Timings",
                    "items": [
                        "Beverages (Coffee, Tea, etc.): Available from 11:00 AM in the cafeteria.",
                        "Lunch: 1:00 PM – 2:30 PM",
                        "Evening Snacks: 5:00 PM – 6:00 PM"
                    ]
                },
                {
                    "title": "3. Food & Snack Guidelines",
                    "items": [
                        "Employees are requested to consume food responsibly and adhere to portion guidelines (e.g., limited number of eggs, chicken pieces, etc.) to ensure availability for all."
                    ]
                },
                {
                    "title": "4. Point of Contact for Issues",
                    "items": [
                        "For any concerns, issues, or requirements, please reach out to the Admin Team."
                    ]
                },
                {
                    "title": "5. Communication Channels",
                    "items": [
                        "Connect with the Admin Team through: G Space, Gmail, G Chat.",
                        "If there is no response online, you may approach the Reception Desk directly."
                    ]
                },
                {
                    "title": "6. Mobile & Bag Deposit Policy",
                    "items": [
                        "All employees are required to deposit mobile phones and bags in the designated cupboards at the reception before entering the work area.",
                        "These can be accessed during break times if required."
                    ]
                },
                {
                    "title": "7. Library & Leisure Area",
                    "items": [
                        "A dedicated library is available on the 1st floor, containing educational, fiction, and non-fiction books.",
                        "Employees are encouraged to utilize this space during their leisure time."
                    ]
                },
                {
                    "title": "8. Stationery Requirements",
                    "items": [
                        "For any stationery needs, please raise a request with the Admin Team through the communication channels.",
                        "Items will be arranged accordingly."
                    ]
                },
                {
                    "title": "9. Workstation & Facility Support",
                    "items": [
                        "For any issues related to chairs, desks, air conditioning, or refreshment areas, the Admin Team will provide prompt support and resolution."
                    ]
                },
                {
                    "title": "10. IT Support",
                    "items": [
                        "For any system or technical issues, please contact the IT Team. The IT Head, Anbu, will assist in resolving all IT-related concerns."
                    ]
                }
            ]
        }
    },
    {
        "id": "policy_support_hr",
        "name": "Support Team — HR & Leave Policy",
        "category": "Department",
        "version": "1.0",
        "effective_date": "2026-01-01",
        "applicable_to": "Support Staff team",
        "content": {
            "overview": "HR Induction, working-hour expectations, EOD reporting, leave format, company protocol and reimbursement rules for the Support team.",
            "sections": [
                {
                    "title": "1. Working Hours Policies",
                    "items": [
                        "Total working hours: 10 hours per day",
                        "Productive learning/work time: 8 hours 30 minutes",
                        "Break time: 1 hour 30 minutes",
                        "Productivity Criteria — Minimum learning hours: 8 hours 30 minutes; Total working hours requirement: 10 hours",
                        "Break Schedule — 1st Break: 11:00 AM | 2nd Break (Lunch): 01:00 PM | 3rd Break: 05:00 PM",
                        "(Break time can be utilized flexibly within the total allowed duration)"
                    ]
                },
                {
                    "title": "2. End of Day (EOD) Reporting",
                    "items": [
                        "EOD Submission is Mandatory.",
                        "All employees must submit a detailed End of Day (EOD) report via email on a daily basis without exception.",
                        "Email Details — To: Manoj | CC: Praveen, Kripa & Team",
                        "Submission Timeline — EOD must be sent before logging off for the day.",
                        "Reports must be clear, structured, and detailed.",
                        "Incomplete or vague updates will be considered non-compliance.",
                        "Failure to submit EOD may impact performance evaluation and accountability."
                    ]
                },
                {
                    "title": "3. Leave Email Format",
                    "items": [
                        "All leave requests must strictly follow the format below.",
                        "Email Recipients — To: Manoj | CC: Praveen, Kripa, HR Team",
                        "Subject Line Format: [Leave Type] – [Full Day/Half Day] – [Date] (e.g. Sick Leave – Full Day – 12 Feb 2026)",
                        "Email Body should specify: Type of Leave (Emergency / Sick / Pre-Planned / Late Coming), Date(s), Reason.",
                        "Ensure all details are accurate and complete.",
                        "Incomplete emails may lead to delay or rejection of leave approval.",
                        "Must be sent within the defined timelines as per leave policy."
                    ]
                },
                {
                    "title": "4. Leave Policy",
                    "items": [
                        "Emergency Leave (Half/Full Day) — for unforeseen or urgent situations. Must inform via email between 9:45 AM – 10:00 AM. Supporting documents are mandatory.",
                        "Sick Leave (Half/Full Day) — Email intimation required before 7:00 AM. If unwell earlier, inform in advance. On recovery, reply to the same email thread mentioning rejoining time.",
                        "Pre-Planned Leave (Half/Full Day) — Must be applied at least 4 days in advance.",
                        "Late Coming (LC) — Must inform before 9:45 AM (latest by 10:00 AM); should include expected arrival time.",
                        "Half Day Criteria — Minimum 5 hours of productive work.",
                        "Leave Approval Process — All leaves must be approved by management (via Teams). Post approval: send email and update in HRMS portal."
                    ]
                },
                {
                    "title": "5. Company Protocol",
                    "items": [
                        "Salary / Stipend — Processed on the 1st day of every month. If the 1st falls on a Sunday or declared holiday, payment is processed on the last working day of the previous month.",
                        "Workplace Discipline — Sleeping on the floor is strictly prohibited. If feeling drowsy, take a short break and resume work.",
                        "Communication — Employees must remain active on Teams and Google Chat. All official communication will be shared through these platforms. Official mail, Teams and HRMS must be logged in on personal mobile.",
                        "Attendance Compliance — Biometric punch and register signing are mandatory. Failure will result in payroll processing issues.",
                        "Device Policy — Mobile phones are restricted on the work floor; must be kept at the reception area.",
                        "Professional Presence — Employees must update their LinkedIn profile after joining.",
                        "Facilities Provided — Lunch, dinner, evening snacks; tea and coffee available."
                    ]
                },
                {
                    "title": "6. Reimbursement Policy (If Any)",
                    "items": [
                        "All reimbursements require prior management approval. No exceptions.",
                        "Without approval, claims will not be processed.",
                        "Submission Process — Send email with: Expense details and Valid bills/invoices (mandatory).",
                        "Email Details — To: Accounts Team, Front Office | CC: HR Team, Kripa, Praveen.",
                        "Claims must be submitted promptly.",
                        "Incomplete or unclear submissions will be rejected.",
                        "Fake or invalid bills will lead to strict disciplinary action."
                    ]
                }
            ]
        }
    },
    {
        "id": "policy_research_hr",
        "name": "Research Team — HR & Leave Policy",
        "category": "Department",
        "version": "1.0",
        "effective_date": "2026-01-01",
        "applicable_to": "Research Unit",
        "content": {
            "overview": "HR Induction, working-hour expectations (11 hrs), EOD reporting, leave format, company protocol and reimbursement rules (incl. PG accommodation) for the Research team.",
            "sections": [
                {
                    "title": "1. Working Hours Policies",
                    "items": [
                        "Total working hours: 11 hours per day",
                        "Productive learning/work time: 9 hours 30 minutes",
                        "Break time: 1 hour 30 minutes",
                        "Productivity Criteria — Minimum learning hours: 9 hours 30 minutes; Total working hours requirement: 11 hours",
                        "Break Schedule — 1st Break: 11:00 AM | 2nd Break (Lunch): 01:00 PM | 3rd Break: 05:00 PM",
                        "(Break time can be utilized flexibly within the total allowed duration)"
                    ]
                },
                {
                    "title": "2. End of Day (EOD) Reporting",
                    "items": [
                        "EOD Submission is Mandatory.",
                        "All employees must submit a detailed End of Day (EOD) report via email on a daily basis without exception.",
                        "Email Details — To: Manoj | CC: Praveen, Kripa & Team",
                        "Submission Timeline — EOD must be sent before logging off for the day.",
                        "Reports must be clear, structured, and detailed.",
                        "Incomplete or vague updates will be considered non-compliance.",
                        "Failure to submit EOD may impact performance evaluation and accountability."
                    ]
                },
                {
                    "title": "3. Leave Email Format",
                    "items": [
                        "All leave requests must strictly follow the format below.",
                        "Email Recipients — To: Manoj | CC: Praveen, Kripa, HR Team",
                        "Subject Line Format: [Leave Type] – [Full Day/Half Day] – [Date] (e.g. Sick Leave – Full Day – 12 Feb 2026)",
                        "Email Body should specify: Type of Leave (Emergency / Sick / Pre-Planned / Late Coming), Date(s), Reason.",
                        "Ensure all details are accurate and complete.",
                        "Incomplete emails may lead to delay or rejection of leave approval.",
                        "Must be sent within the defined timelines as per leave policy."
                    ]
                },
                {
                    "title": "4. Leave Policy",
                    "items": [
                        "Emergency Leave (Half/Full Day) — for unforeseen or urgent situations. Must inform via email between 9:45 AM – 10:00 AM. Supporting documents are mandatory.",
                        "Sick Leave (Half/Full Day) — Email intimation required before 7:00 AM. If unwell earlier, inform in advance. On recovery, reply to the same email thread mentioning rejoining time.",
                        "Pre-Planned Leave (Half/Full Day) — Must be applied at least 4 days in advance.",
                        "Late Coming (LC) — Must inform before 9:45 AM (latest by 10:00 AM); should include expected arrival time.",
                        "Half Day Criteria — Minimum 5 hours of productive work.",
                        "Leave Approval Process — All leaves must be approved by management (via Teams). Post approval: send email and update in HRMS portal."
                    ]
                },
                {
                    "title": "5. Company Protocol",
                    "items": [
                        "Salary / Stipend — Processed on the 1st day of every month. If the 1st falls on a Sunday or declared holiday, payment is processed on the last working day of the previous month.",
                        "Workplace Discipline — Sleeping on the floor is strictly prohibited. If feeling drowsy, take a short break and resume work.",
                        "Communication — Employees must remain active on Teams and Google Chat. All official communication will be shared through these platforms. Official mail, Teams and HRMS must be logged in on personal mobile.",
                        "Attendance Compliance — Biometric punch and register signing are mandatory. Failure will result in payroll processing issues.",
                        "Device Policy — Mobile phones are restricted on the work floor; must be kept at the reception area.",
                        "Professional Presence — Employees must update their LinkedIn profile after joining.",
                        "Facilities Provided — Lunch, dinner, evening snacks; tea and coffee available."
                    ]
                },
                {
                    "title": "6. Reimbursement Policy",
                    "items": [
                        "PG (Paying Guest) Accommodation Reimbursement — The company will cover 50% of PG accommodation expenses, subject to a maximum limit of ₹5,000 per month. This reimbursement is applicable only with valid bills and prior approval.",
                        "Submission Process — Send email with: Expense details and Valid bills/invoices (mandatory).",
                        "Email Details — To: Accounts Team, Front Office | CC: HR Team, Kripa, Praveen.",
                        "Claims must be submitted promptly.",
                        "Incomplete or unclear submissions will be rejected.",
                        "Fake or invalid bills will lead to strict disciplinary action."
                    ]
                }
            ]
        }
    }
]

# ============== MODELS ==============

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    password_hash: str
    name: str
    role: str = UserRole.EMPLOYEE
    employee_id: Optional[str] = None
    department: Optional[str] = None
    team: Optional[str] = None
    onboarding_status: str = OnboardingStatus.PENDING  # NEW: Track onboarding state
    is_first_login: bool = True  # NEW: Track if first login
    created_at: datetime = Field(default_factory=lambda: get_ist_now())
    is_active: bool = True

class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    token: str
    user: dict

# Comprehensive Employee Model
class Employee(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    
    # Personal Information
    emp_id: str  # Auto-generated
    full_name: str
    official_email: str
    phone_number: Optional[str] = None
    gender: Optional[str] = None  # Male, Female, Other
    date_of_birth: Optional[str] = None
    
    # Employment Information
    custom_employee_id: Optional[str] = None  # Admin-defined Employee ID
    date_of_joining: str
    employment_type: str = EmploymentType.FULL_TIME
    employee_status: str = EmployeeStatus.ACTIVE
    designation: str
    tier_level: str = TierLevel.MID
    reporting_manager_id: Optional[str] = None
    
    # Organization Structure
    department: str
    team: str
    work_location: str = WorkLocation.OFFICE
    
    # HR Configuration
    leave_policy: Optional[str] = "Standard"
    shift_type: Optional[str] = "General"
    attendance_tracking_enabled: bool = True
    
    # Custom Shift Configuration (only for shift_type = "Custom")
    custom_login_time: Optional[str] = None  # Format: "HH:MM" (24-hour)
    custom_logout_time: Optional[str] = None  # Format: "HH:MM" (24-hour)
    custom_total_hours: Optional[float] = None  # Auto-calculated from times
    
    # Salary Configuration (for payroll)
    monthly_salary: Optional[float] = 0.0
    
    # System Access
    user_role: str = UserRole.EMPLOYEE
    login_enabled: bool = True
    biometric_id: Optional[str] = None  # For biometric device mapping
    
    # Legacy fields for compatibility
    avatar: Optional[str] = None
    stars: int = 0
    unsafe_count: int = 0
    
    # Onboarding fields
    onboarding_status: str = OnboardingStatus.PENDING
    onboarding_completed_at: Optional[datetime] = None
    
    # Soft delete
    is_deleted: bool = False
    deleted_at: Optional[datetime] = None
    
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: get_ist_now())
    updated_at: datetime = Field(default_factory=lambda: get_ist_now())

class EmployeeCreate(BaseModel):
    # Personal Information
    full_name: str
    official_email: str
    phone_number: Optional[str] = None
    gender: Optional[str] = None
    date_of_birth: Optional[str] = None
    
    # Employment Information
    custom_employee_id: Optional[str] = None  # Admin-defined Employee ID
    date_of_joining: str
    employment_type: str = EmploymentType.FULL_TIME
    designation: str
    tier_level: str = TierLevel.MID
    reporting_manager_id: Optional[str] = None
    
    # Organization Structure
    department: str
    team: str
    work_location: str = WorkLocation.OFFICE
    
    # HR Configuration
    leave_policy: Optional[str] = "Standard"
    shift_type: Optional[str] = "General"
    attendance_tracking_enabled: bool = True
    
    # Custom Shift Configuration
    custom_login_time: Optional[str] = None
    custom_logout_time: Optional[str] = None
    custom_total_hours: Optional[float] = None
    
    # Salary Configuration
    monthly_salary: Optional[float] = 0.0
    
    # System Access
    user_role: str = UserRole.EMPLOYEE
    login_enabled: bool = True
    biometric_id: Optional[str] = None  # For biometric device mapping

class EmployeeUpdate(BaseModel):
    # Personal Information
    full_name: Optional[str] = None
    official_email: Optional[str] = None
    phone_number: Optional[str] = None
    gender: Optional[str] = None
    date_of_birth: Optional[str] = None
    
    # Employment Information
    custom_employee_id: Optional[str] = None
    date_of_joining: Optional[str] = None
    employment_type: Optional[str] = None
    employee_status: Optional[str] = None
    designation: Optional[str] = None
    tier_level: Optional[str] = None
    reporting_manager_id: Optional[str] = None
    
    # Organization Structure
    department: Optional[str] = None
    team: Optional[str] = None
    work_location: Optional[str] = None
    
    # HR Configuration
    leave_policy: Optional[str] = None
    shift_type: Optional[str] = None
    attendance_tracking_enabled: Optional[bool] = None
    
    # Custom Shift Configuration
    custom_login_time: Optional[str] = None
    custom_logout_time: Optional[str] = None
    custom_total_hours: Optional[float] = None
    
    # Salary Configuration
    monthly_salary: Optional[float] = None
    
    # System Access
    user_role: Optional[str] = None
    login_enabled: Optional[bool] = None
    biometric_id: Optional[str] = None

class Attendance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    emp_name: str
    team: str
    department: str
    date: str
    check_in: Optional[str] = None  # Time in "HH:MM AM/PM" format
    check_out: Optional[str] = None  # Time in "HH:MM AM/PM" format
    check_in_24h: Optional[str] = None  # Time in "HH:MM" 24-hour format for calculations
    check_out_24h: Optional[str] = None  # Time in "HH:MM" 24-hour format for calculations
    total_hours: Optional[str] = None
    total_hours_decimal: Optional[float] = None  # For precise calculations
    status: str = "Not Logged"
    is_lop: bool = False  # Loss of Pay flag
    lop_reason: Optional[str] = None  # Reason for LOP
    shift_type: Optional[str] = "General"  # Employee's shift at time of attendance
    expected_login: Optional[str] = None  # Expected login time based on shift
    expected_logout: Optional[str] = None  # Expected logout time based on shift
    created_at: datetime = Field(default_factory=lambda: get_ist_now())

class LeaveRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    emp_name: str
    team: str
    department: str
    leave_type: str
    leave_split: str = "Full Day"  # Full Day, First Half, Second Half
    start_date: str
    end_date: str
    duration: str
    reason: Optional[str] = None
    supporting_document_url: Optional[str] = None
    supporting_document_name: Optional[str] = None
    status: str = "pending"
    is_lop: Optional[bool] = None  # Set by admin on approval: True=LOP, False=No LOP
    lop_remark: Optional[str] = None
    approved_by: Optional[str] = None
    applied_by_admin: Optional[bool] = False
    # Captures any non-standard columns from bulk imports (extended attributes).
    # Additive — does not affect existing leave flows.
    extra_data: Optional[dict] = None
    created_at: datetime = Field(default_factory=lambda: get_ist_now())

class LeaveRequestCreate(BaseModel):
    employee_id: str
    leave_type: str
    leave_split: str = "Full Day"
    start_date: str
    end_date: str
    reason: Optional[str] = None
    supporting_document_url: Optional[str] = None
    supporting_document_name: Optional[str] = None
    is_lop: Optional[bool] = None
    auto_approve: Optional[bool] = False

# ============== LATE REQUEST / EARLY OUT / MISSED PUNCH MODELS ==============

class LateRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    emp_name: str
    team: str
    department: str
    date: str  # YYYY-MM-DD
    expected_time: Optional[str] = None
    actual_time: Optional[str] = None
    reason: str
    status: str = "pending"  # pending, approved, rejected
    is_lop: Optional[bool] = None
    lop_remark: Optional[str] = None
    approved_by: Optional[str] = None
    applied_by_admin: Optional[bool] = False
    created_at: datetime = Field(default_factory=lambda: get_ist_now())

class LateRequestCreate(BaseModel):
    employee_id: Optional[str] = None  # For admin applying on behalf
    date: str
    expected_time: Optional[str] = None
    actual_time: Optional[str] = None
    reason: str
    is_lop: Optional[bool] = None
    auto_approve: Optional[bool] = False

class EarlyOutRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    emp_name: str
    team: str
    department: str
    date: str
    expected_time: Optional[str] = None
    actual_time: Optional[str] = None
    reason: str
    status: str = "pending"
    is_lop: Optional[bool] = None
    lop_remark: Optional[str] = None
    approved_by: Optional[str] = None
    applied_by_admin: Optional[bool] = False
    created_at: datetime = Field(default_factory=lambda: get_ist_now())

class EarlyOutRequestCreate(BaseModel):
    employee_id: Optional[str] = None
    date: str
    expected_time: Optional[str] = None
    actual_time: Optional[str] = None
    reason: str
    is_lop: Optional[bool] = None
    auto_approve: Optional[bool] = False

class MissedPunchRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    emp_name: str
    team: str
    department: str
    date: str
    punch_type: str  # Check-in, Check-out, Both
    check_in_time: Optional[str] = None
    check_out_time: Optional[str] = None
    reason: str
    status: str = "pending"
    approved_by: Optional[str] = None
    applied_by_admin: Optional[bool] = False
    created_at: datetime = Field(default_factory=lambda: get_ist_now())

class MissedPunchCreate(BaseModel):
    employee_id: Optional[str] = None
    date: str
    punch_type: str  # Check-in, Check-out, Both
    check_in_time: Optional[str] = None
    check_out_time: Optional[str] = None
    reason: str
    auto_approve: Optional[bool] = False

class RequestApproveBody(BaseModel):
    is_lop: Optional[bool] = None
    lop_remark: Optional[str] = None
    model_config = ConfigDict(extra="ignore")

class StarRewardCreate(BaseModel):
    employee_id: str
    stars: int
    reason: str
    type: str = "performance"

class StarReward(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    stars: int
    reason: str
    type: str = "performance"  # performance, learning, innovation, unsafe
    awarded_by: str
    month: str
    created_at: datetime = Field(default_factory=lambda: get_ist_now())

class Team(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    department: str
    lead_id: Optional[str] = None
    member_count: int = 0
    created_at: datetime = Field(default_factory=lambda: get_ist_now())

class Department(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    head_id: Optional[str] = None
    team_count: int = 0

class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    action: str
    resource: str
    resource_id: Optional[str] = None
    details: Optional[str] = None
    timestamp: datetime = Field(default_factory=lambda: get_ist_now())

# ============== NOTIFICATION MODEL ==============

class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str  # recipient user id
    title: str
    message: str
    type: str = "info"  # info, warning, success, action
    link: Optional[str] = None  # frontend route to navigate to
    read: bool = False
    created_at: datetime = Field(default_factory=lambda: get_ist_now())

# ============== OPERATIONAL CHECKLIST MODEL ==============

OPERATIONAL_CHECKLIST_ITEMS = [
    {"key": "workstation_setup", "label": "Workstation Ready (desk, chair, system/laptop)", "category": "Infrastructure"},
    {"key": "stationery_issued", "label": "Stationery & Supplies Issued (notebooks, pens)", "category": "Stationery"},
    {"key": "id_card_issued", "label": "ID Card Issued", "category": "Stationery"},
    {"key": "attendance_configured", "label": "Attendance Device Configured (biometric/login)", "category": "Access"},
    {"key": "access_card_setup", "label": "Access Card / Entry Permissions Issued", "category": "Access"},
    {"key": "system_access_verified", "label": "System Access Created & Verified", "category": "IT"},
    {"key": "role_access_confirmed", "label": "Role-based Access Confirmed", "category": "IT"},
    {"key": "hr_coordination_complete", "label": "HR Coordination for Joining Readiness", "category": "Coordination"},
]

class ChecklistItem(BaseModel):
    key: str
    label: str
    category: str
    completed: bool = False
    completed_by: Optional[str] = None
    completed_at: Optional[str] = None
    notes: Optional[str] = None

class OperationalChecklist(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    emp_name: str
    department: Optional[str] = None
    designation: Optional[str] = None
    items: List[dict] = []
    status: str = "pending"  # pending, in_progress, completed
    created_at: datetime = Field(default_factory=lambda: get_ist_now())
    updated_at: datetime = Field(default_factory=lambda: get_ist_now())

# ============== PAYROLL MODEL ==============

class PayrollRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    emp_name: str
    department: str
    team: str
    month: str  # Format: "YYYY-MM"
    monthly_salary: float = 0.0
    total_days: int = 0  # Calendar days in month
    working_days: int = 0  # Non-Sunday, non-Holiday days
    weekoff_pay: float = 0.0  # +1 per Sunday/Holiday
    extra_pay: float = 0.0  # +1/+0.5 for Sunday/Holiday worked
    lop: float = 0.0  # Total LOP days (supports 0.5)
    final_payable_days: float = 0.0  # (Working Days - LOP) + Weekoff + Extra
    present_days: int = 0  # Backward compat
    lop_days: float = 0  # Backward compat (= lop)
    leave_days: int = 0
    absent_days: int = 0
    per_day_salary: float = 0.0
    lop_deduction: float = 0.0
    net_salary: float = 0.0
    attendance_details: List[dict] = []  # Daily attendance breakdown
    created_at: datetime = Field(default_factory=lambda: get_ist_now())
    updated_at: datetime = Field(default_factory=lambda: get_ist_now())

class ShiftConfigCreate(BaseModel):
    shift_type: str
    login_time: Optional[str] = None  # For custom shifts
    logout_time: Optional[str] = None  # For custom shifts

# ============== ONBOARDING MODELS ==============

class OnboardingDocument(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    document_type: str
    document_label: str
    file_url: Optional[str] = None
    file_public_id: Optional[str] = None
    file_name: Optional[str] = None
    status: str = DocumentStatus.NOT_UPLOADED
    uploaded_at: Optional[datetime] = None
    verified_at: Optional[datetime] = None
    verified_by: Optional[str] = None
    rejection_reason: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: get_ist_now())

class OnboardingRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    emp_id: str
    emp_name: str
    department: str
    team: str
    designation: str
    status: str = OnboardingStatus.PENDING
    documents: List[dict] = []  # List of document statuses
    submitted_at: Optional[datetime] = None
    reviewed_at: Optional[datetime] = None
    reviewed_by: Optional[str] = None
    review_notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: get_ist_now())
    updated_at: datetime = Field(default_factory=lambda: get_ist_now())

class DocumentUpload(BaseModel):
    document_type: str
    file_url: str
    file_public_id: Optional[str] = None
    file_name: Optional[str] = None

class DocumentVerification(BaseModel):
    document_id: str
    status: str  # "verified" or "rejected"
    rejection_reason: Optional[str] = None

class OnboardingApproval(BaseModel):
    status: str  # "approved" or "rejected"
    review_notes: Optional[str] = None

# ============== ISSUE TICKET SYSTEM ENUMS & MODELS ==============

class TicketCategory:
    IT_SYSTEM = "IT & System Support"
    HR_SUPPORT = "HR Support"
    FINANCE = "Finance & Accounts"
    ADMIN_STATIONERY = "Admin & Stationery"
    COMPLIANCE = "Compliance & Legal"
    OPERATIONS = "Operations"

class TicketStatus:
    OPEN = "Open"
    IN_PROGRESS = "In Progress"
    WAITING_APPROVAL = "Waiting for Approval"
    ON_HOLD = "On Hold"
    RESOLVED = "Resolved"
    CLOSED = "Closed"
    REJECTED = "Rejected"

class TicketPriority:
    HIGH = "High"
    MEDIUM = "Medium"
    LOW = "Low"

# Department role mapping for ticket assignment
TICKET_DEPARTMENT_ROLES = {
    TicketCategory.IT_SYSTEM: "it_admin",
    TicketCategory.HR_SUPPORT: "hr_admin",
    TicketCategory.FINANCE: "finance_admin",
    TicketCategory.ADMIN_STATIONERY: "admin_dept",
    TicketCategory.COMPLIANCE: "compliance_officer",
    TicketCategory.OPERATIONS: "operations_manager"
}

# Subcategories for each ticket category
TICKET_SUBCATEGORIES = {
    TicketCategory.IT_SYSTEM: [
        "Login Issue", "Password Reset", "Email Not Working", "HRMS Access Issue",
        "Software Installation", "System Slow / Hanging", "Printer Not Working",
        "Network / Internet Issue", "Biometric Attendance Issue", "Other IT Issue"
    ],
    TicketCategory.HR_SUPPORT: [
        "Salary Not Credited", "Payslip Request", "Leave Balance Issue",
        "Leave Approval Delay", "Attendance Correction", "Joining Letter / Experience Letter",
        "Promotion / Designation Change", "Policy Clarification", "Grievance Complaint", "Other HR Issue"
    ],
    TicketCategory.FINANCE: [
        "Expense Reimbursement", "Claim Status", "Salary Calculation Issue",
        "Bonus / Incentive Issue", "Tax / Form 16 Request", "Bank Detail Update", "Other Finance Issue"
    ],
    TicketCategory.ADMIN_STATIONERY: [
        "Stationery Request", "ID Card Issue", "Access Card Issue", "Cabin / Seat Allocation",
        "AC / Electricity Issue", "Office Cleanliness", "Furniture Request",
        "Asset Allocation (Laptop, Mouse, Keyboard)", "Other Admin Issue"
    ],
    TicketCategory.COMPLIANCE: [
        "Code of Conduct Violation", "Harassment Complaint", "Ethics Complaint",
        "Data Privacy Issue", "Policy Violation Report", "Workplace Safety Concern", "Other Compliance Issue"
    ],
    TicketCategory.OPERATIONS: [
        "Shift Change Request", "Work Location Change", "Project Allocation Issue",
        "Client Escalation", "Workload Issue", "Other Operations Issue"
    ]
}

class TicketStatusUpdate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    status: str
    updated_by: str
    updated_by_name: str
    notes: Optional[str] = None
    updated_at: datetime = Field(default_factory=lambda: get_ist_now())

class TicketAttachment(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    file_url: str
    file_name: str
    file_type: str
    file_public_id: Optional[str] = None
    uploaded_at: datetime = Field(default_factory=lambda: get_ist_now())

class TicketFeedback(BaseModel):
    rating: int  # 1-5
    comment: Optional[str] = None
    submitted_at: datetime = Field(default_factory=lambda: get_ist_now())

class TicketCreate(BaseModel):
    category: str
    subcategory: str
    subject: str
    description: str
    priority: str = TicketPriority.MEDIUM
    attachments: Optional[List[dict]] = None  # List of {file_url, file_name, file_type, file_public_id}
    employee_id: Optional[str] = None  # For admin creating on behalf of employee

class Ticket(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ticket_number: str  # Auto-generated: TKT-YYYYMMDD-XXX
    employee_id: str
    emp_name: str
    emp_email: Optional[str] = None
    department: str
    team: Optional[str] = None
    
    # Ticket details
    category: str
    subcategory: str
    subject: str
    description: str
    priority: str = TicketPriority.MEDIUM
    status: str = TicketStatus.OPEN
    
    # Assignment
    assigned_department: str  # Department role responsible
    assigned_to: Optional[str] = None  # User ID of assigned admin
    assigned_to_name: Optional[str] = None
    
    # Resolution
    resolution: Optional[str] = None
    resolved_at: Optional[datetime] = None
    resolved_by: Optional[str] = None
    resolved_by_name: Optional[str] = None
    
    # Attachments
    attachments: List[dict] = []  # List of TicketAttachment
    
    # Status history
    status_history: List[dict] = []  # List of TicketStatusUpdate
    
    # Feedback
    feedback: Optional[dict] = None  # TicketFeedback
    
    # Metadata
    created_by: Optional[str] = None  # If created by admin on behalf
    created_by_name: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: get_ist_now())
    updated_at: datetime = Field(default_factory=lambda: get_ist_now())

class TicketStatusUpdateRequest(BaseModel):
    status: str
    notes: Optional[str] = None
    resolution: Optional[str] = None

class TicketFeedbackRequest(BaseModel):
    rating: int
    comment: Optional[str] = None

class TicketAssignRequest(BaseModel):
    assigned_to: str  # User ID

# ============== SALARY MANAGEMENT MODELS ==============

class SalaryComponentType:
    EARNING = "earning"
    DEDUCTION = "deduction"

class AdjustmentType:
    BONUS = "bonus"
    INCENTIVE = "incentive"
    REIMBURSEMENT = "reimbursement"
    DEDUCTION = "deduction"
    LOP = "lop"
    ADVANCE_RECOVERY = "advance_recovery"
    PENALTY = "penalty"
    OTHER = "other"

class AdjustmentFrequency:
    ONE_TIME = "one_time"
    RECURRING = "recurring"

# Default salary structure percentages (of Basic)
DEFAULT_SALARY_STRUCTURE = {
    "basic_percentage": 40,  # 40% of CTC
    "hra_percentage": 50,    # 50% of Basic
    "da_percentage": 10,     # 10% of Basic
    "conveyance": 1600,      # Fixed amount
    "medical_allowance": 1250,  # Fixed amount
    "special_allowance_percentage": 0,  # Remaining after other components
    "pf_percentage": 12,     # 12% of Basic (employee contribution)
    "esi_percentage": 0.75,  # 0.75% of Gross (if applicable, gross < 21000)
    "professional_tax": 200, # Fixed (varies by state)
    "tds_percentage": 0      # Calculated based on tax slab
}

class SalaryStructure(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    
    # CTC and Basic
    annual_ctc: float = 0
    monthly_ctc: float = 0
    
    # Earnings
    basic: float = 0
    hra: float = 0
    da: float = 0
    conveyance: float = 0
    medical_allowance: float = 0
    special_allowance: float = 0
    other_allowances: float = 0
    
    # Gross
    gross_salary: float = 0
    
    # Deductions
    pf_employee: float = 0      # Employee PF contribution
    pf_employer: float = 0      # Employer PF contribution
    esi_employee: float = 0     # Employee ESI
    esi_employer: float = 0     # Employer ESI
    professional_tax: float = 0
    tds: float = 0
    other_deductions: float = 0
    
    # Net
    total_deductions: float = 0
    net_salary: float = 0
    
    # Metadata
    effective_from: str = ""
    created_at: datetime = Field(default_factory=lambda: get_ist_now())
    updated_at: datetime = Field(default_factory=lambda: get_ist_now())

class SalaryAdjustment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    
    # Adjustment details
    adjustment_type: str  # bonus, incentive, reimbursement, deduction, lop, advance_recovery, penalty
    category: str  # earning or deduction
    description: str
    amount: float
    
    # Frequency
    frequency: str = AdjustmentFrequency.ONE_TIME  # one_time or recurring
    
    # For one-time: specific month
    applicable_month: Optional[str] = None  # Format: "YYYY-MM"
    
    # For recurring: start and end
    start_month: Optional[str] = None
    end_month: Optional[str] = None  # None means ongoing
    is_active: bool = True
    
    # Audit
    created_by: str
    created_by_name: str
    approved_by: Optional[str] = None
    approved_by_name: Optional[str] = None
    
    created_at: datetime = Field(default_factory=lambda: get_ist_now())
    updated_at: datetime = Field(default_factory=lambda: get_ist_now())

class SalaryAdjustmentCreate(BaseModel):
    adjustment_type: str
    description: str
    amount: float
    frequency: str = AdjustmentFrequency.ONE_TIME
    applicable_month: Optional[str] = None
    start_month: Optional[str] = None
    end_month: Optional[str] = None

class SalaryStructureUpdate(BaseModel):
    annual_ctc: Optional[float] = None
    basic: Optional[float] = None
    hra: Optional[float] = None
    da: Optional[float] = None
    conveyance: Optional[float] = None
    medical_allowance: Optional[float] = None
    special_allowance: Optional[float] = None
    other_allowances: Optional[float] = None
    pf_employee: Optional[float] = None
    esi_employee: Optional[float] = None
    professional_tax: Optional[float] = None
    tds: Optional[float] = None
    other_deductions: Optional[float] = None

# ============== HELPERS ==============

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    return hash_password(password) == hashed

def create_token(user_id: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def log_audit(user_id: str, action: str, resource: str, resource_id: str = None, details: str = None):
    log = AuditLog(user_id=user_id, action=action, resource=resource, resource_id=resource_id, details=details)
    doc = log.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.audit_logs.insert_one(doc.copy())

async def create_notification(user_ids: list, title: str, message: str, notif_type: str = "info", link: str = None):
    """Create notifications for multiple users"""
    for uid in user_ids:
        notif = Notification(user_id=uid, title=title, message=message, type=notif_type, link=link)
        doc = notif.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.notifications.insert_one(doc.copy())

async def notify_role(role: str, title: str, message: str, notif_type: str = "info", link: str = None):
    """Send notification to all users with a specific role"""
    users = await db.users.find({"role": role, "is_active": True}, {"_id": 0, "id": 1}).to_list(1000)
    user_ids = [u["id"] for u in users]
    if user_ids:
        await create_notification(user_ids, title, message, notif_type, link)

async def notify_roles(roles: list, title: str, message: str, notif_type: str = "info", link: str = None):
    """Send notification to all users with any of the specified roles"""
    users = await db.users.find({"role": {"$in": roles}, "is_active": True}, {"_id": 0, "id": 1}).to_list(1000)
    user_ids = [u["id"] for u in users]
    if user_ids:
        await create_notification(user_ids, title, message, notif_type, link)

def serialize_doc(doc: dict) -> dict:
    if not doc:
        return doc
    if 'created_at' in doc and doc['created_at'] is not None and not isinstance(doc['created_at'], str):
        doc['created_at'] = doc['created_at'].isoformat()
    if 'updated_at' in doc and not isinstance(doc['updated_at'], str):
        doc['updated_at'] = doc['updated_at'].isoformat()
    if 'deleted_at' in doc and doc['deleted_at'] and not isinstance(doc['deleted_at'], str):
        doc['deleted_at'] = doc['deleted_at'].isoformat()
    if 'timestamp' in doc and not isinstance(doc['timestamp'], str):
        doc['timestamp'] = doc['timestamp'].isoformat()
    return doc

async def generate_emp_id():
    """Generate unique employee ID"""
    count = await db.employees.count_documents({})
    return f"EMP{str(count + 1).zfill(4)}"

async def generate_ticket_number():
    """Generate unique ticket number: TKT-YYYYMMDD-XXX"""
    today = get_ist_now().strftime("%Y%m%d")
    # Count tickets created today
    start_of_day = get_ist_now().replace(hour=0, minute=0, second=0, microsecond=0)
    count = await db.issue_tickets.count_documents({
        "created_at": {"$gte": start_of_day.isoformat()}
    })
    return f"TKT-{today}-{str(count + 1).zfill(3)}"

# ============== SHIFT & LOP CALCULATION HELPERS ==============

def parse_time_12h_to_24h(time_str: str) -> str:
    """Convert 12-hour format (e.g., '10:00 AM') to 24-hour format (e.g., '10:00')"""
    if not time_str:
        return None
    try:
        time_str = time_str.strip().upper()
        if 'AM' in time_str or 'PM' in time_str:
            dt = datetime.strptime(time_str.replace(' ', ''), "%I:%M%p")
            return dt.strftime("%H:%M")
        return time_str  # Already in 24h format
    except:
        return None

def parse_time_24h_to_minutes(time_str: str) -> int:
    """Convert 24-hour time string to minutes since midnight"""
    if not time_str:
        return None
    try:
        hours, minutes = map(int, time_str.split(':'))
        return hours * 60 + minutes
    except:
        return None

def minutes_to_time_24h(minutes: int) -> str:
    """Convert minutes since midnight to 24-hour time string"""
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours:02d}:{mins:02d}"

def get_shift_timings(employee: dict) -> dict:
    """Get shift timings for an employee based on their shift type"""
    shift_type = employee.get('shift_type', 'General')
    # Grace minutes (defaults 0) come from the Settings module via sync
    late_grace = int(employee.get('late_grace_minutes', 0) or 0)
    early_grace = int(employee.get('early_out_grace_minutes', 0) or 0)
    
    if shift_type == "Custom":
        login_time = employee.get('custom_login_time')
        logout_time = employee.get('custom_logout_time')
        
        if login_time and logout_time:
            login_mins = parse_time_24h_to_minutes(login_time)
            logout_mins = parse_time_24h_to_minutes(logout_time)
            
            # Handle overnight shifts
            if logout_mins < login_mins:
                total_hours = (24 * 60 - login_mins + logout_mins) / 60
            else:
                total_hours = (logout_mins - login_mins) / 60
            
            return {
                "login_time": login_time,
                "logout_time": logout_time,
                "total_hours": total_hours,
                "late_grace_minutes": late_grace,
                "early_out_grace_minutes": early_grace,
            }
        return None
    
    shift_def = SHIFT_DEFINITIONS.get(shift_type, SHIFT_DEFINITIONS["General"])
    return {
        "login_time": shift_def["login_time"],
        "logout_time": shift_def["logout_time"],
        "total_hours": shift_def["total_hours"],
        "late_grace_minutes": late_grace,
        "early_out_grace_minutes": early_grace,
    }


async def _apply_settings_shift_to_employee_payload(payload: dict, shift_identifier: Optional[str]) -> dict:
    """If `shift_identifier` matches a shift configured in Settings (by id OR
    by name), expand the payload with all the legacy fields the existing
    attendance engine expects (`shift_type='Custom'`, `custom_login_time`,
    `custom_logout_time`, `custom_total_hours`, grace minutes, `active_shift_*`).

    The Add-Employee form sends the picked shift in `shift_type`; this helper
    transparently turns that name into a fully-resolved Settings-shift binding
    on the employee record. If the identifier doesn't match any Settings
    shift, the payload is returned unchanged (legacy behaviour preserved).
    """
    if not shift_identifier or shift_identifier in ("Custom",):
        if shift_identifier == "Custom":
            # Plain "Custom" means the user is explicitly opting out of central
            # shift config — clear the binding so reports / dashboards don't
            # mis-attribute the previous Settings shift to this employee.
            payload["active_shift_id"] = None
            payload["active_shift_name"] = None
        return payload

    # Try id first, then name. Honour soft-delete + active status from settings.
    shift = await db.shifts.find_one(
        {"$or": [{"id": shift_identifier}, {"name": shift_identifier}],
         "is_deleted": {"$ne": True}},
        {"_id": 0},
    )
    if not shift:
        return payload

    login = shift["start_time"]
    total = float(shift["total_hours"])
    total_mins = int(round(total * 60))
    login_mins = parse_time_24h_to_minutes(login) or 0
    logout_mins = (login_mins + total_mins) % (24 * 60)
    logout = f"{logout_mins // 60:02d}:{logout_mins % 60:02d}"

    payload.update({
        "shift_type": "Custom",
        "custom_login_time": login,
        "custom_logout_time": logout,
        "custom_total_hours": total,
        "late_grace_minutes": int(shift.get("late_grace_minutes", 0) or 0),
        "early_out_grace_minutes": int(shift.get("early_out_grace_minutes", 0) or 0),
        "active_shift_id": shift["id"],
        "active_shift_name": shift["name"],
    })
    return payload

def calculate_attendance_status(check_in_24h: str, check_out_24h: str, shift_timings: dict, attendance_date: Optional[str] = None) -> dict:
    """
    Calculate attendance status with effective-start (early-arrival) handling:
      * effective_start_time = actual punch-in (always)
      * expected_logout = effective_start_time + required_hours (dynamic)
      * Late entry still applies LOP based on existing late_grace_minutes
      * Early-out / short-hours uses TOTAL WORKED HOURS vs required_hours;
        an employee who completes >= required_hours is NEVER flagged
        early-out, regardless of clock time.
      * Sundays: status = "Sunday" (no LOP). If they punch on Sunday →
        "Worked on Sunday" (no LOP).
    """
    result = {
        "status": AttendanceStatus.PRESENT,
        "is_lop": False,
        "lop_reason": None,
        "total_hours_decimal": 0.0,
        "expected_logout": None,
    }

    # Sunday guard: never flag late/early/LOP/absent on Sundays.
    if attendance_date and is_sunday_ddmmyyyy(attendance_date):
        result["expected_logout"] = None
        if check_in_24h:
            result["status"] = AttendanceStatus.WORKED_ON_SUNDAY
            if check_out_24h:
                in_mins = parse_time_24h_to_minutes(check_in_24h)
                out_mins = parse_time_24h_to_minutes(check_out_24h)
                total_mins = (out_mins - in_mins) if out_mins >= in_mins else (24 * 60 - in_mins + out_mins)
                result["total_hours_decimal"] = total_mins / 60
        else:
            result["status"] = AttendanceStatus.SUNDAY
        return result

    required_hours = shift_timings.get("total_hours", 8) if shift_timings else 8

    # Flexible shift - only check total hours (honors early_out grace)
    if not shift_timings or shift_timings.get("login_time") is None:
        if check_in_24h:
            result["expected_logout"] = add_hours_to_24h(check_in_24h, required_hours)
        if check_in_24h and check_out_24h:
            in_mins = parse_time_24h_to_minutes(check_in_24h)
            out_mins = parse_time_24h_to_minutes(check_out_24h)
            
            if out_mins < in_mins:
                total_mins = 24 * 60 - in_mins + out_mins
            else:
                total_mins = out_mins - in_mins
            
            result["total_hours_decimal"] = total_mins / 60
            early_grace = int((shift_timings or {}).get("early_out_grace_minutes", 0) or 0)
            required_worked = required_hours - (early_grace / 60.0)
            
            if result["total_hours_decimal"] < required_worked:
                result["status"] = AttendanceStatus.LOSS_OF_PAY
                result["is_lop"] = True
                result["lop_reason"] = f"Insufficient hours: {result['total_hours_decimal']:.2f}h < {required_worked}h required"
        return result
    
    # Fixed shift - strict rules apply
    expected_login = parse_time_24h_to_minutes(shift_timings["login_time"])
    late_grace = int(shift_timings.get("late_grace_minutes", 0) or 0)
    early_grace = int(shift_timings.get("early_out_grace_minutes", 0) or 0)
    
    if not check_in_24h:
        result["status"] = AttendanceStatus.NOT_LOGGED
        return result
    
    actual_login = parse_time_24h_to_minutes(check_in_24h)
    # Dynamic expected logout: punch-in + required_hours (always)
    result["expected_logout"] = add_hours_to_24h(check_in_24h, required_hours)
    
    # Check for late login (respects late_grace_minutes; 0 means any minute past start is LATE)
    is_late = actual_login > expected_login + late_grace
    
    # Compute total hours if checked out
    total_hours_decimal = 0.0
    if check_out_24h:
        actual_logout = parse_time_24h_to_minutes(check_out_24h)
        if actual_logout < actual_login:
            total_mins = 24 * 60 - actual_login + actual_logout
        else:
            total_mins = actual_logout - actual_login
        total_hours_decimal = total_mins / 60
        result["total_hours_decimal"] = total_hours_decimal

    if is_late:
        late_mins = actual_login - expected_login
        result["status"] = AttendanceStatus.LOSS_OF_PAY
        result["is_lop"] = True
        result["lop_reason"] = f"Late login by {late_mins} minute(s). Expected: {shift_timings['login_time']}, Actual: {check_in_24h}"
        return result
    
    # Not late — check early logout / short hours using TOTAL WORKED vs REQUIRED.
    # Employees who complete >= required hours are NEVER flagged as early-out,
    # regardless of clock time (early-arrival is preserved).
    if check_out_24h:
        required_worked = required_hours - (early_grace / 60.0)
        if total_hours_decimal < required_worked:
            short_mins = int(round((required_worked - total_hours_decimal) * 60))
            result["status"] = AttendanceStatus.LOSS_OF_PAY
            result["is_lop"] = True
            result["lop_reason"] = f"Early out / short hours by {short_mins} minute(s). Worked: {total_hours_decimal:.2f}h, Required: {required_worked:.2f}h"
            return result
        result["status"] = AttendanceStatus.PRESENT
    else:
        # Only logged in, not logged out yet
        result["status"] = AttendanceStatus.LOGIN
    
    return result

def calculate_total_hours_str(total_hours_decimal: float) -> str:
    """Convert decimal hours to string format (e.g., '8h 30m')"""
    if not total_hours_decimal:
        return "-"
    hours = int(total_hours_decimal)
    minutes = int((total_hours_decimal - hours) * 60)
    return f"{hours}h {minutes}m"

# ============== DEPARTMENT WORK HOURS (PAYROLL ENGINE) ==============

DEPARTMENT_WORK_HOURS = {
    "Research Unit": {"full": 11, "half": 6},
    "Business & Product": {"full": 10, "half": 5},
    "Support Staff": {"full": 9, "half": 4.5},
}

def _parse_date_flex(date_str):
    """Parse date string from multiple formats to a date object."""
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(date_str).strip(), fmt).date()
        except ValueError:
            continue
    return None

def _calc_hours_worked(att_record: dict) -> float:
    """Calculate hours worked from an attendance record."""
    hw = att_record.get("total_hours_decimal", 0) or 0
    if hw > 0:
        return hw
    ci = att_record.get("check_in_24h")
    co = att_record.get("check_out_24h")
    if ci and co:
        in_mins = parse_time_24h_to_minutes(ci)
        out_mins = parse_time_24h_to_minutes(co)
        if in_mins is not None and out_mins is not None:
            diff = out_mins - in_mins if out_mins > in_mins else (1440 - in_mins + out_mins)
            return diff / 60
    return 0


def _leave_code_for_status(leave_type: str, leave_split: str) -> str:
    """Map a (leave_type, leave_split) pair to the user-facing payroll grid code.

    Used only for approved-non-LOP leave-day rows in the payroll Attendance View
    grid. The mapping (per HR instruction 2026-05-07) is:

        Pre-Planned   Full → PF | Half → PH
        Sick          Full → SF | Half → SH
        Emergency     Full → EF | Half → EH
        Optional      Any  → OH        (Optional Holiday — no half/full split)
        Casual / Earned / Annual / Maternity / Paternity / Bereavement /
        General Leave Full → PA | Half → PP   (Paid Leave bucket)
    """
    lt = (leave_type or "").strip().lower()
    is_half = (leave_split or "").strip().lower() in (
        "first half", "second half", "half day", "half"
    )
    if lt.startswith("pre-planned") or lt.startswith("preplanned"):
        return "PH" if is_half else "PF"
    if lt.startswith("sick"):
        return "SH" if is_half else "SF"
    if lt.startswith("emergency"):
        return "EH" if is_half else "EF"
    if lt.startswith("optional"):
        # Optional Holiday — single code regardless of split
        return "OH"
    # Everything else → Paid Leave bucket
    return "PP" if is_half else "PA"

async def _prefetch_payroll_data(employee_ids: list, year: int, month_num: int, days_in_month: int) -> dict:
    """Batch-prefetch all payroll-related data for a set of employees in one month."""
    # 1. Attendance (both date formats)
    all_att = await db.attendance.find({
        "employee_id": {"$in": employee_ids},
        "$or": [
            {"date": {"$regex": f"^\\d{{2}}-{month_num:02d}-{year}$"}},
            {"date": {"$regex": f"^{year}-{month_num:02d}-"}}
        ]
    }, {"_id": 0}).to_list(len(employee_ids) * 35)
    att_by_emp = {}
    for r in all_att:
        att_by_emp.setdefault(r["employee_id"], []).append(r)

    # 2. Leaves
    all_leaves = await db.leaves.find({
        "employee_id": {"$in": employee_ids},
        "$or": [
            {"start_date": {"$regex": f"^{year}-{month_num:02d}"}},
            {"end_date": {"$regex": f"^{year}-{month_num:02d}"}}
        ]
    }, {"_id": 0}).to_list(len(employee_ids) * 10)
    leaves_by_emp = {}
    for lv in all_leaves:
        leaves_by_emp.setdefault(lv["employee_id"], []).append(lv)

    # 3. Approved late requests
    all_late = await db.late_requests.find({
        "employee_id": {"$in": employee_ids},
        "status": "approved",
        "date": {"$regex": f"^{year}-{month_num:02d}"}
    }, {"_id": 0}).to_list(len(employee_ids) * 10)
    late_by_emp = {}
    for lr in all_late:
        late_by_emp.setdefault(lr["employee_id"], []).append(lr)

    # 4. Missed punches
    all_mp = await db.missed_punches.find({
        "employee_id": {"$in": employee_ids},
        "date": {"$regex": f"^{year}-{month_num:02d}"}
    }, {"_id": 0}).to_list(len(employee_ids) * 10)
    mp_by_emp = {}
    for mp in all_mp:
        mp_by_emp.setdefault(mp["employee_id"], []).append(mp)

    # 5. Holidays
    holiday_records = await db.holidays.find({
        "date": {
            "$gte": f"{year}-{month_num:02d}-01",
            "$lte": f"{year}-{month_num:02d}-{days_in_month:02d}"
        }
    }, {"_id": 0}).to_list(50)
    holiday_dates = set()
    for h in holiday_records:
        hd = _parse_date_flex(h.get("date"))
        if hd:
            holiday_dates.add(hd)

    return {
        "att_by_emp": att_by_emp,
        "leaves_by_emp": leaves_by_emp,
        "late_by_emp": late_by_emp,
        "mp_by_emp": mp_by_emp,
        "holiday_dates": holiday_dates,
    }

async def calculate_payroll_for_employee(employee_id: str, month: str, employee: dict = None, prefetched: dict = None) -> dict:
    """
    Payroll engine aligned to strict specification.
    - Department-based work hour mapping (Research 11h, Business 10h, Support 9h)
    - Sunday/Holiday: Weekoff Pay +1; if worked Full→PF Extra+1, Half→PH Extra+0.5
    - Working day classification cross-references leaves, late requests, missed punches
    - LOP: A=1, LC=0.5, PH(working day)=0.5, pending/LOP leave
    - Payable Days = (Working Days - LOP) + Weekoff Pay + Extra Pay
    - Relieved employee: if last_day_payable=0 → subtract 1 day
    """
    import calendar
    from datetime import date, timedelta
    from datetime import timezone as tz

    if not employee:
        employee = await db.employees.find_one(
            {"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0}
        )
    if not employee:
        return None

    year, month_num = map(int, month.split('-'))
    days_in_month = calendar.monthrange(year, month_num)[1]

    # --- Parse employee dates ---
    joining_date = _parse_date_flex(employee.get("date_of_joining"))
    emp_status = employee.get("employee_status", "Active")
    relieving_date = None
    if emp_status == EmployeeStatus.INACTIVE:
        relieving_date = _parse_date_flex(employee.get("inactive_date"))
        # Inactive without a date → cannot calculate payroll
        if not relieving_date:
            return None

    payroll_start = date(year, month_num, 1)
    payroll_end = date(year, month_num, days_in_month)

    if joining_date and joining_date > payroll_end:
        return None
    if relieving_date and relieving_date < payroll_start:
        return None

    dept = employee.get("department", "")
    work_cfg = DEPARTMENT_WORK_HOURS.get(dept, {"full": 10, "half": 5})
    full_hours = work_cfg["full"]
    half_hours = work_cfg["half"]

    # --- Load data (prefetched or individual queries) ---
    if prefetched:
        holiday_dates = prefetched["holiday_dates"]
        att_records = prefetched["att_by_emp"].get(employee_id, [])
        leave_records = prefetched["leaves_by_emp"].get(employee_id, [])
        late_records = prefetched["late_by_emp"].get(employee_id, [])
        all_missed = prefetched["mp_by_emp"].get(employee_id, [])
    else:
        holiday_records_q = await db.holidays.find({
            "date": {"$gte": f"{year}-{month_num:02d}-01", "$lte": f"{year}-{month_num:02d}-{days_in_month:02d}"}
        }, {"_id": 0}).to_list(50)
        holiday_dates = set()
        for h in holiday_records_q:
            hd = _parse_date_flex(h.get("date"))
            if hd:
                holiday_dates.add(hd)
        att_records = await db.attendance.find({
            "employee_id": employee_id,
            "$or": [
                {"date": {"$regex": f"^\\d{{2}}-{month_num:02d}-{year}$"}},
                {"date": {"$regex": f"^{year}-{month_num:02d}-"}}
            ]
        }, {"_id": 0}).to_list(50)
        leave_records = await db.leaves.find({
            "employee_id": employee_id,
            "$or": [
                {"start_date": {"$regex": f"^{year}-{month_num:02d}"}},
                {"end_date": {"$regex": f"^{year}-{month_num:02d}"}}
            ]
        }, {"_id": 0}).to_list(100)
        late_records = await db.late_requests.find({
            "employee_id": employee_id, "status": "approved",
            "date": {"$regex": f"^{year}-{month_num:02d}"}
        }, {"_id": 0}).to_list(50)
        all_missed = await db.missed_punches.find({
            "employee_id": employee_id,
            "date": {"$regex": f"^{year}-{month_num:02d}"}
        }, {"_id": 0}).to_list(50)
    # --- Build attendance map with merging ---
    attendance_map = {}
    for r in att_records:
        d = r.get("date", "")
        if len(d) >= 10 and d[4] == '-':
            parts = d.split('-')
            normalized = f"{parts[2]}-{parts[1]}-{parts[0]}"
        else:
            normalized = d
        if normalized in attendance_map:
            existing = attendance_map[normalized]
            for fld in ("check_in", "check_in_24h", "check_out", "check_out_24h", "total_hours", "total_hours_decimal"):
                if r.get(fld) and not existing.get(fld):
                    existing[fld] = r[fld]
            # Recalculate hours if both timestamps now present
            ci_ = existing.get("check_in_24h")
            co_ = existing.get("check_out_24h")
            if ci_ and co_ and not existing.get("total_hours_decimal"):
                try:
                    i_m = parse_time_24h_to_minutes(ci_)
                    o_m = parse_time_24h_to_minutes(co_)
                    diff = o_m - i_m if o_m > i_m else (1440 - i_m + o_m)
                    existing["total_hours_decimal"] = round(diff / 60, 2)
                    existing["total_hours"] = f"{int(diff // 60)}h {int(diff % 60)}m"
                except Exception:
                    pass
        else:
            attendance_map[normalized] = r

    # --- Build leave map ---
    leave_map = {}
    for lv in leave_records:
        s = _parse_date_flex(lv.get("start_date"))
        e = _parse_date_flex(lv.get("end_date"))
        if s and e:
            cur = s
            while cur <= e:
                if payroll_start <= cur <= payroll_end:
                    leave_map[cur.strftime("%Y-%m-%d")] = lv
                cur += timedelta(days=1)

    late_approved_dates = {lr["date"] for lr in late_records}

    mp_approved_dates = set()
    mp_pending_dates = set()
    for mp in all_missed:
        md = _parse_date_flex(mp.get("date"))
        if md and payroll_start <= md <= payroll_end:
            iso = md.strftime("%Y-%m-%d")
            if mp.get("status") == "approved":
                mp_approved_dates.add(iso)
            elif mp.get("status") == "pending":
                mp_pending_dates.add(iso)

    # IST today
    ist = tz(timedelta(hours=5, minutes=30))
    today = datetime.now(ist).date()

    # --- Payroll accumulators ---
    total_days = days_in_month
    working_days = 0
    weekoff_pay = 0.0
    extra_pay = 0.0
    lop = 0.0
    attendance_details = []

    for day in range(1, days_in_month + 1):
        current_date = date(year, month_num, day)
        date_dd = f"{day:02d}-{month_num:02d}-{year}"
        date_iso = current_date.strftime("%Y-%m-%d")
        day_name = current_date.strftime("%a")
        is_sun = current_date.weekday() == 6
        is_hol = current_date in holiday_dates
        is_future = current_date > today

        detail = {
            "date": date_dd,
            "day_name": day_name,
            "is_sunday": is_sun,
            "is_holiday": is_hol,
            "status": "NA",
            "lop_value": 0,
            "weekoff_value": 0,
            "extra_value": 0,
            "check_in": None,
            "check_out": None,
            "total_hours": None,
            "is_lop": False,
        }

        # --- Before joining → BLANK ---
        if joining_date and current_date < joining_date:
            detail["status"] = "BLANK"
            attendance_details.append(detail)
            continue

        # --- After relieving → R ---
        if relieving_date and current_date > relieving_date:
            detail["status"] = "R"
            attendance_details.append(detail)
            continue

        # --- Future dates ---
        if is_future:
            if is_sun:
                detail["status"] = "Su"
                weekoff_pay += 1
                detail["weekoff_value"] = 1
            elif is_hol:
                detail["status"] = "H"
                # Holidays do NOT contribute to weekoff — only extra_pay if worked
            else:
                detail["status"] = "NA"
                working_days += 1
            attendance_details.append(detail)
            continue

        att = attendance_map.get(date_dd)
        leave = leave_map.get(date_iso)

        # ===== SECTION 5A: SUNDAY =====
        if is_sun:
            weekoff_pay += 1
            detail["weekoff_value"] = 1

            if att:
                detail["check_in"] = att.get("check_in")
                detail["check_out"] = att.get("check_out")
                detail["total_hours"] = att.get("total_hours")
                hw = _calc_hours_worked(att)
                if hw >= full_hours:
                    detail["status"] = "P"
                    extra_pay += 1
                    detail["extra_value"] = 1
                elif hw >= half_hours:
                    detail["status"] = "HD"
                    extra_pay += 0.5
                    detail["extra_value"] = 0.5
                else:
                    detail["status"] = "WO"
            else:
                detail["status"] = "WO"

            attendance_details.append(detail)
            continue

        # ===== SECTION 5B: HOLIDAY (non-Sunday) =====
        if is_hol:
            # Holiday: NO weekoff pay. Only extra_pay if employee worked.
            if att:
                detail["check_in"] = att.get("check_in")
                detail["check_out"] = att.get("check_out")
                detail["total_hours"] = att.get("total_hours")
                hw = _calc_hours_worked(att)
                if hw >= full_hours:
                    detail["status"] = "P"
                    extra_pay += 1
                    detail["extra_value"] = 1
                elif hw >= half_hours:
                    detail["status"] = "HD"
                    extra_pay += 0.5
                    detail["extra_value"] = 0.5
                else:
                    detail["status"] = "H"
            else:
                detail["status"] = "H"

            attendance_details.append(detail)
            continue

        # ===== WORKING DAY =====
        working_days += 1

        # --- SECTION 6A: LEAVE ONLY (no attendance) ---
        if leave and not att:
            ls = leave.get("status", "pending")
            is_lop_flag = leave.get("is_lop")
            split = leave.get("leave_split", "Full Day")

            if ls == "approved":
                if is_lop_flag is True:
                    if split == "Full Day":
                        detail["status"] = "LOP"
                        detail["is_lop"] = True
                        lop += 1
                        detail["lop_value"] = 1
                    else:
                        detail["status"] = "LOP"
                        detail["is_lop"] = True
                        lop += 0.5
                        detail["lop_value"] = 0.5
                else:
                    # Map leave-type + split to granular abbreviation.
                    detail["status"] = _leave_code_for_status(leave.get("leave_type"), split)
            elif ls == "pending":
                if split == "Full Day":
                    lop += 1
                    detail["lop_value"] = 1
                else:
                    lop += 0.5
                    detail["lop_value"] = 0.5
                detail["status"] = "LOP"
                detail["is_lop"] = True
            else:
                detail["status"] = "A"
                lop += 1
                detail["lop_value"] = 1

            attendance_details.append(detail)
            continue

        # --- SECTION 6B: LEAVE + ATTENDANCE ---
        if leave and att:
            detail["check_in"] = att.get("check_in")
            detail["check_out"] = att.get("check_out")
            detail["total_hours"] = att.get("total_hours")
            ls = leave.get("status", "pending")
            is_lop_flag = leave.get("is_lop")
            split = leave.get("leave_split", "Full Day")

            if split in ("First Half", "Second Half"):
                if ls == "approved" and is_lop_flag is not True:
                    # Half-day approved leave + worked the other half → present full
                    detail["status"] = "P"
                else:
                    # Half-day leave was rejected/pending/LOP → 0.5 LOP, show as HD
                    detail["status"] = "HD"
                    detail["is_lop"] = True
                    lop += 0.5
                    detail["lop_value"] = 0.5
            else:
                detail["status"] = "P"

            attendance_details.append(detail)
            continue

        # --- SECTION 6C: ATTENDANCE ONLY (no leave) ---
        if att:
            detail["check_in"] = att.get("check_in")
            detail["check_out"] = att.get("check_out")
            detail["total_hours"] = att.get("total_hours")
            ci24 = att.get("check_in_24h")
            co24 = att.get("check_out_24h")
            hw = _calc_hours_worked(att)

            if not co24:
                # --- Without Checkout ---
                if current_date == today:
                    if date_iso in late_approved_dates:
                        detail["status"] = "P"
                    else:
                        detail["status"] = "P"
                else:
                    # Past day, no checkout = missed out-punch
                    if ci24:
                        ci_mins = parse_time_24h_to_minutes(ci24)
                        if ci_mins is not None and ci_mins < 600:
                            detail["status"] = "P"
                        else:
                            detail["status"] = "HD"
                            detail["is_lop"] = True
                            lop += 0.5
                            detail["lop_value"] = 0.5
                    else:
                        detail["status"] = "MP"
            else:
                # --- With Checkout ---
                is_late = False
                if ci24:
                    shift_timings = get_shift_timings(employee)
                    exp_login = shift_timings.get("login_time") if shift_timings else None
                    if exp_login:
                        act_mins = parse_time_24h_to_minutes(ci24)
                        exp_mins = parse_time_24h_to_minutes(exp_login)
                        if act_mins is not None and exp_mins is not None and act_mins > exp_mins:
                            is_late = True

                if hw >= full_hours and not is_late:
                    detail["status"] = "P"
                elif hw >= full_hours and is_late:
                    if date_iso in late_approved_dates:
                        detail["status"] = "P"
                    else:
                        detail["status"] = "LC"
                        detail["is_lop"] = True
                        lop += 0.5
                        detail["lop_value"] = 0.5
                elif hw >= half_hours:
                    detail["status"] = "HD"
                    detail["is_lop"] = True
                    lop += 0.5
                    detail["lop_value"] = 0.5
                else:
                    detail["status"] = "A"
                    lop += 1
                    detail["lop_value"] = 1

            attendance_details.append(detail)
            continue

        # --- SECTION 6D: NO ATTENDANCE (and no leave) ---
        if date_iso in mp_approved_dates:
            detail["status"] = "P"
        elif date_iso in mp_pending_dates:
            detail["status"] = "MP"
        else:
            detail["status"] = "A"
            lop += 1
            detail["lop_value"] = 1

        attendance_details.append(detail)

    # ===== SECTION 8: FINAL PAY FORMULA =====
    # Payable Days = (Working Days - LOP) + Weekoff Pay + Extra Pay
    final_payable_days = (working_days - lop) + weekoff_pay + extra_pay

    # ===== SECTION 9: RELIEVED EMPLOYEE ADJUSTMENT =====
    last_day_payable = employee.get("last_day_payable", True)
    if emp_status == EmployeeStatus.INACTIVE and last_day_payable in (False, 0, "0", "false", "False"):
        final_payable_days -= 1

    final_payable_days = max(0, final_payable_days)

    # Salary calculation
    monthly_salary = employee.get("monthly_salary", 0.0) or 0.0
    per_day_salary = monthly_salary / total_days if total_days > 0 else 0
    net_salary = per_day_salary * final_payable_days
    lop_deduction = monthly_salary - net_salary

    # Derived counts for backward compatibility
    _present_codes = ("P", "HD", "PF", "PH", "SF", "SH", "EF", "EH", "PA", "PP", "OH")
    _leave_codes = ("PF", "PH", "SF", "SH", "EF", "EH", "PA", "PP", "OH")
    present_count = sum(1 for d in attendance_details if d["status"] in _present_codes)
    leave_count = sum(1 for d in attendance_details if d["status"] in _leave_codes)
    absent_count = sum(1 for d in attendance_details if d["status"] == "A")

    return {
        "employee_id": employee_id,
        "emp_name": employee.get("full_name"),
        "emp_id": employee.get("emp_id"),
        "department": dept,
        "team": employee.get("team"),
        "designation": employee.get("designation"),
        "shift_type": employee.get("shift_type"),
        "date_of_joining": employee.get("date_of_joining"),
        "employee_status": emp_status,
        "month": month,
        "monthly_salary": monthly_salary,
        # Spec-defined output fields
        "total_days": total_days,
        "working_days": working_days,
        "weekoff_pay": weekoff_pay,
        "extra_pay": extra_pay,
        "lop": lop,
        "final_payable_days": final_payable_days,
        # Salary
        "per_day_salary": round(per_day_salary, 2),
        "lop_deduction": round(max(0, lop_deduction), 2),
        "net_salary": round(max(0, net_salary), 2),
        # Daily breakdown
        "attendance_details": attendance_details,
        # Backward compatibility fields
        "present_days": present_count,
        "lop_days": lop,
        "leave_days": leave_count,
        "absent_days": absent_count,
    }

# ============== EMAIL HELPERS ==============

async def send_email_notification(to_email: str, subject: str, html_content: str):
    """Send email notification using Resend"""
    if not resend.api_key:
        logger.warning("Resend API key not configured, skipping email")
        return None
    
    try:
        params = {
            "from": SENDER_EMAIL,
            "to": [to_email],
            "subject": subject,
            "html": html_content
        }
        result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Email sent to {to_email}: {result.get('id')}")
        return result
    except Exception as e:
        logger.error(f"Failed to send email: {str(e)}")
        return None

def get_leave_approval_email(emp_name: str, leave_type: str, start_date: str, end_date: str, status: str):
    """Generate HTML email for leave approval/rejection"""
    status_color = "#10b981" if status == "approved" else "#ef4444"
    status_text = "Approved" if status == "approved" else "Rejected"
    
    return f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: #0b1f3b; padding: 20px; border-radius: 8px 8px 0 0;">
            <h1 style="color: white; margin: 0; font-size: 24px;">BluBridge HRMS</h1>
        </div>
        <div style="background: #fffdf7; padding: 30px; border: 1px solid #e5e5e5; border-top: none; border-radius: 0 0 8px 8px;">
            <h2 style="color: #0b1f3b; margin-top: 0;">Leave Request {status_text}</h2>
            <p style="color: #666;">Dear {emp_name},</p>
            <p style="color: #666;">Your leave request has been <span style="color: {status_color}; font-weight: bold;">{status_text.lower()}</span>.</p>
            <div style="background: #f5f5f5; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <p style="margin: 5px 0;"><strong>Leave Type:</strong> {leave_type}</p>
                <p style="margin: 5px 0;"><strong>From:</strong> {start_date}</p>
                <p style="margin: 5px 0;"><strong>To:</strong> {end_date}</p>
                <p style="margin: 5px 0;"><strong>Status:</strong> <span style="color: {status_color};">{status_text}</span></p>
            </div>
            <p style="color: #999; font-size: 12px; margin-top: 30px;">This is an automated notification from BluBridge HRMS.</p>
        </div>
    </div>
    """

def get_star_reward_email(emp_name: str, stars: int, reason: str, awarded_by: str):
    """Generate HTML email for star reward notification"""
    star_color = "#10b981" if stars > 0 else "#ef4444"
    star_text = f"+{stars}" if stars > 0 else str(stars)
    
    return f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: #0b1f3b; padding: 20px; border-radius: 8px 8px 0 0;">
            <h1 style="color: white; margin: 0; font-size: 24px;">BluBridge HRMS</h1>
        </div>
        <div style="background: #fffdf7; padding: 30px; border: 1px solid #e5e5e5; border-top: none; border-radius: 0 0 8px 8px;">
            <h2 style="color: #0b1f3b; margin-top: 0;">Star Reward Notification</h2>
            <p style="color: #666;">Dear {emp_name},</p>
            <p style="color: #666;">You have received a star reward!</p>
            <div style="background: #f5f5f5; padding: 15px; border-radius: 8px; margin: 20px 0; text-align: center;">
                <p style="font-size: 48px; margin: 0; color: {star_color};">{star_text} ⭐</p>
                <p style="margin: 10px 0 0 0; color: #666;">{reason}</p>
            </div>
            <p style="color: #999; font-size: 12px;">Awarded by: {awarded_by}</p>
            <p style="color: #999; font-size: 12px; margin-top: 30px;">This is an automated notification from BluBridge HRMS.</p>
        </div>
    </div>
    """

def get_welcome_email(emp_name: str, emp_id: str, username: str, password: str, login_url: str):
    """Generate HTML email for new employee welcome with credentials"""
    return f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: linear-gradient(135deg, #0b1f3b 0%, #1e3a5f 100%); padding: 30px; border-radius: 12px 12px 0 0; text-align: center;">
            <h1 style="color: white; margin: 0; font-size: 28px; font-weight: 600;">Welcome to BluBridge!</h1>
            <p style="color: rgba(255,255,255,0.8); margin: 10px 0 0 0; font-size: 14px;">Your Employee Account Has Been Created</p>
        </div>
        <div style="background: #fffdf7; padding: 35px; border: 1px solid #e5e5e5; border-top: none; border-radius: 0 0 12px 12px;">
            <p style="color: #333; font-size: 16px; margin-top: 0;">Dear <strong>{emp_name}</strong>,</p>
            <p style="color: #666; line-height: 1.6;">Welcome to the BluBridge team! Your employee account has been successfully created. Below are your login credentials to access the BluBridge HRMS portal.</p>
            
            <div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); padding: 25px; border-radius: 10px; margin: 25px 0; border-left: 4px solid #0b1f3b;">
                <h3 style="color: #0b1f3b; margin: 0 0 15px 0; font-size: 16px;">🔐 Your Login Credentials</h3>
                <table style="width: 100%; border-collapse: collapse;">
                    <tr>
                        <td style="padding: 8px 0; color: #666; width: 120px;">Employee ID:</td>
                        <td style="padding: 8px 0; color: #333; font-weight: 600;">{emp_id}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #666;">Username:</td>
                        <td style="padding: 8px 0; color: #333; font-weight: 600;">{username}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #666;">Password:</td>
                        <td style="padding: 8px 0; color: #333; font-weight: 600; font-family: monospace; background: #fff; padding: 5px 10px; border-radius: 4px;">{password}</td>
                    </tr>
                </table>
            </div>
            
            <div style="text-align: center; margin: 30px 0;">
                <a href="{login_url}" style="display: inline-block; background: #0b1f3b; color: white; padding: 14px 35px; text-decoration: none; border-radius: 8px; font-weight: 600; font-size: 15px;">Login to HRMS Portal</a>
            </div>
            
            <div style="background: #fff3cd; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #ffc107;">
                <p style="color: #856404; margin: 0; font-size: 13px;">⚠️ <strong>Security Reminder:</strong> Please change your password after your first login for security purposes.</p>
            </div>
            
            <hr style="border: none; border-top: 1px solid #e5e5e5; margin: 25px 0;">
            
            <p style="color: #999; font-size: 12px; margin-bottom: 5px;">If you have any questions, please contact HR at hrms@blubridge.ai</p>
            <p style="color: #999; font-size: 12px; margin: 0;">This is an automated notification from BluBridge HRMS.</p>
        </div>
    </div>
    """

async def send_welcome_email(emp_name: str, emp_id: str, email: str, username: str, password: str, login_url: str):
    """Send welcome email with credentials to new employee"""
    resend_api_key = os.environ.get('RESEND_API_KEY')
    sender_email = os.environ.get('SENDER_EMAIL', 'hrms@blubridge.ai')
    
    if not resend_api_key:
        logger.warning("RESEND_API_KEY not configured - skipping welcome email")
        return None
    
    resend.api_key = resend_api_key
    
    html_content = get_welcome_email(emp_name, emp_id, username, password, login_url)
    
    params = {
        "from": sender_email,
        "to": [email],
        "subject": f"Welcome to BluBridge - Your Login Credentials ({emp_id})",
        "html": html_content
    }
    
    try:
        email_result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Welcome email sent to {email}")
        return email_result
    except Exception as e:
        logger.error(f"Failed to send welcome email to {email}: {str(e)}")
        return None

# ============== CLOUDINARY ROUTES ==============

@api_router.get("/cloudinary/signature")
async def get_cloudinary_signature(
    resource_type: str = Query("auto", enum=["image", "video", "auto", "raw"]),
    folder: str = Query("employees"),
    current_user: dict = Depends(get_current_user)
):
    """Generate signed upload params for Cloudinary"""
    ALLOWED_FOLDERS = ("employees", "documents", "avatars")
    if folder not in ALLOWED_FOLDERS:
        raise HTTPException(status_code=400, detail="Invalid folder path")
    
    timestamp = int(time.time())
    # Simple params for public upload
    params = {
        "timestamp": timestamp,
        "folder": f"blubridge/{folder}",
        "type": "upload"  # Ensure public upload type
    }
    
    signature = cloudinary.utils.api_sign_request(
        params,
        os.environ.get("CLOUDINARY_API_SECRET")
    )
    
    return {
        "signature": signature,
        "timestamp": timestamp,
        "cloud_name": os.environ.get("CLOUDINARY_CLOUD_NAME"),
        "api_key": os.environ.get("CLOUDINARY_API_KEY"),
        "folder": f"blubridge/{folder}",
        "resource_type": resource_type,
        "type": "upload"  # Ensure public access
    }

@api_router.delete("/cloudinary/{public_id:path}")
async def delete_cloudinary_asset(
    public_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete asset from Cloudinary"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    try:
        result = await asyncio.to_thread(
            cloudinary.uploader.destroy,
            public_id,
            invalidate=True
        )
        return {"success": True, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============== AUTH ROUTES ==============

@api_router.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest):
    # Case-insensitive username/email match + trim whitespace (users often type with capital letter / autofill trailing space)
    raw_identifier = (request.username or "").strip()
    if not raw_identifier:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    safe_regex = re.escape(raw_identifier)
    user = await db.users.find_one(
        {"$or": [
            {"username": {"$regex": f"^{safe_regex}$", "$options": "i"}},
            {"email": {"$regex": f"^{safe_regex}$", "$options": "i"}},
        ]},
        {"_id": 0},
    )
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Your account is deactivated. Contact admin.")
    if not verify_password(request.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"], user["role"])
    await log_audit(user["id"], "login", "auth")
    
    user_response = {k: v for k, v in user.items() if k != "password_hash"}
    
    # Add onboarding info for employees
    if user.get("role") == UserRole.EMPLOYEE and user.get("employee_id"):
        onboarding = await db.onboarding.find_one({"employee_id": user["employee_id"]}, {"_id": 0})
        user_response["onboarding_status"] = onboarding.get("status") if onboarding else user.get("onboarding_status", OnboardingStatus.PENDING)
        user_response["is_first_login"] = user.get("is_first_login", True)
        user_response["onboarding_completed"] = user_response["onboarding_status"] == OnboardingStatus.APPROVED
    
    return {"token": token, "user": user_response}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return {k: v for k, v in current_user.items() if k != "password_hash"}

@api_router.put("/auth/update-profile")
async def update_admin_profile(
    name: Optional[str] = None,
    email: Optional[str] = None,
    phone: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Update admin user profile"""
    update_data = {}
    if name:
        update_data["name"] = name
    if email:
        update_data["email"] = email
    if phone:
        update_data["phone"] = phone
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = get_ist_now().isoformat()
    
    await db.users.update_one({"id": current_user["id"]}, {"$set": update_data})
    
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password_hash": 0})
    return serialize_doc(user)

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

@api_router.post("/auth/change-password")
async def change_admin_password(
    data: ChangePasswordRequest,
    current_user: dict = Depends(get_current_user)
):
    """Change admin user password"""
    user = await db.users.find_one({"id": current_user["id"]})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify current password
    current_hash = hashlib.sha256(data.current_password.encode()).hexdigest()
    if user.get("password_hash") != current_hash:
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Update to new password
    new_hash = hashlib.sha256(data.new_password.encode()).hexdigest()
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {
            "password_hash": new_hash,
            "updated_at": get_ist_now().isoformat()
        }}
    )
    
    return {"message": "Password changed successfully"}


# ============== CREDENTIAL RESET — TWO INDEPENDENT FLOWS ==============
# Flow A: Admin Direct Reset    (no email; admin sets/auto-generates password)
# Flow B: Employee Self Reset   (login → "Forgot Password?" → username/email →
#                                token sent to registered email → set new pwd)
# Both flows reuse the existing SHA-256 hash_password() so the login engine and
# password storage shape are UNCHANGED. The reset TOKEN itself is a separate
# random secret (secrets.token_urlsafe), stored in `password_reset_tokens`.
import secrets as _secrets


def _generate_strong_password(length: int = 12) -> str:
    """Generate a strong password: ≥1 lower, upper, digit, symbol; ≥12 chars."""
    if length < 12:
        length = 12
    upper = "ABCDEFGHJKLMNPQRSTUVWXYZ"     # excludes I, O for readability
    lower = "abcdefghijkmnopqrstuvwxyz"   # excludes l
    digits = "23456789"                   # excludes 0, 1
    symbols = "!@#$%^&*?-_"
    pools = [upper, lower, digits, symbols]
    pwd = [_secrets.choice(p) for p in pools]
    all_chars = upper + lower + digits + symbols
    pwd += [_secrets.choice(all_chars) for _ in range(length - len(pwd))]
    _secrets.SystemRandom().shuffle(pwd)
    return "".join(pwd)


def _validate_password_strength(pw: str) -> tuple[bool, str]:
    if not pw or len(pw) < 8:
        return False, "Password must be at least 8 characters long"
    if len(pw) > 128:
        return False, "Password is too long (max 128 chars)"
    has_letter = any(c.isalpha() for c in pw)
    has_digit = any(c.isdigit() for c in pw)
    if not (has_letter and has_digit):
        return False, "Password must contain at least one letter and one digit"
    return True, ""


class AdminResetCredentialsRequest(BaseModel):
    password: Optional[str] = None
    confirm_password: Optional[str] = None
    auto_generate: bool = False
    force_change_on_next_login: bool = True


@api_router.post("/admin/employees/{employee_id}/reset-credentials")
async def admin_reset_employee_credentials(
    employee_id: str,
    payload: AdminResetCredentialsRequest,
    current_user: dict = Depends(get_current_user),
):
    """ADMIN DIRECT RESET — no email is sent. Admin either supplies a password
    or asks the system to auto-generate one. The generated/new password is
    returned ONCE in the response so the admin can hand it to the employee.

    Behavior:
      • Old password is invalidated immediately (hash overwritten).
      • `force_change_on_next_login=True` (default) sets `must_change_password`
        on the user, which the FE can use to redirect to /change-password.
      • Audit log entry recorded.
      • Password is NEVER stored or logged in plain text.
    """
    if current_user.get("role") not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    user = await db.users.find_one({"employee_id": employee_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="No login account exists for this employee")

    # Decide the new password
    if payload.auto_generate:
        new_password = _generate_strong_password(12)
        generated = True
    else:
        if not payload.password:
            raise HTTPException(status_code=400, detail="Password is required (or set auto_generate=true)")
        if payload.confirm_password is not None and payload.password != payload.confirm_password:
            raise HTTPException(status_code=400, detail="Passwords do not match")
        ok, msg = _validate_password_strength(payload.password)
        if not ok:
            raise HTTPException(status_code=400, detail=msg)
        new_password = payload.password
        generated = False

    new_hash = hash_password(new_password)
    update = {
        "password_hash": new_hash,
        "must_change_password": bool(payload.force_change_on_next_login),
        "password_updated_at": get_ist_now().isoformat(),
        "password_updated_by": current_user.get("id"),
        "password_updated_method": "admin_reset",
    }
    await db.users.update_one({"id": user["id"]}, {"$set": update})

    # Invalidate any pending self-service reset tokens for this user
    try:
        await db.password_reset_tokens.update_many(
            {"user_id": user["id"], "used": False},
            {"$set": {"used": True, "invalidated_reason": "admin_reset", "used_at": get_ist_now().isoformat()}},
        )
    except Exception:
        pass

    await log_audit(
        current_user.get("id"),
        action="admin_reset_credentials",
        resource="user",
        resource_id=user["id"],
        details=f"Admin reset password for employee {employee_id} ({employee.get('full_name','')}); generated={generated}; force_change={payload.force_change_on_next_login}",
    )

    response: dict = {
        "ok": True,
        "employee_id": employee_id,
        "username": user.get("username"),
        "force_change_on_next_login": bool(payload.force_change_on_next_login),
        "generated": generated,
    }
    # Per spec: surface the password ONCE only when auto-generated. When the
    # admin typed it themselves, they already know it — never echo it back.
    if generated:
        response["password"] = new_password
    return response


# ---------------- Flow B: Employee Self Reset -------------------------------
RESET_TOKEN_TTL_MINUTES = 30


class ForgotPasswordRequest(BaseModel):
    identifier: str  # username OR email


class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str
    confirm_password: Optional[str] = None


async def _ensure_password_reset_token_indexes() -> None:
    try:
        await db.password_reset_tokens.create_index([("token", 1)], unique=True)
        # TTL index: Mongo deletes the doc once expires_at is in the past
        await db.password_reset_tokens.create_index("expires_at", expireAfterSeconds=0)
    except Exception as _e:  # noqa: F841
        pass


def _resolve_employee_email(user: dict) -> Optional[str]:
    """Pick the best email for sending a reset link. Prefer official_email
    on the linked employee record, fall back to the user's own email."""
    email = (user.get("email") or "").strip()
    return email if email and "@" in email else None


@api_router.post("/auth/forgot-password")
async def forgot_password(payload: ForgotPasswordRequest):
    """EMPLOYEE SELF RESET — accepts either username or email.

    Per product spec: when no match is found, we explicitly tell the user
    "User not found" (NOT the typical generic "if the account exists" message).
    A token is created (single-use, 30 min TTL) and a reset link is emailed
    to the REGISTERED EMAIL ONLY.
    """
    raw = (payload.identifier or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="Please enter your username or email")
    safe = re.escape(raw)
    user = await db.users.find_one(
        {"$or": [
            {"username": {"$regex": f"^{safe}$", "$options": "i"}},
            {"email": {"$regex": f"^{safe}$", "$options": "i"}},
        ]},
        {"_id": 0},
    )
    if not user:
        raise HTTPException(status_code=404, detail="User not found. Please check your username or email.")
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Your account is deactivated. Contact admin.")

    # Prefer the linked employee's official_email; fall back to user.email.
    email_to: Optional[str] = None
    employee_id = user.get("employee_id")
    if employee_id:
        emp = await db.employees.find_one({"id": employee_id}, {"_id": 0, "official_email": 1, "email": 1, "full_name": 1})
        if emp:
            email_to = emp.get("official_email") or emp.get("email")
    if not email_to:
        email_to = _resolve_employee_email(user)
    if not email_to:
        raise HTTPException(status_code=400, detail="No registered email on file. Please contact admin.")

    # Issue a single-use token (32 bytes ≈ 43 url-safe chars)
    token = _secrets.token_urlsafe(32)
    now_utc = datetime.utcnow()
    expires_at = now_utc + timedelta(minutes=RESET_TOKEN_TTL_MINUTES)
    await db.password_reset_tokens.insert_one({
        "id": str(uuid.uuid4()),
        "token": token,
        "user_id": user["id"],
        "username": user.get("username"),
        "email": email_to,
        "created_at": now_utc.isoformat(),
        # Mongo TTL index needs a real BSON datetime in UTC (naive is treated as UTC).
        "expires_at": expires_at,
        "used": False,
    })

    base = (os.environ.get("FRONTEND_BASE_URL") or os.environ.get("REACT_APP_BACKEND_URL") or "").rstrip("/")
    reset_url = f"{base}/reset-password?token={token}"

    # Send via existing email engine. NOT routed through cron audit dedup —
    # this is transactional, must always go.
    try:
        from email_templates import password_reset_email
        from email_service import send_hrms_email
        display_name = user.get("name") or (user.get("username") or "there").title()
        if employee_id:
            emp = await db.employees.find_one({"id": employee_id}, {"_id": 0, "full_name": 1})
            if emp and emp.get("full_name"):
                display_name = emp["full_name"]
        html = password_reset_email(display_name, reset_url, RESET_TOKEN_TTL_MINUTES)
        await send_hrms_email(
            db,
            email_type="password_reset",
            scope_key=f"password_reset:{token}",
            to_email=email_to,
            subject="Reset Your BluBridge HRMS Password",
            html=html,
            employee_id=employee_id,
            force=True,  # always send; never dedup transactional resets
        )
    except Exception as e:
        logger.error("forgot-password: email send failed: %s", e)
        raise HTTPException(status_code=502, detail="Could not send reset email — please try again or contact admin.")

    await log_audit(
        user["id"],
        action="forgot_password_request",
        resource="user",
        resource_id=user["id"],
        details=f"Reset email sent to {email_to[:3]}***@{email_to.split('@',1)[1] if '@' in email_to else 'unknown'}",
    )

    # Mask the email in the response so we don't enumerate addresses to a third party
    if "@" in email_to:
        local, domain = email_to.split("@", 1)
        masked = f"{local[:2]}***@{domain}" if len(local) > 2 else f"***@{domain}"
    else:
        masked = "your registered email"
    return {"ok": True, "message": "Reset link sent. Please check your inbox.", "sent_to": masked}


@api_router.get("/auth/reset-password/validate")
async def validate_reset_token(token: str):
    """Lightweight pre-flight check used by the FE before showing the form."""
    if not token:
        raise HTTPException(status_code=400, detail="Token is required")
    doc = await db.password_reset_tokens.find_one({"token": token}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=400, detail="Invalid or expired reset link")
    if doc.get("used"):
        raise HTTPException(status_code=400, detail="This reset link has already been used")
    expires_at = doc.get("expires_at")
    if isinstance(expires_at, str):
        try:
            expires_at = datetime.fromisoformat(expires_at)
        except Exception:
            expires_at = None
    if expires_at:
        # Normalize to naive UTC to compare with utcnow()
        if expires_at.tzinfo is not None:
            expires_at = expires_at.astimezone(timezone.utc).replace(tzinfo=None)
        if expires_at < datetime.utcnow():
            raise HTTPException(status_code=400, detail="This reset link has expired")
    username = doc.get("username") or "your account"
    return {"ok": True, "username": username}


@api_router.post("/auth/reset-password")
async def reset_password(payload: ResetPasswordRequest):
    """Complete the self-service password reset using a one-time token."""
    if not payload.token:
        raise HTTPException(status_code=400, detail="Reset token is required")
    if payload.confirm_password is not None and payload.new_password != payload.confirm_password:
        raise HTTPException(status_code=400, detail="Passwords do not match")
    ok, msg = _validate_password_strength(payload.new_password)
    if not ok:
        raise HTTPException(status_code=400, detail=msg)

    doc = await db.password_reset_tokens.find_one({"token": payload.token}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=400, detail="Invalid or expired reset link")
    if doc.get("used"):
        raise HTTPException(status_code=400, detail="This reset link has already been used")
    expires_at = doc.get("expires_at")
    if isinstance(expires_at, str):
        try:
            expires_at = datetime.fromisoformat(expires_at)
        except Exception:
            expires_at = None
    if expires_at:
        if expires_at.tzinfo is not None:
            expires_at = expires_at.astimezone(timezone.utc).replace(tzinfo=None)
        if expires_at < datetime.utcnow():
            raise HTTPException(status_code=400, detail="This reset link has expired")

    user = await db.users.find_one({"id": doc.get("user_id")}, {"_id": 0})
    if not user or not user.get("is_active", True):
        raise HTTPException(status_code=400, detail="Account not found or deactivated")

    new_hash = hash_password(payload.new_password)
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {
            "password_hash": new_hash,
            "must_change_password": False,  # they just chose it themselves
            "password_updated_at": get_ist_now().isoformat(),
            "password_updated_method": "self_reset",
        }},
    )
    # Mark token used (one-time enforcement)
    await db.password_reset_tokens.update_one(
        {"token": payload.token},
        {"$set": {"used": True, "used_at": get_ist_now().isoformat()}},
    )
    # Invalidate any other outstanding tokens for the same user
    await db.password_reset_tokens.update_many(
        {"user_id": user["id"], "used": False},
        {"$set": {"used": True, "invalidated_reason": "superseded_by_reset", "used_at": get_ist_now().isoformat()}},
    )

    await log_audit(
        user["id"],
        action="self_reset_password",
        resource="user",
        resource_id=user["id"],
        details="Password reset via self-service flow",
    )
    return {"ok": True, "message": "Password reset successfully. Please login."}


# ============== EMPLOYEE MASTER ROUTES ==============

@api_router.get("/employees")
async def get_employees(
    department: Optional[str] = None,
    team: Optional[str] = None,
    status: Optional[str] = None,
    employment_type: Optional[str] = None,
    tier_level: Optional[str] = None,
    work_location: Optional[str] = None,
    inactive_type: Optional[str] = None,
    search: Optional[str] = None,
    include_deleted: bool = False,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    # Exclude soft-deleted by default, but include them when filtering by inactive status or inactive type
    if not include_deleted:
        if (status and status == EmployeeStatus.INACTIVE) or (inactive_type and inactive_type != "All"):
            pass  # Don't exclude — we want to see inactive employees
        else:
            query["is_deleted"] = {"$ne": True}
    
    if department and department != "All":
        query["department"] = department
    if team and team != "All":
        query["team"] = team
    if status and status != "All":
        query["employee_status"] = status
    if employment_type and employment_type != "All":
        query["employment_type"] = employment_type
    if tier_level and tier_level != "All":
        query["tier_level"] = tier_level
    if work_location and work_location != "All":
        query["work_location"] = work_location
    if inactive_type and inactive_type != "All":
        query["inactive_type"] = inactive_type
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"official_email": {"$regex": search, "$options": "i"}},
            {"emp_id": {"$regex": search, "$options": "i"}},
            {"designation": {"$regex": search, "$options": "i"}},
            {"custom_employee_id": {"$regex": search, "$options": "i"}},
            {"biometric_id": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * limit
    total = await db.employees.count_documents(query)
    employees = await db.employees.find(query, {"_id": 0}).skip(skip).limit(limit).sort("created_at", -1).to_list(limit)
    
    # Add reporting manager name
    for emp in employees:
        if emp.get("reporting_manager_id"):
            manager = await db.employees.find_one({"id": emp["reporting_manager_id"]}, {"_id": 0, "full_name": 1})
            emp["reporting_manager_name"] = manager.get("full_name") if manager else None
    
    return {
        "employees": [serialize_doc(e) for e in employees],
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/employees/all")
async def get_all_employees(current_user: dict = Depends(get_current_user)):
    """Get all active employees (for dropdowns)"""
    query = {"is_deleted": {"$ne": True}, "employee_status": EmployeeStatus.ACTIVE}
    employees = await db.employees.find(query, {"_id": 0, "id": 1, "emp_id": 1, "full_name": 1, "department": 1, "team": 1, "custom_employee_id": 1, "biometric_id": 1}).to_list(1000)
    return employees

@api_router.get("/employees/autocomplete")
async def employee_autocomplete(
    q: str = Query("", min_length=0),
    current_user: dict = Depends(get_current_user)
):
    """Lightweight autocomplete: returns max 10 matches by name/email/emp_id."""
    if not q or len(q.strip()) < 1:
        return []
    q = q.strip()
    query = {
        "is_deleted": {"$ne": True},
        "$or": [
            {"full_name": {"$regex": q, "$options": "i"}},
            {"official_email": {"$regex": q, "$options": "i"}},
            {"emp_id": {"$regex": q, "$options": "i"}},
            {"custom_employee_id": {"$regex": q, "$options": "i"}},
        ]
    }
    results = await db.employees.find(
        query,
        {"_id": 0, "id": 1, "full_name": 1, "official_email": 1, "emp_id": 1, "department": 1}
    ).limit(10).to_list(10)
    return results

@api_router.get("/employees/stats")
async def get_employee_stats(current_user: dict = Depends(get_current_user)):
    """Get employee statistics for dashboard"""
    base_query = {"is_deleted": {"$ne": True}}
    
    total = await db.employees.count_documents(base_query)
    active = await db.employees.count_documents({**base_query, "employee_status": EmployeeStatus.ACTIVE})
    inactive = await db.employees.count_documents({**base_query, "employee_status": EmployeeStatus.INACTIVE})
    resigned = await db.employees.count_documents({**base_query, "employee_status": EmployeeStatus.RESIGNED})
    
    # By department
    by_department = {}
    departments = await db.departments.find({}, {"_id": 0, "name": 1}).to_list(100)
    for dept in departments:
        count = await db.employees.count_documents({**base_query, "department": dept["name"], "employee_status": EmployeeStatus.ACTIVE})
        by_department[dept["name"]] = count
    
    # By employment type
    by_type = {}
    for emp_type in [EmploymentType.FULL_TIME, EmploymentType.PART_TIME, EmploymentType.CONTRACT, EmploymentType.INTERN]:
        count = await db.employees.count_documents({**base_query, "employment_type": emp_type, "employee_status": EmployeeStatus.ACTIVE})
        by_type[emp_type] = count
    
    # By work location
    by_location = {}
    for loc in [WorkLocation.REMOTE, WorkLocation.OFFICE, WorkLocation.HYBRID]:
        count = await db.employees.count_documents({**base_query, "work_location": loc, "employee_status": EmployeeStatus.ACTIVE})
        by_location[loc] = count
    
    return {
        "total": total,
        "active": active,
        "inactive": inactive,
        "resigned": resigned,
        "by_department": by_department,
        "by_employment_type": by_type,
        "by_work_location": by_location
    }

# ============== BULK IMPORT ENDPOINTS ==============

IMPORT_TEMPLATE_COLUMNS = [
    "Employee Name", "Employee ID", "Biometric ID", "Email", "Phone",
    "Gender", "Date of Birth", "Date of Joining", "Department", "Team",
    "Designation", "Employment Type", "Tier Level", "Work Location",
    "Shift Type", "Monthly Salary", "User Role"
]

@api_router.get("/employees/import-template")
async def get_import_template(current_user: dict = Depends(get_current_user)):
    """Download sample Excel template for bulk employee import"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Import"
    
    # Header row
    for col_idx, header in enumerate(IMPORT_TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="063c88", end_color="063c88", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    # Sample row
    sample_data = [
        "John Doe", "EID-001", "BIO-001", "john.doe@company.com", "9876543210",
        "Male", "1990-01-15", "2026-01-01", "Technology", "Backend",
        "Software Engineer", "Full-time", "Mid", "Office",
        "General", "50000", "employee"
    ]
    for col_idx, val in enumerate(sample_data, 1):
        ws.cell(row=2, column=col_idx, value=val)
    
    # Auto-adjust column widths
    for col_idx in range(1, len(IMPORT_TEMPLATE_COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
    
    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["Field", "Required", "Notes"],
        ["Employee Name", "Yes", "Full name of the employee"],
        ["Employee ID", "Yes", "Unique alphanumeric ID (e.g., EID-001)"],
        ["Biometric ID", "Yes", "Unique ID for biometric device mapping"],
        ["Email", "Yes", "Must be unique, used for login credentials"],
        ["Phone", "No", "10-digit phone number"],
        ["Gender", "No", "Male / Female / Other"],
        ["Date of Birth", "No", "Format: YYYY-MM-DD"],
        ["Date of Joining", "Yes", "Format: YYYY-MM-DD"],
        ["Department", "Yes", "Must match existing department"],
        ["Team", "Yes", "Must match existing team"],
        ["Designation", "Yes", "Job title"],
        ["Employment Type", "No", "Full-time / Part-time / Contract / Intern (default: Full-time)"],
        ["Tier Level", "No", "Junior / Mid / Senior / Lead (default: Mid)"],
        ["Work Location", "No", "Remote / Office / Hybrid (default: Office)"],
        ["Shift Type", "No", "General / Morning / Evening / Night / Flexible (default: General)"],
        ["Monthly Salary", "No", "Numeric value in INR (default: 0)"],
        ["User Role", "No", "employee / hr / system_admin / office_admin (default: employee)"]
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = openpyxl.styles.Font(bold=True)
    for col_idx in range(1, 4):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_import_template.xlsx"}
    )

@api_router.get("/employees/export")
async def export_employees(
    department: Optional[str] = None,
    team: Optional[str] = None,
    status: Optional[str] = None,
    employment_type: Optional[str] = None,
    tier_level: Optional[str] = None,
    work_location: Optional[str] = None,
    inactive_type: Optional[str] = None,
    search: Optional[str] = None,
    include_deleted: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Export ALL employees matching the given filters as a styled .xlsx file.

    Columns match the bulk import template (17 fields) so the export is a
    valid round-trip with the import flow.
    """
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN, UserRole.OFFICE_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    # Build the same query as GET /employees so the export honors current filters.
    query = {}
    if not include_deleted:
        if (status and status == EmployeeStatus.INACTIVE) or (inactive_type and inactive_type != "All"):
            pass
        else:
            query["is_deleted"] = {"$ne": True}

    if department and department != "All":
        query["department"] = department
    if team and team != "All":
        query["team"] = team
    if status and status != "All":
        query["employee_status"] = status
    if employment_type and employment_type != "All":
        query["employment_type"] = employment_type
    if tier_level and tier_level != "All":
        query["tier_level"] = tier_level
    if work_location and work_location != "All":
        query["work_location"] = work_location
    if inactive_type and inactive_type != "All":
        query["inactive_type"] = inactive_type
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"official_email": {"$regex": search, "$options": "i"}},
            {"emp_id": {"$regex": search, "$options": "i"}},
            {"designation": {"$regex": search, "$options": "i"}},
            {"custom_employee_id": {"$regex": search, "$options": "i"}},
            {"biometric_id": {"$regex": search, "$options": "i"}}
        ]

    employees = await db.employees.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Styled header (dark-blue background, bold white) — matches import template
    header_fill = openpyxl.styles.PatternFill(start_color="063c88", end_color="063c88", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(IMPORT_TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Data rows (mapped exactly to IMPORT_TEMPLATE_COLUMNS order)
    for row_idx, e in enumerate(employees, start=2):
        salary = e.get("monthly_salary")
        try:
            salary_val = float(salary) if salary is not None else 0
        except (TypeError, ValueError):
            salary_val = 0
        row_values = [
            e.get("full_name") or "",
            e.get("custom_employee_id") or "",
            e.get("biometric_id") or "",
            e.get("official_email") or "",
            e.get("phone_number") or "",
            e.get("gender") or "",
            e.get("date_of_birth") or "",
            e.get("date_of_joining") or "",
            e.get("department") or "",
            e.get("team") or "",
            e.get("designation") or "",
            e.get("employment_type") or "",
            e.get("tier_level") or "",
            e.get("work_location") or "",
            e.get("shift_type") or "",
            salary_val,
            e.get("user_role") or "",
        ]
        for col_idx, val in enumerate(row_values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Column widths sized for content
    widths = [22, 14, 14, 28, 14, 10, 14, 14, 18, 16, 22, 14, 12, 14, 12, 14, 14]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"employees-{get_ist_now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@api_router.post("/employees/bulk-import")
async def bulk_import_employees(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk import employees from CSV or Excel file"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Read file content
    content = await file.read()
    filename = file.filename.lower() if file.filename else ""
    
    rows = []
    if filename.endswith('.csv'):
        # Parse CSV
        try:
            text_content = content.decode('utf-8-sig')  # Handle BOM
            reader = csv.DictReader(io.StringIO(text_content))
            for row in reader:
                rows.append(row)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {str(e)}")
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Parse Excel
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(cell is None for cell in row):
                    continue
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    if header and col_idx < len(row):
                        row_dict[header] = row[col_idx]
                if row_dict:
                    rows.append(row_dict)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Please upload .csv or .xlsx file")
    
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in file")
    
    # Static allowed values (matching frontend dropdowns)
    VALID_DEPARTMENTS = {"research unit", "support staff", "business & product"}
    VALID_TEAMS = {"data", "parallelism", "quantization", "compiler", "tensor & ops", "hardware", "administation", "it", "product team", "unknown"}
    VALID_DESIGNATIONS = {"ai research scientist", "ai research - intern", "research", "front office", "junior admin", "junior system admin", "business & product - product team", "system engineer"}
    
    # Canonical name mapping (lowercase -> proper case)
    DEPT_CANONICAL = {d.lower(): d for d in ["Research Unit", "Support Staff", "Business & Product"]}
    TEAM_CANONICAL = {t.lower(): t for t in ["Data", "Parallelism", "Quantization", "Compiler", "Tensor & Ops", "Hardware", "Administation", "IT", "Product Team", "Unknown"]}
    DESIG_CANONICAL = {d.lower(): d for d in ["AI Research scientist", "AI Research - Intern", "Research", "Front Office", "Junior Admin", "Junior System admin", "Business & Product - Product Team", "System Engineer"]}
    
    # Ensure department and team records exist in DB (auto-create if missing)
    existing_depts = {d["name"].lower() for d in await db.departments.find({}, {"_id": 0, "name": 1}).to_list(100)}
    existing_teams = {t["name"].lower() for t in await db.teams.find({}, {"_id": 0, "name": 1}).to_list(500)}
    
    for dept_lower, dept_name in DEPT_CANONICAL.items():
        if dept_lower not in existing_depts:
            await db.departments.insert_one({"id": str(uuid.uuid4()), "name": dept_name, "description": "", "head_id": None, "member_count": 0, "created_at": get_ist_now().isoformat()})
    
    for team_lower, team_name in TEAM_CANONICAL.items():
        if team_lower not in existing_teams:
            await db.teams.insert_one({"id": str(uuid.uuid4()), "name": team_name, "department": "", "description": "", "lead_id": None, "member_count": 0, "created_at": get_ist_now().isoformat()})
    
    # Get existing employee IDs and biometric IDs for uniqueness check
    existing_employees = await db.employees.find(
        {"is_deleted": {"$ne": True}},
        {"_id": 0, "official_email": 1, "custom_employee_id": 1, "biometric_id": 1}
    ).to_list(10000)
    
    existing_emails = {e.get("official_email", "").lower() for e in existing_employees if e.get("official_email")}
    existing_emp_ids = {e.get("custom_employee_id", "").lower() for e in existing_employees if e.get("custom_employee_id")}
    existing_bio_ids = {e.get("biometric_id", "").lower() for e in existing_employees if e.get("biometric_id")}
    
    # Track IDs in current batch for duplicates within file
    batch_emails = set()
    batch_emp_ids = set()
    batch_bio_ids = set()
    
    results = {
        "total": len(rows),
        "success": 0,
        "failed": 0,
        "errors": []
    }
    
    # Column name mapping (flexible)
    def get_value(row, *keys):
        for key in keys:
            for row_key in row.keys():
                if row_key and row_key.strip().lower() == key.lower():
                    return row.get(row_key)
        return None
    
    def parse_date(val):
        """Parse date from various formats to YYYY-MM-DD. Handles Excel numeric/datetime dates too."""
        if val is None or str(val).strip() == "" or str(val).strip().lower() == "none":
            return None
        
        # Handle Python datetime objects (openpyxl may return these directly)
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        
        # Handle Excel numeric date serial (float/int)
        if isinstance(val, (int, float)):
            try:
                from datetime import datetime as dt_cls, timedelta as td_cls
                excel_epoch = dt_cls(1899, 12, 30)
                return (excel_epoch + td_cls(days=int(val))).strftime("%Y-%m-%d")
            except Exception:
                return None
        
        raw = str(val).strip()
        # If already ISO YYYY-MM-DD
        if len(raw) == 10 and raw[4] == '-' and raw[7] == '-':
            try:
                datetime.strptime(raw, "%Y-%m-%d")
                return raw
            except ValueError:
                pass
        
        # MM/DD/YYYY (primary), then other formats
        for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        
        return "__INVALID__"
    
    for idx, row in enumerate(rows, start=2):
        row_errors = []
        
        # Extract values with flexible column names
        full_name = get_value(row, "Employee Name", "Full Name", "Name")
        custom_employee_id = get_value(row, "Employee ID", "Custom Employee ID", "Emp ID")
        biometric_id = get_value(row, "Biometric ID", "Bio ID", "Biometric")
        email = get_value(row, "Email", "Official Email", "Email Address")
        phone = get_value(row, "Phone", "Phone Number", "Contact")
        gender = get_value(row, "Gender")
        dob = get_value(row, "Date of Birth", "DOB", "Birth Date")
        doj = get_value(row, "Date of Joining", "DOJ", "Joining Date", "Start Date")
        department = get_value(row, "Department", "Dept")
        team = get_value(row, "Team")
        designation = get_value(row, "Designation", "Title", "Job Title", "Position")
        employment_type = get_value(row, "Employment Type", "Type") or "Full-time"
        tier_level = get_value(row, "Tier Level", "Tier", "Level") or "Mid"
        work_location = get_value(row, "Work Location", "Location") or "Office"
        shift_type = get_value(row, "Shift Type", "Shift") or "General"
        monthly_salary = get_value(row, "Monthly Salary", "Salary")
        user_role = get_value(row, "User Role", "Role") or "employee"
        
        # Convert values
        full_name = str(full_name).strip() if full_name else ""
        custom_employee_id = str(custom_employee_id).strip() if custom_employee_id else ""
        biometric_id = str(biometric_id).strip() if biometric_id else ""
        email = str(email).strip() if email else ""
        phone = str(phone).strip() if phone else None
        gender = str(gender).strip() if gender else None
        
        # Parse dates BEFORE string conversion (preserves datetime/numeric types from Excel)
        parsed_dob = parse_date(dob)
        if dob is not None and parsed_dob == "__INVALID__":
            row_errors.append(f"Invalid Date Format for Date of Birth: '{dob}'")
            parsed_dob = None
        dob = parsed_dob
        
        parsed_doj = parse_date(doj)
        if doj is not None and parsed_doj == "__INVALID__":
            row_errors.append(f"Invalid Date Format for Date of Joining: '{doj}'")
            parsed_doj = ""
        doj = parsed_doj if parsed_doj else ""
        
        department = str(department).strip() if department else ""
        team = str(team).strip() if team else ""
        designation = str(designation).strip() if designation else ""
        
        try:
            monthly_salary = float(monthly_salary) if monthly_salary else 0.0
        except:
            monthly_salary = 0.0
        
        # Validate required fields
        if not full_name:
            row_errors.append("Employee Name is required")
        if not custom_employee_id:
            row_errors.append("Employee ID is required")
        if not biometric_id:
            row_errors.append("Biometric ID is required")
        if not email:
            row_errors.append("Email is required")
        if not doj:
            row_errors.append("Date of Joining is required")
        if not department:
            row_errors.append("Department is required")
        if not team:
            row_errors.append("Team is required")
        if not designation:
            row_errors.append("Designation is required")
        
        # Validate department (case-insensitive, static list)
        if department:
            dept_lower = department.lower().strip()
            if dept_lower in DEPT_CANONICAL:
                department = DEPT_CANONICAL[dept_lower]
            # Accept as-is if not in static list (no rejection)
        
        # Validate team (case-insensitive, static list)
        if team:
            team_lower = team.lower().strip()
            if team_lower in TEAM_CANONICAL:
                team = TEAM_CANONICAL[team_lower]
        
        # Normalize designation (case-insensitive)
        if designation:
            desig_lower = designation.lower().strip()
            if desig_lower in DESIG_CANONICAL:
                designation = DESIG_CANONICAL[desig_lower]
        
        # Normalize employment type (flexible matching)
        emp_type_map = {"full time": "Full-time", "full-time": "Full-time", "fulltime": "Full-time",
                        "part time": "Part-time", "part-time": "Part-time", "parttime": "Part-time",
                        "contract": "Contract", "intern": "Intern", "internship": "Intern"}
        employment_type = emp_type_map.get(str(employment_type).lower().strip(), employment_type)
        
        # Normalize tier level
        tier_map = {"junior": "Junior", "mid": "Mid", "senior": "Senior", "lead": "Lead"}
        tier_level = tier_map.get(str(tier_level).lower().strip(), tier_level)
        
        # Normalize user role
        user_role = str(user_role).lower().strip()
        role_map = {"employee": "employee", "hr": "hr", "system_admin": "system_admin",
                    "office_admin": "office_admin"}
        user_role = role_map.get(user_role, "employee")
        
        # Extract deactivation fields (optional)
        row_status = get_value(row, "Status", "Employee Status")
        row_inactive_type = get_value(row, "Inactive Type", "Inactive_Type", "InactiveType")
        row_inactive_date = get_value(row, "Inactive Date", "Inactive_Date", "Deactivation Date")
        row_inactive_reason = get_value(row, "Reason", "Inactive Reason", "Deactivation Reason")
        row_last_day_payable = get_value(row, "Last Day Payable", "Last_Day_Payable")
        
        # Check for duplicate email
        if email:
            email_lower = email.lower()
            if email_lower in existing_emails:
                row_errors.append(f"Email '{email}' already exists")
            elif email_lower in batch_emails:
                row_errors.append(f"Duplicate email '{email}' in file")
            else:
                batch_emails.add(email_lower)
        
        # Check for duplicate Employee ID
        if custom_employee_id:
            emp_id_lower = custom_employee_id.lower()
            if emp_id_lower in existing_emp_ids:
                row_errors.append(f"Employee ID '{custom_employee_id}' already exists")
            elif emp_id_lower in batch_emp_ids:
                row_errors.append(f"Duplicate Employee ID '{custom_employee_id}' in file")
            else:
                batch_emp_ids.add(emp_id_lower)
        
        # Check for duplicate Biometric ID
        if biometric_id:
            bio_id_lower = biometric_id.lower()
            if bio_id_lower in existing_bio_ids:
                row_errors.append(f"Biometric ID '{biometric_id}' already exists")
            elif bio_id_lower in batch_bio_ids:
                row_errors.append(f"Duplicate Biometric ID '{biometric_id}' in file")
            else:
                batch_bio_ids.add(bio_id_lower)
        
        if row_errors:
            results["failed"] += 1
            results["errors"].append({
                "row": idx,
                "employee_name": full_name or f"Row {idx}",
                "errors": row_errors
            })
            continue
        
        # Create employee
        try:
            emp_id = await generate_emp_id()
            
            employee = Employee(
                emp_id=emp_id,
                full_name=full_name,
                official_email=email,
                phone_number=phone,
                gender=gender,
                date_of_birth=dob,
                custom_employee_id=custom_employee_id,
                date_of_joining=doj,
                employment_type=employment_type,
                designation=designation,
                tier_level=tier_level,
                department=department,
                team=team,
                work_location=work_location,
                shift_type=shift_type,
                monthly_salary=monthly_salary,
                user_role=user_role.lower() if user_role else "employee",
                login_enabled=True,
                biometric_id=biometric_id
            )
            
            emp_doc = employee.model_dump()
            emp_doc['created_at'] = emp_doc['created_at'].isoformat()
            emp_doc['updated_at'] = emp_doc['updated_at'].isoformat()
            
            await db.employees.insert_one(emp_doc.copy())
            
            # Create user account for login
            username = email.split('@')[0]
            name_part = full_name.replace(' ', '').lower()[:4]
            phone_part = str(phone)[-4:] if phone and len(str(phone)) >= 4 else str(uuid.uuid4())[:4]
            temp_password = f"{name_part}@{phone_part}"
            
            existing_user = await db.users.find_one({"username": username})
            if not existing_user:
                new_user = User(
                    username=username,
                    email=email,
                    password_hash=hash_password(temp_password),
                    name=full_name,
                    role=user_role.lower() if user_role else "employee",
                    employee_id=employee.id,
                    department=department,
                    team=team
                )
                user_doc = new_user.model_dump()
                user_doc['created_at'] = user_doc['created_at'].isoformat()
                await db.users.insert_one(user_doc.copy())
            
            # Send welcome email with credentials (non-blocking)
            asyncio.create_task(
                send_welcome_email(
                    emp_name=full_name,
                    emp_id=emp_id,
                    email=email,
                    username=username,
                    password=temp_password,
                    login_url=f"{os.environ.get('FRONTEND_URL', 'https://leave-code-mapper.preview.emergentagent.com')}/login"
                )
            )

            # Temporary policy: skip onboarding within configured window so
            # bulk-imported employees can log in directly.
            if should_skip_onboarding_now():
                _now_iso = get_ist_now().isoformat()
                await db.employees.update_one(
                    {"id": employee.id},
                    {"$set": {"onboarding_status": OnboardingStatus.APPROVED, "onboarding_completed_at": _now_iso}}
                )
                await db.users.update_many(
                    {"employee_id": employee.id},
                    {"$set": {"onboarding_status": OnboardingStatus.APPROVED, "is_first_login": False}}
                )
            
            # Add to existing sets to prevent duplicates in same batch
            existing_emails.add(email.lower())
            existing_emp_ids.add(custom_employee_id.lower())
            existing_bio_ids.add(biometric_id.lower())
            
            # Handle bulk deactivation if status=inactive
            if row_status and str(row_status).lower().strip() == "inactive":
                deact_type = str(row_inactive_type).strip() if row_inactive_type else "Terminated"
                parsed_inactive_date = parse_date(row_inactive_date) if row_inactive_date else get_ist_now().strftime("%Y-%m-%d")
                if parsed_inactive_date == "__INVALID__":
                    parsed_inactive_date = get_ist_now().strftime("%Y-%m-%d")
                deact_reason = str(row_inactive_reason).strip() if row_inactive_reason else ""
                ldp = str(row_last_day_payable).strip().lower() in ("yes", "true", "1") if row_last_day_payable else False
                
                await db.employees.update_one({"id": employee.id}, {"$set": {
                    "employee_status": EmployeeStatus.INACTIVE,
                    "login_enabled": False,
                    "inactive_type": deact_type,
                    "inactive_date": parsed_inactive_date,
                    "inactive_reason": deact_reason,
                    "last_day_payable": ldp,
                    "deactivated_by": current_user["id"]
                }})
                await db.users.update_one({"employee_id": employee.id}, {"$set": {"is_active": False}})
            
            results["success"] += 1
            
        except Exception as e:
            results["failed"] += 1
            results["errors"].append({
                "row": idx,
                "employee_name": full_name,
                "errors": [str(e)]
            })
    
    return results

@api_router.get("/employees/{employee_id}")
async def get_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    employee = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Add reporting manager details
    if employee.get("reporting_manager_id"):
        manager = await db.employees.find_one({"id": employee["reporting_manager_id"]}, {"_id": 0, "full_name": 1, "emp_id": 1})
        employee["reporting_manager_name"] = manager.get("full_name") if manager else None
        employee["reporting_manager_emp_id"] = manager.get("emp_id") if manager else None

    # Attach the linked login username (read-only, detail-view only).
    # Auth tables are NOT modified — we only project the username for display.
    user_doc = await db.users.find_one({"employee_id": employee_id}, {"_id": 0, "username": 1})
    employee["username"] = user_doc.get("username") if user_doc else None

    return serialize_doc(employee)

@api_router.post("/employees")
async def create_employee(data: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Case-insensitive duplicate guards (Email + Biometric ID)
    email_norm = (data.official_email or "").strip()
    bio_norm = (data.biometric_id or "").strip() if data.biometric_id else ""

    # Check for duplicate email among active employees (case-insensitive)
    if email_norm:
        existing_active = await db.employees.find_one({
            "official_email": {"$regex": f"^{re.escape(email_norm)}$", "$options": "i"},
            "is_deleted": {"$ne": True},
        })
        if existing_active:
            raise HTTPException(status_code=400, detail="Employee with this Email already exists")
    
    # Validate uniqueness of custom_employee_id
    if data.custom_employee_id:
        existing_cid = await db.employees.find_one({"custom_employee_id": data.custom_employee_id, "is_deleted": {"$ne": True}})
        if existing_cid:
            raise HTTPException(status_code=400, detail=f"Employee ID '{data.custom_employee_id}' already exists")
    
    # Validate uniqueness of biometric_id (case-insensitive)
    if bio_norm:
        existing_bio = await db.employees.find_one({
            "biometric_id": {"$regex": f"^{re.escape(bio_norm)}$", "$options": "i"},
            "is_deleted": {"$ne": True},
        })
        if existing_bio:
            raise HTTPException(status_code=400, detail="Employee with this Biometric ID already exists")
    
    # Check if there's a deleted employee with same email - reactivate instead (case-insensitive)
    existing_deleted = await db.employees.find_one({
        "official_email": {"$regex": f"^{re.escape(email_norm)}$", "$options": "i"},
        "is_deleted": True,
    }) if email_norm else None
    
    username = data.official_email.split('@')[0]
    name_part = data.full_name.replace(' ', '').lower()[:4]
    if data.phone_number and len(data.phone_number) >= 4:
        phone_part = data.phone_number[-4:]
    else:
        phone_part = str(uuid.uuid4())[:4]
    temp_password = f"{name_part}@{phone_part}"
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://blubrg.com')
    login_url = f"{frontend_url}/login"
    
    if existing_deleted:
        # Reactivate the deleted employee with updated info
        emp_id = existing_deleted.get("emp_id")
        employee_id = existing_deleted.get("id")
        
        update_data = {
            "is_deleted": False,
            "deleted_at": None,
            "employee_status": EmployeeStatus.ACTIVE,
            "full_name": data.full_name,
            "phone_number": data.phone_number,
            "gender": data.gender,
            "date_of_birth": data.date_of_birth,
            "date_of_joining": data.date_of_joining,
            "employment_type": data.employment_type,
            "designation": data.designation,
            "tier_level": data.tier_level,
            "reporting_manager_id": data.reporting_manager_id,
            "department": data.department,
            "team": data.team,
            "work_location": data.work_location,
            "leave_policy": data.leave_policy,
            "shift_type": data.shift_type,
            "attendance_tracking_enabled": data.attendance_tracking_enabled,
            "user_role": data.user_role,
            "login_enabled": data.login_enabled,
            "custom_employee_id": data.custom_employee_id,
            "biometric_id": data.biometric_id,
            "updated_at": get_ist_now().isoformat()
        }
        # If the picked shift_type matches a Settings shift, expand all derived fields
        update_data = await _apply_settings_shift_to_employee_payload(update_data, data.shift_type)
        
        await db.employees.update_one({"id": employee_id}, {"$set": update_data})
        
        # Update team member count
        await db.teams.update_one({"name": data.team}, {"$inc": {"member_count": 1}})
        
        # Update or create user account with new credentials
        if data.login_enabled:
            existing_user = await db.users.find_one({"username": username})
            if existing_user:
                # Update existing user with new password and reactivate
                await db.users.update_one(
                    {"username": username},
                    {"$set": {
                        "password_hash": hash_password(temp_password),
                        "is_active": True,
                        "name": data.full_name,
                        "role": data.user_role if data.user_role else UserRole.EMPLOYEE,
                        "department": data.department,
                        "team": data.team
                    }}
                )
            else:
                # Create new user
                new_user = User(
                    username=username,
                    email=data.official_email,
                    password_hash=hash_password(temp_password),
                    name=data.full_name,
                    role=data.user_role if data.user_role else UserRole.EMPLOYEE,
                    employee_id=employee_id,
                    department=data.department,
                    team=data.team
                )
                user_doc = new_user.model_dump()
                user_doc['created_at'] = user_doc['created_at'].isoformat()
                await db.users.insert_one(user_doc.copy())
            
            # Send welcome email with new credentials
            asyncio.create_task(
                send_welcome_email(
                    emp_name=data.full_name,
                    emp_id=emp_id,
                    email=data.official_email,
                    username=username,
                    password=temp_password,
                    login_url=login_url
                )
            )
        
        await log_audit(current_user["id"], "reactivate", "employee", employee_id, f"Reactivated employee: {data.full_name}")
        
        employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
        result = serialize_doc(employee)
        if data.login_enabled:
            result['temp_password'] = temp_password
            result['username'] = username
            result['reactivated'] = True
        
        return result
    
    # Create new employee
    emp_id = await generate_emp_id()
    
    employee = Employee(
        emp_id=emp_id,
        full_name=data.full_name,
        official_email=data.official_email,
        phone_number=data.phone_number,
        gender=data.gender,
        date_of_birth=data.date_of_birth,
        custom_employee_id=data.custom_employee_id,
        date_of_joining=data.date_of_joining,
        employment_type=data.employment_type,
        designation=data.designation,
        tier_level=data.tier_level,
        reporting_manager_id=data.reporting_manager_id,
        department=data.department,
        team=data.team,
        work_location=data.work_location,
        leave_policy=data.leave_policy,
        shift_type=data.shift_type,
        attendance_tracking_enabled=data.attendance_tracking_enabled,
        user_role=data.user_role,
        login_enabled=data.login_enabled,
        biometric_id=data.biometric_id
    )
    
    doc = employee.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    # If the picked shift_type matches a Settings shift, expand all derived fields
    doc = await _apply_settings_shift_to_employee_payload(doc, data.shift_type)
    await db.employees.insert_one(doc.copy())
    
    # Update team member count
    await db.teams.update_one({"name": data.team}, {"$inc": {"member_count": 1}})
    
    # Create user account if login is enabled
    if data.login_enabled:
        # Ensure username is unique by appending a suffix if needed
        existing_user = await db.users.find_one({"username": username})
        if existing_user:
            suffix = 1
            while await db.users.find_one({"username": f"{username}{suffix}"}):
                suffix += 1
            username = f"{username}{suffix}"
        if True:
            new_user = User(
                username=username,
                email=data.official_email,
                password_hash=hash_password(temp_password),
                name=data.full_name,
                role=data.user_role if data.user_role else UserRole.EMPLOYEE,
                employee_id=employee.id,
                department=data.department,
                team=data.team
            )
            user_doc = new_user.model_dump()
            user_doc['created_at'] = user_doc['created_at'].isoformat()
            await db.users.insert_one(user_doc.copy())
        
        # Send welcome email with credentials
        asyncio.create_task(
            send_welcome_email(
                emp_name=data.full_name,
                emp_id=emp_id,
                email=data.official_email,
                username=username,
                password=temp_password,
                login_url=login_url
            )
        )
    
    await log_audit(current_user["id"], "create", "employee", employee.id, f"Created employee: {data.full_name}")
    
    # Create onboarding record for new employee
    onboarding_record = OnboardingRecord(
        employee_id=employee.id,
        emp_id=emp_id,
        emp_name=data.full_name,
        department=data.department,
        team=data.team,
        designation=data.designation
    )
    onboarding_doc = onboarding_record.model_dump()
    onboarding_doc['created_at'] = onboarding_doc['created_at'].isoformat()
    onboarding_doc['updated_at'] = onboarding_doc['updated_at'].isoformat()
    await db.onboarding.insert_one(onboarding_doc.copy())
    
    # Create document placeholders for onboarding
    for req_doc in REQUIRED_DOCUMENTS:
        doc_record = OnboardingDocument(
            employee_id=employee.id,
            document_type=req_doc["type"],
            document_label=req_doc["label"]
        )
        doc_data = doc_record.model_dump()
        doc_data['created_at'] = doc_data['created_at'].isoformat()
        await db.onboarding_documents.insert_one(doc_data.copy())
    
    # Create operational checklist for Office Admin
    checklist = OperationalChecklist(
        employee_id=employee.id,
        emp_name=data.full_name,
        department=data.department,
        designation=data.designation,
        items=[item.copy() for item in OPERATIONAL_CHECKLIST_ITEMS]
    )
    checklist_doc = checklist.model_dump()
    checklist_doc['created_at'] = checklist_doc['created_at'].isoformat()
    checklist_doc['updated_at'] = checklist_doc['updated_at'].isoformat()
    await db.operational_checklists.insert_one(checklist_doc.copy())

    # Temporary policy: during the configured window, new employees skip the
    # onboarding flow entirely and land on the dashboard right after login.
    if should_skip_onboarding_now():
        now_iso = get_ist_now().isoformat()
        await db.employees.update_one(
            {"id": employee.id},
            {"$set": {"onboarding_status": OnboardingStatus.APPROVED, "onboarding_completed_at": now_iso}}
        )
        await db.users.update_many(
            {"employee_id": employee.id},
            {"$set": {"onboarding_status": OnboardingStatus.APPROVED, "is_first_login": False}}
        )
        await db.onboarding.update_one(
            {"employee_id": employee.id},
            {"$set": {"status": OnboardingStatus.APPROVED, "approved_at": now_iso, "updated_at": now_iso}}
        )
    
    result = serialize_doc(doc)
    if data.login_enabled:
        result['temp_password'] = temp_password
        result['username'] = username
    
    # Notify HR about new employee onboarding
    asyncio.create_task(notify_role(
        UserRole.HR,
        "New Employee Onboarded",
        f"{data.full_name} has been added to the system. Start verification & induction.",
        "action",
        "/verification"
    ))
    
    # Notify Office Admin about operational setup needed
    asyncio.create_task(notify_role(
        UserRole.OFFICE_ADMIN,
        "Operational Setup Required",
        f"New employee {data.full_name} joining. Prepare workstation, ID card, stationery & access setup.",
        "action",
        "/operational-checklist"
    ))
    
    return result

@api_router.put("/employees/{employee_id}")
async def update_employee(employee_id: str, data: EmployeeUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    existing = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check for duplicate email if changing (case-insensitive)
    if data.official_email and (data.official_email or "").strip().lower() != (existing.get("official_email") or "").strip().lower():
        dup = await db.employees.find_one({
            "official_email": {"$regex": f"^{re.escape(data.official_email.strip())}$", "$options": "i"},
            "id": {"$ne": employee_id},
            "is_deleted": {"$ne": True},
        })
        if dup:
            raise HTTPException(status_code=400, detail="Employee with this Email already exists")
    
    # Check uniqueness of custom_employee_id if changing
    if data.custom_employee_id and data.custom_employee_id != existing.get("custom_employee_id"):
        dup_cid = await db.employees.find_one({"custom_employee_id": data.custom_employee_id, "id": {"$ne": employee_id}, "is_deleted": {"$ne": True}})
        if dup_cid:
            raise HTTPException(status_code=400, detail=f"Employee ID '{data.custom_employee_id}' already exists")
    
    # Check uniqueness of biometric_id if changing (case-insensitive)
    if data.biometric_id and (data.biometric_id or "").strip().lower() != (existing.get("biometric_id") or "").strip().lower():
        dup_bio = await db.employees.find_one({
            "biometric_id": {"$regex": f"^{re.escape(data.biometric_id.strip())}$", "$options": "i"},
            "id": {"$ne": employee_id},
            "is_deleted": {"$ne": True},
        })
        if dup_bio:
            raise HTTPException(status_code=400, detail="Employee with this Biometric ID already exists")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = get_ist_now().isoformat()

    # If shift_type matches a Settings shift, expand all derived fields so
    # the employee record stays consistent with central shift configuration.
    update_data = await _apply_settings_shift_to_employee_payload(update_data, update_data.get("shift_type"))

    # Handle team change
    old_team = existing.get("team")
    new_team = update_data.get("team")
    if new_team and old_team and new_team != old_team:
        await db.teams.update_one({"name": old_team}, {"$inc": {"member_count": -1}})
        await db.teams.update_one({"name": new_team}, {"$inc": {"member_count": 1}})
    
    result = await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    await log_audit(current_user["id"], "update", "employee", employee_id, f"Updated fields: {list(update_data.keys())}")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return serialize_doc(employee)

@api_router.delete("/employees/{employee_id}")
async def deactivate_employee(employee_id: str, request: Request, current_user: dict = Depends(get_current_user)):
    """Soft deactivate - marks employee as inactive without deleting the record"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    existing = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    if existing.get("employee_status") == EmployeeStatus.INACTIVE:
        raise HTTPException(status_code=400, detail="Employee is already inactive")
    
    # Parse optional body (form modal data)
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    
    inactive_type = body.get("inactive_type", "Terminated")
    inactive_date = body.get("inactive_date", get_ist_now().strftime("%Y-%m-%d"))
    reason = body.get("reason", "")
    last_day_payable = body.get("last_day_payable", False)
    
    update_data = {
        "employee_status": EmployeeStatus.INACTIVE,
        "login_enabled": False,
        "inactive_type": inactive_type,
        "inactive_date": inactive_date,
        "inactive_reason": reason,
        "last_day_payable": last_day_payable,
        "deactivated_by": current_user["id"],
        "updated_at": get_ist_now().isoformat()
    }
    
    await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    
    # Update team member count
    await db.teams.update_one({"name": existing.get("team")}, {"$inc": {"member_count": -1}})
    
    # Deactivate the user account as well
    await db.users.update_one(
        {"employee_id": employee_id},
        {"$set": {"is_active": False}}
    )
    
    await log_audit(current_user["id"], "deactivate", "employee", employee_id, f"Deactivated employee: {existing.get('full_name')} | Type: {inactive_type} | Reason: {reason}")
    return {"message": "Employee deactivated successfully"}


# ============== EMPLOYEE PERMANENT DELETE (SAFE + FORCE) ==============
# Linked collections that hold per-employee data — used by impact preview AND
# by both delete modes. Keep this list central and explicit; never cascade by
# heuristic. New per-employee collections must be added here intentionally.
EMPLOYEE_LINKED_COLLECTIONS = [
    "attendance",
    "leaves",
    "late_requests",
    "early_out_requests",
    "missed_punches",
    "payroll",
    "salary_adjustments",
    "employee_documents",
    "performance_reviews",
    "timesheets",
    "star_records",
    "biometric_devices_map",
    "shift_overrides",
    "employee_warnings",
]


async def _employee_record_counts(employee_id: str) -> dict:
    counts: dict = {}
    for coll in EMPLOYEE_LINKED_COLLECTIONS:
        try:
            counts[coll] = await db[coll].count_documents({"employee_id": employee_id})
        except Exception:
            counts[coll] = 0
    counts["total"] = sum(v for k, v in counts.items() if k != "total")
    return counts


@api_router.get("/employees/{employee_id}/deletion-impact")
async def employee_deletion_impact(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Counts of related records — drives the safe vs force delete UX.
    HR + system_admin only."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")
    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0, "id": 1, "full_name": 1, "employee_status": 1})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    counts = await _employee_record_counts(employee_id)
    return {
        "employee_id": employee_id,
        "full_name": emp.get("full_name"),
        "employee_status": emp.get("employee_status"),
        "counts": counts,
        "can_permanent_delete": counts["total"] == 0,
    }


@api_router.delete("/employees/{employee_id}/permanent")
async def employee_permanent_delete(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Safe permanent delete — hard-deletes the employee row ONLY when there
    are zero linked records anywhere. HR + system_admin only."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")
    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    counts = await _employee_record_counts(employee_id)
    if counts["total"] > 0:
        raise HTTPException(
            status_code=409,
            detail="Cannot permanently delete. Employee has existing records. Please use Deactivate instead.",
        )

    # Hard delete employee + their user account (no linked-data destruction
    # possible because we just verified counts are zero).
    await db.employees.delete_one({"id": employee_id})
    await db.users.delete_many({"employee_id": employee_id})

    await db.employee_deletion_audit.insert_one({
        "employee_id": employee_id,
        "full_name": emp.get("full_name"),
        "action": "permanent_delete",
        "performed_by": current_user["id"],
        "performed_by_username": current_user.get("username"),
        "performed_by_role": current_user.get("role"),
        "counts": counts,
        "timestamp": get_ist_now().isoformat(),
    })
    await log_audit(current_user["id"], "permanent_delete", "employee", employee_id, f"Permanently deleted: {emp.get('full_name')}")
    return {"message": "Employee permanently deleted", "employee_id": employee_id}


@api_router.delete("/employees/{employee_id}/force")
async def employee_force_delete(
    employee_id: str,
    payload: dict = Body(default={}),
    current_user: dict = Depends(get_current_user),
):
    """DESTRUCTIVE: deletes the employee AND every linked record across all
    modules. Super Admin (system_admin) only.

    Body must include `confirmation_text: "DELETE"` to proceed — mirrors the
    forced-input safety gate enforced by the UI.
    """
    if current_user["role"] != UserRole.SYSTEM_ADMIN:
        raise HTTPException(status_code=403, detail="Force delete is restricted to Super Admin")

    if (payload or {}).get("confirmation_text", "").strip() != "DELETE":
        raise HTTPException(status_code=400, detail='Force delete requires confirmation_text="DELETE"')

    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    counts_before = await _employee_record_counts(employee_id)

    deleted_per_collection: dict = {}
    for coll in EMPLOYEE_LINKED_COLLECTIONS:
        try:
            res = await db[coll].delete_many({"employee_id": employee_id})
            deleted_per_collection[coll] = res.deleted_count
        except Exception as e:
            deleted_per_collection[coll] = f"error:{e}"

    await db.employees.delete_one({"id": employee_id})
    await db.users.delete_many({"employee_id": employee_id})

    await db.employee_deletion_audit.insert_one({
        "employee_id": employee_id,
        "full_name": emp.get("full_name"),
        "action": "force_delete",
        "performed_by": current_user["id"],
        "performed_by_username": current_user.get("username"),
        "performed_by_role": current_user.get("role"),
        "counts_before": counts_before,
        "deleted_per_collection": deleted_per_collection,
        "timestamp": get_ist_now().isoformat(),
    })
    await log_audit(
        current_user["id"], "force_delete", "employee", employee_id,
        f"FORCE deleted: {emp.get('full_name')} | total_records_destroyed={counts_before['total']}",
    )
    return {
        "message": "Employee and all related records permanently deleted",
        "employee_id": employee_id,
        "counts_before": counts_before,
        "deleted_per_collection": deleted_per_collection,
    }


# ============== EMPLOYEE BULK DEACTIVATE ==============

INACTIVE_IMPORT_COLUMNS = [
    "Email", "Inactive Date", "Reason", "Type", "Last Day Payable"
]

INACTIVE_COLUMN_ALIASES: dict[str, str] = {
    # email
    "Email": "email",
    "Emp Mail ID": "email",
    "Employee Email": "email",
    "Email ID": "email",
    "Mail": "email",
    # inactive date
    "Inactive Date": "inactive_date",
    "Inactive_Date": "inactive_date",
    "Relived_date": "inactive_date",
    "Relieved Date": "inactive_date",
    "Relieving Date": "inactive_date",
    "Last Working Day": "inactive_date",
    "Termination Date": "inactive_date",
    "Date": "inactive_date",
    # type (Relieved / Terminated)
    "Type": "inactive_type",
    "Inactive Type": "inactive_type",
    "Inactive_Type": "inactive_type",
    "Status Type": "inactive_type",
    "Action": "inactive_type",
    # reason
    "Reason": "reason",
    "Remarks": "reason",
    "Notes": "reason",
    "Comments": "reason",
    # last day payable
    "Last Day Payable": "last_day_payable",
    "Last_day_payable": "last_day_payable",
    "Last_Day_Payable": "last_day_payable",
    "LDP": "last_day_payable",
    "Payable": "last_day_payable",
}


def _build_inactive_alias_index() -> dict[str, str]:
    return {_normalize_header(k): v for k, v in INACTIVE_COLUMN_ALIASES.items()}


def _normalize_inactive_type(value) -> str:
    """Map Type column → 'Relieved' or 'Terminated'. Default 'Terminated'."""
    if value in (None, ""):
        return "Terminated"
    s = str(value).strip().lower()
    if s in ("relieved", "relived", "resigned", "resignation", "retired", "retirement"):
        return "Relieved"
    if s in ("terminated", "termination", "fired", "dismissed"):
        return "Terminated"
    return "Terminated"


def _normalize_payable(value) -> bool:
    """Map Last Day Payable cell → bool. 1/yes/true → True; 0/no/false/blank → False."""
    if value in (None, ""):
        return False
    s = str(value).strip().lower()
    if s in ("1", "yes", "y", "true", "t"):
        return True
    return False


@api_router.get("/employees/bulk-deactivate/template")
async def download_inactive_template(current_user: dict = Depends(get_current_user)):
    """Download styled .xlsx template for bulk-deactivation."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inactive Employees"
    header_fill = openpyxl.styles.PatternFill(start_color="063c88", end_color="063c88", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(INACTIVE_IMPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    sample = [
        ["employee@blubridge.com", "2026-04-22", "Resigned voluntarily", "Relieved", 0],
        ["another@blubridge.com", "22/04/2026", "Performance issues", "Terminated", 0],
    ]
    for r_idx, row_vals in enumerate(sample, start=2):
        for c_idx, v in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    widths = [32, 14, 36, 14, 18]
    for c_idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w

    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Sheet Column", "Required", "Maps To", "Notes"])
    ws2.append(["Email", "Yes", "employee_id (via email)", "Aliases: Emp Mail ID, Employee Email, Email ID, Mail"])
    ws2.append(["Inactive Date", "Yes", "inactive_date", "Aliases: Relived_date, Relieving Date, Last Working Day. Accepts YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY, or Excel datetime"])
    ws2.append(["Reason", "Yes", "inactive_reason", "Aliases: Remarks, Notes, Comments. Free text"])
    ws2.append(["Type", "No", "inactive_type", "Aliases: Inactive Type, Action. 'Relieved' or 'Terminated' (default: Terminated)"])
    ws2.append(["Last Day Payable", "No", "last_day_payable", "Aliases: LDP, Payable. 1/Yes/True → payable; 0/No/False/blank → not payable"])
    ws2.append(["", "", "", ""])
    ws2.append(["⚠️ This action will:", "", "", ""])
    ws2.append(["• Set employee_status = 'Inactive'", "", "", ""])
    ws2.append(["• Disable login (login_enabled = false, user.is_active = false)", "", "", ""])
    ws2.append(["• Decrement team member count", "", "", ""])
    ws2.append(["• Log to audit trail", "", "", ""])
    ws2.append(["• Skip employees not found by email (silent)", "", "", ""])
    ws2.append(["• Skip employees already inactive (silent)", "", "", ""])
    for col_idx in range(1, 5):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 36
    for c in range(1, 5):
        ws2.cell(1, c).font = openpyxl.styles.Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inactive_employees_template.xlsx"}
    )


@api_router.post("/employees/bulk-deactivate/preview")
async def preview_bulk_deactivate(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview detected columns + sample rows + match counts for bulk deactivation."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")
    headers = [h for h in headers if h]

    alias_index = _build_inactive_alias_index()
    column_mapping = {h: alias_index.get(_normalize_header(h)) for h in headers}
    mapped = {h: c for h, c in column_mapping.items() if c}
    ignored = [h for h, c in column_mapping.items() if not c]

    required_fields = ["email", "inactive_date", "reason"]
    detected_canonical = set(mapped.values())
    missing_required = [f for f in required_fields if f not in detected_canonical]

    # Match check
    will_deactivate = 0
    already_inactive = 0
    not_found = 0
    if not missing_required:
        for r in rows:
            row = _remap_row(r, alias_index)
            email = (str(row.get("email") or "")).strip().lower()
            if not email:
                not_found += 1
                continue
            emp = await db.employees.find_one(
                {"official_email": email, "is_deleted": {"$ne": True}},
                {"_id": 0, "id": 1, "employee_status": 1}
            )
            if not emp:
                not_found += 1
            elif emp.get("employee_status") == EmployeeStatus.INACTIVE:
                already_inactive += 1
            else:
                will_deactivate += 1

    sample = []
    for r in rows[:5]:
        sample.append({k: (v.isoformat() if isinstance(v, datetime) else (str(v) if v is not None else None)) for k, v in r.items()})

    return {
        "filename": file.filename,
        "total_rows": len(rows),
        "headers": headers,
        "column_mapping": mapped,
        "ignored_columns": ignored,
        "missing_required": missing_required,
        "ready_to_import": len(missing_required) == 0,
        "will_deactivate": will_deactivate,
        "already_inactive": already_inactive,
        "not_found": not_found,
        "sample_rows": sample,
    }


@api_router.post("/employees/bulk-deactivate")
async def bulk_deactivate_employees(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk-deactivate employees from an .xlsx/.csv file. HR-only.

    For each row:
    - Look up employee by email (case-insensitive)
    - If not found OR already inactive → skip silently (counted in summary)
    - Otherwise: set status=Inactive, login disabled, store inactive_type/date/reason/last_day_payable
    - Decrement team member count, deactivate user account, write audit log
    """
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")

    alias_index = _build_inactive_alias_index()
    headers_clean = [h for h in headers if h]
    sheet_to_canon = {h: alias_index.get(_normalize_header(h)) for h in headers_clean}
    detected_canonical = {c for c in sheet_to_canon.values() if c}
    ignored_columns = [h for h, c in sheet_to_canon.items() if not c]

    missing = []
    if "email" not in detected_canonical:
        missing.append("Email")
    if "inactive_date" not in detected_canonical:
        missing.append("Inactive Date")
    if "reason" not in detected_canonical:
        missing.append("Reason")
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing mandatory column(s): {', '.join(missing)}")

    total = len(rows)
    deactivated = 0
    skipped_not_found = 0
    skipped_already_inactive = 0
    failed = 0
    errors: List[dict] = []

    for idx, raw_row in enumerate(rows, start=2):
        row = _remap_row(raw_row, alias_index)

        email = (str(row.get("email") or "")).strip().lower()
        if not email:
            skipped_not_found += 1
            continue

        emp = await db.employees.find_one(
            {"official_email": email, "is_deleted": {"$ne": True}},
            {"_id": 0}
        )
        if not emp:
            skipped_not_found += 1
            continue
        if emp.get("employee_status") == EmployeeStatus.INACTIVE:
            skipped_already_inactive += 1
            continue

        # Parse date - prefer combined dt then plain date
        inactive_date_iso, _ = _parse_import_datetime(row.get("inactive_date")) if row.get("inactive_date") not in (None, "") else (None, None)
        if not inactive_date_iso:
            inactive_date_iso = _parse_import_date(row.get("inactive_date"))
        if not inactive_date_iso:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": f"Invalid Inactive Date: '{row.get('inactive_date')}'"})
            continue

        reason = str(row.get("reason")).strip() if row.get("reason") not in (None, "") else ""
        if not reason:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "Reason is required"})
            continue

        inactive_type = _normalize_inactive_type(row.get("inactive_type"))
        last_day_payable = _normalize_payable(row.get("last_day_payable"))

        update_data = {
            "employee_status": EmployeeStatus.INACTIVE,
            "login_enabled": False,
            "inactive_type": inactive_type,
            "inactive_date": inactive_date_iso,
            "inactive_reason": reason,
            "last_day_payable": last_day_payable,
            "deactivated_by": current_user["id"],
            "updated_at": get_ist_now().isoformat(),
        }
        await db.employees.update_one({"id": emp["id"]}, {"$set": update_data})
        if emp.get("team"):
            await db.teams.update_one({"name": emp["team"]}, {"$inc": {"member_count": -1}})
        await db.users.update_one({"employee_id": emp["id"]}, {"$set": {"is_active": False}})
        await log_audit(
            current_user["id"], "deactivate", "employee", emp["id"],
            f"Bulk-deactivated: {emp.get('full_name')} | Type: {inactive_type} | Date: {inactive_date_iso} | Reason: {reason}"
        )
        deactivated += 1

    return {
        "total": total,
        "deactivated": deactivated,
        "skipped_not_found": skipped_not_found,
        "skipped_already_inactive": skipped_already_inactive,
        "failed": failed,
        "column_mapping": {h: c for h, c in sheet_to_canon.items() if c},
        "ignored_columns": ignored_columns,
        "errors": errors,
    }

@api_router.put("/employees/{employee_id}/restore")
async def restore_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Restore soft-deleted employee"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    existing = await db.employees.find_one({"id": employee_id, "is_deleted": True}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Deleted employee not found")
    
    update_data = {
        "is_deleted": False,
        "deleted_at": None,
        "employee_status": EmployeeStatus.ACTIVE,
        "updated_at": get_ist_now().isoformat()
    }
    
    await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    
    # Update team member count
    await db.teams.update_one({"name": existing.get("team")}, {"$inc": {"member_count": 1}})
    
    await log_audit(current_user["id"], "restore", "employee", employee_id, f"Restored employee: {existing.get('full_name')}")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return serialize_doc(employee)

@api_router.put("/employees/{employee_id}/reactivate")
async def reactivate_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Reactivate an inactive employee — restores login access, preserves deactivation history."""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")

    existing = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    if existing.get("employee_status") != EmployeeStatus.INACTIVE:
        raise HTTPException(status_code=400, detail="Employee is not inactive")

    update_data = {
        "employee_status": EmployeeStatus.ACTIVE,
        "login_enabled": True,
        "updated_at": get_ist_now().isoformat()
    }

    await db.employees.update_one({"id": employee_id}, {"$set": update_data})

    # Re-enable user login
    await db.users.update_one(
        {"employee_id": employee_id},
        {"$set": {"is_active": True}}
    )

    # Update team member count
    await db.teams.update_one({"name": existing.get("team")}, {"$inc": {"member_count": 1}})

    await log_audit(current_user["id"], "reactivate", "employee", employee_id,
                    f"Reactivated employee: {existing.get('full_name')} (was {existing.get('inactive_type', 'N/A')} since {existing.get('inactive_date', 'N/A')})")

    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return serialize_doc(employee)

class AvatarUpdate(BaseModel):
    avatar_url: str
    avatar_public_id: Optional[str] = None

@api_router.put("/employees/{employee_id}/avatar")
async def update_employee_avatar(employee_id: str, data: AvatarUpdate, current_user: dict = Depends(get_current_user)):
    """Update employee avatar/photo"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    existing = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Delete old avatar from Cloudinary if exists
    old_public_id = existing.get("avatar_public_id")
    if old_public_id:
        try:
            await asyncio.to_thread(cloudinary.uploader.destroy, old_public_id, invalidate=True)
        except Exception as e:
            logger.warning(f"Failed to delete old avatar: {e}")
    
    await db.employees.update_one(
        {"id": employee_id},
        {"$set": {
            "avatar": data.avatar_url,
            "avatar_public_id": data.avatar_public_id,
            "updated_at": get_ist_now().isoformat()
        }}
    )
    
    await log_audit(current_user["id"], "update_avatar", "employee", employee_id)
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return serialize_doc(employee)

# ============== ATTENDANCE ROUTES ==============

@api_router.get("/attendance")
async def get_attendance(
    employee_name: Optional[str] = None,
    team: Optional[str] = None,
    department: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if employee_name:
        query["emp_name"] = {"$regex": employee_name, "$options": "i"}
    if team and team != "All":
        query["team"] = team
    if department and department != "All":
        query["department"] = department
    if status and status != "All":
        query["status"] = status
    
    attendance = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(5000)
    
    # Filter by date range in Python (DD-MM-YYYY strings don't sort lexicographically)
    def parse_ddmmyyyy(ds):
        try:
            parts = ds.split("-")
            return int(parts[2]) * 10000 + int(parts[1]) * 100 + int(parts[0])
        except:
            return 0

    if from_date or to_date:
        from_val = parse_ddmmyyyy(from_date) if from_date else 0
        to_val = parse_ddmmyyyy(to_date) if to_date else 99999999
        
        attendance = [a for a in attendance if from_val <= parse_ddmmyyyy(a.get("date", "")) <= to_val]
    
    # Gap-fill missing dates with "Absent" stub records (Issue 2 fix).
    # Only applies when both from_date & to_date are provided AND no specific status
    # filter is requested (an "Absent" filter would defeat its purpose). The
    # response shape (field names / order) is preserved exactly.
    if from_date and to_date and (not status or status in ("All", "", "Absent")):
        try:
            f_d = datetime.strptime(from_date, "%d-%m-%Y").date()
            t_d = datetime.strptime(to_date, "%d-%m-%Y").date()
        except (ValueError, TypeError):
            f_d = t_d = None

        if f_d and t_d and f_d <= t_d:
            # Determine candidate employees: union of those already in `attendance`
            # plus any employees matching the supplied filters (so we still
            # surface "Absent" days for filtered employees with NO records).
            emp_query = {"is_deleted": {"$ne": True}}
            if employee_name:
                emp_query["full_name"] = {"$regex": employee_name, "$options": "i"}
            if team and team != "All":
                emp_query["team"] = team
            if department and department != "All":
                emp_query["department"] = department

            employees_for_fill = await db.employees.find(
                emp_query,
                {"_id": 0, "id": 1, "full_name": 1, "team": 1, "department": 1,
                 "shift_type": 1, "date_of_joining": 1, "employee_status": 1,
                 "inactive_date": 1, "attendance_tracking_enabled": 1}
            ).to_list(5000)

            existing_keys = {(a.get("employee_id"), a.get("date")) for a in attendance}

            def to_ddmmyyyy(d):
                return d.strftime("%d-%m-%Y")

            def parse_doj(s):
                if not s:
                    return None
                # Accept ISO strings or DD-MM-YYYY
                try:
                    if "T" in s or "-" in s and len(s.split("-")[0]) == 4:
                        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
                except Exception:
                    pass
                try:
                    return datetime.strptime(s.split("T")[0], "%Y-%m-%d").date()
                except Exception:
                    pass
                try:
                    return datetime.strptime(s, "%d-%m-%Y").date()
                except Exception:
                    return None

            stubs = []
            for emp in employees_for_fill:
                if emp.get("attendance_tracking_enabled") is False:
                    continue
                doj = parse_doj(emp.get("date_of_joining"))
                inactive = parse_doj(emp.get("inactive_date")) if emp.get("employee_status") == "Inactive" else None
                cur = f_d
                while cur <= t_d:
                    if doj and cur < doj:
                        cur += timedelta(days=1)
                        continue
                    if inactive and cur > inactive:
                        cur += timedelta(days=1)
                        continue
                    key = (emp["id"], to_ddmmyyyy(cur))
                    if key not in existing_keys:
                        # Sundays are non-working — mark as "Sunday", never Absent.
                        stub_status = AttendanceStatus.SUNDAY if cur.weekday() == 6 else AttendanceStatus.ABSENT
                        stubs.append({
                            "id": f"stub-{emp['id']}-{to_ddmmyyyy(cur)}",
                            "employee_id": emp["id"],
                            "emp_name": emp.get("full_name", ""),
                            "team": emp.get("team", ""),
                            "department": emp.get("department", ""),
                            "date": to_ddmmyyyy(cur),
                            "check_in": None,
                            "check_out": None,
                            "check_in_24h": None,
                            "check_out_24h": None,
                            "total_hours": None,
                            "total_hours_decimal": None,
                            "status": stub_status,
                            "is_lop": False,
                            "lop_reason": None,
                            "shift_type": emp.get("shift_type", "General"),
                            "expected_login": None,
                            "expected_logout": None,
                            "created_at": None,
                            "_synthetic": True,
                        })
                    cur += timedelta(days=1)

            if stubs:
                # Apply same status filter to stubs if user explicitly asked for "Absent"
                if status == "Absent":
                    attendance = [s for s in stubs]
                else:
                    attendance.extend(stubs)
    
    # Sort by date descending (proper chronological order)
    def date_sort_key(a):
        try:
            parts = a.get("date", "01-01-1970").split("-")
            return int(parts[2]) * 10000 + int(parts[1]) * 100 + int(parts[0])
        except:
            return 0
    
    attendance.sort(key=date_sort_key, reverse=True)
    
    return [serialize_doc(a) for a in attendance]

@api_router.post("/attendance/check-in")
async def check_in(employee_id: str, current_user: dict = Depends(get_current_user)):
    employee = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    if not employee.get("attendance_tracking_enabled", True):
        raise HTTPException(status_code=400, detail="Attendance tracking disabled for this employee")
    
    today = get_ist_today()
    existing = await db.attendance.find_one({"employee_id": employee_id, "date": today})
    
    if existing and existing.get("check_in"):
        raise HTTPException(status_code=400, detail="Already checked in today")
    
    now = get_ist_now()
    check_in_time = now.strftime("%I:%M %p")
    check_in_24h = now.strftime("%H:%M")
    
    # Get shift timings for the employee
    shift_timings = get_shift_timings(employee)
    expected_login = shift_timings.get("login_time") if shift_timings else None
    expected_logout = shift_timings.get("logout_time") if shift_timings else None
    
    # Initial status calculation (will be finalized on check-out)
    initial_status = AttendanceStatus.LOGIN
    is_lop = False
    lop_reason = None
    
    # For fixed shifts, check if late login
    if expected_login:
        expected_mins = parse_time_24h_to_minutes(expected_login)
        actual_mins = parse_time_24h_to_minutes(check_in_24h)
        
        if actual_mins > expected_mins:
            late_mins = actual_mins - expected_mins
            initial_status = AttendanceStatus.LOSS_OF_PAY
            is_lop = True
            lop_reason = f"Late login by {late_mins} minute(s). Expected: {expected_login}, Actual: {check_in_24h}"
    
    attendance = Attendance(
        employee_id=employee_id,
        emp_name=employee["full_name"],
        team=employee["team"],
        department=employee["department"],
        date=today,
        check_in=check_in_time,
        check_in_24h=check_in_24h,
        status=initial_status,
        is_lop=is_lop,
        lop_reason=lop_reason,
        shift_type=employee.get("shift_type", "General"),
        expected_login=expected_login,
        expected_logout=expected_logout
    )
    doc = attendance.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.attendance.insert_one(doc.copy())
    
    return serialize_doc(doc)

@api_router.post("/attendance/check-out")
async def check_out(employee_id: str, current_user: dict = Depends(get_current_user)):
    today = get_ist_today()
    attendance = await db.attendance.find_one({"employee_id": employee_id, "date": today}, {"_id": 0})
    
    if not attendance:
        raise HTTPException(status_code=404, detail="No check-in found for today")
    if attendance.get("check_out"):
        raise HTTPException(status_code=400, detail="Already checked out")
    
    # Get employee for shift info
    employee = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    shift_timings = get_shift_timings(employee) if employee else None
    
    now = get_ist_now()
    check_out_time = now.strftime("%I:%M %p")
    check_out_24h = now.strftime("%H:%M")
    
    check_in_24h = attendance.get("check_in_24h")
    if not check_in_24h and attendance.get("check_in"):
        check_in_24h = parse_time_12h_to_24h(attendance.get("check_in"))
    
    # Calculate attendance status with strict LOP rules
    if shift_timings:
        status_result = calculate_attendance_status(check_in_24h, check_out_24h, shift_timings)
    else:
        # No shift timings - just mark as completed
        status_result = {
            "status": AttendanceStatus.COMPLETED,
            "is_lop": False,
            "lop_reason": None,
            "total_hours_decimal": 0.0
        }
        # Calculate hours manually
        if check_in_24h:
            in_mins = parse_time_24h_to_minutes(check_in_24h)
            out_mins = parse_time_24h_to_minutes(check_out_24h)
            if out_mins < in_mins:
                total_mins = 24 * 60 - in_mins + out_mins
            else:
                total_mins = out_mins - in_mins
            status_result["total_hours_decimal"] = total_mins / 60
    
    total_hours_str = calculate_total_hours_str(status_result.get("total_hours_decimal", 0))
    
    # If already marked as LOP from late login, keep it
    final_is_lop = attendance.get("is_lop", False) or status_result.get("is_lop", False)
    final_lop_reason = attendance.get("lop_reason") or status_result.get("lop_reason")
    final_status = AttendanceStatus.LOSS_OF_PAY if final_is_lop else status_result.get("status", AttendanceStatus.COMPLETED)
    
    update_data = {
        "check_out": check_out_time,
        "check_out_24h": check_out_24h,
        "total_hours": total_hours_str,
        "total_hours_decimal": status_result.get("total_hours_decimal", 0),
        "status": final_status,
        "is_lop": final_is_lop,
        "lop_reason": final_lop_reason
    }
    
    await db.attendance.update_one(
        {"employee_id": employee_id, "date": today},
        {"$set": update_data}
    )
    
    updated = await db.attendance.find_one({"employee_id": employee_id, "date": today}, {"_id": 0})
    return serialize_doc(updated)

# ============== BIOMETRIC ATTENDANCE IMPORT ==============

@api_router.post("/attendance/import-biometric")
async def import_biometric_attendance(
    records: list = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """Ingest biometric attendance data from external device sync service."""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    if not records or not isinstance(records, list):
        raise HTTPException(status_code=400, detail="Request body must be a non-empty JSON array")
    
    total_records = len(records)
    processed = 0
    skipped = 0
    unmapped = 0
    unmapped_ids = set()
    
    # Pre-fetch all employees with biometric_id for mapping
    emp_cursor = db.employees.find(
        {"biometric_id": {"$ne": None}, "is_deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "biometric_id": 1, "full_name": 1, "team": 1, "department": 1, "shift_type": 1,
         "custom_login_time": 1, "custom_logout_time": 1, "custom_total_hours": 1}
    )
    bio_map = {}
    async for emp in emp_cursor:
        if emp.get("biometric_id"):
            bio_map[str(emp["biometric_id"])] = emp
    
    # Group valid punches: { (employee_id, date_str) : [datetime, ...] }
    grouped = {}
    device_ips = {}
    raw_logs = []
    
    for rec in records:
        if not isinstance(rec, dict):
            skipped += 1
            continue
        
        device_user_id = rec.get("deviceUserId")
        record_time = rec.get("recordTime")
        device_ip = rec.get("ip", "")
        
        if not device_user_id or not record_time:
            skipped += 1
            continue
        
        device_user_id = str(device_user_id).strip()
        
        # Map to employee
        employee = bio_map.get(device_user_id)
        if not employee:
            unmapped += 1
            unmapped_ids.add(device_user_id)
            raw_logs.append({
                "deviceUserId": device_user_id,
                "recordTime": record_time,
                "ip": device_ip,
                "status": "unmapped"
            })
            continue
        
        # Parse recordTime
        try:
            if isinstance(record_time, str):
                # Handle ISO format with or without timezone
                punch_dt = datetime.fromisoformat(record_time.replace("Z", "+00:00"))
            else:
                skipped += 1
                continue
        except (ValueError, TypeError):
            skipped += 1
            continue
        
        # Convert to IST
        punch_ist = punch_dt.astimezone(IST)
        # Cross-midnight handling: punches <= threshold belong to previous day.
        date_str = get_effective_attendance_date(punch_ist)
        emp_id = employee["id"]
        
        key = (emp_id, date_str)
        if key not in grouped:
            grouped[key] = {"punches": [], "employee": employee, "ip": device_ip}
        grouped[key]["punches"].append(punch_ist)
        if device_ip:
            grouped[key]["ip"] = device_ip
        
        raw_logs.append({
            "deviceUserId": device_user_id,
            "recordTime": record_time,
            "ip": device_ip,
            "employee_id": emp_id,
            "date": date_str,
            "status": "mapped"
        })
    
    # Store raw punch log for audit
    if raw_logs:
        await db.biometric_punch_logs.insert_one({
            "_id": str(uuid.uuid4()),
            "imported_at": get_ist_now().isoformat(),
            "imported_by": current_user["id"],
            "total_punches": len(raw_logs),
            "logs": raw_logs
        })
    
    # Process grouped punches - atomic upsert per employee+date
    for (emp_id, date_str), data in grouped.items():
        # Dedupe by minute — biometric devices commonly re-sync the same punch
        # many times. Without dedupe a single physical punch can be incorrectly
        # treated as both IN and OUT.
        punches = sorted({p.replace(second=0, microsecond=0) for p in data["punches"]})
        employee = data["employee"]
        device_ip = data["ip"]

        # Within a batch, the true chronological first punch is IN and the
        # last is OUT (accounting for cross-midnight punches which may appear
        # as small HH:MM numerically but are actually later in shift time).
        in_time = punches[0]
        out_time = punches[-1] if len(punches) > 1 else None

        in_24h = in_time.strftime("%H:%M")
        in_12h = in_time.strftime("%I:%M %p")
        out_24h = out_time.strftime("%H:%M") if out_time else None
        out_12h = out_time.strftime("%I:%M %p") if out_time else None

        # Check existing record for this employee+date
        existing = await db.attendance.find_one(
            {"employee_id": emp_id, "date": date_str},
            {"_id": 0, "check_in_24h": 1, "check_out_24h": 1}
        )

        # Merge with existing record using shift-aware offsets so that
        # cross-midnight OUT punches (00:00-05:00) are correctly recognized
        # as the latest point in the shift, not as the earliest IN.
        if existing:
            existing_in = existing.get("check_in_24h")
            existing_out = existing.get("check_out_24h")

            candidate_times = []
            for t in (in_24h, out_24h, existing_in, existing_out):
                if t:
                    off = attendance_shift_offset(t)
                    if off is not None:
                        candidate_times.append((t, off))

            if candidate_times:
                # Deduplicate identical time strings (keep one entry per value)
                seen = {}
                for t, off in candidate_times:
                    seen[t] = off
                uniq = list(seen.items())
                # IN = earliest offset; OUT = latest offset (only if distinct)
                in_pick = min(uniq, key=lambda c: c[1])
                out_pick = max(uniq, key=lambda c: c[1])
                in_24h = in_pick[0]
                out_24h = out_pick[0] if out_pick[0] != in_pick[0] else None
                in_12h = None
                out_12h = None

        # Recalculate 12h from final 24h values
        if in_24h and not in_12h:
            t = datetime.strptime(in_24h, "%H:%M")
            in_12h = t.strftime("%I:%M %p")
        if out_24h and not out_12h:
            t = datetime.strptime(out_24h, "%H:%M")
            out_12h = t.strftime("%I:%M %p")

        # Calculate total hours and status
        total_hours_decimal = 0.0
        total_hours_str = None
        status = AttendanceStatus.LOGIN
        is_lop = False
        lop_reason = None
        dynamic_expected_logout = None

        shift_timings = get_shift_timings(employee)

        # Sunday handling
        if is_sunday_ddmmyyyy(date_str):
            if in_24h and out_24h:
                in_m = parse_time_24h_to_minutes(in_24h)
                out_m = parse_time_24h_to_minutes(out_24h)
                total_hours_decimal = round((out_m - in_m) / 60, 2) if out_m > in_m else round((24 * 60 - in_m + out_m) / 60, 2)
                total_hours_str = calculate_total_hours_str(total_hours_decimal)
            status = AttendanceStatus.WORKED_ON_SUNDAY if in_24h else AttendanceStatus.SUNDAY
            is_lop = False
            lop_reason = None
        elif in_24h and out_24h:
            in_mins = parse_time_24h_to_minutes(in_24h)
            out_mins = parse_time_24h_to_minutes(out_24h)
            if out_mins > in_mins:
                total_hours_decimal = round((out_mins - in_mins) / 60, 2)
            else:
                # Cross-midnight: OUT is on next calendar day
                total_hours_decimal = round((24 * 60 - in_mins + out_mins) / 60, 2)
            total_hours_str = calculate_total_hours_str(total_hours_decimal)
            status = AttendanceStatus.PRESENT

            if shift_timings:
                status_result = calculate_attendance_status(in_24h, out_24h, shift_timings, attendance_date=date_str)
                status = status_result.get("status", status)
                is_lop = status_result.get("is_lop", False)
                lop_reason = status_result.get("lop_reason")
                dynamic_expected_logout = status_result.get("expected_logout")
                if status_result.get("total_hours_decimal"):
                    total_hours_decimal = status_result["total_hours_decimal"]
                    total_hours_str = calculate_total_hours_str(total_hours_decimal)
        elif in_24h and not out_24h:
            # Single punch - no out time. Compute dynamic expected_logout = IN + required_hours
            if shift_timings:
                req_h = shift_timings.get("total_hours", 8)
                dynamic_expected_logout = add_hours_to_24h(in_24h, req_h)
                expected_login = shift_timings.get("login_time")
                if expected_login:
                    expected_mins = parse_time_24h_to_minutes(expected_login)
                    actual_mins = parse_time_24h_to_minutes(in_24h)
                    if actual_mins > expected_mins:
                        late_mins = actual_mins - expected_mins
                        is_lop = True
                        lop_reason = f"Late login by {late_mins} minute(s). Expected: {expected_login}, Actual: {in_24h}"
                        status = AttendanceStatus.LOSS_OF_PAY
        
        # Atomic upsert: update_one with upsert=True prevents duplicates
        update_doc = {
            "$set": {
                "check_in": in_12h,
                "check_in_24h": in_24h,
                "check_out": out_12h,
                "check_out_24h": out_24h,
                "total_hours": total_hours_str,
                "total_hours_decimal": total_hours_decimal,
                "status": status,
                "is_lop": is_lop,
                "lop_reason": lop_reason,
                "source": "biometric",
                "device_ip": device_ip,
                # Dynamic expected_logout (punch_in + required_hours) overrides
                # any static shift logout previously stored. Falls back to the
                # shift's logout_time when no IN punch yet.
                "expected_logout": dynamic_expected_logout if dynamic_expected_logout else (shift_timings.get("logout_time") if shift_timings else None),
            },
            "$setOnInsert": {
                "id": str(uuid.uuid4()),
                "employee_id": emp_id,
                "emp_name": employee["full_name"],
                "team": employee.get("team", ""),
                "department": employee.get("department", ""),
                "date": date_str,
                "shift_type": employee.get("shift_type", "General"),
                "expected_login": shift_timings.get("login_time") if shift_timings else None,
                "created_at": get_ist_now().isoformat()
            }
        }
        
        await db.attendance.update_one(
            {"employee_id": emp_id, "date": date_str},
            update_doc,
            upsert=True
        )
        processed += 1
    
    await log_audit(
        current_user["id"], "biometric_import", "attendance", None,
        f"Biometric import: {processed} processed, {skipped} skipped, {unmapped} unmapped out of {total_records}"
    )
    
    return {
        "totalRecords": total_records,
        "processed": processed,
        "skipped": skipped,
        "unmapped": unmapped,
        "unmappedDeviceUserIds": list(unmapped_ids) if unmapped_ids else []
    }

@api_router.post("/attendance/recompute-from-punches")
async def recompute_attendance_from_punches(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Re-derive check_in (MIN) and check_out (MAX) for each (employee, date) by
    scanning raw biometric_punch_logs. Useful when historical attendance rows
    show only an IN time even though multiple punches exist on the device.
    Single-punch days keep check_out = NULL (do NOT duplicate IN as OUT).
    Does NOT change schema, response shape, or status calculation rules."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN, UserRole.OFFICE_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    # Aggregate raw logs into (emp_id, date) -> [punch_24h, ...]
    match_stage = {"logs.status": "mapped"}
    if employee_id:
        match_stage["logs.employee_id"] = employee_id

    pipeline = [
        {"$unwind": "$logs"},
        {"$match": match_stage},
        {"$group": {
            "_id": {"employee_id": "$logs.employee_id", "date": "$logs.date"},
            "punches": {"$push": "$logs.recordTime"},
        }},
    ]
    cursor = db.biometric_punch_logs.aggregate(pipeline, allowDiskUse=True)

    def _parse(rt):
        if not isinstance(rt, str):
            return None
        try:
            dt = datetime.fromisoformat(rt.replace("Z", "+00:00"))
            return dt.astimezone(IST)
        except Exception:
            return None

    def parse_ddmmyyyy(ds):
        try:
            parts = ds.split("-")
            return int(parts[2]) * 10000 + int(parts[1]) * 100 + int(parts[0])
        except Exception:
            return 0

    f_val = parse_ddmmyyyy(from_date) if from_date else 0
    t_val = parse_ddmmyyyy(to_date) if to_date else 99999999

    # Pre-fetch employee data for shift recalculation
    emp_cache = {}

    updated = 0
    skipped_locked = 0
    no_change = 0
    examined = 0

    async for grp in cursor:
        emp_id = grp["_id"]["employee_id"]
        date_str = grp["_id"]["date"]
        if from_date or to_date:
            v = parse_ddmmyyyy(date_str)
            if not (f_val <= v <= t_val):
                continue
        examined += 1

        parsed = [p for p in (_parse(rt) for rt in grp.get("punches", [])) if p]
        # Dedupe by minute (devices commonly re-sync the same punch many times).
        unique = sorted({p.replace(second=0, microsecond=0) for p in parsed})
        if not unique:
            continue
        in_dt = unique[0]
        out_dt = unique[-1] if len(unique) > 1 else None

        in_24h = in_dt.strftime("%H:%M")
        in_12h = in_dt.strftime("%I:%M %p")
        out_24h = out_dt.strftime("%H:%M") if out_dt else None
        out_12h = out_dt.strftime("%I:%M %p") if out_dt else None

        existing = await db.attendance.find_one(
            {"employee_id": emp_id, "date": date_str},
            {"_id": 0}
        )

        # Respect manual overrides / approved corrections
        if existing and (existing.get("is_manual_override") or existing.get("is_approved_correction")):
            skipped_locked += 1
            continue

        # No-op short-circuit
        if existing and existing.get("check_in_24h") == in_24h and existing.get("check_out_24h") == out_24h:
            no_change += 1
            continue

        # Compute total hours and status
        total_hours_decimal = 0.0
        total_hours_str = None
        status = AttendanceStatus.LOGIN
        is_lop = False
        lop_reason = None
        dynamic_expected_logout = None

        if emp_id not in emp_cache:
            emp_cache[emp_id] = await db.employees.find_one({"id": emp_id}, {"_id": 0}) or {}
        employee = emp_cache[emp_id]
        shift_timings = get_shift_timings(employee) if employee else None

        # Sunday handling: never LOP, mark "Sunday" / "Worked on Sunday"
        if is_sunday_ddmmyyyy(date_str):
            if in_24h and out_24h:
                im = parse_time_24h_to_minutes(in_24h)
                om = parse_time_24h_to_minutes(out_24h)
                total_hours_decimal = round((om - im) / 60, 2) if om > im else round((24 * 60 - im + om) / 60, 2)
                total_hours_str = calculate_total_hours_str(total_hours_decimal)
            status = AttendanceStatus.WORKED_ON_SUNDAY if in_24h else AttendanceStatus.SUNDAY
            is_lop = False
            lop_reason = None
        elif in_24h and out_24h:
            in_mins = parse_time_24h_to_minutes(in_24h)
            out_mins = parse_time_24h_to_minutes(out_24h)
            if out_mins > in_mins:
                total_hours_decimal = round((out_mins - in_mins) / 60, 2)
            else:
                total_hours_decimal = round((24 * 60 - in_mins + out_mins) / 60, 2)
            total_hours_str = calculate_total_hours_str(total_hours_decimal)
            status = AttendanceStatus.PRESENT
            if shift_timings:
                sr = calculate_attendance_status(in_24h, out_24h, shift_timings, attendance_date=date_str)
                status = sr.get("status", status)
                is_lop = sr.get("is_lop", False)
                lop_reason = sr.get("lop_reason")
                dynamic_expected_logout = sr.get("expected_logout")
                if sr.get("total_hours_decimal"):
                    total_hours_decimal = sr["total_hours_decimal"]
                    total_hours_str = calculate_total_hours_str(total_hours_decimal)
        elif in_24h and not out_24h:
            if shift_timings:
                req_h = shift_timings.get("total_hours", 8)
                dynamic_expected_logout = add_hours_to_24h(in_24h, req_h)
                expected_login = shift_timings.get("login_time")
                if expected_login:
                    em = parse_time_24h_to_minutes(expected_login)
                    am = parse_time_24h_to_minutes(in_24h)
                    if am > em:
                        is_lop = True
                        lop_reason = f"Late login by {am - em} minute(s). Expected: {expected_login}, Actual: {in_24h}"
                        status = AttendanceStatus.LOSS_OF_PAY

        update_doc = {
            "$set": {
                "check_in": in_12h,
                "check_in_24h": in_24h,
                "check_out": out_12h,
                "check_out_24h": out_24h,
                "total_hours": total_hours_str,
                "total_hours_decimal": total_hours_decimal,
                "status": status,
                "is_lop": is_lop,
                "lop_reason": lop_reason,
                "source": "biometric",
                "expected_logout": dynamic_expected_logout if dynamic_expected_logout else (shift_timings.get("logout_time") if shift_timings else None),
            },
            "$setOnInsert": {
                "id": str(uuid.uuid4()),
                "employee_id": emp_id,
                "emp_name": (employee or {}).get("full_name", ""),
                "team": (employee or {}).get("team", ""),
                "department": (employee or {}).get("department", ""),
                "date": date_str,
                "shift_type": (employee or {}).get("shift_type", "General"),
                "expected_login": shift_timings.get("login_time") if shift_timings else None,
                "created_at": get_ist_now().isoformat(),
            }
        }
        await db.attendance.update_one(
            {"employee_id": emp_id, "date": date_str},
            update_doc,
            upsert=True,
        )
        updated += 1

    await log_audit(
        current_user["id"], "recompute_attendance_in_out", "attendance", None,
        f"Recompute IN/OUT from punches: examined={examined}, updated={updated}, unchanged={no_change}, locked={skipped_locked}"
    )

    return {
        "examined": examined,
        "updated": updated,
        "unchanged": no_change,
        "skipped_locked": skipped_locked,
    }

def classify_attendance_bucket(rec: dict) -> str:
    """Classify attendance record into a STRICTLY mutually exclusive bucket
    that drives the Dashboard "Today's Attendance Status" tiles.

    Buckets (strict, mutually exclusive):
      - 'logged_in'  → has IN, no OUT (still working)
      - 'completed'  → has IN + OUT, full hours (no LOP, no Early Out flag)
      - 'early_out'  → has IN + OUT, hours short (Early Out / Loss of Pay)
      - 'no_login'   → no IN at all (synthetic Absent / Sunday / no record)

    'Late Login' is a SECONDARY flag (see is_late_login_record) and overlays
    on top of these buckets — it is NOT a primary category here.
    """
    has_in = bool(rec.get("check_in") or rec.get("check_in_24h"))
    has_out = bool(rec.get("check_out") or rec.get("check_out_24h"))
    status = (rec.get("status") or "").strip()
    is_lop = bool(rec.get("is_lop"))

    if not has_in:
        return "no_login"
    if not has_out:
        return "logged_in"
    # Has both IN + OUT — split by short-hours / LOP
    if status in (AttendanceStatus.EARLY_OUT, AttendanceStatus.LOSS_OF_PAY) or is_lop:
        return "early_out"
    return "completed"


def is_late_login_record(rec: dict) -> bool:
    """Secondary flag — true if the record represents a late-login day,
    regardless of whether the employee subsequently completed the day."""
    if not rec:
        return False
    if (rec.get("status") or "").strip() == AttendanceStatus.LATE_LOGIN:
        return True
    if "late login" in (rec.get("lop_reason") or "").lower():
        return True
    return False


@api_router.get("/attendance/stats")
async def get_attendance_stats(
    date: Optional[str] = None, 
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # Use single date or date range
    if not date and not from_date:
        date = get_ist_today()
    
    # Only count active employees with attendance tracking enabled
    total_employees = await db.employees.count_documents({
        "employee_status": EmployeeStatus.ACTIVE,
        "is_deleted": {"$ne": True},
        "attendance_tracking_enabled": True
    })
    
    # Helper to parse DD-MM-YYYY to sortable integer
    def parse_ddmmyyyy(ds):
        try:
            parts = ds.split("-")
            return int(parts[2]) * 10000 + int(parts[1]) * 100 + int(parts[0])
        except:
            return 0
    
    # For date range, fetch all and filter in Python (DD-MM-YYYY strings don't sort lexicographically in MongoDB)
    if from_date and to_date:
        all_records = await db.attendance.find({}, {"_id": 0}).to_list(10000)
        from_val = parse_ddmmyyyy(from_date)
        to_val = parse_ddmmyyyy(to_date)
        filtered = [a for a in all_records if from_val <= parse_ddmmyyyy(a.get("date", "")) <= to_val]
    elif date:
        filtered = await db.attendance.find({"date": date}, {"_id": 0}).to_list(10000)
    else:
        filtered = []
    
    # Strict mutually-exclusive bucketing — single source of truth.
    # The frontend tile clicks fetch the same records and apply the same
    # classification so counts and detail lists ALWAYS line up.
    logged_in = 0
    completed = 0
    early_out = 0
    late_login = 0
    lop_count = 0
    employees_with_in = set()

    for a in filtered:
        bucket = classify_attendance_bucket(a)
        if bucket == "logged_in":
            logged_in += 1
            employees_with_in.add(a.get("employee_id"))
        elif bucket == "completed":
            completed += 1
            employees_with_in.add(a.get("employee_id"))
        elif bucket == "early_out":
            early_out += 1
            employees_with_in.add(a.get("employee_id"))
        if is_late_login_record(a):
            late_login += 1
        if a.get("is_lop"):
            lop_count += 1

    # "Leaves / No Login" count — MUST equal the size of /dashboard/leave-list
    # so the tile number and the table row count always line up.
    #   = |employees on approved/pending leave in window ∪ employees without any IN punch|
    # Leaves are stored as YYYY-MM-DD; attendance/filter dates are DD-MM-YYYY.
    def _dmy_to_ymd(ds: Optional[str]) -> Optional[str]:
        if not ds:
            return None
        try:
            return datetime.strptime(ds, "%d-%m-%Y").strftime("%Y-%m-%d")
        except ValueError:
            return ds

    q_from = from_date or date
    q_to = to_date or date
    ymd_from = _dmy_to_ymd(q_from)
    ymd_to = _dmy_to_ymd(q_to)

    # Resolve the SAME cohort used by /dashboard/leave-list so counts align.
    active_tracking_emps = await db.employees.find(
        {
            "employee_status": EmployeeStatus.ACTIVE,
            "is_deleted": {"$ne": True},
            "attendance_tracking_enabled": True,
        },
        {"_id": 0, "id": 1},
    ).to_list(5000)
    active_ids = {e["id"] for e in active_tracking_emps}

    # Restrict "with IN" to the active+tracking cohort (employees_with_in can
    # include inactive employees whose attendance records still live on).
    active_with_in = employees_with_in & active_ids

    on_leave_ids: set = set()
    if ymd_from and ymd_to:
        leave_docs = await db.leaves.find(
            {
                "status": {"$in": ["approved", "pending"]},
                "start_date": {"$lte": ymd_to},
                "end_date": {"$gte": ymd_from},
            },
            {"_id": 0, "employee_id": 1},
        ).to_list(5000)
        on_leave_ids = {lv.get("employee_id") for lv in leave_docs if lv.get("employee_id")}
    on_leave_ids &= active_ids

    no_in_ids = active_ids - active_with_in
    # Union: employees on leave OR with no IN punch
    not_logged = len(on_leave_ids | no_in_ids)

    return {
        "total_employees": total_employees,
        "logged_in": logged_in,
        "not_logged": not_logged,
        "early_out": early_out,
        "late_login": late_login,
        "logout": completed,  # frontend "Completed" tile reads `logout`
        "lop_count": lop_count,
        "present": completed,
    }

# ============== LEAVE ROUTES ==============

@api_router.get("/leaves")
async def get_leaves(
    employee_name: Optional[str] = None,
    team: Optional[str] = None,
    department: Optional[str] = None,
    leave_type: Optional[str] = None,
    status: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if employee_name:
        query["emp_name"] = {"$regex": employee_name, "$options": "i"}
    if team and team != "All":
        query["team"] = team
    if department and department != "All":
        query["department"] = department
    if leave_type and leave_type != "All":
        query["leave_type"] = leave_type
    if status and status != "All":
        query["status"] = status
    if from_date:
        query["start_date"] = {"$gte": from_date}
    if to_date:
        if "start_date" in query:
            query["start_date"]["$lte"] = to_date
        else:
            query["start_date"] = {"$lte": to_date}
    
    leaves = await db.leaves.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [serialize_doc(l) for l in leaves]

@api_router.post("/leaves")
async def create_leave(data: LeaveRequestCreate, current_user: dict = Depends(get_current_user)):
    employee = await db.employees.find_one({"id": data.employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    start = datetime.strptime(data.start_date, "%Y-%m-%d")
    end = datetime.strptime(data.end_date, "%Y-%m-%d")
    
    # JOB 7: Single leave per day check
    current_date = start
    while current_date <= end:
        date_str = current_date.strftime("%Y-%m-%d")
        existing = await db.leaves.find_one({
            "employee_id": data.employee_id,
            "status": {"$ne": "rejected"},
            "start_date": {"$lte": date_str},
            "end_date": {"$gte": date_str}
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"Leave request already exists for {date_str}")
        current_date += timedelta(days=1)
    
    start = datetime.strptime(data.start_date, "%Y-%m-%d")
    end = datetime.strptime(data.end_date, "%Y-%m-%d")
    
    # Calculate duration based on leave_split
    if data.leave_split in ["First Half", "Second Half"]:
        duration_str = "0.5 day(s)"
    else:
        duration = (end - start).days + 1
        duration_str = f"{duration} day(s)"
    
    is_admin = current_user["role"] in ALL_ADMIN_ROLES
    
    leave = LeaveRequest(
        employee_id=data.employee_id,
        emp_name=employee["full_name"],
        team=employee["team"],
        department=employee["department"],
        leave_type=data.leave_type,
        leave_split=data.leave_split,
        start_date=data.start_date,
        end_date=data.end_date,
        duration=duration_str,
        reason=data.reason,
        supporting_document_url=data.supporting_document_url,
        supporting_document_name=data.supporting_document_name,
        applied_by_admin=is_admin,
        status="approved" if (data.auto_approve and is_admin) else "pending",
        is_lop=data.is_lop if (data.auto_approve and is_admin) else None,
        approved_by=current_user["id"] if (data.auto_approve and is_admin) else None
    )
    doc = leave.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.leaves.insert_one(doc.copy())
    
    await log_audit(current_user["id"], "create", "leave", leave.id)
    return serialize_doc(doc)

# ============== LEAVE BULK IMPORT ==============

LEAVE_IMPORT_COLUMNS = [
    "Emp Mail ID", "Leave Type", "From Date", "To Date",
    "Leave Split", "Reason", "Status", "Applied Date", "Approved By",
    "Approval Date", "Comments", "Auto Approve & Set LOP Status"
]

# Sheet-column aliases → canonical internal field name. Matching is case-insensitive
# and whitespace-tolerant (see _normalize_header).
# Multiple sheet headers can map to the same canonical field — first match wins.
LEAVE_COLUMN_ALIASES: dict[str, str] = {
    # email
    "Employee Email": "email",
    "Emp Mail ID": "email",
    "Email": "email",
    "Email ID": "email",
    "Mail": "email",
    # leave type
    "Leave Type": "leave_type",
    "Type": "leave_type",
    "Leave Category": "leave_type",
    # dates
    "From Date": "start_date",
    "Start Date": "start_date",
    "Leave From": "start_date",
    "To Date": "end_date",
    "End Date": "end_date",
    "Leave To": "end_date",
    # leave split (replaces No of Days — system auto-calculates days)
    "Leave Split": "leave_split",
    "Split": "leave_split",
    "Day Type": "leave_split",
    # reason
    "Reason": "reason",
    "Purpose": "reason",
    # status
    "Status": "status",
    "Leave Status": "status",
    # applied at
    "Applied Date": "applied_at",
    "Applied On": "applied_at",
    "Date Applied": "applied_at",
    # approved by
    "Approved By": "approved_by",
    "Approver": "approved_by",
    "Approving Authority": "approved_by",
    # approval date
    "Approval Date": "approved_at",
    "Approved On": "approved_at",
    "Date Approved": "approved_at",
    # comments
    "Comments": "comments",
    "Remark": "comments",
    "Remarks": "comments",
    "Notes": "comments",
    # auto-approve & LOP
    "Auto Approve & Set LOP Status": "auto_approve_lop",
    "Auto Approve": "auto_approve_lop",
    "Auto Approve LOP": "auto_approve_lop",
}


def _normalize_header(s) -> str:
    """Lowercase + collapse internal whitespace + strip for tolerant matching."""
    if s is None:
        return ""
    return " ".join(str(s).strip().split()).lower()


def _build_alias_index() -> dict[str, str]:
    """Lowercased+trimmed alias → canonical field key."""
    return {_normalize_header(k): v for k, v in LEAVE_COLUMN_ALIASES.items()}


def _remap_row(row: dict, alias_index: dict[str, str]) -> dict:
    """Return a dict keyed by canonical field names. Unknown columns are dropped.

    If multiple sheet columns map to the same canonical field, the first non-empty
    value wins.
    """
    out: dict = {}
    for header, value in row.items():
        if header is None:
            continue
        canon = alias_index.get(_normalize_header(header))
        if not canon:
            continue
        if canon in out and out[canon] not in (None, ""):
            continue  # first non-empty value wins
        out[canon] = value
    return out


# Canonical leave types used in the system. Anything not matching falls back to "General Leave".
# "Optional" is intentionally a constraint-free type — see create_leave for behaviour.
CANONICAL_LEAVE_TYPES = [
    "Sick", "Casual", "Earned", "Maternity", "Annual",
    "Emergency", "Preplanned", "Optional", "General Leave"
]


def _normalize_leave_type(raw) -> str:
    """Map an arbitrary leave-type string to a canonical value via case-insensitive
    fuzzy match. Falls back to "General Leave" if no reasonable match is found."""
    if raw is None:
        return "General Leave"
    s = str(raw).strip()
    if not s:
        return "General Leave"
    # Strip common suffixes like "(SL)", "(CL)" etc.
    cleaned = s.split("(")[0].strip()
    cleaned_lower = cleaned.lower()

    # Direct case-insensitive match (handles "sick", "Sick Leave", "sl" etc.)
    for canon in CANONICAL_LEAVE_TYPES:
        if cleaned_lower == canon.lower():
            return canon
    # "Sick Leave" → starts with "Sick"
    for canon in CANONICAL_LEAVE_TYPES:
        if cleaned_lower.startswith(canon.lower()) or canon.lower() in cleaned_lower:
            return canon
    # Fuzzy ratio
    from difflib import SequenceMatcher
    best = ("General Leave", 0.0)
    for canon in CANONICAL_LEAVE_TYPES:
        ratio = SequenceMatcher(None, cleaned_lower, canon.lower()).ratio()
        if ratio > best[1]:
            best = (canon, ratio)
    return best[0] if best[1] >= 0.7 else "General Leave"


def _parse_import_date(value) -> Optional[str]:
    """Parse a date cell (datetime, date, or string in DD-MM-YYYY / DD/MM/YYYY /
    YYYY-MM-DD) into canonical "YYYY-MM-DD". Returns None if invalid."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _normalize_status(value) -> Optional[str]:
    """Normalize Status to one of approved/pending/rejected. Returns None if blank."""
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s:
        return None
    if s in ("approved", "approve", "accept", "accepted"):
        return "approved"
    if s in ("pending", "open", "new", "in progress", "in-progress"):
        return "pending"
    if s in ("rejected", "reject", "denied", "declined"):
        return "rejected"
    return None  # unrecognized


def _normalize_leave_split(value) -> Optional[str]:
    """Normalize Leave Split → 'Full Day' | 'First Half'. Returns None when blank,
    or 'INVALID' sentinel when the value cannot be recognized so caller can reject."""
    if value in (None, ""):
        return None
    s = str(value).strip().lower()
    if s in ("full day", "full", "fullday", "full_day", "f"):
        return "Full Day"
    if s in ("half day", "half", "halfday", "half_day", "h", "first half", "1st half"):
        return "First Half"
    if s in ("second half", "2nd half"):
        return "Second Half"
    return "INVALID"


def _normalize_bool(value) -> Optional[bool]:
    """Convert an input cell to bool. Returns None when blank, 'INVALID' sentinel
    string when the value can't be recognized."""
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value in (0, 1):
            return bool(value)
        return "INVALID"
    s = str(value).strip().lower()
    if s in ("yes", "y", "true", "t", "1"):
        return True
    if s in ("no", "n", "false", "f", "0"):
        return False
    return "INVALID"


def _read_leave_import_rows(file_bytes: bytes, filename: str) -> tuple[List[str], List[dict]]:
    """Read .xlsx or .csv into (headers, rows). Each row is a dict header→cell."""
    name_lower = (filename or "").lower()
    rows: List[dict] = []
    if name_lower.endswith(".xlsx") or name_lower.endswith(".xlsm"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else "" for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        return headers, rows
    if name_lower.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = [h.strip() for h in (reader.fieldnames or [])]
        for raw in reader:
            row = {(k or "").strip(): (v.strip() if isinstance(v, str) else v) for k, v in raw.items()}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        return headers, rows
    raise HTTPException(status_code=400, detail="Unsupported file format. Use .xlsx or .csv")


@api_router.get("/leaves/import-template")
async def download_leave_import_template(current_user: dict = Depends(get_current_user)):
    """Download an .xlsx template for bulk leave import."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leaves Import"
    header_fill = openpyxl.styles.PatternFill(start_color="063c88", end_color="063c88", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(LEAVE_IMPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    sample = [
        ["employee@blubridge.com", "Sick", "2026-04-10", "2026-04-12", "Full Day", "Fever", "Approved", "2026-04-09", "Admin", "2026-04-09", "Auto-approved by HR", "No"],
        ["another@blubridge.com", "Casual Leave (CL)", "10/04/2026", "10/04/2026", "Half Day", "Personal", "Pending", "2026-04-09", "", "", "", "Yes"],
    ]
    for r_idx, row_vals in enumerate(sample, start=2):
        for c_idx, v in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    widths = [30, 18, 14, 14, 14, 30, 12, 14, 18, 14, 30, 28]
    for c_idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Sheet Column", "Required", "Maps To", "Notes"])
    ws2.append(["Emp Mail ID", "Yes", "email", "Aliases: Employee Email, Email, Email ID, Mail"])
    ws2.append(["Leave Type", "Yes", "leave_type", "Aliases: Type, Leave Category. Sick/Casual/Earned/Maternity/Annual/Emergency/Preplanned. Anything else → 'General Leave'."])
    ws2.append(["From Date", "Yes", "start_date", "Aliases: Start Date, Leave From. Accepts YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY"])
    ws2.append(["To Date", "Yes", "end_date", "Aliases: End Date, Leave To. Must be >= From Date"])
    ws2.append(["Leave Split", "No", "leave_split", "Aliases: Split, Day Type. Accepted: 'Full Day' or 'Half Day' (default Full Day). System auto-calculates total days."])
    ws2.append(["Reason", "No", "reason", "Aliases: Purpose. Free text"])
    ws2.append(["Status", "No", "status", "Aliases: Leave Status. Approved / Pending / Rejected (default: Pending)"])
    ws2.append(["Applied Date", "No", "applied_at", "Aliases: Applied On, Date Applied. Stored in extra_data"])
    ws2.append(["Approved By", "No", "approved_by", "Aliases: Approver, Approving Authority. 'Admin' or admin's username/email/full name → resolved to user ID"])
    ws2.append(["Approval Date", "No", "approved_at", "Aliases: Approved On, Date Approved. Stored in extra_data"])
    ws2.append(["Comments", "No", "comments", "Aliases: Remark, Remarks, Notes. Stored as the leave's lop_remark"])
    ws2.append(["Auto Approve & Set LOP Status", "No", "auto_approve_lop", "Aliases: Auto Approve, Auto Approve LOP. Yes/No, TRUE/FALSE, 1/0. If TRUE → force-approve + set LOP flag."])
    for col_idx in range(1, 5):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 36
    for c in range(1, 5):
        ws2.cell(1, c).font = openpyxl.styles.Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leaves_import_template.xlsx"}
    )


@api_router.post("/leaves/import/preview")
async def preview_leave_import(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview detected columns + first 5 sample rows from an import file.

    Returns the alias mapping (sheet column → DB field) so admin can verify
    what will happen before committing the import.
    """
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")
    headers = [h for h in headers if h]

    alias_index = _build_alias_index()
    column_mapping = {h: alias_index.get(_normalize_header(h)) for h in headers}
    mapped = {h: c for h, c in column_mapping.items() if c}
    ignored = [h for h, c in column_mapping.items() if not c]

    required_fields = ["email", "leave_type", "start_date", "end_date"]
    detected_canonical = set(mapped.values())
    missing_required = [f for f in required_fields if f not in detected_canonical]

    sample = []
    for r in rows[:5]:
        sample.append({k: (v.isoformat() if isinstance(v, datetime) else v) for k, v in r.items()})

    return {
        "filename": file.filename,
        "total_rows": len(rows),
        "headers": headers,
        "column_mapping": mapped,           # sheet-header → canonical field
        "ignored_columns": ignored,         # unrecognized sheet columns
        "missing_required": missing_required,
        "ready_to_import": len(missing_required) == 0,
        "sample_rows": sample,
    }


@api_router.post("/leaves/bulk-import")
async def bulk_import_leaves(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk-import leave records from .xlsx or .csv. Admin-only.

    Behaviour (per agreed spec):
    - Map by Employee Email; unknown emails are skipped with an error.
    - Leave Type is fuzzy-matched against canonical types; unmatched → 'General Leave'.
    - Validate date formats; auto-calc Number of Days when blank.
    - Skip overlapping date ranges with existing non-rejected leaves (duplicate guard).
    - Default Status when blank = 'pending'.
    - No emails are sent for imported rows (silent bulk insert).
    - Approved rows are marked applied_by_admin=true, approved_by=current admin.
    """
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")

    # Build alias index and remap headers → canonical fields
    alias_index = _build_alias_index()
    headers_clean = [h for h in headers if h]
    sheet_to_canon = {h: alias_index.get(_normalize_header(h)) for h in headers_clean}
    detected_canonical = {c for c in sheet_to_canon.values() if c}
    ignored_columns = [h for h, c in sheet_to_canon.items() if not c]

    # Validate mandatory canonical fields
    required_fields = ["email", "leave_type", "start_date", "end_date"]
    missing_fields = [f for f in required_fields if f not in detected_canonical]
    if missing_fields:
        # Reverse-map for friendly error message (show common alias names)
        canon_to_friendly = {
            "email": "Employee Email / Emp Mail ID",
            "leave_type": "Leave Type",
            "start_date": "From Date",
            "end_date": "To Date",
        }
        raise HTTPException(
            status_code=400,
            detail=f"Missing mandatory column(s): {', '.join(canon_to_friendly[f] for f in missing_fields)}"
        )

    # Pre-fetch employees by email for O(1) lookup
    emp_cursor = db.employees.find(
        {"is_deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "official_email": 1, "full_name": 1, "team": 1, "department": 1}
    )
    email_to_emp: dict = {}
    async for e in emp_cursor:
        email = (e.get("official_email") or "").strip().lower()
        if email:
            email_to_emp[email] = e

    # Pre-fetch users for "Approved By" resolution (admins/HR + employees)
    approver_index: dict[str, str] = {}  # normalized name/email → user_id
    async for u in db.users.find({}, {"_id": 0, "id": 1, "username": 1, "email": 1, "role": 1}):
        if u.get("username"):
            approver_index[_normalize_header(u["username"])] = u["id"]
        if u.get("email"):
            approver_index[_normalize_header(u["email"])] = u["id"]
    # Add employees by full_name and email so a manager's display name resolves too
    async for e in db.employees.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "full_name": 1, "official_email": 1}):
        if e.get("full_name"):
            key = _normalize_header(e["full_name"])
            approver_index.setdefault(key, e["id"])
        if e.get("official_email"):
            approver_index.setdefault(_normalize_header(e["official_email"]), e["id"])

    def _resolve_approver(value) -> tuple[Optional[str], Optional[str]]:
        """Return (resolved_id, raw_unresolved). For 'Admin' or known names/usernames/
        emails, returns (id, None). Otherwise returns (None, raw) so caller can
        keep raw text as a fallback note."""
        if value in (None, ""):
            return (None, None)
        s = str(value).strip()
        if not s:
            return (None, None)
        norm = _normalize_header(s)
        if norm in ("admin", "administrator", "hr admin", "hr"):
            # Default to current admin doing the import
            return (current_user["id"], None)
        if norm in approver_index:
            return (approver_index[norm], None)
        return (None, s)

    total = len(rows)
    success = 0
    skipped_duplicates = 0
    failed = 0
    errors: List[dict] = []
    inserts: List[dict] = []
    inbatch_ranges: dict[str, list[tuple[str, str]]] = {}

    for idx, raw_row in enumerate(rows, start=2):  # row 1 is the header
        # Remap sheet headers to canonical field names; unknown columns are dropped
        row = _remap_row(raw_row, alias_index)

        email = (str(row.get("email") or "")).strip().lower()
        if not email:
            failed += 1
            errors.append({"row": idx, "email": "", "reason": "Employee Email is blank"})
            continue

        emp = email_to_emp.get(email)
        if not emp:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "No employee found with this email"})
            continue

        from_date = _parse_import_date(row.get("start_date"))
        to_date = _parse_import_date(row.get("end_date"))
        if not from_date or not to_date:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "Invalid From/To date format (use YYYY-MM-DD or DD-MM-YYYY)"})
            continue

        try:
            sd = datetime.strptime(from_date, "%Y-%m-%d")
            ed = datetime.strptime(to_date, "%Y-%m-%d")
        except ValueError:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "Invalid date values"})
            continue

        if sd > ed:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "From Date must be <= To Date"})
            continue

        # Leave Split (replaces No of Days). Days are auto-calculated from dates + split.
        split_norm = _normalize_leave_split(row.get("leave_split"))
        if split_norm == "INVALID":
            failed += 1
            errors.append({"row": idx, "email": email, "reason": f"Invalid Leave Split value '{row.get('leave_split')}' (use Full Day or Half Day)"})
            continue
        leave_split = split_norm or "Full Day"  # default to Full Day when blank

        # Auto-calc days from dates + split
        if leave_split == "Full Day":
            days = (ed - sd).days + 1
        else:  # Half Day variants count as 0.5
            days = ((ed - sd).days + 1) * 0.5
        duration_str = f"{int(days) if days == int(days) else days} day(s)"

        # Auto Approve & Set LOP flag
        auto_lop_raw = row.get("auto_approve_lop")
        auto_lop = _normalize_bool(auto_lop_raw)
        if auto_lop == "INVALID":
            failed += 1
            errors.append({"row": idx, "email": email, "reason": f"Invalid 'Auto Approve & Set LOP Status' value '{auto_lop_raw}' (use Yes/No, TRUE/FALSE, 1/0)"})
            continue

        leave_type = _normalize_leave_type(row.get("leave_type"))
        status_norm = _normalize_status(row.get("status")) or "pending"
        reason = (str(row.get("reason")).strip() if row.get("reason") not in (None, "") else None)
        comments = (str(row.get("comments")).strip() if row.get("comments") not in (None, "") else None)
        applied_at = _parse_import_date(row.get("applied_at"))
        approved_at = _parse_import_date(row.get("approved_at"))

        # Resolve Approved By (name / username / email / "Admin" → user_id)
        approver_raw = row.get("approved_by")
        approver_id, approver_unresolved = _resolve_approver(approver_raw)

        # If Auto Approve & Set LOP Status = TRUE, force-approve and apply LOP
        is_lop_flag: Optional[bool] = None
        if auto_lop is True:
            status_norm = "approved"
            is_lop_flag = True

        # Duplicate guard: any non-rejected overlapping leave for this employee?
        overlap = await db.leaves.find_one({
            "employee_id": emp["id"],
            "status": {"$ne": "rejected"},
            "start_date": {"$lte": to_date},
            "end_date": {"$gte": from_date}
        }, {"_id": 0, "id": 1})
        if overlap:
            skipped_duplicates += 1
            errors.append({"row": idx, "email": email, "reason": f"Overlaps with existing leave between {from_date} and {to_date}"})
            continue

        # Also check overlaps with rows added earlier in THIS batch
        in_batch_overlap = False
        for prev_from, prev_to in inbatch_ranges.get(emp["id"], []):
            if prev_from <= to_date and prev_to >= from_date:
                in_batch_overlap = True
                break
        if in_batch_overlap:
            skipped_duplicates += 1
            errors.append({"row": idx, "email": email, "reason": f"Overlaps with another row in this upload between {from_date} and {to_date}"})
            continue
        inbatch_ranges.setdefault(emp["id"], []).append((from_date, to_date))

        # Final approved_by: resolved user id > current admin (when status approved) > None
        final_approved_by = approver_id
        if final_approved_by is None and status_norm == "approved":
            final_approved_by = current_user["id"]

        leave_doc = {
            "id": str(uuid.uuid4()),
            "employee_id": emp["id"],
            "emp_name": emp.get("full_name", ""),
            "team": emp.get("team", ""),
            "department": emp.get("department", ""),
            "leave_type": leave_type,
            "leave_split": leave_split,
            "start_date": from_date,
            "end_date": to_date,
            "duration": duration_str,
            "reason": reason,
            "supporting_document_url": None,
            "supporting_document_name": None,
            "status": status_norm,
            "is_lop": is_lop_flag,
            "lop_remark": comments,  # Map sheet "Comments" → existing lop_remark column
            "approved_by": final_approved_by,
            "applied_by_admin": True,
            # Preserve original sheet metadata that can't be mapped to a primary column
            "extra_data": {
                k: v for k, v in {
                    "applied_at": applied_at,
                    "approved_at": approved_at,
                    "approver_unresolved": approver_unresolved,
                }.items() if v is not None
            } or None,
            "created_at": get_ist_now().isoformat(),
        }
        inserts.append(leave_doc)
        success += 1

    # Batch insert in chunks of 500 for performance with 1000+ rows
    if inserts:
        chunk = 500
        for i in range(0, len(inserts), chunk):
            await db.leaves.insert_many([d.copy() for d in inserts[i:i + chunk]])

    await log_audit(
        current_user["id"], "bulk_import", "leaves", None,
        f"Leave bulk import: total={total} success={success} duplicates={skipped_duplicates} failed={failed}"
    )

    return {
        "total": total,
        "success": success,
        "skipped_duplicates": skipped_duplicates,
        "failed": failed,
        "column_mapping": {h: c for h, c in sheet_to_canon.items() if c},
        "ignored_columns": ignored_columns,
        "errors": errors[:1000],  # cap to keep response manageable
    }


class LeaveApproveRequest(BaseModel):
    is_lop: Optional[bool] = None  # True=LOP, False=No LOP
    lop_remark: Optional[str] = None

class LeaveAdminEdit(BaseModel):
    """Admin/HR-side edit payload for a leave record. All fields optional —
    only supplied keys are written. Status is intentionally NOT editable here
    (use approve/reject endpoints) so approval logic stays the single source
    of truth."""
    leave_type: Optional[str] = None
    leave_split: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    reason: Optional[str] = None


@api_router.put("/leaves/{leave_id}")
async def admin_edit_leave(leave_id: str, data: LeaveAdminEdit, current_user: dict = Depends(get_current_user)):
    """HR/Admin can edit ANY leave (Pending / Approved / Rejected). Updates
    the same record in place — no duplication. Audit fields are written so
    edits are traceable. Approval status is preserved untouched."""
    if current_user["role"] not in ALL_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")

    update = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")

    # Recompute duration if dates / split changed
    new_split = update.get("leave_split", leave.get("leave_split", "Full Day"))
    new_start = update.get("start_date", leave.get("start_date"))
    new_end = update.get("end_date", leave.get("end_date"))
    try:
        if new_split in ("First Half", "Second Half"):
            update["duration"] = "0.5 day(s)"
        elif new_start and new_end:
            sd = datetime.strptime(new_start, "%Y-%m-%d")
            ed = datetime.strptime(new_end, "%Y-%m-%d")
            update["duration"] = f"{(ed - sd).days + 1} day(s)"
    except ValueError:
        pass  # leave duration as-is on bad date format

    update["edited_by"] = current_user["id"]
    update["edited_at"] = get_ist_now().isoformat()

    await db.leaves.update_one({"id": leave_id}, {"$set": update})
    await log_audit(current_user["id"], "edit", "leave", leave_id)
    updated = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    return serialize_doc(updated)



@api_router.put("/leaves/{leave_id}/approve")
async def approve_leave(leave_id: str, data: Optional[LeaveApproveRequest] = None, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
    
    update_fields = {
        "status": "approved",
        "approved_by": current_user["id"],
        "approved_at": get_ist_now().isoformat()
    }
    if data and data.is_lop is not None:
        update_fields["is_lop"] = data.is_lop
    if data and data.lop_remark:
        update_fields["lop_remark"] = data.lop_remark
    
    result = await db.leaves.update_one(
        {"id": leave_id},
        {"$set": update_fields}
    )
    
    await log_audit(current_user["id"], "approve", "leave", leave_id)
    
    # Send email notification
    employee = await db.employees.find_one({"id": leave["employee_id"]}, {"_id": 0})
    if employee and employee.get("official_email"):
        email_html = get_leave_approval_email(
            employee["full_name"],
            leave["leave_type"],
            leave["start_date"],
            leave["end_date"],
            "approved"
        )
        asyncio.create_task(send_email_notification(
            employee["official_email"],
            "Leave Request Approved - BluBridge HRMS",
            email_html
        ))
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    
    # Notify employee about leave approval
    emp_user = await db.users.find_one({"employee_id": leave.get("employee_id")}, {"_id": 0})
    if emp_user:
        asyncio.create_task(create_notification(
            [emp_user["id"]],
            "Leave Approved",
            f"Your {leave.get('leave_type', '')} leave request has been approved.",
            "success",
            "/employee/leave"
        ))
    
    return serialize_doc(leave)

@api_router.put("/leaves/{leave_id}/reject")
async def reject_leave(leave_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
    
    result = await db.leaves.update_one(
        {"id": leave_id},
        {"$set": {"status": "rejected", "approved_by": current_user["id"], "approved_at": get_ist_now().isoformat()}}
    )
    
    await log_audit(current_user["id"], "reject", "leave", leave_id)
    
    # Send email notification
    employee = await db.employees.find_one({"id": leave["employee_id"]}, {"_id": 0})
    if employee and employee.get("official_email"):
        email_html = get_leave_approval_email(
            employee["full_name"],
            leave["leave_type"],
            leave["start_date"],
            leave["end_date"],
            "rejected"
        )
        asyncio.create_task(send_email_notification(
            employee["official_email"],
            "Leave Request Rejected - BluBridge HRMS",
            email_html
        ))
    
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    
    # Notify employee about leave rejection
    emp_user = await db.users.find_one({"employee_id": leave.get("employee_id")}, {"_id": 0})
    if emp_user:
        asyncio.create_task(create_notification(
            [emp_user["id"]],
            "Leave Rejected",
            f"Your {leave.get('leave_type', '')} leave request has been rejected.",
            "warning",
            "/employee/leave"
        ))
    
    return serialize_doc(leave)


# ============== UNIVERSAL REQUEST RESET ENGINE ==============
#
# `Reset to Pending` flips a processed request back to its initial state and
# fully reverses any side-effects (attendance corrections, LOP flags, etc.).
# Then HR can re-process the request from a clean slate — Approve / Reject /
# edit LOP — without stacking effects on top of an already-applied state.
#
# Audit collection: `request_resets`. Every reset writes a row containing
# `request_type`, `request_id`, `previous_status`, `previous_snapshot`,
# `reset_by`, `reset_at`, `reason`, and (for missed-punches) the attendance
# rollback details.

class RequestResetBody(BaseModel):
    reason: Optional[str] = None


async def _log_request_reset(request_type: str, request_id: str, prev_snapshot: dict,
                              attendance_rollback: Optional[dict], current_user: dict, reason: Optional[str]):
    """Single source of truth for the reset audit trail. Failures are logged
    but never propagate — audit must NEVER block the actual reset."""
    try:
        doc = {
            "id": str(uuid.uuid4()),
            "request_type": request_type,
            "request_id": request_id,
            "previous_status": prev_snapshot.get("status"),
            "previous_snapshot": prev_snapshot,
            "attendance_rollback": attendance_rollback,
            "reset_by": current_user["id"],
            "reset_at": get_ist_now().isoformat(),
            "reason": reason,
        }
        await db.request_resets.insert_one(doc.copy())
    except Exception as e:
        logger.warning(f"Failed to write request_resets audit row for {request_type}/{request_id}: {e}")


@api_router.post("/leaves/{leave_id}/reset")
async def reset_leave(leave_id: str, body: RequestResetBody = RequestResetBody(),
                       current_user: dict = Depends(get_current_user)):
    """Reset a Leave request back to `pending`. Side-effects on payroll /
    attendance are computed dynamically from `is_lop` and `status`, so
    clearing those fields is sufficient to undo the impact."""
    if current_user["role"] not in ALL_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")

    if leave.get("status") == "pending" and leave.get("is_lop") is None and not leave.get("approved_by"):
        # Already at initial state — idempotent no-op
        return serialize_doc(leave)

    snapshot = dict(leave)
    await db.leaves.update_one(
        {"id": leave_id},
        {
            "$set": {
                "status": "pending",
                "reset_at": get_ist_now().isoformat(),
                "reset_by": current_user["id"],
            },
            "$unset": {
                "approved_by": "",
                "approved_at": "",
                "is_lop": "",
                "lop_remark": "",
                "rejection_reason": "",
            },
        },
    )

    await _log_request_reset("leave", leave_id, snapshot, None, current_user, body.reason)
    await log_audit(current_user["id"], "reset", "leave", leave_id)
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    return serialize_doc(leave)


@api_router.post("/late-requests/{request_id}/reset")
async def reset_late_request(request_id: str, body: RequestResetBody = RequestResetBody(),
                              current_user: dict = Depends(get_current_user)):
    """Reset a Late Request back to `pending`."""
    if current_user["role"] not in ALL_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    rec = await db.late_requests.find_one({"id": request_id}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Request not found")

    if rec.get("status") == "pending" and rec.get("is_lop") is None and not rec.get("approved_by"):
        return serialize_doc(rec)

    snapshot = dict(rec)
    await db.late_requests.update_one(
        {"id": request_id},
        {
            "$set": {"status": "pending", "reset_at": get_ist_now().isoformat(), "reset_by": current_user["id"]},
            "$unset": {"approved_by": "", "approved_at": "", "is_lop": "", "lop_remark": "", "rejection_reason": ""},
        },
    )
    await _log_request_reset("late_request", request_id, snapshot, None, current_user, body.reason)
    await log_audit(current_user["id"], "reset", "late_request", request_id)
    rec = await db.late_requests.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(rec)


@api_router.post("/early-out-requests/{request_id}/reset")
async def reset_early_out_request(request_id: str, body: RequestResetBody = RequestResetBody(),
                                   current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ALL_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    rec = await db.early_out_requests.find_one({"id": request_id}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Request not found")

    if rec.get("status") == "pending" and rec.get("is_lop") is None and not rec.get("approved_by"):
        return serialize_doc(rec)

    snapshot = dict(rec)
    await db.early_out_requests.update_one(
        {"id": request_id},
        {
            "$set": {"status": "pending", "reset_at": get_ist_now().isoformat(), "reset_by": current_user["id"]},
            "$unset": {"approved_by": "", "approved_at": "", "is_lop": "", "lop_remark": "", "rejection_reason": ""},
        },
    )
    await _log_request_reset("early_out_request", request_id, snapshot, None, current_user, body.reason)
    await log_audit(current_user["id"], "reset", "early_out_request", request_id)
    rec = await db.early_out_requests.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(rec)


@api_router.post("/missed-punches/{request_id}/reset")
async def reset_missed_punch(request_id: str, body: RequestResetBody = RequestResetBody(),
                              current_user: dict = Depends(get_current_user)):
    """Reset a Missed Punch request back to `pending` AND restore the
    pre-correction attendance state.

    Uses the most recent `attendance_corrections` row (keyed by request_id)
    as the source of truth for old check_in / check_out / status. Recomputes
    total_hours via `calculate_attendance_status` from the restored values
    so the attendance row is in a fully consistent post-rollback state.
    """
    if current_user["role"] not in ALL_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    rec = await db.missed_punches.find_one({"id": request_id}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Request not found")

    snapshot = dict(rec)
    rollback_info: Optional[dict] = None

    # Find the most recent applied audit row for this request
    audit = await db.attendance_corrections.find_one(
        {"request_id": request_id}, {"_id": 0}, sort=[("created_at", -1)]
    )

    if audit:
        emp_id = audit.get("employee_id") or rec.get("employee_id")
        date = audit.get("date") or rec.get("date")
        existing_att = await db.attendance.find_one({"employee_id": emp_id, "date": date}, {"_id": 0})

        old_in_24h = audit.get("old_check_in_24h")
        old_out_24h = audit.get("old_check_out_24h")
        old_in = audit.get("old_check_in")
        old_out = audit.get("old_check_out")
        old_status = audit.get("old_status")
        source_before = audit.get("source_before") or "biometric"

        employee = await db.employees.find_one({"id": emp_id, "is_deleted": {"$ne": True}}, {"_id": 0})
        shift_timings = get_shift_timings(employee) if employee else None

        # Recompute status / total_hours from the restored times so dashboards / payroll are consistent
        if old_in_24h:
            recomputed = calculate_attendance_status(
                old_in_24h, old_out_24h, shift_timings or {}, attendance_date=date
            )
            new_status_after_rollback = recomputed.get("status") or old_status or AttendanceStatus.PRESENT
            total_hours_decimal = float(recomputed.get("total_hours_decimal") or 0.0)
            total_hours_str = calculate_total_hours_str(total_hours_decimal) if total_hours_decimal else None
            is_lop = bool(recomputed.get("is_lop"))
            lop_reason = recomputed.get("lop_reason")
        else:
            new_status_after_rollback = AttendanceStatus.NOT_LOGGED
            total_hours_decimal = 0.0
            total_hours_str = None
            is_lop = False
            lop_reason = None

        if existing_att and source_before == "none":
            # Attendance row only existed BECAUSE of this correction — delete it
            await db.attendance.delete_one({"employee_id": emp_id, "date": date})
            rollback_info = {"action": "deleted", "employee_id": emp_id, "date": date}
        elif existing_att:
            payload = {
                "check_in": old_in,
                "check_in_24h": old_in_24h,
                "check_out": old_out,
                "check_out_24h": old_out_24h,
                "status": new_status_after_rollback,
                "total_hours": total_hours_str,
                "total_hours_decimal": round(total_hours_decimal, 2),
                "is_lop": is_lop,
                "lop_reason": lop_reason,
                "source": source_before,
                "missed_punch_corrected": False,
                "missed_punch_reverted_at": get_ist_now().isoformat(),
                "missed_punch_reverted_by": current_user["id"],
            }
            await db.attendance.update_one(
                {"employee_id": emp_id, "date": date},
                {
                    "$set": payload,
                    "$unset": {
                        "missed_punch_request_id": "",
                        "missed_punch_corrected_at": "",
                        "missed_punch_corrected_by": "",
                    },
                },
            )
            rollback_info = {"action": "restored", "employee_id": emp_id, "date": date,
                             "from_state": {"check_in_24h": existing_att.get("check_in_24h"),
                                            "check_out_24h": existing_att.get("check_out_24h"),
                                            "status": existing_att.get("status")},
                             "to_state": {"check_in_24h": old_in_24h,
                                          "check_out_24h": old_out_24h,
                                          "status": new_status_after_rollback}}

        # Mark all audit rows for this request as reverted (don't delete — keep history)
        await db.attendance_corrections.update_many(
            {"request_id": request_id, "reverted_at": {"$exists": False}},
            {"$set": {"reverted_at": get_ist_now().isoformat(), "reverted_by": current_user["id"]}},
        )

    # Reset the request itself — clear approval + correction-applied flags
    await db.missed_punches.update_one(
        {"id": request_id},
        {
            "$set": {"status": "pending", "reset_at": get_ist_now().isoformat(), "reset_by": current_user["id"]},
            "$unset": {
                "approved_by": "",
                "approved_at": "",
                "rejection_reason": "",
                "correction_applied_at": "",
                "correction_applied_by": "",
                "is_applied": "",
            },
        },
    )

    await _log_request_reset("missed_punch", request_id, snapshot, rollback_info, current_user, body.reason)
    await log_audit(current_user["id"], "reset", "missed_punch", request_id)
    rec = await db.missed_punches.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(rec)




# ============== STAR REWARDS ROUTES ==============

@api_router.get("/star-rewards")
async def get_star_rewards(
    team: Optional[str] = None,
    department: Optional[str] = None,
    month: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": {"$ne": True}, "employee_status": EmployeeStatus.ACTIVE}
    if team and team != "All":
        query["team"] = team
    if department and department != "All":
        query["department"] = department
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"official_email": {"$regex": search, "$options": "i"}},
            {"emp_id": {"$regex": search, "$options": "i"}}
        ]
    
    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)
    
    # Map full_name to name for backward compatibility
    for emp in employees:
        emp["name"] = emp.get("full_name", "")
        emp["email"] = emp.get("official_email", "")
    
    return [serialize_doc(e) for e in employees]

@api_router.post("/star-rewards")
async def add_star_reward(data: StarRewardCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    employee = await db.employees.find_one({"id": data.employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Restrict to Research Unit employees only
    if employee.get("department") != "Research Unit":
        raise HTTPException(status_code=400, detail="Star rewards can only be given to Research Unit employees")
    
    current_month = get_ist_now().strftime("%Y-%m")
    
    reward = StarReward(
        employee_id=data.employee_id,
        stars=data.stars,
        reason=data.reason,
        type=data.type,
        awarded_by=current_user["id"],
        month=current_month
    )
    doc = reward.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.star_rewards.insert_one(doc.copy())
    
    # Update employee stars and unsafe count
    new_stars = employee.get("stars", 0) + data.stars
    update_data = {"stars": new_stars}
    
    # If type is unsafe, increment unsafe count
    if data.type == "unsafe":
        new_unsafe = employee.get("unsafe_count", 0) + 1
        update_data["unsafe_count"] = new_unsafe
    
    await db.employees.update_one({"id": data.employee_id}, {"$set": update_data})
    
    await log_audit(current_user["id"], "award_stars", "star_reward", reward.id)
    
    # Send email notification
    if employee.get("official_email"):
        awarder = await db.users.find_one({"id": current_user["id"]}, {"_id": 0})
        awarder_name = awarder.get("name", "Admin") if awarder else "Admin"
        email_html = get_star_reward_email(
            employee["full_name"],
            data.stars,
            data.reason,
            awarder_name
        )
        asyncio.create_task(send_email_notification(
            employee["official_email"],
            "Star Reward Notification - BluBridge HRMS",
            email_html
        ))
    
    return {"message": "Stars awarded", "new_total": new_stars}

@api_router.get("/star-rewards/history/{employee_id}")
async def get_star_history(employee_id: str, current_user: dict = Depends(get_current_user)):
    rewards = await db.star_rewards.find({"employee_id": employee_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return [serialize_doc(r) for r in rewards]

# ============== TEAM ROUTES ==============

@api_router.get("/teams")
async def get_teams(department: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"is_deleted": {"$ne": True}}
    if department and department != "All":
        query["department"] = department
    
    teams = await db.teams.find(query, {"_id": 0}).to_list(100)
    
    # Calculate actual member count from employees
    for team in teams:
        count = await db.employees.count_documents({
            "team": team["name"],
            "is_deleted": {"$ne": True},
            "employee_status": EmployeeStatus.ACTIVE
        })
        team["member_count"] = count
    
    return [serialize_doc(t) for t in teams]

@api_router.get("/teams/{team_id}")
async def get_team(team_id: str, current_user: dict = Depends(get_current_user)):
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    members = await db.employees.find({
        "team": team["name"],
        "is_deleted": {"$ne": True},
        "employee_status": EmployeeStatus.ACTIVE
    }, {"_id": 0}).to_list(100)
    
    return {"team": serialize_doc(team), "members": [serialize_doc(m) for m in members]}

@api_router.get("/departments")
async def get_departments(current_user: dict = Depends(get_current_user)):
    departments = await db.departments.find({"is_deleted": {"$ne": True}}, {"_id": 0}).to_list(100)
    
    # Calculate actual counts
    for dept in departments:
        emp_count = await db.employees.count_documents({
            "department": dept["name"],
            "is_deleted": {"$ne": True},
            "employee_status": EmployeeStatus.ACTIVE
        })
        team_count = await db.teams.count_documents({"department": dept["name"]})
        dept["employee_count"] = emp_count
        dept["team_count"] = team_count
    
    return [serialize_doc(d) for d in departments]

# ============== DASHBOARD ROUTES ==============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    today = get_ist_now().strftime("%d-%m-%Y")
    
    # Use provided dates or default to today
    query_from = from_date if from_date else today
    query_to = to_date if to_date else today
    
    # Get counts from employee master
    base_query = {"is_deleted": {"$ne": True}, "employee_status": EmployeeStatus.ACTIVE}
    
    total_research = await db.employees.count_documents({**base_query, "department": "Research Unit"})
    total_support = await db.employees.count_documents({**base_query, "department": "Support Staff"})
    pending_approvals = await db.leaves.count_documents({"status": "pending"})
    upcoming_leaves = await db.leaves.count_documents({
        "status": "approved",
        "start_date": {"$gte": get_ist_now().strftime("%Y-%m-%d")}
    })
    
    # Get attendance stats with date range support
    attendance_stats = await get_attendance_stats(
        date=None, 
        from_date=query_from, 
        to_date=query_to, 
        current_user=current_user
    )
    employee_stats = await get_employee_stats(current_user)
    
    return {
        "total_research_unit": total_research,
        "total_support_staff": total_support,
        "pending_approvals": pending_approvals,
        "upcoming_leaves": upcoming_leaves,
        "attendance": attendance_stats,
        "employee_stats": employee_stats
    }

@api_router.get("/dashboard/leave-list")
async def get_dashboard_leave_list(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    today = get_ist_now().strftime("%d-%m-%Y")
    query_date = from_date if from_date else today

    # Active employees with attendance tracking
    all_employees = await db.employees.find({
        "employee_status": EmployeeStatus.ACTIVE,
        "is_deleted": {"$ne": True},
        "attendance_tracking_enabled": True
    }, {"_id": 0}).to_list(1000)

    # Logged-in employees for the date / range (attendance dates are DD-MM-YYYY)
    if from_date and to_date:
        logged_records = await db.attendance.find(
            {"date": {"$gte": from_date, "$lte": to_date}},
            {"_id": 0}
        ).to_list(10000)
    else:
        logged_records = await db.attendance.find({"date": query_date}, {"_id": 0}).to_list(1000)

    logged_ids = {a["employee_id"] for a in logged_records}

    # --- Leave priority lookup -------------------------------------------------
    # Leaves use YYYY-MM-DD for start_date / end_date. Convert DD-MM-YYYY query
    # date(s) to that format so we can cross-reference reliably.
    def _dmy_to_ymd(ds: Optional[str]) -> Optional[str]:
        if not ds:
            return None
        try:
            return datetime.strptime(ds, "%d-%m-%Y").strftime("%Y-%m-%d")
        except ValueError:
            return ds  # assume already YYYY-MM-DD

    ymd_from = _dmy_to_ymd(from_date or query_date)
    ymd_to = _dmy_to_ymd(to_date or query_date)

    # Pull any approved/pending leave overlapping the window (approved takes
    # priority, pending is shown as fall-back so admins can see the full leave
    # picture — only "rejected" is excluded).
    leave_docs = await db.leaves.find(
        {
            "status": {"$in": ["approved", "pending"]},
            "start_date": {"$lte": ymd_to},
            "end_date": {"$gte": ymd_from},
        },
        {"_id": 0},
    ).to_list(5000)

    # Index best leave per employee (approved wins over pending)
    leave_by_emp: dict = {}
    status_rank = {"approved": 2, "pending": 1}
    for lv in leave_docs:
        emp_id = lv.get("employee_id")
        if not emp_id:
            continue
        current = leave_by_emp.get(emp_id)
        if not current or status_rank.get(lv.get("status"), 0) > status_rank.get(current.get("status"), 0):
            leave_by_emp[emp_id] = lv

    # --- Build response ---------------------------------------------------------
    # Priority 1: on leave (regardless of attendance punch — a leave always wins
    # for dashboard display purposes). Priority 2: no attendance + no leave.
    result = []

    # 1) Employees on leave for the window
    for emp in all_employees:
        lv = leave_by_emp.get(emp["id"])
        if not lv:
            continue
        # Display leave date in DD-MM-YYYY for UI consistency
        try:
            lv_start_display = datetime.strptime(lv.get("start_date", ""), "%Y-%m-%d").strftime("%d-%m-%Y")
        except (ValueError, TypeError):
            lv_start_display = lv.get("start_date") or query_date
        result.append({
            "emp_name": emp["full_name"],
            "team": emp.get("team"),
            "department": emp.get("department"),
            "leave_type": lv.get("leave_type") or "Leave",
            "leave_split": lv.get("leave_split"),
            "date": lv_start_display,
            "status": "Leave" if lv.get("status") == "approved" else "Pending Leave",
            "reason": lv.get("reason"),
        })

    # 2) No attendance AND no leave
    for emp in all_employees:
        if emp["id"] in logged_ids:
            continue
        if emp["id"] in leave_by_emp:
            continue
        result.append({
            "emp_name": emp["full_name"],
            "team": emp.get("team"),
            "department": emp.get("department"),
            "leave_type": "-",
            "date": query_date,
            "status": "Not Login",
        })

    return result[:50]

# ============== REPORTS ROUTES ==============

def _normalize_date_to_int(ds: Optional[str]) -> Optional[int]:
    """Convert any common date string to YYYYMMDD int for sortable comparison.
    Accepts DD-MM-YYYY, YYYY-MM-DD, DD/MM/YYYY, YYYY/MM/DD, ISO strings."""
    if not ds:
        return None
    s = ds.strip().split("T")[0].replace("/", "-")
    parts = s.split("-")
    if len(parts) != 3:
        return None
    try:
        a, b, c = int(parts[0]), int(parts[1]), int(parts[2])
    except ValueError:
        return None
    if len(parts[0]) == 4:        # YYYY-MM-DD
        return a * 10000 + b * 100 + c
    if len(parts[2]) == 4:        # DD-MM-YYYY
        return c * 10000 + b * 100 + a
    return None


@api_router.get("/reports/attendance")
async def get_attendance_report(
    from_date: str,
    to_date: str,
    department: Optional[str] = None,
    team: Optional[str] = None,
    employee_name: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if team and team != "All":
        query["team"] = team
    if department and department != "All":
        query["department"] = department
    if employee_name and employee_name.strip():
        # Anchor and case-insensitive — matches both partial typing and exact
        # autocomplete selections of full name.
        query["emp_name"] = {"$regex": re.escape(employee_name.strip()), "$options": "i"}
    if status and status != "All":
        query["status"] = status

    # Date format mismatch fix: attendance.date is stored DD-MM-YYYY but the
    # frontend's DatePicker emits YYYY-MM-DD. Compare normalized integer keys.
    f_int = _normalize_date_to_int(from_date)
    t_int = _normalize_date_to_int(to_date)

    records = await db.attendance.find(query, {"_id": 0}).to_list(20000)
    if f_int is not None or t_int is not None:
        lo = f_int if f_int is not None else 0
        hi = t_int if t_int is not None else 99999999
        records = [r for r in records if (lambda v: v is not None and lo <= v <= hi)(_normalize_date_to_int(r.get("date")))]

    # Stable sort by date asc for predictable export order
    records.sort(key=lambda r: _normalize_date_to_int(r.get("date")) or 0)
    return [serialize_doc(r) for r in records]


@api_router.get("/reports/attendance/export")
async def export_attendance_report(
    from_date: str,
    to_date: str,
    department: Optional[str] = None,
    team: Optional[str] = None,
    employee_name: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export Attendance report in the wide pivot format that matches the
    reference template, with each attendance field in its OWN dedicated cell:

      Row 1 (date header, merged across 6 cols per day):
          S.No | Name | Team | Department | <date> | <date> | ...
      Row 2 (sub-header per date, 6 columns):
          (blank x4) | In Time | Out Time | Late-in | Early Out | Total Hours | Status | ...
      Row 3+ (data rows, one per employee).

    No multi-line cells. No combined values. Each atomic field is a discrete
    column. Compatible with day-wise / month-wise / custom range exports.
    """
    f_int = _normalize_date_to_int(from_date)
    t_int = _normalize_date_to_int(to_date)
    if f_int is None or t_int is None or f_int > t_int:
        raise HTTPException(status_code=400, detail="Invalid date range")

    # Resolve target employees with status-aware filtering:
    #   - Always include currently Active employees.
    #   - Include employees who left during/after the report start date.
    #   - Exclude employees whose last working day is BEFORE the report start.
    emp_query = {"is_deleted": {"$ne": True}}
    if team and team != "All":
        emp_query["team"] = team
    if department and department != "All":
        emp_query["department"] = department
    if employee_name and employee_name.strip():
        emp_query["full_name"] = {"$regex": re.escape(employee_name.strip()), "$options": "i"}

    employees = await db.employees.find(
        emp_query,
        {"_id": 0, "id": 1, "full_name": 1, "team": 1, "department": 1,
         "employee_status": 1, "inactive_date": 1, "last_day_payable": 1, "date_of_joining": 1}
    ).to_list(5000)

    def _is_employee_in_range(emp):
        # Active employees: always include
        if emp.get("employee_status") != "Inactive":
            return True
        # Inactive employees: include only if their last working day is on/after report start
        last_day_raw = emp.get("last_day_payable") or emp.get("inactive_date")
        last_day_int = _normalize_date_to_int(last_day_raw)
        if last_day_int is None:
            # No reliable exit date — fall back to including (data integrity safety)
            return True
        return last_day_int >= f_int

    employees = [e for e in employees if _is_employee_in_range(e)]
    employees.sort(key=lambda e: (e.get("full_name") or "").lower())

    emp_ids = [e["id"] for e in employees]
    att_query = {"employee_id": {"$in": emp_ids}}
    if status and status != "All":
        att_query["status"] = status
    att_records = await db.attendance.find(att_query, {"_id": 0}).to_list(50000)

    att_map = {}
    for r in att_records:
        v = _normalize_date_to_int(r.get("date"))
        if v is None or not (f_int <= v <= t_int):
            continue
        att_map[(r["employee_id"], r["date"])] = r

    # Build the date list (inclusive)
    def _parse(ds):
        ds = ds.split("T")[0].replace("/", "-")
        if len(ds.split("-")[0]) == 4:
            return datetime.strptime(ds, "%Y-%m-%d").date()
        return datetime.strptime(ds, "%d-%m-%Y").date()
    f_d, t_d = _parse(from_date), _parse(to_date)
    date_list = []
    cur = f_d
    while cur <= t_d:
        date_list.append(cur)
        cur += timedelta(days=1)

    SUB_HEADERS = ["In Time", "Out Time", "Late-in", "Early Out", "Total Hours", "Status"]
    PER_DATE_COLS = len(SUB_HEADERS)
    FIXED_COLS = 4

    def cells_for_day(emp_id: str, d):
        """Return a 6-tuple (in, out, late, early, hours, status) for one day."""
        date_dd = d.strftime("%d-%m-%Y")
        rec = att_map.get((emp_id, date_dd))
        is_sun = d.weekday() == 6
        if not rec or not (rec.get("check_in") or rec.get("check_in_24h")):
            status_text = "Sunday (Not Login)" if is_sun else "Not Login"
            return ("", "", "", "", "", status_text)
        in_t = rec.get("check_in") or ""
        out_t = rec.get("check_out") or ""
        hours = rec.get("total_hours") or ""
        # Convert "11h 36m" → "11:36"
        h_clean = hours
        if hours and "h" in hours:
            try:
                hh = int(hours.split("h")[0].strip())
                mm_part = hours.split("h")[1]
                mm = int(mm_part.replace("m", "").strip()) if mm_part.strip() else 0
                h_clean = f"{hh:02d}:{mm:02d}"
            except Exception:
                h_clean = hours
        lop_reason = rec.get("lop_reason") or ""
        is_late = bool(rec.get("is_lop")) and "Late login" in lop_reason
        is_early = bool(rec.get("is_lop")) and ("Early out" in lop_reason or "short hours" in lop_reason)
        if not out_t:
            return (in_t, "", "Yes" if is_late else "", "", "", "Login")
        return (in_t, out_t, "Yes" if is_late else "", "Yes" if is_early else "", h_clean, "Logout")

    # Build XLSX
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="FFEFEFEF")
    # Subtle, professional flag colors that read well in Excel.
    LATE_FILL = PatternFill("solid", fgColor="FFFCE4D6")     # light orange
    EARLY_FILL = PatternFill("solid", fgColor="FFFFF2CC")    # light yellow
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1 — fixed column headers + per-date merged date headers
    fixed_headers = ["S.No", "Name", "Team", "Department"]
    for i, h in enumerate(fixed_headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold; c.fill = fill; c.alignment = center; c.border = border
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
        ws.cell(row=2, column=i).border = border

    for d_idx, d in enumerate(date_list):
        start_col = FIXED_COLS + 1 + d_idx * PER_DATE_COLS
        end_col = start_col + PER_DATE_COLS - 1
        c = ws.cell(row=1, column=start_col, value=d.strftime("%Y-%m-%d"))
        c.font = bold; c.fill = fill; c.alignment = center; c.border = border
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        # Sub-headers in row 2
        for i, sh in enumerate(SUB_HEADERS):
            sc = ws.cell(row=2, column=start_col + i, value=sh)
            sc.font = bold; sc.fill = fill; sc.alignment = center; sc.border = border

    # Data rows starting row 3
    for idx, emp in enumerate(employees, 1):
        row = 2 + idx  # row 3 onwards
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=emp.get("full_name", "")).border = border
        ws.cell(row=row, column=3, value=emp.get("team") or "Unassigned").border = border
        ws.cell(row=row, column=4, value=emp.get("department") or "").border = border
        for d_idx, d in enumerate(date_list):
            start_col = FIXED_COLS + 1 + d_idx * PER_DATE_COLS
            vals = cells_for_day(emp["id"], d)
            in_val, out_val, late_val, early_val, _hours_val, _status_val = vals
            is_late = late_val == "Yes"
            is_early = early_val == "Yes"
            for i, v in enumerate(vals):
                c = ws.cell(row=row, column=start_col + i, value=v if v != "" else None)
                c.alignment = center
                c.border = border
                # Cell-level shading only on the relevant columns:
                #   - Late entry  → light orange on In Time + Late-in marker
                #   - Early exit  → light yellow on Out Time + Early Out marker
                if is_late and i in (0, 2):
                    c.fill = LATE_FILL
                if is_early and i in (1, 3):
                    c.fill = EARLY_FILL

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 16
    for i in range(FIXED_COLS + 1, FIXED_COLS + len(date_list) * PER_DATE_COLS + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "E3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"attendance_report_{from_date}_to_{to_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@api_router.get("/reports/leaves")
async def get_leave_report(
    from_date: str,
    to_date: str,
    department: Optional[str] = None,
    team: Optional[str] = None,
    employee_name: Optional[str] = None,
    leave_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if team and team != "All":
        query["team"] = team
    if department and department != "All":
        query["department"] = department
    if employee_name and employee_name.strip():
        query["emp_name"] = {"$regex": re.escape(employee_name.strip()), "$options": "i"}
    if leave_type and leave_type != "All":
        query["leave_type"] = leave_type

    f_int = _normalize_date_to_int(from_date)
    t_int = _normalize_date_to_int(to_date)

    records = await db.leaves.find(query, {"_id": 0}).to_list(20000)
    if f_int is not None or t_int is not None:
        lo = f_int if f_int is not None else 0
        hi = t_int if t_int is not None else 99999999
        # A leave overlaps the window if its [start_date, end_date] intersects [lo, hi]
        def _in_window(rec):
            s = _normalize_date_to_int(rec.get("start_date"))
            e = _normalize_date_to_int(rec.get("end_date") or rec.get("start_date"))
            if s is None and e is None:
                return False
            s = s if s is not None else e
            e = e if e is not None else s
            return s <= hi and e >= lo
        records = [r for r in records if _in_window(r)]

    records.sort(key=lambda r: _normalize_date_to_int(r.get("start_date")) or 0)
    return [serialize_doc(r) for r in records]

@api_router.get("/reports/employees")
async def get_employee_report(
    department: Optional[str] = None,
    team: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": {"$ne": True}}
    if department and department != "All":
        query["department"] = department
    if team and team != "All":
        query["team"] = team
    if status and status != "All":
        query["employee_status"] = status
    
    records = await db.employees.find(query, {"_id": 0}).to_list(10000)
    return [serialize_doc(r) for r in records]

# ============== PAYROLL ROUTES ==============

@api_router.get("/payroll")
async def get_payroll_data(
    month: str,  # Format: "YYYY-MM"
    department: Optional[str] = None,
    team: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get payroll data for all employees for a given month"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Include Active + Inactive employees (relieved within period handled by engine)
    query = {"is_deleted": {"$ne": True}, "employee_status": {"$in": [EmployeeStatus.ACTIVE, EmployeeStatus.INACTIVE]}}
    if department and department != "All":
        query["department"] = department
    if team and team != "All":
        query["team"] = team
    
    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)
    
    year, month_num = map(int, month.split('-'))
    import calendar
    days_in_month = calendar.monthrange(year, month_num)[1]
    emp_ids = [emp["id"] for emp in employees]
    prefetched = await _prefetch_payroll_data(emp_ids, year, month_num, days_in_month)

    payroll_data = []
    for emp in employees:
        payroll = await calculate_payroll_for_employee(emp["id"], month, employee=emp, prefetched=prefetched)
        if payroll:
            payroll_data.append(payroll)
    
    return payroll_data

@api_router.get("/payroll/{employee_id}")
async def get_employee_payroll(
    employee_id: str,
    month: str,  # Format: "YYYY-MM"
    current_user: dict = Depends(get_current_user)
):
    """Get detailed payroll for a specific employee"""
    payroll = await calculate_payroll_for_employee(employee_id, month)
    if not payroll:
        raise HTTPException(status_code=404, detail="Employee not found")
    return payroll

@api_router.get("/payroll/summary/{month}")
async def get_payroll_summary(
    month: str,  # Format: "YYYY-MM"
    department: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get payroll summary for a month"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {"is_deleted": {"$ne": True}, "employee_status": {"$in": [EmployeeStatus.ACTIVE, EmployeeStatus.INACTIVE]}}
    if department and department != "All":
        query["department"] = department
    
    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)
    
    year, month_num = map(int, month.split('-'))
    import calendar
    days_in_month = calendar.monthrange(year, month_num)[1]
    emp_ids = [emp["id"] for emp in employees]
    prefetched = await _prefetch_payroll_data(emp_ids, year, month_num, days_in_month)

    total_employees = 0
    total_salary = 0.0
    total_deductions = 0.0
    total_net_salary = 0.0
    total_lop_days = 0.0
    total_present_days = 0
    total_weekoff_pay = 0.0
    total_extra_pay = 0.0
    total_payable_days = 0.0
    
    for emp in employees:
        payroll = await calculate_payroll_for_employee(emp["id"], month, employee=emp, prefetched=prefetched)
        if payroll:
            total_employees += 1
            total_salary += payroll.get("monthly_salary", 0)
            total_deductions += payroll.get("lop_deduction", 0)
            total_net_salary += payroll.get("net_salary", 0)
            total_lop_days += payroll.get("lop", 0)
            total_present_days += payroll.get("present_days", 0)
            total_weekoff_pay += payroll.get("weekoff_pay", 0)
            total_extra_pay += payroll.get("extra_pay", 0)
            total_payable_days += payroll.get("final_payable_days", 0)
    
    return {
        "month": month,
        "total_employees": total_employees,
        "total_salary": round(total_salary, 2),
        "total_deductions": round(total_deductions, 2),
        "total_net_salary": round(total_net_salary, 2),
        "total_lop_days": total_lop_days,
        "total_present_days": total_present_days,
        "total_weekoff_pay": total_weekoff_pay,
        "total_extra_pay": total_extra_pay,
        "total_payable_days": round(total_payable_days, 2),
    }

# ============== SHIFT CONFIGURATION ROUTES ==============

@api_router.get("/config/shifts")
async def get_shift_configurations(current_user: dict = Depends(get_current_user)):
    """Get all available shift configurations"""
    shifts = []
    for shift_type, config in SHIFT_DEFINITIONS.items():
        shifts.append({
            "type": shift_type,
            "login_time": config.get("login_time"),
            "logout_time": config.get("logout_time"),
            "total_hours": config.get("total_hours"),
            "description": config.get("description")
        })
    return shifts

@api_router.get("/config/shift/{shift_type}")
async def get_shift_details(shift_type: str, current_user: dict = Depends(get_current_user)):
    """Get details for a specific shift type"""
    if shift_type not in SHIFT_DEFINITIONS:
        raise HTTPException(status_code=404, detail="Shift type not found")
    
    config = SHIFT_DEFINITIONS[shift_type]
    return {
        "type": shift_type,
        "login_time": config.get("login_time"),
        "logout_time": config.get("logout_time"),
        "total_hours": config.get("total_hours"),
        "description": config.get("description")
    }

@api_router.put("/employees/{employee_id}/shift")
async def update_employee_shift(
    employee_id: str,
    shift_data: ShiftConfigCreate,
    current_user: dict = Depends(get_current_user)
):
    """Update employee's shift configuration"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    employee = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    update_data = {"shift_type": shift_data.shift_type}
    
    if shift_data.shift_type == "Custom":
        if not shift_data.login_time or not shift_data.logout_time:
            raise HTTPException(status_code=400, detail="Custom shift requires login_time and logout_time")
        
        update_data["custom_login_time"] = shift_data.login_time
        update_data["custom_logout_time"] = shift_data.logout_time
        
        # Calculate total hours
        login_mins = parse_time_24h_to_minutes(shift_data.login_time)
        logout_mins = parse_time_24h_to_minutes(shift_data.logout_time)
        
        if logout_mins < login_mins:
            total_hours = (24 * 60 - login_mins + logout_mins) / 60
        else:
            total_hours = (logout_mins - login_mins) / 60
        
        update_data["custom_total_hours"] = total_hours
    else:
        # Clear custom fields for predefined shifts
        update_data["custom_login_time"] = None
        update_data["custom_logout_time"] = None
        update_data["custom_total_hours"] = None
    
    update_data["updated_at"] = get_ist_now().isoformat()
    
    await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    
    await log_audit(current_user["id"], "update_shift", "employee", employee_id, f"Updated shift to: {shift_data.shift_type}")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return serialize_doc(employee)

@api_router.put("/employees/{employee_id}/monthly-salary")
async def update_employee_monthly_salary_legacy(
    employee_id: str,
    monthly_salary: float = Query(..., ge=0),
    current_user: dict = Depends(get_current_user)
):
    """Update employee's monthly salary (Legacy route)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    employee = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    update_data = {
        "monthly_salary": monthly_salary,
        "updated_at": get_ist_now().isoformat()
    }
    
    await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    
    await log_audit(current_user["id"], "update_salary", "employee", employee_id, f"Updated salary to: {monthly_salary}")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return serialize_doc(employee)

# ============== AUDIT LOGS ==============

@api_router.get("/audit-logs")
async def get_audit_logs(
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    resource: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in SYSTEM_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {}
    if user_id:
        query["user_id"] = user_id
    if action:
        query["action"] = action
    if resource:
        query["resource"] = resource
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).to_list(500)
    return [serialize_doc(l) for l in logs]

# ============== CONFIG/LOOKUP ROUTES ==============

@api_router.get("/config/employment-types")
async def get_employment_types():
    return [EmploymentType.FULL_TIME, EmploymentType.PART_TIME, EmploymentType.CONTRACT, EmploymentType.INTERN]

@api_router.get("/config/employee-statuses")
async def get_employee_statuses():
    return [EmployeeStatus.ACTIVE, EmployeeStatus.INACTIVE, EmployeeStatus.RESIGNED]

@api_router.get("/config/tier-levels")
async def get_tier_levels():
    return [TierLevel.JUNIOR, TierLevel.MID, TierLevel.SENIOR, TierLevel.LEAD]

@api_router.get("/config/work-locations")
async def get_work_locations():
    return [WorkLocation.REMOTE, WorkLocation.OFFICE, WorkLocation.HYBRID]

@api_router.get("/config/user-roles")
async def get_user_roles():
    return [UserRole.HR, UserRole.SYSTEM_ADMIN, UserRole.OFFICE_ADMIN, UserRole.EMPLOYEE]

# ============== EMPLOYEE PORTAL MODELS ==============

class EmployeeLeaveCreate(BaseModel):
    leave_type: str  # Sick, Emergency, Preplanned, Casual, Annual
    leave_split: str = "Full Day"  # Full Day, First Half, Second Half
    start_date: str  # YYYY-MM-DD
    end_date: str  # YYYY-MM-DD
    reason: str
    supporting_document_url: Optional[str] = None
    supporting_document_name: Optional[str] = None

class EmployeeAttendanceRecord(BaseModel):
    date: str
    login: Optional[str] = None
    logout: Optional[str] = None
    total_hours: Optional[str] = None
    status: str  # Present, Late, Early Out, Absent, Leave, NA, Sunday

class PasswordChangeRequest(BaseModel):
    current_password: str
    new_password: str
    confirm_password: str

# ============== EMPLOYEE PORTAL ROUTES ==============

@api_router.post("/employee/change-password")
async def change_employee_password(data: PasswordChangeRequest, current_user: dict = Depends(get_current_user)):
    """Change employee's own password"""
    # Validate new password matches confirmation
    if data.new_password != data.confirm_password:
        raise HTTPException(status_code=400, detail="New passwords do not match")
    
    # Validate new password length
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters long")
    
    # Verify current password
    current_hash = hash_password(data.current_password)
    if current_user.get("password_hash") != current_hash:
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Update password in database
    new_hash = hash_password(data.new_password)
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"password_hash": new_hash}}
    )
    
    # Log the action
    await log_audit(current_user["id"], "change_password", "user", current_user["id"])
    
    return {"message": "Password changed successfully"}

@api_router.get("/employee/profile")
async def get_employee_profile(current_user: dict = Depends(get_current_user)):
    """Get logged-in employee's profile"""
    if not current_user.get("employee_id"):
        raise HTTPException(status_code=404, detail="No employee profile linked")
    
    employee = await db.employees.find_one({"id": current_user["employee_id"], "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee profile not found")
    
    return serialize_doc(employee)

@api_router.get("/employee/dashboard")
async def get_employee_dashboard(current_user: dict = Depends(get_current_user)):
    """Get employee dashboard data with summary cards and clock info"""
    if not current_user.get("employee_id"):
        raise HTTPException(status_code=404, detail="No employee profile linked")
    
    employee_id = current_user["employee_id"]
    employee = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get current month attendance stats (in IST)
    now = get_ist_now()
    current_month = now.strftime("%m-%Y")
    today_str = now.strftime("%d-%m-%Y")
    
    # Calculate month boundaries
    first_day = now.replace(day=1)
    if now.month == 12:
        next_month = now.replace(year=now.year + 1, month=1, day=1)
    else:
        next_month = now.replace(month=now.month + 1, day=1)
    last_day = next_month - timedelta(days=1)
    
    # Get all attendance records for this month
    attendance_records = await db.attendance.find({
        "employee_id": employee_id,
        "date": {"$regex": f"-{now.strftime('%m-%Y')}$"}
    }, {"_id": 0}).to_list(31)
    
    # Calculate summary stats
    # According to requirements: if "Early Out" has late login time, count in both Late AND Early Out
    active_days = 0
    inactive_days = 0
    late_arrivals = 0
    early_outs = 0
    
    # Define late threshold (10:00 AM)
    late_threshold_hour = 10
    late_threshold_minute = 0
    
    for record in attendance_records:
        status = record.get("status", "")
        check_in = record.get("check_in", "")
        
        # Check if login was late (after 10 AM)
        is_late_login = False
        if check_in:
            try:
                # Parse check-in time (format: "10:30 AM" or "09:15 AM")
                time_parts = check_in.upper().replace('.', ':').strip()
                if 'AM' in time_parts or 'PM' in time_parts:
                    is_pm = 'PM' in time_parts
                    time_str = time_parts.replace('AM', '').replace('PM', '').strip()
                    parts = time_str.split(':')
                    hour = int(parts[0])
                    minute = int(parts[1]) if len(parts) > 1 else 0
                    
                    # Convert to 24-hour format
                    if is_pm and hour != 12:
                        hour += 12
                    elif not is_pm and hour == 12:
                        hour = 0
                    
                    # Check if late (after 10 AM)
                    if hour > late_threshold_hour or (hour == late_threshold_hour and minute > late_threshold_minute):
                        is_late_login = True
            except:
                pass
        
        # Count based on status
        if status in ["Login", "Completed", "Present"]:
            active_days += 1
            if is_late_login:
                late_arrivals += 1
        elif status in ["Absent", "NA", "Not Logged"]:
            inactive_days += 1
        elif status == "Late Login" or status == "Late":
            late_arrivals += 1
            active_days += 1
        elif status == "Early Out":
            early_outs += 1
            active_days += 1
            # If early out AND also late login, count in late as well
            if is_late_login:
                late_arrivals += 1
    
    # Get today's attendance
    today_attendance = await db.attendance.find_one({
        "employee_id": employee_id,
        "date": today_str
    }, {"_id": 0})
    
    login_time = None
    logout_time = None
    hours_today = None
    
    if today_attendance:
        login_time = today_attendance.get("check_in")
        logout_time = today_attendance.get("check_out")
        hours_today = today_attendance.get("total_hours")
    
    return {
        "employee_name": employee.get("full_name"),
        "employee_id": employee.get("emp_id"),
        "department": employee.get("department"),
        "team": employee.get("team"),
        "summary": {
            "active_days": active_days,
            "inactive_days": inactive_days,
            "late_arrivals": late_arrivals,
            "early_outs": early_outs
        },
        "today": {
            "date": today_str,
            "login_time": login_time,
            "logout_time": logout_time,
            "hours_today": hours_today,
            "is_logged_in": login_time is not None,
            "is_logged_out": logout_time is not None
        },
        "current_month": now.strftime("%B %Y")
    }

@api_router.post("/employee/clock-in")
async def employee_clock_in(current_user: dict = Depends(get_current_user)):
    """Employee self clock-in"""
    if not current_user.get("employee_id"):
        raise HTTPException(status_code=404, detail="No employee profile linked")
    
    employee_id = current_user["employee_id"]
    employee = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    if not employee.get("attendance_tracking_enabled", True):
        raise HTTPException(status_code=400, detail="Attendance tracking disabled")
    
    now = get_ist_now()
    today_str = now.strftime("%d-%m-%Y")
    check_in_time = now.strftime("%I:%M %p")
    
    existing = await db.attendance.find_one({"employee_id": employee_id, "date": today_str})
    if existing and existing.get("check_in"):
        raise HTTPException(status_code=400, detail="Already clocked in today")
    
    # Determine status (Late if after 9:30 AM)
    status = "Login"
    if now.hour > 9 or (now.hour == 9 and now.minute > 30):
        status = "Late Login"
    
    attendance = Attendance(
        employee_id=employee_id,
        emp_name=employee["full_name"],
        team=employee["team"],
        department=employee["department"],
        date=today_str,
        check_in=check_in_time,
        status=status
    )
    doc = attendance.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.attendance.insert_one(doc.copy())
    
    return {"message": "Clocked in successfully", "time": check_in_time, "status": status}

@api_router.post("/employee/clock-out")
async def employee_clock_out(current_user: dict = Depends(get_current_user)):
    """Employee self clock-out"""
    if not current_user.get("employee_id"):
        raise HTTPException(status_code=404, detail="No employee profile linked")
    
    employee_id = current_user["employee_id"]
    now = get_ist_now()
    today_str = now.strftime("%d-%m-%Y")
    check_out_time = now.strftime("%I:%M %p")
    
    attendance = await db.attendance.find_one({"employee_id": employee_id, "date": today_str}, {"_id": 0})
    if not attendance:
        raise HTTPException(status_code=404, detail="No clock-in found for today")
    if attendance.get("check_out"):
        raise HTTPException(status_code=400, detail="Already clocked out")
    
    # Calculate total hours
    try:
        check_in_str = attendance.get("check_in", "")
        check_in_dt = datetime.strptime(f"{today_str} {check_in_str}", "%d-%m-%Y %I:%M %p")
        check_out_dt = datetime.strptime(f"{today_str} {check_out_time}", "%d-%m-%Y %I:%M %p")
        diff = check_out_dt - check_in_dt
        hours = diff.seconds // 3600
        minutes = (diff.seconds % 3600) // 60
        total_hours = f"{hours}h {minutes}m"
    except:
        total_hours = None
    
    # Determine status (Early Out if before 6:00 PM)
    status = "Completed"
    if now.hour < 18:
        status = "Early Out"
    
    await db.attendance.update_one(
        {"employee_id": employee_id, "date": today_str},
        {"$set": {"check_out": check_out_time, "total_hours": total_hours, "status": status}}
    )
    
    return {"message": "Clocked out successfully", "time": check_out_time, "total_hours": total_hours, "status": status}

@api_router.get("/employee/attendance")
async def get_employee_attendance(
    duration: Optional[str] = "this_week",  # this_week, last_week, this_month, last_month, custom
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status_filter: Optional[str] = "All",
    current_user: dict = Depends(get_current_user)
):
    """Get employee's own attendance records with filters"""
    if not current_user.get("employee_id"):
        raise HTTPException(status_code=404, detail="No employee profile linked")
    
    employee_id = current_user["employee_id"]
    now = get_ist_now()
    
    # If from_date/to_date are provided, treat as custom date range
    custom_from = from_date or start_date
    custom_to = to_date or end_date
    if custom_from and custom_to:
        duration = "custom"
    
    # Calculate date range based on duration
    if duration == "custom" and custom_from and custom_to:
        try:
            start = datetime.strptime(custom_from, "%d-%m-%Y")
            end = datetime.strptime(custom_to, "%d-%m-%Y")
        except:
            start = now - timedelta(days=now.weekday())
            end = start + timedelta(days=6)
    elif duration == "this_week":
        # Monday to Sunday of current week
        start = now - timedelta(days=now.weekday())
        end = start + timedelta(days=6)
    elif duration == "last_week":
        start = now - timedelta(days=now.weekday() + 7)
        end = start + timedelta(days=6)
    elif duration == "this_month":
        start = now.replace(day=1)
        if now.month == 12:
            end = now.replace(year=now.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = now.replace(month=now.month + 1, day=1) - timedelta(days=1)
    elif duration == "last_month":
        first_this_month = now.replace(day=1)
        end = first_this_month - timedelta(days=1)
        start = end.replace(day=1)
    else:
        # Default to this week
        start = now - timedelta(days=now.weekday())
        end = start + timedelta(days=6)
    
    # Generate all dates in range
    records = []
    current_date = start
    while current_date <= end:
        date_str = current_date.strftime("%d-%m-%Y")
        day_name = current_date.strftime("%A")
        
        # Check if Sunday
        if day_name == "Sunday":
            record = {
                "date": date_str,
                "day": day_name,
                "login": "-",
                "logout": "-",
                "total_hours": "-",
                "status": "Sunday"
            }
        elif current_date > now:
            # Future date
            record = {
                "date": date_str,
                "day": day_name,
                "login": "-",
                "logout": "-",
                "total_hours": "-",
                "status": "NA"
            }
        else:
            # Look up actual attendance (query both date formats for compatibility)
            date_str_iso = current_date.strftime("%Y-%m-%d")
            att = await db.attendance.find_one({
                "employee_id": employee_id,
                "date": {"$in": [date_str, date_str_iso]}
            }, {"_id": 0})
            
            # Check for approved leave
            leave = await db.leaves.find_one({
                "employee_id": employee_id,
                "status": "approved",
                "start_date": {"$lte": current_date.strftime("%Y-%m-%d")},
                "end_date": {"$gte": current_date.strftime("%Y-%m-%d")}
            }, {"_id": 0})
            
            if leave:
                # Show leave type instead of generic "Leave"
                leave_type = leave.get("leave_type", "Leave")
                record = {
                    "date": date_str,
                    "day": day_name,
                    "login": "-",
                    "logout": "-",
                    "total_hours": "-",
                    "status": f"{leave_type} Leave"
                }
            elif att:
                # Determine display status
                display_status = "Present"
                if att.get("status") == "Late Login":
                    display_status = "Late"
                elif att.get("status") == "Early Out":
                    display_status = "Early Out"
                elif att.get("status") == "Completed":
                    display_status = "Present"
                elif att.get("status") == "Login":
                    display_status = "Present"
                
                record = {
                    "date": date_str,
                    "day": day_name,
                    "login": att.get("check_in", "-"),
                    "logout": att.get("check_out", "-"),
                    "total_hours": att.get("total_hours", "-"),
                    "status": display_status
                }
            else:
                # No attendance record - absent
                record = {
                    "date": date_str,
                    "day": day_name,
                    "login": "-",
                    "logout": "-",
                    "total_hours": "-",
                    "status": "Absent"
                }
        
        # Apply status filter
        if status_filter == "All" or record["status"] == status_filter:
            records.append(record)
        
        current_date += timedelta(days=1)
    
    return records

@api_router.get("/employee/leaves")
async def get_employee_leaves(current_user: dict = Depends(get_current_user)):
    """Get employee's leave requests and history"""
    if not current_user.get("employee_id"):
        raise HTTPException(status_code=404, detail="No employee profile linked")
    
    employee_id = current_user["employee_id"]
    now = get_ist_now()
    today_str = now.strftime("%Y-%m-%d")
    
    # Get all leaves for this employee
    all_leaves = await db.leaves.find({"employee_id": employee_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Separate into requests (future/current) and history (past)
    requests = []
    history = []
    
    for leave in all_leaves:
        leave_data = serialize_doc(leave)
        # Convert date format for display
        if leave.get("start_date"):
            try:
                dt = datetime.strptime(leave["start_date"], "%Y-%m-%d")
                leave_data["display_date"] = dt.strftime("%d-%m-%Y")
            except:
                leave_data["display_date"] = leave["start_date"]
        
        if leave.get("start_date", "") >= today_str:
            requests.append(leave_data)
        else:
            history.append(leave_data)
    
    return {
        "requests": requests,
        "history": history,
        "requests_count": len(requests),
        "history_count": len(history)
    }

@api_router.post("/employee/leaves/apply")
async def apply_employee_leave(data: EmployeeLeaveCreate, current_user: dict = Depends(get_current_user)):
    """Apply for leave with optional document upload"""
    if not current_user.get("employee_id"):
        raise HTTPException(status_code=404, detail="No employee profile linked")
    
    employee_id = current_user["employee_id"]
    employee = await db.employees.find_one({"id": employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Validate reason length
    if not data.reason or len(data.reason.strip()) < 10:
        raise HTTPException(status_code=400, detail="Reason must be at least 10 characters")
    
    # Parse and validate leave dates
    try:
        start_dt = datetime.strptime(data.start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(data.end_date, "%Y-%m-%d")
        start_date = data.start_date
        end_date = data.end_date
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="End date cannot be before start date")
    
    today = get_ist_now().date()
    
    # JOB 6: Leave type-specific rules
    if data.leave_type == "Sick":
        # Sick leave: past + current only, no future
        if start_dt.date() > today:
            raise HTTPException(status_code=400, detail="Sick leave can only be applied for past or current dates")
    elif data.leave_type == "Casual":
        # Casual leave: minimum 4 working days before (exclude Sundays)
        if start_dt.date() <= today:
            raise HTTPException(status_code=400, detail="Casual leave cannot be applied for past or current dates")
        # Count working days between today and start date (excluding Sundays)
        working_days = 0
        check_date = today + timedelta(days=1)
        while check_date < start_dt.date():
            if check_date.weekday() != 6:  # 6 = Sunday
                working_days += 1
            check_date += timedelta(days=1)
        if working_days < 4:
            raise HTTPException(status_code=400, detail="Casual leave must be applied at least 4 working days in advance (excluding Sundays)")
    # Emergency leave: no restrictions
    
    # JOB 7: Single leave per day - check all dates in range
    current_date = start_dt
    while current_date <= end_dt:
        date_str = current_date.strftime("%Y-%m-%d")
        existing = await db.leaves.find_one({
            "employee_id": employee_id,
            "status": {"$ne": "rejected"},
            "$or": [
                {"start_date": {"$lte": date_str}, "end_date": {"$gte": date_str}},
            ]
        })
        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"Leave request already exists for {date_str}. Please edit existing request instead."
            )
        current_date += timedelta(days=1)
    
    # Calculate duration based on leave_split
    if data.leave_split in ["First Half", "Second Half"]:
        duration = "0.5 day(s)"
    else:
        duration_days = (end_dt - start_dt).days + 1
        duration = f"{duration_days} day{'s' if duration_days > 1 else ''}"
    
    # Create leave request
    leave = LeaveRequest(
        employee_id=employee_id,
        emp_name=employee["full_name"],
        team=employee["team"],
        department=employee["department"],
        leave_type=data.leave_type,
        leave_split=data.leave_split,
        start_date=start_date,
        end_date=end_date,
        duration=duration,
        reason=data.reason,
        supporting_document_url=data.supporting_document_url,
        supporting_document_name=data.supporting_document_name,
        status="pending"
    )
    
    doc = leave.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.leaves.insert_one(doc.copy())
    
    await log_audit(current_user["id"], "apply_leave", "leave", leave.id)
    
    # Notify HR about leave request
    asyncio.create_task(notify_role(
        UserRole.HR,
        "New Leave Request",
        f"{employee.get('full_name', 'Employee')} has submitted a {data.leave_type} leave request ({data.start_date} to {data.end_date}).",
        "action",
        "/leave"
    ))
    
    return {"message": "Leave request submitted successfully", "leave_id": leave.id}

@api_router.put("/employee/leaves/{leave_id}")
async def update_employee_leave(leave_id: str, data: EmployeeLeaveCreate, current_user: dict = Depends(get_current_user)):
    """Update a pending leave request"""
    if not current_user.get("employee_id"):
        raise HTTPException(status_code=404, detail="No employee profile linked")
    
    employee_id = current_user["employee_id"]
    
    leave = await db.leaves.find_one({"id": leave_id, "employee_id": employee_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
    
    if leave.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Can only edit pending leave requests")
    
    if not data.reason or len(data.reason.strip()) < 10:
        raise HTTPException(status_code=400, detail="Reason must be at least 10 characters")
    
    start_dt = datetime.strptime(data.start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(data.end_date, "%Y-%m-%d")
    
    if data.leave_split in ["First Half", "Second Half"]:
        duration_str = "0.5 day(s)"
    else:
        duration_str = f"{(end_dt - start_dt).days + 1} day(s)"
    
    update_data = {
        "leave_type": data.leave_type,
        "leave_split": data.leave_split,
        "start_date": data.start_date,
        "end_date": data.end_date,
        "duration": duration_str,
        "reason": data.reason,
        "supporting_document_url": data.supporting_document_url,
        "supporting_document_name": data.supporting_document_name
    }
    
    await db.leaves.update_one({"id": leave_id}, {"$set": update_data})
    await log_audit(current_user["id"], "update_leave", "leave", leave_id)
    
    return {"message": "Leave request updated successfully"}

# ============== ONBOARDING ROUTES ==============

@api_router.get("/onboarding/stats")
async def get_onboarding_stats(current_user: dict = Depends(get_current_user)):
    """Get onboarding statistics for dashboard"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    total = await db.onboarding.count_documents({})
    pending = await db.onboarding.count_documents({"status": OnboardingStatus.PENDING})
    in_progress = await db.onboarding.count_documents({"status": OnboardingStatus.IN_PROGRESS})
    under_review = await db.onboarding.count_documents({"status": OnboardingStatus.UNDER_REVIEW})
    approved = await db.onboarding.count_documents({"status": OnboardingStatus.APPROVED})
    rejected = await db.onboarding.count_documents({"status": OnboardingStatus.REJECTED})
    
    # Get pending document verifications
    pending_verifications = await db.onboarding_documents.count_documents({"status": DocumentStatus.UPLOADED})
    rejected_documents = await db.onboarding_documents.count_documents({"status": DocumentStatus.REJECTED})
    
    return {
        "total_employees": total,
        "pending": pending,
        "in_progress": in_progress,
        "under_review": under_review,
        "approved": approved,
        "rejected": rejected,
        "pending_verifications": pending_verifications,
        "rejected_documents": rejected_documents,
        "completion_rate": round((approved / total * 100) if total > 0 else 0, 1)
    }

@api_router.get("/verification/pending-count")
async def get_verification_pending_count(current_user: dict = Depends(get_current_user)):
    """Lightweight endpoint for sidebar badge - returns count of pending verifications"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    under_review = await db.onboarding.count_documents({"status": OnboardingStatus.UNDER_REVIEW})
    pending = await db.onboarding.count_documents({"status": OnboardingStatus.PENDING})
    in_progress = await db.onboarding.count_documents({"status": OnboardingStatus.IN_PROGRESS})
    return {"count": under_review + pending + in_progress}

@api_router.get("/onboarding/list")
async def get_onboarding_list(
    status: Optional[str] = None,
    department: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all onboarding records for HR review"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {}
    if status and status != "All":
        query["status"] = status
    if department and department != "All":
        query["department"] = department
    if search:
        query["$or"] = [
            {"emp_name": {"$regex": search, "$options": "i"}},
            {"emp_id": {"$regex": search, "$options": "i"}}
        ]
    
    records = await db.onboarding.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [serialize_doc(r) for r in records]

@api_router.get("/onboarding/employee/{employee_id}")
async def get_employee_onboarding(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get onboarding status and documents for an employee"""
    # Employees can only view their own, HR can view all
    if current_user["role"] == UserRole.EMPLOYEE:
        if current_user.get("employee_id") != employee_id:
            raise HTTPException(status_code=403, detail="Permission denied")
    
    onboarding = await db.onboarding.find_one({"employee_id": employee_id}, {"_id": 0})
    if not onboarding:
        raise HTTPException(status_code=404, detail="Onboarding record not found")
    
    # Get all documents
    documents = await db.onboarding_documents.find({"employee_id": employee_id}, {"_id": 0}).to_list(20)
    
    return {
        "onboarding": serialize_doc(onboarding),
        "documents": [serialize_doc(d) for d in documents],
        "required_documents": REQUIRED_DOCUMENTS
    }

@api_router.get("/onboarding/my-status")
async def get_my_onboarding_status(current_user: dict = Depends(get_current_user)):
    """Get current user's onboarding status"""
    employee_id = current_user.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="No employee linked to this user")
    
    onboarding = await db.onboarding.find_one({"employee_id": employee_id}, {"_id": 0})
    if not onboarding:
        # Create onboarding record if doesn't exist
        employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        onboarding = OnboardingRecord(
            employee_id=employee_id,
            emp_id=employee.get("emp_id"),
            emp_name=employee.get("full_name"),
            department=employee.get("department"),
            team=employee.get("team"),
            designation=employee.get("designation")
        )
        doc = onboarding.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.onboarding.insert_one(doc.copy())
        
        # Create document placeholders
        for req_doc in REQUIRED_DOCUMENTS:
            doc_record = OnboardingDocument(
                employee_id=employee_id,
                document_type=req_doc["type"],
                document_label=req_doc["label"]
            )
            doc_data = doc_record.model_dump()
            doc_data['created_at'] = doc_data['created_at'].isoformat()
            await db.onboarding_documents.insert_one(doc_data.copy())
        
        onboarding = doc
    
    # Get documents
    documents = await db.onboarding_documents.find({"employee_id": employee_id}, {"_id": 0}).to_list(20)
    
    # Check if user is in onboarding flow
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0})
    
    return {
        "onboarding": serialize_doc(onboarding) if isinstance(onboarding, dict) else onboarding,
        "documents": [serialize_doc(d) for d in documents],
        "required_documents": REQUIRED_DOCUMENTS,
        "is_first_login": user.get("is_first_login", True),
        "onboarding_completed": user.get("onboarding_status") == OnboardingStatus.APPROVED
    }

@api_router.post("/onboarding/upload-document")
async def upload_onboarding_document(data: DocumentUpload, current_user: dict = Depends(get_current_user)):
    """Upload a document for onboarding"""
    employee_id = current_user.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="No employee linked to this user")
    
    # Find existing document record
    doc_record = await db.onboarding_documents.find_one({
        "employee_id": employee_id,
        "document_type": data.document_type
    }, {"_id": 0})
    
    if not doc_record:
        raise HTTPException(status_code=404, detail="Document type not found")
    
    # Update document
    update_data = {
        "file_url": data.file_url,
        "file_public_id": data.file_public_id,
        "file_name": data.file_name,
        "status": DocumentStatus.UPLOADED,
        "uploaded_at": get_ist_now().isoformat()
    }
    
    await db.onboarding_documents.update_one(
        {"employee_id": employee_id, "document_type": data.document_type},
        {"$set": update_data}
    )
    
    # Update onboarding status to in_progress
    await db.onboarding.update_one(
        {"employee_id": employee_id},
        {"$set": {"status": OnboardingStatus.IN_PROGRESS, "updated_at": get_ist_now().isoformat()}}
    )
    
    await log_audit(current_user["id"], "upload_document", "onboarding", employee_id, f"Uploaded {data.document_type}")
    
    return {"message": "Document uploaded successfully"}

@api_router.post("/onboarding/submit")
async def submit_onboarding(current_user: dict = Depends(get_current_user)):
    """Submit onboarding for HR review"""
    employee_id = current_user.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="No employee linked to this user")
    
    # Check all required documents are uploaded
    documents = await db.onboarding_documents.find({"employee_id": employee_id}, {"_id": 0}).to_list(20)
    
    required_types = [d["type"] for d in REQUIRED_DOCUMENTS if d["required"]]
    uploaded_types = [d["document_type"] for d in documents if d.get("status") in [DocumentStatus.UPLOADED, DocumentStatus.VERIFIED]]
    
    missing = [t for t in required_types if t not in uploaded_types]
    if missing:
        missing_labels = [d["label"] for d in REQUIRED_DOCUMENTS if d["type"] in missing]
        raise HTTPException(status_code=400, detail=f"Missing required documents: {', '.join(missing_labels)}")
    
    # Update onboarding status
    await db.onboarding.update_one(
        {"employee_id": employee_id},
        {"$set": {
            "status": OnboardingStatus.UNDER_REVIEW,
            "submitted_at": get_ist_now().isoformat(),
            "updated_at": get_ist_now().isoformat()
        }}
    )
    
    await log_audit(current_user["id"], "submit_onboarding", "onboarding", employee_id)
    
    return {"message": "Onboarding submitted for review"}

@api_router.post("/onboarding/verify-document")
async def verify_onboarding_document(data: DocumentVerification, current_user: dict = Depends(get_current_user)):
    """HR verifies/rejects a document"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    doc = await db.onboarding_documents.find_one({"id": data.document_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    if data.status not in [DocumentStatus.VERIFIED, DocumentStatus.REJECTED]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    update_data = {
        "status": data.status,
        "verified_at": get_ist_now().isoformat(),
        "verified_by": current_user["id"]
    }
    
    if data.status == DocumentStatus.REJECTED:
        update_data["rejection_reason"] = data.rejection_reason
    
    await db.onboarding_documents.update_one(
        {"id": data.document_id},
        {"$set": update_data}
    )
    
    await log_audit(current_user["id"], f"verify_document_{data.status}", "onboarding", doc["employee_id"], data.document_type if hasattr(data, 'document_type') else None)
    
    return {"message": f"Document {data.status}"}

@api_router.post("/onboarding/approve/{employee_id}")
async def approve_onboarding(employee_id: str, data: OnboardingApproval, current_user: dict = Depends(get_current_user)):
    """HR approves/rejects entire onboarding"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    onboarding = await db.onboarding.find_one({"employee_id": employee_id}, {"_id": 0})
    if not onboarding:
        raise HTTPException(status_code=404, detail="Onboarding record not found")
    
    if data.status not in [OnboardingStatus.APPROVED, OnboardingStatus.REJECTED]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    # Update onboarding record
    update_data = {
        "status": data.status,
        "reviewed_at": get_ist_now().isoformat(),
        "reviewed_by": current_user["id"],
        "review_notes": data.review_notes,
        "updated_at": get_ist_now().isoformat()
    }
    
    await db.onboarding.update_one({"employee_id": employee_id}, {"$set": update_data})
    
    # Update employee's onboarding status
    employee_update = {"onboarding_status": data.status, "updated_at": get_ist_now().isoformat()}
    if data.status == OnboardingStatus.APPROVED:
        employee_update["onboarding_completed_at"] = get_ist_now().isoformat()
    
    await db.employees.update_one({"id": employee_id}, {"$set": employee_update})
    
    # Update user's onboarding status
    await db.users.update_one(
        {"employee_id": employee_id},
        {"$set": {"onboarding_status": data.status, "is_first_login": False}}
    )
    
    # Send notification email
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if employee:
        status_text = "approved" if data.status == OnboardingStatus.APPROVED else "requires attention"
        subject = f"Onboarding {status_text.title()} - {employee.get('full_name')}"
        html = get_onboarding_status_email(
            employee.get("full_name"),
            data.status,
            data.review_notes
        )
        asyncio.create_task(send_email_notification(employee.get("official_email"), subject, html))
    
    await log_audit(current_user["id"], f"approve_onboarding_{data.status}", "onboarding", employee_id, data.review_notes)
    
    return {"message": f"Onboarding {data.status}"}

@api_router.post("/onboarding/request-reupload/{employee_id}")
async def request_document_reupload(employee_id: str, document_type: str, reason: str, current_user: dict = Depends(get_current_user)):
    """HR requests re-upload of a specific document"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    doc = await db.onboarding_documents.find_one({
        "employee_id": employee_id,
        "document_type": document_type
    }, {"_id": 0})
    
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    await db.onboarding_documents.update_one(
        {"employee_id": employee_id, "document_type": document_type},
        {"$set": {
            "status": DocumentStatus.REJECTED,
            "rejection_reason": reason,
            "verified_at": get_ist_now().isoformat(),
            "verified_by": current_user["id"]
        }}
    )
    
    # Update onboarding status back to in_progress
    await db.onboarding.update_one(
        {"employee_id": employee_id},
        {"$set": {"status": OnboardingStatus.IN_PROGRESS, "updated_at": get_ist_now().isoformat()}}
    )
    
    return {"message": "Re-upload requested"}

# ============== ISSUE TICKET ROUTES ==============

@api_router.get("/issue-tickets/categories")
async def get_ticket_categories(current_user: dict = Depends(get_current_user)):
    """Get all ticket categories and subcategories"""
    categories = []
    for category, subcategories in TICKET_SUBCATEGORIES.items():
        categories.append({
            "category": category,
            "subcategories": subcategories,
            "assigned_role": TICKET_DEPARTMENT_ROLES.get(category)
        })
    return categories

@api_router.get("/issue-tickets")
async def get_issue_tickets(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    category: Optional[str] = None,
    assigned_department: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    current_user: dict = Depends(get_current_user)
):
    """Get issue tickets based on user role"""
    query = {}
    
    # Role-based filtering
    user_role = current_user.get("role")
    
    # Employees can only see their own tickets
    if user_role == UserRole.EMPLOYEE:
        query["employee_id"] = current_user.get("employee_id")
    # Department admins see tickets assigned to their department
    elif user_role in ["it_admin", "hr_admin", "finance_admin", "admin_dept", "compliance_officer", "operations_manager"]:
        query["assigned_department"] = user_role
    # HR, system_admin, office_admin can see all tickets
    
    # Apply filters
    if status and status != "All":
        query["status"] = status
    if priority and priority != "All":
        query["priority"] = priority
    if category and category != "All":
        query["category"] = category
    if assigned_department and assigned_department != "All":
        query["assigned_department"] = assigned_department
    if search:
        query["$or"] = [
            {"ticket_number": {"$regex": search, "$options": "i"}},
            {"subject": {"$regex": search, "$options": "i"}},
            {"emp_name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * limit
    total = await db.issue_tickets.count_documents(query)
    tickets = await db.issue_tickets.find(query, {"_id": 0}).skip(skip).limit(limit).sort("created_at", -1).to_list(limit)
    
    return {
        "tickets": [serialize_doc(t) for t in tickets],
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/issue-tickets/stats")
async def get_issue_ticket_stats(current_user: dict = Depends(get_current_user)):
    """Get comprehensive ticket statistics"""
    user_role = current_user.get("role")
    query = {}
    
    # Role-based filtering for stats
    if user_role == UserRole.EMPLOYEE:
        query["employee_id"] = current_user.get("employee_id")
    elif user_role in ["it_admin", "hr_admin", "finance_admin", "admin_dept", "compliance_officer", "operations_manager"]:
        query["assigned_department"] = user_role
    
    # Get counts by status
    total = await db.issue_tickets.count_documents(query)
    by_status = {}
    for status in [TicketStatus.OPEN, TicketStatus.IN_PROGRESS, TicketStatus.WAITING_APPROVAL, 
                   TicketStatus.ON_HOLD, TicketStatus.RESOLVED, TicketStatus.CLOSED, TicketStatus.REJECTED]:
        count = await db.issue_tickets.count_documents({**query, "status": status})
        by_status[status] = count
    
    # Get counts by priority
    by_priority = {}
    for priority in [TicketPriority.HIGH, TicketPriority.MEDIUM, TicketPriority.LOW]:
        count = await db.issue_tickets.count_documents({**query, "priority": priority})
        by_priority[priority] = count
    
    # Get counts by category (admin only)
    by_category = {}
    if user_role not in [UserRole.EMPLOYEE]:
        for category in TICKET_SUBCATEGORIES.keys():
            count = await db.issue_tickets.count_documents({**query, "category": category})
            by_category[category] = count
    
    # Get average resolution time (only for resolved/closed tickets)
    avg_resolution_hours = None
    resolved_tickets = await db.issue_tickets.find({
        **query, 
        "status": {"$in": [TicketStatus.RESOLVED, TicketStatus.CLOSED]},
        "resolved_at": {"$ne": None}
    }, {"_id": 0, "created_at": 1, "resolved_at": 1}).to_list(500)
    
    if resolved_tickets:
        total_hours = 0
        for ticket in resolved_tickets:
            try:
                created = datetime.fromisoformat(ticket["created_at"].replace("Z", "+00:00")) if isinstance(ticket["created_at"], str) else ticket["created_at"]
                resolved = datetime.fromisoformat(ticket["resolved_at"].replace("Z", "+00:00")) if isinstance(ticket["resolved_at"], str) else ticket["resolved_at"]
                total_hours += (resolved - created).total_seconds() / 3600
            except:
                pass
        avg_resolution_hours = round(total_hours / len(resolved_tickets), 1) if resolved_tickets else None
    
    return {
        "total": total,
        "by_status": by_status,
        "by_priority": by_priority,
        "by_category": by_category,
        "avg_resolution_hours": avg_resolution_hours
    }

@api_router.get("/issue-tickets/{ticket_id}")
async def get_issue_ticket(ticket_id: str, current_user: dict = Depends(get_current_user)):
    """Get single ticket details"""
    ticket = await db.issue_tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Check access
    user_role = current_user.get("role")
    if user_role == UserRole.EMPLOYEE:
        if ticket.get("employee_id") != current_user.get("employee_id"):
            raise HTTPException(status_code=403, detail="Access denied")
    
    return serialize_doc(ticket)

@api_router.post("/issue-tickets")
async def create_issue_ticket(data: TicketCreate, current_user: dict = Depends(get_current_user)):
    """Create a new issue ticket"""
    # Validate category and subcategory
    if data.category not in TICKET_SUBCATEGORIES:
        raise HTTPException(status_code=400, detail="Invalid category")
    if data.subcategory not in TICKET_SUBCATEGORIES[data.category]:
        raise HTTPException(status_code=400, detail="Invalid subcategory for this category")
    
    # Determine employee info
    employee_id = data.employee_id if data.employee_id else current_user.get("employee_id")
    employee = None
    created_by = None
    created_by_name = None
    
    if employee_id:
        employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    
    # If admin creating on behalf of employee
    if data.employee_id and current_user.get("role") in [UserRole.HR]:
        created_by = current_user["id"]
        created_by_name = current_user.get("name")
    
    # Generate ticket number
    ticket_number = await generate_ticket_number()
    
    # Determine assigned department
    assigned_department = TICKET_DEPARTMENT_ROLES.get(data.category, "hr_admin")
    
    # Process attachments
    attachments = []
    if data.attachments:
        for att in data.attachments:
            attachments.append({
                "id": str(uuid.uuid4()),
                "file_url": att.get("file_url"),
                "file_name": att.get("file_name"),
                "file_type": att.get("file_type"),
                "file_public_id": att.get("file_public_id"),
                "uploaded_at": get_ist_now().isoformat()
            })
    
    # Create initial status history
    status_history = [{
        "id": str(uuid.uuid4()),
        "status": TicketStatus.OPEN,
        "updated_by": current_user["id"],
        "updated_by_name": current_user.get("name"),
        "notes": "Ticket created",
        "updated_at": get_ist_now().isoformat()
    }]
    
    ticket = Ticket(
        ticket_number=ticket_number,
        employee_id=employee_id or current_user["id"],
        emp_name=employee.get("full_name") if employee else current_user.get("name"),
        emp_email=employee.get("official_email") if employee else current_user.get("email"),
        department=employee.get("department") if employee else current_user.get("department", "N/A"),
        team=employee.get("team") if employee else current_user.get("team"),
        category=data.category,
        subcategory=data.subcategory,
        subject=data.subject,
        description=data.description,
        priority=data.priority,
        assigned_department=assigned_department,
        attachments=attachments,
        status_history=status_history,
        created_by=created_by,
        created_by_name=created_by_name
    )
    
    doc = ticket.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.issue_tickets.insert_one(doc.copy())
    
    await log_audit(current_user["id"], "create_issue_ticket", "issue_ticket", ticket.id, f"Ticket {ticket_number} created")
    
    return serialize_doc(doc)

@api_router.put("/issue-tickets/{ticket_id}/status")
async def update_issue_ticket_status(
    ticket_id: str, 
    data: TicketStatusUpdateRequest,
    current_user: dict = Depends(get_current_user)
):
    """Update ticket status"""
    ticket = await db.issue_tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Validate status
    valid_statuses = [TicketStatus.OPEN, TicketStatus.IN_PROGRESS, TicketStatus.WAITING_APPROVAL,
                     TicketStatus.ON_HOLD, TicketStatus.RESOLVED, TicketStatus.CLOSED, TicketStatus.REJECTED]
    if data.status not in valid_statuses:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    # Check permissions
    user_role = current_user.get("role")
    # Employees can only close their own resolved tickets
    if user_role == UserRole.EMPLOYEE:
        if ticket.get("employee_id") != current_user.get("employee_id"):
            raise HTTPException(status_code=403, detail="Access denied")
        if data.status != TicketStatus.CLOSED or ticket.get("status") != TicketStatus.RESOLVED:
            raise HTTPException(status_code=403, detail="Employees can only close resolved tickets")
    
    # Create status update entry
    status_update = {
        "id": str(uuid.uuid4()),
        "status": data.status,
        "updated_by": current_user["id"],
        "updated_by_name": current_user.get("name"),
        "notes": data.notes,
        "updated_at": get_ist_now().isoformat()
    }
    
    update_data = {
        "status": data.status,
        "updated_at": get_ist_now().isoformat()
    }
    
    # Add resolution info for resolved/closed statuses
    if data.status in [TicketStatus.RESOLVED, TicketStatus.CLOSED]:
        if data.resolution:
            update_data["resolution"] = data.resolution
        update_data["resolved_at"] = get_ist_now().isoformat()
        update_data["resolved_by"] = current_user["id"]
        update_data["resolved_by_name"] = current_user.get("name")
    
    await db.issue_tickets.update_one(
        {"id": ticket_id},
        {
            "$set": update_data,
            "$push": {"status_history": status_update}
        }
    )
    
    await log_audit(current_user["id"], "update_ticket_status", "issue_ticket", ticket_id, f"Status: {data.status}")
    
    return {"message": f"Ticket status updated to {data.status}"}

@api_router.put("/issue-tickets/{ticket_id}/assign")
async def assign_issue_ticket(
    ticket_id: str,
    data: TicketAssignRequest,
    current_user: dict = Depends(get_current_user)
):
    """Assign ticket to a specific user"""
    if current_user["role"] not in [UserRole.HR, 
                                    "it_admin", "hr_admin", "finance_admin", "admin_dept", 
                                    "compliance_officer", "operations_manager"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    ticket = await db.issue_tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Get assignee info
    assignee = await db.users.find_one({"id": data.assigned_to}, {"_id": 0})
    if not assignee:
        raise HTTPException(status_code=404, detail="Assignee not found")
    
    # Create status update entry
    status_update = {
        "id": str(uuid.uuid4()),
        "status": ticket.get("status"),
        "updated_by": current_user["id"],
        "updated_by_name": current_user.get("name"),
        "notes": f"Ticket assigned to {assignee.get('name')}",
        "updated_at": get_ist_now().isoformat()
    }
    
    await db.issue_tickets.update_one(
        {"id": ticket_id},
        {
            "$set": {
                "assigned_to": data.assigned_to,
                "assigned_to_name": assignee.get("name"),
                "updated_at": get_ist_now().isoformat()
            },
            "$push": {"status_history": status_update}
        }
    )
    
    return {"message": f"Ticket assigned to {assignee.get('name')}"}

@api_router.post("/issue-tickets/{ticket_id}/feedback")
async def submit_ticket_feedback(
    ticket_id: str,
    data: TicketFeedbackRequest,
    current_user: dict = Depends(get_current_user)
):
    """Submit feedback for a resolved/closed ticket"""
    ticket = await db.issue_tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Only ticket owner can submit feedback
    if ticket.get("employee_id") != current_user.get("employee_id"):
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Only for resolved/closed tickets
    if ticket.get("status") not in [TicketStatus.RESOLVED, TicketStatus.CLOSED]:
        raise HTTPException(status_code=400, detail="Can only submit feedback for resolved/closed tickets")
    
    # Validate rating
    if data.rating < 1 or data.rating > 5:
        raise HTTPException(status_code=400, detail="Rating must be between 1 and 5")
    
    feedback = {
        "rating": data.rating,
        "comment": data.comment,
        "submitted_at": get_ist_now().isoformat()
    }
    
    await db.issue_tickets.update_one(
        {"id": ticket_id},
        {"$set": {"feedback": feedback, "updated_at": get_ist_now().isoformat()}}
    )
    
    return {"message": "Feedback submitted successfully"}

@api_router.post("/issue-tickets/{ticket_id}/attachment")
async def add_ticket_attachment(
    ticket_id: str,
    file_url: str,
    file_name: str,
    file_type: str,
    file_public_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Add attachment to an existing ticket"""
    ticket = await db.issue_tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Check access
    if current_user.get("role") == UserRole.EMPLOYEE:
        if ticket.get("employee_id") != current_user.get("employee_id"):
            raise HTTPException(status_code=403, detail="Access denied")
    
    attachment = {
        "id": str(uuid.uuid4()),
        "file_url": file_url,
        "file_name": file_name,
        "file_type": file_type,
        "file_public_id": file_public_id,
        "uploaded_at": get_ist_now().isoformat()
    }
    
    await db.issue_tickets.update_one(
        {"id": ticket_id},
        {
            "$push": {"attachments": attachment},
            "$set": {"updated_at": get_ist_now().isoformat()}
        }
    )
    
    return {"message": "Attachment added", "attachment": attachment}

# Keep legacy ticket routes for backward compatibility
@api_router.get("/tickets")
async def get_tickets(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all tickets (HR) or own tickets (employee) - Legacy endpoint"""
    query = {}
    
    if current_user["role"] == UserRole.EMPLOYEE:
        query["employee_id"] = current_user.get("employee_id")
    
    if status and status != "All":
        query["status"] = status
    if priority and priority != "All":
        query["priority"] = priority
    
    tickets = await db.tickets.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [serialize_doc(t) for t in tickets]

@api_router.post("/tickets")
async def create_ticket_legacy(data: dict, current_user: dict = Depends(get_current_user)):
    """Create a new support ticket - Legacy endpoint"""
    employee_id = current_user.get("employee_id")
    employee = None
    
    if employee_id:
        employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    
    ticket_doc = {
        "id": str(uuid.uuid4()),
        "employee_id": employee_id or current_user["id"],
        "emp_name": employee.get("full_name") if employee else current_user.get("name"),
        "department": employee.get("department") if employee else current_user.get("department", "N/A"),
        "subject": data.get("subject", ""),
        "description": data.get("description", ""),
        "priority": data.get("priority", "medium"),
        "status": "open",
        "assigned_to": None,
        "resolution": None,
        "created_at": get_ist_now().isoformat(),
        "updated_at": get_ist_now().isoformat(),
        "resolved_at": None
    }
    
    await db.tickets.insert_one(ticket_doc.copy())
    await log_audit(current_user["id"], "create_ticket", "ticket", ticket_doc["id"])
    
    return serialize_doc(ticket_doc)

@api_router.put("/tickets/{ticket_id}/status")
async def update_ticket_status(ticket_id: str, status: str, resolution: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Update ticket status (HR only) - Legacy endpoint"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    update_data = {
        "status": status,
        "updated_at": get_ist_now().isoformat()
    }
    
    if status in ["resolved", "closed"]:
        update_data["resolved_at"] = get_ist_now().isoformat()
        update_data["resolution"] = resolution
    
    await db.tickets.update_one({"id": ticket_id}, {"$set": update_data})
    
    return {"message": f"Ticket status updated to {status}"}

@api_router.get("/tickets/stats")
async def get_ticket_stats(current_user: dict = Depends(get_current_user)):
    """Get ticket statistics - Legacy endpoint"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    total = await db.tickets.count_documents({})
    open_tickets = await db.tickets.count_documents({"status": "open"})
    in_progress = await db.tickets.count_documents({"status": "in_progress"})
    resolved = await db.tickets.count_documents({"status": "resolved"})
    
    return {
        "total": total,
        "open": open_tickets,
        "in_progress": in_progress,
        "resolved": resolved
    }

# ============== AUDIT LOG ROUTES ==============

@api_router.get("/audit-logs")
async def get_audit_logs(
    resource: Optional[str] = None,
    action: Optional[str] = None,
    limit: int = Query(100, ge=1, le=500),
    current_user: dict = Depends(get_current_user)
):
    """Get audit logs (admin only)"""
    if current_user["role"] not in SYSTEM_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {}
    if resource:
        query["resource"] = resource
    if action:
        query["action"] = action
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    
    # Add user names
    for log in logs:
        user = await db.users.find_one({"id": log["user_id"]}, {"_id": 0, "name": 1})
        log["user_name"] = user.get("name") if user else "Unknown"
    
    return [serialize_doc(l) for l in logs]

# ============== SALARY MANAGEMENT ROUTES ==============

def calculate_salary_structure(annual_ctc: float) -> dict:
    """Calculate salary breakdown from annual CTC based on professional structure"""
    monthly_ctc = annual_ctc / 12
    
    # Base Components (A)
    basic = round(monthly_ctc * 0.30, 2)  # 30% of Monthly CTC
    hra = round(basic * 0.50, 2)  # 50% of Basic
    base_components = basic + hra
    
    # Variable Compensation (20% of CTC)
    variable_compensation = round(monthly_ctc * 0.20, 2)
    
    # Retirement Benefits (C)
    pf_basic = min(basic, 15000)
    pf_employee = round(pf_basic * 0.12, 2)
    pf_employer = round(pf_basic * 0.12, 2)
    gratuity = round(basic * 15 / 26 / 12, 2)  # Gratuity Act formula
    retirement_benefits = pf_employer + gratuity
    
    # Fixed known allowances in Basket (B)
    lta = round(basic * 0.0567, 2)
    phone_internet = 1100
    bonus = round(basic * 0.10, 2)
    stay_travel = round(basic * 0.30, 2)
    food_reimbursement = 1210
    medical_allowance = 1250
    conveyance = 1600
    
    known_allowances = lta + phone_internet + bonus + stay_travel + food_reimbursement + medical_allowance + conveyance
    
    # Special Allowance = balancing figure so total equals CTC exactly
    fixed_target = monthly_ctc - variable_compensation
    special_allowance = round(fixed_target - base_components - known_allowances - retirement_benefits, 2)
    if special_allowance < 0:
        special_allowance = 0
    
    basket_allowances = known_allowances + special_allowance
    fixed_compensation = base_components + basket_allowances + retirement_benefits
    
    gross_salary = base_components + basket_allowances
    
    # ESI
    if gross_salary < 21000:
        esi_employee = round(gross_salary * 0.0075, 2)
        esi_employer = round(gross_salary * 0.0325, 2)
    else:
        esi_employee = 0
        esi_employer = 0
    
    # Professional Tax (standard 200, varies by state)
    professional_tax = 200
    
    # Total Deductions (Employee side)
    total_deductions = round(pf_employee + esi_employee + professional_tax, 2)
    
    # Net Salary (Take Home)
    net_salary = round(gross_salary - total_deductions, 2)
    
    return {
        "annual_ctc": annual_ctc,
        "monthly_ctc": round(monthly_ctc, 2),
        
        # Base Components (A)
        "basic": basic,
        "hra": hra,
        "base_components_total": round(base_components, 2),
        
        # Basket of Allowances (B)
        "lta": lta,
        "phone_internet": phone_internet,
        "bonus": bonus,
        "stay_travel": stay_travel,
        "special_allowance": special_allowance,
        "food_reimbursement": food_reimbursement,
        "medical_allowance": medical_allowance,
        "conveyance": conveyance,
        "basket_allowances_total": round(basket_allowances, 2),
        "da": 0,  # Keep for backward compatibility
        "other_allowances": 0,
        
        # Retirement Benefits (C)
        "pf_employer": pf_employer,
        "gratuity": gratuity,
        "retirement_benefits_total": round(retirement_benefits, 2),
        
        # Totals
        "fixed_compensation": round(fixed_compensation, 2),
        "variable_compensation": round(variable_compensation, 2),
        "gross_salary": round(gross_salary, 2),
        
        # Deductions
        "pf_employee": pf_employee,
        "esi_employee": esi_employee,
        "esi_employer": esi_employer,
        "professional_tax": professional_tax,
        "tds": 0,
        "other_deductions": 0,
        "total_deductions": total_deductions,
        
        # Net
        "net_salary": net_salary
    }

def calculate_salary_structure_research(annual_ctc: float) -> dict:
    """CTC-based salary breakdown for Research designation employees.
    Uses B-percentage allocation: LTA=5.6%B, Bonus=9.9%B, Stay=30%B, Special=remainder.
    No Medical Allowance or Conveyance. PF Employer & Employee fixed at 1800."""
    monthly_ctc = annual_ctc / 12

    # Variable = 20% of CTC, Fixed = 80% of CTC
    variable_compensation = round(monthly_ctc * 0.20, 2)
    fixed = round(monthly_ctc * 0.80, 2)

    # A: Base Components (Basic = 30% CTC, HRA = 50% Basic → A = 45% CTC)
    basic = round(monthly_ctc * 0.30, 2)
    hra = round(basic * 0.50, 2)
    base_components = round(basic + hra, 2)

    # C: Retirement Benefits (PF fixed 1800, Gratuity via Gratuity Act)
    pf_employer = 1800
    pf_employee = 1800
    gratuity = round(basic * 15 / 26 / 12, 2)
    retirement_benefits = round(pf_employer + gratuity, 2)

    # B: Basket of Allowances = Fixed - A - C
    basket_total = round(fixed - base_components - retirement_benefits, 2)
    if basket_total < 0:
        basket_total = 0

    # B components: percentages of B + fixed items
    lta = round(basket_total * 0.056, 2)
    phone_internet = 1100
    bonus = round(basket_total * 0.099, 2)
    stay_travel = round(basket_total * 0.30, 2)
    food_reimbursement = 1210
    # Special Allowance = balancing figure within B
    special_allowance = round(basket_total - lta - phone_internet - bonus - stay_travel - food_reimbursement, 2)
    if special_allowance < 0:
        special_allowance = 0

    fixed_compensation = round(base_components + basket_total + retirement_benefits, 2)
    gross_salary = round(base_components + basket_total, 2)

    # ESI (only if gross < 21000)
    esi_employee = round(gross_salary * 0.0075, 2) if gross_salary < 21000 else 0
    esi_employer = round(gross_salary * 0.0325, 2) if gross_salary < 21000 else 0

    professional_tax = 200
    total_deductions = round(pf_employee + esi_employee + professional_tax, 2)
    net_salary = round(gross_salary - total_deductions, 2)

    return {
        "annual_ctc": annual_ctc,
        "monthly_ctc": round(monthly_ctc, 2),
        "basic": basic,
        "hra": hra,
        "base_components_total": base_components,
        "lta": lta,
        "phone_internet": phone_internet,
        "bonus": bonus,
        "stay_travel": stay_travel,
        "special_allowance": special_allowance,
        "food_reimbursement": food_reimbursement,
        "medical_allowance": 0,
        "conveyance": 0,
        "basket_allowances_total": basket_total,
        "da": 0,
        "other_allowances": 0,
        "pf_employer": pf_employer,
        "gratuity": gratuity,
        "retirement_benefits_total": retirement_benefits,
        "fixed_compensation": fixed_compensation,
        "variable_compensation": variable_compensation,
        "gross_salary": gross_salary,
        "pf_employee": pf_employee,
        "esi_employee": esi_employee,
        "esi_employer": esi_employer,
        "professional_tax": professional_tax,
        "tds": 0,
        "other_deductions": 0,
        "total_deductions": total_deductions,
        "net_salary": net_salary
    }

def get_salary_calculator(designation: str):
    """Return the appropriate salary calculator based on employee designation."""
    if designation and "research" in designation.lower():
        return calculate_salary_structure_research
    return calculate_salary_structure

@api_router.get("/employees/{employee_id}/salary")
async def get_employee_salary(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get employee salary structure"""
    # Check permissions
    if current_user["role"] == UserRole.EMPLOYEE:
        if current_user.get("employee_id") != employee_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get salary structure
    salary = await db.salary_structures.find_one({"employee_id": employee_id}, {"_id": 0})
    
    if not salary:
        # Create default salary structure from monthly_salary
        monthly_salary = employee.get("monthly_salary", 0)
        if monthly_salary > 0:
            annual_ctc = monthly_salary * 12
            calc_fn = get_salary_calculator(employee.get("designation", ""))
            salary_data = calc_fn(annual_ctc)
            salary_data["id"] = str(uuid.uuid4())
            salary_data["employee_id"] = employee_id
            salary_data["effective_from"] = employee.get("date_of_joining", get_ist_today())
            salary_data["created_at"] = get_ist_now().isoformat()
            salary_data["updated_at"] = get_ist_now().isoformat()
            await db.salary_structures.insert_one(salary_data.copy())
            salary = salary_data
        else:
            return {"employee_id": employee_id, "salary": None, "message": "No salary configured"}
    
    return {
        "employee_id": employee_id,
        "employee_name": employee.get("full_name"),
        "salary": serialize_doc(salary)
    }

@api_router.put("/employees/{employee_id}/salary")
async def update_employee_salary(
    employee_id: str, 
    data: SalaryStructureUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update employee salary structure (Admin only)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # If annual_ctc is provided, recalculate entire structure
    if data.annual_ctc is not None:
        calc_fn = get_salary_calculator(employee.get("designation", ""))
        salary_data = calc_fn(data.annual_ctc)
        salary_data["updated_at"] = get_ist_now().isoformat()
        
        existing = await db.salary_structures.find_one({"employee_id": employee_id})
        if existing:
            await db.salary_structures.update_one(
                {"employee_id": employee_id},
                {"$set": salary_data}
            )
        else:
            salary_data["id"] = str(uuid.uuid4())
            salary_data["employee_id"] = employee_id
            salary_data["effective_from"] = get_ist_today()
            salary_data["created_at"] = get_ist_now().isoformat()
            await db.salary_structures.insert_one(salary_data.copy())
        
        # Update employee's monthly_salary field
        await db.employees.update_one(
            {"id": employee_id},
            {"$set": {"monthly_salary": salary_data["gross_salary"]}}
        )
        
        await log_audit(current_user["id"], "update_salary", "salary", employee_id, f"Updated CTC to {data.annual_ctc}")
        
        return {"message": "Salary structure updated", "salary": salary_data}
    
    # Otherwise update individual components
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        update_data["updated_at"] = get_ist_now().isoformat()
        
        # Recalculate totals
        existing = await db.salary_structures.find_one({"employee_id": employee_id}, {"_id": 0})
        if existing:
            merged = {**existing, **update_data}
            
            # Recalculate gross
            gross = (merged.get("basic", 0) + merged.get("hra", 0) + merged.get("da", 0) + 
                    merged.get("conveyance", 0) + merged.get("medical_allowance", 0) + 
                    merged.get("special_allowance", 0) + merged.get("other_allowances", 0))
            
            # Recalculate deductions
            total_deductions = (merged.get("pf_employee", 0) + merged.get("esi_employee", 0) + 
                              merged.get("professional_tax", 0) + merged.get("tds", 0) + 
                              merged.get("other_deductions", 0))
            
            update_data["gross_salary"] = round(gross, 2)
            update_data["total_deductions"] = round(total_deductions, 2)
            update_data["net_salary"] = round(gross - total_deductions, 2)
            
            await db.salary_structures.update_one(
                {"employee_id": employee_id},
                {"$set": update_data}
            )
            
            await log_audit(current_user["id"], "update_salary_components", "salary", employee_id)
            
            return {"message": "Salary components updated"}
    
    return {"message": "No changes made"}

@api_router.get("/employees/{employee_id}/salary/adjustments")
async def get_salary_adjustments(
    employee_id: str,
    month: Optional[str] = None,  # Format: YYYY-MM
    current_user: dict = Depends(get_current_user)
):
    """Get salary adjustments for an employee"""
    if current_user["role"] == UserRole.EMPLOYEE:
        if current_user.get("employee_id") != employee_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"employee_id": employee_id}
    
    if month:
        # Get adjustments applicable for this month
        query["$or"] = [
            {"frequency": "one_time", "applicable_month": month},
            {
                "frequency": "recurring",
                "is_active": True,
                "$or": [
                    {"start_month": {"$lte": month}, "end_month": {"$gte": month}},
                    {"start_month": {"$lte": month}, "end_month": None}
                ]
            }
        ]
    
    adjustments = await db.salary_adjustments.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    return {"employee_id": employee_id, "adjustments": [serialize_doc(a) for a in adjustments]}

@api_router.post("/employees/{employee_id}/salary/adjustments")
async def create_salary_adjustment(
    employee_id: str,
    data: SalaryAdjustmentCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a salary adjustment (Admin only)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Determine category based on type
    earning_types = ["bonus", "incentive", "reimbursement"]
    category = SalaryComponentType.EARNING if data.adjustment_type in earning_types else SalaryComponentType.DEDUCTION
    
    adjustment = SalaryAdjustment(
        employee_id=employee_id,
        adjustment_type=data.adjustment_type,
        category=category,
        description=data.description,
        amount=data.amount,
        frequency=data.frequency,
        applicable_month=data.applicable_month if data.frequency == "one_time" else None,
        start_month=data.start_month if data.frequency == "recurring" else None,
        end_month=data.end_month if data.frequency == "recurring" else None,
        created_by=current_user["id"],
        created_by_name=current_user.get("name")
    )
    
    doc = adjustment.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["updated_at"] = doc["updated_at"].isoformat()
    await db.salary_adjustments.insert_one(doc.copy())
    
    await log_audit(current_user["id"], "create_salary_adjustment", "salary_adjustment", adjustment.id, 
                   f"{data.adjustment_type}: {data.amount} for {employee.get('full_name')}")
    
    return {"message": "Adjustment created", "adjustment": serialize_doc(doc)}

@api_router.delete("/employees/{employee_id}/salary/adjustments/{adjustment_id}")
async def delete_salary_adjustment(
    employee_id: str,
    adjustment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a salary adjustment (Admin only)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    result = await db.salary_adjustments.delete_one({"id": adjustment_id, "employee_id": employee_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Adjustment not found")
    
    await log_audit(current_user["id"], "delete_salary_adjustment", "salary_adjustment", adjustment_id)
    
    return {"message": "Adjustment deleted"}

@api_router.put("/employees/{employee_id}/salary/adjustments/{adjustment_id}")
async def update_salary_adjustment(
    employee_id: str,
    adjustment_id: str,
    data: SalaryAdjustmentCreate,
    current_user: dict = Depends(get_current_user)
):
    """Update a salary adjustment (Admin only)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    existing = await db.salary_adjustments.find_one({"id": adjustment_id, "employee_id": employee_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Adjustment not found")
    
    earning_types = ["bonus", "incentive", "reimbursement"]
    category = SalaryComponentType.EARNING if data.adjustment_type in earning_types else SalaryComponentType.DEDUCTION
    
    update_data = {
        "adjustment_type": data.adjustment_type,
        "category": category,
        "description": data.description,
        "amount": data.amount,
        "frequency": data.frequency,
        "applicable_month": data.applicable_month if data.frequency == "one_time" else None,
        "start_month": data.start_month if data.frequency == "recurring" else None,
        "end_month": data.end_month if data.frequency == "recurring" else None,
        "updated_at": get_ist_now().isoformat()
    }
    
    await db.salary_adjustments.update_one({"id": adjustment_id}, {"$set": update_data})
    
    await log_audit(current_user["id"], "update_salary_adjustment", "salary_adjustment", adjustment_id)
    
    return {"message": "Adjustment updated"}

@api_router.get("/employees/{employee_id}/payslip/{month}")
async def get_payslip(
    employee_id: str,
    month: str,  # Format: YYYY-MM
    current_user: dict = Depends(get_current_user)
):
    """Get payslip for a specific month - Compensation & Benefits Structure format"""
    if current_user["role"] == UserRole.EMPLOYEE:
        if current_user.get("employee_id") != employee_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get base salary structure
    salary = await db.salary_structures.find_one({"employee_id": employee_id}, {"_id": 0})
    if not salary:
        raise HTTPException(status_code=404, detail="Salary structure not found")
    
    # Get adjustments for this month
    adjustments = await db.salary_adjustments.find({
        "employee_id": employee_id,
        "$or": [
            {"frequency": "one_time", "applicable_month": month},
            {
                "frequency": "recurring",
                "is_active": True,
                "$or": [
                    {"start_month": {"$lte": month}, "end_month": {"$gte": month}},
                    {"start_month": {"$lte": month}, "end_month": None}
                ]
            }
        ]
    }, {"_id": 0}).to_list(50)
    
    # Calculate adjusted salary
    total_earnings_adjustment = sum(a["amount"] for a in adjustments if a["category"] == "earning")
    total_deductions_adjustment = sum(a["amount"] for a in adjustments if a["category"] == "deduction")
    
    adjusted_gross = salary.get("gross_salary", 0) + total_earnings_adjustment
    adjusted_deductions = salary.get("total_deductions", 0) + total_deductions_adjustment
    
    # Get LOP days from attendance if available
    lop_days = 0
    attendance_records = await db.attendances.find({
        "emp_id": employee.get("emp_id"),
        "status": "LOP"
    }, {"_id": 0}).to_list(31)
    for record in attendance_records:
        if record.get("date", "").startswith(month.replace("-", "")[:6]) or month in record.get("date", ""):
            lop_days += 1
    
    # Calculate LOP deduction
    per_day_salary = salary.get("gross_salary", 0) / 30
    lop_deduction = round(per_day_salary * lop_days, 2)
    
    # Net Pay
    net_pay = round(adjusted_gross - adjusted_deductions - lop_deduction, 2)
    
    payslip = {
        "employee_id": employee_id,
        "employee_name": employee.get("full_name"),
        "emp_id": employee.get("emp_id"),
        "designation": employee.get("designation"),
        "department": employee.get("department"),
        "tier_level": employee.get("tier_level", "Tier 1"),
        "date_of_joining": employee.get("date_of_joining"),
        "month": month,
        "pay_period": month,
        
        # Base Components (A)
        "basic": salary.get("basic", 0),
        "hra": salary.get("hra", 0),
        "base_components_total": salary.get("base_components_total", salary.get("basic", 0) + salary.get("hra", 0)),
        
        # Basket of Allowances (B)
        "lta": salary.get("lta", 0),
        "phone_internet": salary.get("phone_internet", 0),
        "bonus": salary.get("bonus", 0),
        "stay_travel": salary.get("stay_travel", 0),
        "special_allowance": salary.get("special_allowance", 0),
        "food_reimbursement": salary.get("food_reimbursement", 0),
        "medical_allowance": salary.get("medical_allowance", 0),
        "conveyance": salary.get("conveyance", 0),
        "basket_allowances_total": salary.get("basket_allowances_total", 0),
        
        # Retirement Benefits (C)
        "pf_employer": salary.get("pf_employer", 0),
        "gratuity": salary.get("gratuity", 0),
        "retirement_benefits_total": salary.get("retirement_benefits_total", 0),
        
        # Compensation Totals
        "fixed_compensation": salary.get("fixed_compensation", 0),
        "variable_compensation": salary.get("variable_compensation", 0),
        "gross_salary": adjusted_gross,
        "annual_ctc": salary.get("annual_ctc", 0),
        "monthly_ctc": salary.get("monthly_ctc", 0),
        
        # Deductions
        "pf_employee": salary.get("pf_employee", 0),
        "esi_employee": salary.get("esi_employee", 0),
        "professional_tax": salary.get("professional_tax", 0),
        "tds": salary.get("tds", 0),
        "other_deductions": salary.get("other_deductions", 0),
        "total_deductions": adjusted_deductions + lop_deduction,
        
        # Adjustments
        "earnings_adjustments": [a for a in adjustments if a["category"] == "earning"],
        "deduction_adjustments": [a for a in adjustments if a["category"] == "deduction"],
        "total_earnings_adjustment": total_earnings_adjustment,
        "total_deductions_adjustment": total_deductions_adjustment,
        
        # LOP
        "lop_days": lop_days,
        "lop_deduction": lop_deduction,
        
        # Net Pay
        "net_pay": net_pay,
        
        # Company info
        "company_name": "BluBridge Technologies Pvt Ltd",
        "company_address": "Chennai, Tamil Nadu, India",
        
        # Insurance Coverage
        "medical_insurance": 300000,
        "accident_insurance": "1x CTC (Min ₹5 Lakhs)",
        "life_insurance": 500000
    }
    
    return payslip

@api_router.get("/employee-profile/salary")
async def get_my_salary(current_user: dict = Depends(get_current_user)):
    """Get current employee's salary structure"""
    if current_user["role"] != UserRole.EMPLOYEE:
        raise HTTPException(status_code=403, detail="This endpoint is for employees only")
    
    employee_id = current_user.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=404, detail="Employee profile not found")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    salary = await db.salary_structures.find_one({"employee_id": employee_id}, {"_id": 0})
    
    if not salary and employee:
        monthly_salary = employee.get("monthly_salary", 0)
        if monthly_salary > 0:
            annual_ctc = monthly_salary * 12
            calc_fn = get_salary_calculator(employee.get("designation", ""))
            salary_data = calc_fn(annual_ctc)
            salary_data["id"] = str(uuid.uuid4())
            salary_data["employee_id"] = employee_id
            salary_data["effective_from"] = employee.get("date_of_joining", get_ist_today())
            salary_data["created_at"] = get_ist_now().isoformat()
            salary_data["updated_at"] = get_ist_now().isoformat()
            await db.salary_structures.insert_one(salary_data.copy())
            salary = salary_data
    
    # Get current month adjustments
    current_month = get_ist_now().strftime("%Y-%m")
    adjustments = await db.salary_adjustments.find({
        "employee_id": employee_id,
        "$or": [
            {"frequency": "one_time", "applicable_month": current_month},
            {"frequency": "recurring", "is_active": True}
        ]
    }, {"_id": 0}).to_list(20)
    
    return {
        "salary": serialize_doc(salary) if salary else None,
        "adjustments": [serialize_doc(a) for a in adjustments]
    }

# ============== EMPLOYEE DOCUMENTS ROUTES ==============

class EmployeeDocumentUpload(BaseModel):
    file_url: str
    file_name: str
    file_type: str
    file_public_id: Optional[str] = None
    document_type: str = "offer_letter"  # offer_letter, appointment_letter, etc.

@api_router.get("/employees/{employee_id}/documents")
async def get_employee_documents(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get employee's official documents (Admin can view any, Employee can view own)"""
    # Check permissions
    if current_user["role"] == UserRole.EMPLOYEE:
        if current_user.get("employee_id") != employee_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get documents from employee_documents collection
    documents = await db.employee_documents.find(
        {"employee_id": employee_id}, 
        {"_id": 0}
    ).to_list(50)
    
    return {
        "employee_id": employee_id,
        "employee_name": employee.get("full_name"),
        "documents": [serialize_doc(d) for d in documents]
    }

@api_router.post("/employees/{employee_id}/documents")
async def upload_employee_document(
    employee_id: str, 
    data: EmployeeDocumentUpload,
    current_user: dict = Depends(get_current_user)
):
    """Upload official document for an employee (Admin only)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check if document of this type already exists
    existing = await db.employee_documents.find_one({
        "employee_id": employee_id,
        "document_type": data.document_type
    })
    
    doc = {
        "id": str(uuid.uuid4()),
        "employee_id": employee_id,
        "document_type": data.document_type,
        "file_url": data.file_url,
        "file_name": data.file_name,
        "file_type": data.file_type,
        "file_public_id": data.file_public_id,
        "uploaded_by": current_user["id"],
        "uploaded_by_name": current_user.get("name"),
        "uploaded_at": get_ist_now().isoformat(),
        "updated_at": get_ist_now().isoformat()
    }
    
    if existing:
        # Update existing document
        await db.employee_documents.update_one(
            {"id": existing["id"]},
            {"$set": {
                "file_url": data.file_url,
                "file_name": data.file_name,
                "file_type": data.file_type,
                "file_public_id": data.file_public_id,
                "uploaded_by": current_user["id"],
                "uploaded_by_name": current_user.get("name"),
                "updated_at": get_ist_now().isoformat()
            }}
        )
        doc["id"] = existing["id"]
        await log_audit(current_user["id"], "update_employee_document", "employee_document", existing["id"], f"Updated {data.document_type} for {employee.get('full_name')}")
    else:
        # Insert new document
        await db.employee_documents.insert_one(doc.copy())
        await log_audit(current_user["id"], "upload_employee_document", "employee_document", doc["id"], f"Uploaded {data.document_type} for {employee.get('full_name')}")
    
    return {"message": f"Document uploaded successfully", "document": serialize_doc(doc)}

@api_router.delete("/employees/{employee_id}/documents/{document_id}")
async def delete_employee_document(
    employee_id: str,
    document_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete an employee document (Admin only)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    document = await db.employee_documents.find_one({"id": document_id, "employee_id": employee_id})
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    await db.employee_documents.delete_one({"id": document_id})
    await log_audit(current_user["id"], "delete_employee_document", "employee_document", document_id)
    
    return {"message": "Document deleted successfully"}

@api_router.get("/employee-profile/documents")
async def get_my_documents(current_user: dict = Depends(get_current_user)):
    """Get current employee's official documents"""
    if current_user["role"] != UserRole.EMPLOYEE:
        raise HTTPException(status_code=403, detail="This endpoint is for employees only")
    
    employee_id = current_user.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=404, detail="Employee profile not found")
    
    documents = await db.employee_documents.find(
        {"employee_id": employee_id}, 
        {"_id": 0}
    ).to_list(50)
    
    return {
        "documents": [serialize_doc(d) for d in documents]
    }

# ============== POLICIES ROUTES ==============

# Policies whose visibility is restricted by department.
# Map: policy_id -> set of department names allowed (admins always see them)
DEPARTMENT_RESTRICTED_POLICIES = {
    "policy_research": {"Research Unit"},
    "policy_research_hr": {"Research Unit"},
    "policy_support_hr": {"Support Staff"},
}

# Policies that are completely hidden from every user (including admins).
# To re-enable a policy, simply remove its id from this set.
HIDDEN_POLICIES: set = {"policy_research"}

# Policies that bypass ALL access checks (department / club / role) and must
# always be visible to every authenticated user. Used for organisation-wide
# documents like the IT and Communication Policy and the Admin Induction.
GLOBAL_POLICIES = {"policy_it", "policy_admin_induction"}

async def _is_policy_visible_to_user(policy_id: str, current_user: dict) -> bool:
    """Decide if a policy is visible to the current user.
    Hidden policies are invisible to everyone.
    Global policies are visible to every authenticated user, regardless of
    role / department / club.
    Admin roles see all non-hidden policies.
    Employees only see department-restricted policies whose allowlist
    contains their employee.department.
    """
    if policy_id in HIDDEN_POLICIES:
        return False
    # Global override — explicit allow for organisation-wide docs.
    if policy_id in GLOBAL_POLICIES:
        return True
    restricted_to = DEPARTMENT_RESTRICTED_POLICIES.get(policy_id)
    if not restricted_to:
        return True
    # Admins (hr, system_admin, office_admin) always see restricted policies
    if current_user.get("role") in ALL_ADMIN_ROLES:
        return True
    employee_id = current_user.get("employee_id")
    if not employee_id:
        return False
    employee = await db.employees.find_one(
        {"id": employee_id, "is_deleted": {"$ne": True}},
        {"_id": 0, "department": 1},
    )
    if not employee:
        return False
    return employee.get("department") in restricted_to


@api_router.get("/policies")
async def get_policies(current_user: dict = Depends(get_current_user)):
    """Get all company policies"""
    # Check if policies exist in database - dedup first
    # Remove duplicates keeping only first per id
    pipeline = [
        {"$group": {"_id": "$id", "docs": {"$push": "$_id"}, "count": {"$sum": 1}}},
        {"$match": {"count": {"$gt": 1}}}
    ]
    async for group in db.policies.aggregate(pipeline):
        ids_to_delete = group["docs"][1:]
        if ids_to_delete:
            await db.policies.delete_many({"_id": {"$in": ids_to_delete}})
    
    policies = await db.policies.find({}, {"_id": 0}).to_list(20)
    
    if not policies:
        # Seed default policies using upsert to prevent duplicates
        for policy in COMPANY_POLICIES:
            policy_doc = {
                **policy,
                "created_at": get_ist_now().isoformat(),
                "updated_at": get_ist_now().isoformat()
            }
            await db.policies.update_one(
                {"id": policy["id"]},
                {"$setOnInsert": policy_doc},
                upsert=True
            )
        policies = await db.policies.find({}, {"_id": 0}).to_list(20)
    
    # Apply department-based visibility filter
    visible_policies = []
    for policy in policies:
        if await _is_policy_visible_to_user(policy.get("id"), current_user):
            visible_policies.append(policy)
    return [serialize_doc(p) for p in visible_policies]

@api_router.get("/policies/{policy_id}")
async def get_policy(policy_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific policy by ID"""
    policy = await db.policies.find_one({"id": policy_id}, {"_id": 0})
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")
    if not await _is_policy_visible_to_user(policy_id, current_user):
        raise HTTPException(status_code=403, detail="You do not have access to this policy")
    return serialize_doc(policy)

@api_router.put("/policies/{policy_id}")
async def update_policy(policy_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Update a policy (admin only)"""
    if current_user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    policy = await db.policies.find_one({"id": policy_id}, {"_id": 0})
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")
    
    data["updated_at"] = get_ist_now().isoformat()
    await db.policies.update_one({"id": policy_id}, {"$set": data})
    
    await log_audit(current_user["id"], "update_policy", "policy", policy_id)
    return {"message": "Policy updated successfully"}

# ============== EMPLOYEE EDUCATION & EXPERIENCE ROUTES ==============

class EducationEntry(BaseModel):
    level: str  # Class X, Class XII, Graduation, Post Graduation, Doctorate
    institution: str
    board_university: str
    year_of_passing: str
    percentage_cgpa: str

class ExperienceEntry(BaseModel):
    company_name: str
    designation: str
    start_date: str
    end_date: Optional[str] = None
    is_current: bool = False
    responsibilities: Optional[str] = None

class EducationExperienceUpdate(BaseModel):
    education: Optional[List[dict]] = None
    experience: Optional[List[dict]] = None

@api_router.get("/employee-profile/education-experience")
async def get_my_education_experience(current_user: dict = Depends(get_current_user)):
    """Get current employee's education and experience"""
    employee_id = current_user.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="No employee linked to this user")
    
    # Get employee record
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    return {
        "education": employee.get("education", []),
        "experience": employee.get("experience", []),
        "education_verified": employee.get("education_verified", False),
        "experience_verified": employee.get("experience_verified", False)
    }

@api_router.put("/employee-profile/education-experience")
async def update_my_education_experience(
    data: EducationExperienceUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update current employee's education and experience"""
    employee_id = current_user.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="No employee linked to this user")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check if already verified - cannot edit verified data
    if data.education and employee.get("education_verified"):
        raise HTTPException(status_code=400, detail="Education details are verified and cannot be modified")
    if data.experience and employee.get("experience_verified"):
        raise HTTPException(status_code=400, detail="Experience details are verified and cannot be modified")
    
    update_data = {}
    if data.education is not None:
        update_data["education"] = data.education
    if data.experience is not None:
        update_data["experience"] = data.experience
    
    if update_data:
        update_data["updated_at"] = get_ist_now().isoformat()
        await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    
    await log_audit(current_user["id"], "update_education_experience", "employee", employee_id)
    return {"message": "Details updated successfully"}

@api_router.get("/employees/{employee_id}/education-experience")
async def get_employee_education_experience(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get employee's education and experience (admin view)"""
    if current_user["role"] not in [UserRole.HR]:
        # Employees can only view their own
        if current_user.get("employee_id") != employee_id:
            raise HTTPException(status_code=403, detail="Permission denied")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    return {
        "employee_id": employee_id,
        "emp_name": employee.get("full_name"),
        "education": employee.get("education", []),
        "experience": employee.get("experience", []),
        "education_verified": employee.get("education_verified", False),
        "experience_verified": employee.get("experience_verified", False)
    }

@api_router.post("/employees/{employee_id}/verify-education")
async def verify_employee_education(employee_id: str, current_user: dict = Depends(get_current_user)):
    """HR verifies employee's education details"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    await db.employees.update_one(
        {"id": employee_id},
        {"$set": {
            "education_verified": True,
            "education_verified_by": current_user["id"],
            "education_verified_at": get_ist_now().isoformat()
        }}
    )
    
    await log_audit(current_user["id"], "verify_education", "employee", employee_id)
    return {"message": "Education verified successfully"}

@api_router.post("/employees/{employee_id}/verify-experience")
async def verify_employee_experience(employee_id: str, current_user: dict = Depends(get_current_user)):
    """HR verifies employee's experience details"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    await db.employees.update_one(
        {"id": employee_id},
        {"$set": {
            "experience_verified": True,
            "experience_verified_by": current_user["id"],
            "experience_verified_at": get_ist_now().isoformat()
        }}
    )
    
    await log_audit(current_user["id"], "verify_experience", "employee", employee_id)
    return {"message": "Experience verified successfully"}

# ============== EMAIL HELPERS FOR ONBOARDING ==============

def get_onboarding_status_email(emp_name: str, status: str, notes: Optional[str] = None):
    """Generate HTML email for onboarding status notification"""
    status_color = "#10b981" if status == "approved" else "#f59e0b"
    status_text = "Approved" if status == "approved" else "Requires Attention"
    
    message = "Congratulations! Your onboarding has been approved. You now have full access to the HRMS portal." if status == "approved" else "Your onboarding requires some attention. Please log in to the portal to review the feedback and make necessary updates."
    
    return f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: #0b1f3b; padding: 20px; border-radius: 8px 8px 0 0;">
            <h1 style="color: white; margin: 0; font-size: 24px;">BluBridge HRMS</h1>
        </div>
        <div style="background: #fffdf7; padding: 30px; border: 1px solid #e5e5e5; border-top: none; border-radius: 0 0 8px 8px;">
            <h2 style="color: #0b1f3b; margin-top: 0;">Onboarding Status Update</h2>
            <p style="color: #666;">Dear {emp_name},</p>
            <p style="color: #666;">{message}</p>
            <div style="background: #f5f5f5; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid {status_color};">
                <p style="margin: 5px 0;"><strong>Status:</strong> <span style="color: {status_color};">{status_text}</span></p>
                {f'<p style="margin: 5px 0;"><strong>Notes:</strong> {notes}</p>' if notes else ''}
            </div>
            <p style="color: #999; font-size: 12px; margin-top: 30px;">This is an automated notification from BluBridge HRMS.</p>
        </div>
    </div>
    """

# ============== HOLIDAY ROUTES ==============

class HolidayCreate(BaseModel):
    name: str
    date: str  # YYYY-MM-DD format
    day: str
    type: str = "company"  # national, regional, religious, company
    note: Optional[str] = None

class HolidayUpdate(BaseModel):
    name: Optional[str] = None
    date: Optional[str] = None
    day: Optional[str] = None
    type: Optional[str] = None
    note: Optional[str] = None

@api_router.get("/holidays")
async def get_holidays(
    year: int = Query(2026),
    current_user: dict = Depends(get_current_user)
):
    """Get all holidays for a given year"""
    # First check if we have holidays in database
    holidays = await db.holidays.find({"year": year}, {"_id": 0}).sort("date", 1).to_list(50)
    
    if not holidays:
        # Seed default holidays for 2026 if not present
        if year == 2026:
            for h in COMPANY_HOLIDAYS_2026:
                holiday_doc = {
                    "id": h["id"],
                    "name": h["name"],
                    "date": h["date"],
                    "day": h["day"],
                    "type": h["type"],
                    "note": h.get("note"),
                    "year": 2026,
                    "created_at": get_ist_now().isoformat()
                }
                await db.holidays.update_one(
                    {"id": h["id"], "year": 2026},
                    {"$setOnInsert": holiday_doc},
                    upsert=True
                )
            holidays = await db.holidays.find({"year": year}, {"_id": 0}).sort("date", 1).to_list(50)
    
    # Calculate stats
    total = len(holidays)
    upcoming = sum(1 for h in holidays if h["date"] >= datetime.now().strftime("%Y-%m-%d"))
    by_type = {}
    for h in holidays:
        t = h.get("type", "company")
        by_type[t] = by_type.get(t, 0) + 1
    
    return {
        "holidays": [serialize_doc(h) for h in holidays],
        "stats": {
            "total": total,
            "upcoming": upcoming,
            "by_type": by_type
        }
    }

@api_router.get("/holidays/upcoming")
async def get_upcoming_holidays(
    limit: int = Query(5, ge=1, le=10),
    current_user: dict = Depends(get_current_user)
):
    """Get next upcoming holidays"""
    today = datetime.now().strftime("%Y-%m-%d")
    holidays = await db.holidays.find(
        {"date": {"$gte": today}}, 
        {"_id": 0}
    ).sort("date", 1).limit(limit).to_list(limit)
    
    return [serialize_doc(h) for h in holidays]

@api_router.post("/holidays")
async def create_holiday(data: HolidayCreate, current_user: dict = Depends(get_current_user)):
    """Create a new holiday (admin only)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Parse year from date
    year = int(data.date.split("-")[0])
    
    holiday_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "date": data.date,
        "day": data.day,
        "type": data.type,
        "note": data.note,
        "year": year,
        "created_at": get_ist_now().isoformat()
    }
    
    await db.holidays.insert_one(holiday_doc.copy())
    await log_audit(current_user["id"], "create_holiday", "holiday", holiday_doc["id"], data.name)
    
    return serialize_doc(holiday_doc)

@api_router.put("/holidays/{holiday_id}")
async def update_holiday(holiday_id: str, data: HolidayUpdate, current_user: dict = Depends(get_current_user)):
    """Update a holiday (admin only)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    holiday = await db.holidays.find_one({"id": holiday_id}, {"_id": 0})
    if not holiday:
        raise HTTPException(status_code=404, detail="Holiday not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        if "date" in update_data:
            update_data["year"] = int(update_data["date"].split("-")[0])
        await db.holidays.update_one({"id": holiday_id}, {"$set": update_data})
    
    await log_audit(current_user["id"], "update_holiday", "holiday", holiday_id)
    
    return {"message": "Holiday updated successfully"}

@api_router.delete("/holidays/{holiday_id}")
async def delete_holiday(holiday_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a holiday (admin only)"""
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    result = await db.holidays.delete_one({"id": holiday_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Holiday not found")
    
    await log_audit(current_user["id"], "delete_holiday", "holiday", holiday_id)
    
    return {"message": "Holiday deleted successfully"}

# ============== SEED DATA ==============

@api_router.post("/seed")
async def seed_database():
    # Check if already seeded
    admin_exists = await db.users.find_one({"username": "admin"})
    if admin_exists:
        # Also ensure employee user exists
        employee_user_exists = await db.users.find_one({"username": "user"})
        if not employee_user_exists:
            # Find first employee to link
            first_emp = await db.employees.find_one({"is_deleted": {"$ne": True}}, {"_id": 0})
            if first_emp:
                emp_user = User(
                    username="user",
                    email="user@blubridge.com",
                    password_hash=hash_password("user"),
                    name=first_emp.get("full_name", "Employee User"),
                    role=UserRole.EMPLOYEE,
                    employee_id=first_emp["id"],
                    department=first_emp.get("department"),
                    team=first_emp.get("team")
                )
                doc = emp_user.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.users.insert_one(doc.copy())
        return {"message": "Database already seeded"}
    
    # Create HR admin user
    hr_admin = User(
        username="admin",
        email="admin@blubridge.com",
        password_hash=hash_password("pass123"),
        name="HR Admin",
        role=UserRole.HR,
        department="Human Resources",
        is_first_login=False,
        onboarding_status="completed"
    )
    doc = hr_admin.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc.copy())
    
    # Create System Admin user
    sys_admin = User(
        username="sysadmin",
        email="sysadmin@blubridge.com",
        password_hash=hash_password("pass123"),
        name="System Admin",
        role=UserRole.SYSTEM_ADMIN,
        department="IT",
        is_first_login=False,
        onboarding_status="completed"
    )
    doc = sys_admin.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc.copy())
    
    # Create Office Admin user
    off_admin = User(
        username="offadmin",
        email="offadmin@blubridge.com",
        password_hash=hash_password("pass123"),
        name="Office Admin",
        role=UserRole.OFFICE_ADMIN,
        department="Administration",
        is_first_login=False,
        onboarding_status="completed"
    )
    doc = off_admin.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc.copy())
    
    # Create departments
    departments = [
        {"id": str(uuid.uuid4()), "name": "Research Unit", "team_count": 8},
        {"id": str(uuid.uuid4()), "name": "Support Staff", "team_count": 1},
        {"id": str(uuid.uuid4()), "name": "Business & Product", "team_count": 5}
    ]
    await db.departments.insert_many(departments)
    
    # Create teams
    teams = [
        {"id": str(uuid.uuid4()), "name": "Compiler - Auto Differentiation", "department": "Research Unit", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Compiler - Computation Graph", "department": "Research Unit", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Data", "department": "Research Unit", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Framework - Graph & Auto-differentiation", "department": "Research Unit", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Framework - parallelism", "department": "Research Unit", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Framework - Tensor & Ops", "department": "Research Unit", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Framework - Quantz", "department": "Research Unit", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Tokenizer", "department": "Research Unit", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Administration", "department": "Support Staff", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Product Management", "department": "Business & Product", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Sales", "department": "Business & Product", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Marketing", "department": "Business & Product", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Business Development", "department": "Business & Product", "member_count": 0},
        {"id": str(uuid.uuid4()), "name": "Customer Success", "department": "Business & Product", "member_count": 0}
    ]
    await db.teams.insert_many(teams)
    
    # Create employees with full schema
    employees_data = [
        {"full_name": "Adhitya Charan", "official_email": "adhitya.blubridge@evoplus.in", "team": "Framework - parallelism", "stars": -11, "unsafe_count": 3, "gender": "Male", "tier_level": TierLevel.MID},
        {"full_name": "Adwaid Suresh", "official_email": "suresh.blubridge@evoplus.in", "team": "Compiler - Auto Differentiation", "stars": -2, "unsafe_count": 0, "gender": "Male", "tier_level": TierLevel.SENIOR},
        {"full_name": "Amarnath V S", "official_email": "amarnath.blubridge@evoplus.in", "team": "Framework - Quantz", "stars": -5, "unsafe_count": 2, "gender": "Male", "tier_level": TierLevel.MID},
        {"full_name": "Anuj Kumar", "official_email": "anuj.blubridge@evoplus.in", "team": "Framework - parallelism", "stars": -11, "unsafe_count": 3, "gender": "Male", "tier_level": TierLevel.JUNIOR},
        {"full_name": "Aravind P", "official_email": "aravind.blubridge@evoplus.in", "team": "Tokenizer", "stars": -5, "unsafe_count": 2, "gender": "Male", "tier_level": TierLevel.MID},
        {"full_name": "Aravind S", "official_email": "aravinds.blubridge@evoplus.in", "team": "Compiler - Auto Differentiation", "stars": -5, "unsafe_count": 2, "gender": "Male", "tier_level": TierLevel.SENIOR},
        {"full_name": "Chaithanya", "official_email": "chaithanya.blubridge@evoplus.in", "team": "Data", "stars": 0, "unsafe_count": 0, "gender": "Female", "tier_level": TierLevel.MID},
        {"full_name": "Dinesh", "official_email": "dinesh.blubridge@evoplus.in", "team": "Framework - parallelism", "stars": 2, "unsafe_count": 0, "gender": "Male", "tier_level": TierLevel.LEAD},
        {"full_name": "Gowtham", "official_email": "gowtham.blubridge@evoplus.in", "team": "Data", "stars": 5, "unsafe_count": 0, "gender": "Male", "tier_level": TierLevel.SENIOR},
        {"full_name": "Gowthamkumar", "official_email": "gowthamkumar.blubridge@evoplus.in", "team": "Framework - Tensor & Ops", "stars": 3, "unsafe_count": 0, "gender": "Male", "tier_level": TierLevel.MID},
        {"full_name": "Grishma", "official_email": "grishma.blubridge@evoplus.in", "team": "Framework - Graph & Auto-differentiation", "stars": 1, "unsafe_count": 0, "gender": "Female", "tier_level": TierLevel.JUNIOR},
        {"full_name": "Hamza", "official_email": "hamza.blubridge@evoplus.in", "team": "Administration", "stars": 0, "unsafe_count": 0, "gender": "Male", "tier_level": TierLevel.MID},
        {"full_name": "Harshini", "official_email": "harshini.blubridge@evoplus.in", "team": "Compiler - Auto Differentiation", "stars": 2, "unsafe_count": 0, "gender": "Female", "tier_level": TierLevel.JUNIOR},
        {"full_name": "Jenifa", "official_email": "jenifa.blubridge@evoplus.in", "team": "Compiler - Auto Differentiation", "stars": 4, "unsafe_count": 0, "gender": "Female", "tier_level": TierLevel.MID},
        {"full_name": "Jona", "official_email": "jona.blubridge@evoplus.in", "team": "Compiler - Auto Differentiation", "stars": 1, "unsafe_count": 0, "gender": "Male", "tier_level": TierLevel.SENIOR},
        {"full_name": "Kota", "official_email": "kota.blubridge@evoplus.in", "team": "Framework - Tensor & Ops", "stars": -3, "unsafe_count": 1, "gender": "Male", "tier_level": TierLevel.MID},
        {"full_name": "Pragathi V", "official_email": "pragathi.blubridge@evoplus.in", "team": "Administration", "stars": 0, "unsafe_count": 0, "gender": "Female", "tier_level": TierLevel.JUNIOR},
        {"full_name": "Suresh K", "official_email": "sureshk.blubridge@evoplus.in", "team": "Compiler - Auto Differentiation", "stars": 2, "unsafe_count": 0, "gender": "Male", "tier_level": TierLevel.LEAD},
    ]
    
    for i, emp_data in enumerate(employees_data):
        dept = "Support Staff" if emp_data["team"] == "Administration" else "Research Unit"
        emp = Employee(
            emp_id=f"EMP{str(i + 1).zfill(4)}",
            full_name=emp_data["full_name"],
            official_email=emp_data["official_email"],
            gender=emp_data.get("gender"),
            department=dept,
            team=emp_data["team"],
            designation="Software Engineer",
            tier_level=emp_data.get("tier_level", TierLevel.MID),
            date_of_joining="2024-01-15",
            employment_type=EmploymentType.FULL_TIME,
            employee_status=EmployeeStatus.ACTIVE,
            work_location=WorkLocation.OFFICE,
            stars=emp_data["stars"],
            unsafe_count=emp_data["unsafe_count"]
        )
        doc = emp.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.employees.insert_one(doc.copy())
    
    # Create sample attendance
    today = get_ist_now().strftime("%d-%m-%Y")
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    
    for emp in employees[:15]:
        check_in_hour = 8 + (hash(emp["id"]) % 3)
        check_in_min = hash(emp["id"]) % 60
        check_in_time = f"{str(check_in_hour).zfill(2)}:{str(check_in_min).zfill(2)} AM"
        
        att = Attendance(
            employee_id=emp["id"],
            emp_name=emp["full_name"],
            team=emp["team"],
            department=emp["department"],
            date=today,
            check_in=check_in_time,
            status="Login"
        )
        doc = att.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.attendance.insert_one(doc.copy())
    
    # Create employee user account (user/user)
    first_emp = employees[0] if employees else None
    if first_emp:
        emp_user = User(
            username="user",
            email="user@blubridge.com",
            password_hash=hash_password("user"),
            name=first_emp.get("full_name", "Employee User"),
            role=UserRole.EMPLOYEE,
            employee_id=first_emp["id"],
            department=first_emp.get("department"),
            team=first_emp.get("team")
        )
        doc = emp_user.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc.copy())
    
    return {"message": "Database seeded successfully"}

# ============== NOTIFICATION ROUTES ==============

@api_router.get("/notifications")
async def get_notifications(
    limit: int = Query(50, ge=1, le=200),
    current_user: dict = Depends(get_current_user)
):
    """Get notifications for current user"""
    notifications = await db.notifications.find(
        {"user_id": current_user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(limit)
    return [serialize_doc(n) for n in notifications]

@api_router.get("/notifications/unread-count")
async def get_unread_notification_count(current_user: dict = Depends(get_current_user)):
    """Get count of unread notifications"""
    count = await db.notifications.count_documents({"user_id": current_user["id"], "read": False})
    return {"count": count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Mark a single notification as read"""
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user["id"]},
        {"$set": {"read": True}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification marked as read"}

@api_router.put("/notifications/mark-all-read")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    """Mark all notifications as read for current user"""
    await db.notifications.update_many(
        {"user_id": current_user["id"], "read": False},
        {"$set": {"read": True}}
    )
    return {"message": "All notifications marked as read"}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a notification"""
    result = await db.notifications.delete_one({"id": notification_id, "user_id": current_user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification deleted"}

# ============== ROLE MANAGEMENT ROUTES ==============

@api_router.get("/roles/users")
async def get_all_users_with_roles(
    role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all users with their roles (system admin + HR)"""
    if current_user["role"] not in SYSTEM_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    query = {}
    if role:
        query["role"] = role
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(1000)
    return [serialize_doc(u) for u in users]

@api_router.put("/roles/users/{user_id}/role")
async def update_user_role(
    user_id: str,
    data: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """Update a user's role (system admin only)"""
    if current_user["role"] not in SYSTEM_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    new_role = data.get("role")
    if new_role not in [UserRole.HR, UserRole.SYSTEM_ADMIN, UserRole.OFFICE_ADMIN, UserRole.EMPLOYEE]:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    target_user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    old_role = target_user.get("role")
    await db.users.update_one({"id": user_id}, {"$set": {"role": new_role}})
    
    await log_audit(current_user["id"], "update_role", "user", user_id, f"Role changed from {old_role} to {new_role}")
    
    # Notify user about role change
    role_labels = {"hr": "HR Team", "system_admin": "System Admin", "office_admin": "Office Admin", "employee": "Employee"}
    asyncio.create_task(create_notification(
        [user_id],
        "Role Updated",
        f"Your role has been changed to {role_labels.get(new_role, new_role)}. Your access permissions have been updated.",
        "info",
        "/dashboard"
    ))
    
    return {"message": f"User role updated to {new_role}"}

@api_router.get("/roles/permissions")
async def get_role_permissions(current_user: dict = Depends(get_current_user)):
    """Get the permission matrix for all roles"""
    if current_user["role"] not in SYSTEM_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    return {
        "hr": {
            "label": "HR Team",
            "description": "Full control over all HR modules",
            "permissions": [
                "employees.full", "attendance.full", "leave.full", "payroll.full",
                "holidays.full", "policies.full", "star_reward.full", "team.full",
                "tickets.full", "reports.full", "audit_logs.view", "verification.full",
                "late_requests.full", "early_out.full", "missed_punch.full", "onboarding.full"
            ]
        },
        "system_admin": {
            "label": "System Admin",
            "description": "System control + limited HRMS view access",
            "permissions": [
                "employees.view", "attendance.view", "leave.view",
                "roles.manage", "audit_logs.view", "notifications.view"
            ]
        },
        "office_admin": {
            "label": "Office Admin",
            "description": "View employee data + limited operational access",
            "permissions": [
                "employees.view", "attendance.view", "leave.view",
                "holidays.view", "notifications.view"
            ]
        },
        "employee": {
            "label": "Employee",
            "description": "Self-service employee portal",
            "permissions": [
                "self.attendance", "self.leave", "self.documents",
                "self.salary", "self.profile", "self.tickets"
            ]
        }
    }

# ============== OPERATIONAL CHECKLIST ROUTES ==============

OFFICE_ADMIN_ROLES = ["hr", "office_admin"]

@api_router.get("/operational-checklists")
async def get_operational_checklists(
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all operational checklists (Office Admin + HR)"""
    if current_user["role"] not in OFFICE_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {}
    if status and status != "all":
        query["status"] = status
    if search:
        query["$or"] = [
            {"emp_name": {"$regex": search, "$options": "i"}},
            {"employee_id": {"$regex": search, "$options": "i"}}
        ]
    
    checklists = await db.operational_checklists.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [serialize_doc(c) for c in checklists]

@api_router.get("/operational-checklists/stats")
async def get_operational_checklist_stats(current_user: dict = Depends(get_current_user)):
    """Get stats for operational checklists"""
    if current_user["role"] not in OFFICE_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    total = await db.operational_checklists.count_documents({})
    pending = await db.operational_checklists.count_documents({"status": "pending"})
    in_progress = await db.operational_checklists.count_documents({"status": "in_progress"})
    completed = await db.operational_checklists.count_documents({"status": "completed"})
    
    return {
        "total": total,
        "pending": pending,
        "in_progress": in_progress,
        "completed": completed
    }

@api_router.get("/operational-checklists/pending-count")
async def get_operational_checklist_pending_count(current_user: dict = Depends(get_current_user)):
    """Lightweight count for sidebar badge"""
    if current_user["role"] not in OFFICE_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    pending = await db.operational_checklists.count_documents({"status": {"$ne": "completed"}})
    return {"count": pending}

@api_router.get("/operational-checklists/{employee_id}")
async def get_operational_checklist(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get checklist for a specific employee"""
    if current_user["role"] not in OFFICE_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    checklist = await db.operational_checklists.find_one({"employee_id": employee_id}, {"_id": 0})
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist not found")
    return serialize_doc(checklist)

@api_router.put("/operational-checklists/{employee_id}/item/{item_key}")
async def update_checklist_item(
    employee_id: str,
    item_key: str,
    data: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """Toggle a checklist item completed/uncompleted"""
    if current_user["role"] not in OFFICE_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    checklist = await db.operational_checklists.find_one({"employee_id": employee_id}, {"_id": 0})
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist not found")
    
    completed = data.get("completed", False)
    notes = data.get("notes", "")
    now = get_ist_now().isoformat()
    
    items = checklist.get("items", [])
    found = False
    for item in items:
        if item["key"] == item_key:
            item["completed"] = completed
            item["completed_by"] = current_user.get("name", current_user.get("username")) if completed else None
            item["completed_at"] = now if completed else None
            item["notes"] = notes
            found = True
            break
    
    if not found:
        raise HTTPException(status_code=404, detail="Checklist item not found")
    
    # Compute overall status
    total_items = len(items)
    completed_count = sum(1 for i in items if i.get("completed"))
    if completed_count == 0:
        new_status = "pending"
    elif completed_count == total_items:
        new_status = "completed"
    else:
        new_status = "in_progress"
    
    await db.operational_checklists.update_one(
        {"employee_id": employee_id},
        {"$set": {"items": items, "status": new_status, "updated_at": now}}
    )
    
    # If fully completed, notify HR
    if new_status == "completed":
        asyncio.create_task(notify_role(
            UserRole.HR,
            "Operational Setup Complete",
            f"All operational items for {checklist.get('emp_name', 'employee')} have been completed by {current_user.get('name', 'Office Admin')}.",
            "success",
            "/verification"
        ))
    
    await log_audit(current_user["id"], "update_checklist_item", "operational_checklist", employee_id, f"Item {item_key} {'completed' if completed else 'unchecked'}")
    
    updated = await db.operational_checklists.find_one({"employee_id": employee_id}, {"_id": 0})
    return serialize_doc(updated)

# ============== LATE REQUEST ROUTES ==============

async def _resolve_employee(data_employee_id, current_user):
    """Helper to resolve employee for admin-applied or self-applied requests"""
    is_hr = current_user["role"] == UserRole.HR
    if data_employee_id and is_hr:
        emp = await db.employees.find_one({"id": data_employee_id, "is_deleted": {"$ne": True}}, {"_id": 0})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        return emp, is_hr
    elif current_user.get("employee_id"):
        emp = await db.employees.find_one({"id": current_user["employee_id"], "is_deleted": {"$ne": True}}, {"_id": 0})
        if not emp:
            raise HTTPException(status_code=404, detail="Employee profile not found")
        return emp, is_hr
    raise HTTPException(status_code=400, detail="Employee ID required")

@api_router.get("/late-requests")
async def get_late_requests(
    status: Optional[str] = None,
    employee_name: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    is_admin = current_user["role"] in ALL_ADMIN_ROLES
    if not is_admin:
        query["employee_id"] = current_user.get("employee_id", "")
    if status and status != "All":
        query["status"] = status
    if employee_name:
        query["emp_name"] = {"$regex": employee_name, "$options": "i"}
    records = await db.late_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [serialize_doc(r) for r in records]

@api_router.post("/late-requests")
async def create_late_request(data: LateRequestCreate, current_user: dict = Depends(get_current_user)):
    emp, is_admin = await _resolve_employee(data.employee_id, current_user)
    
    # JOB 4: Prevent duplicate late request (same employee + date)
    existing = await db.late_requests.find_one({
        "employee_id": emp["id"],
        "date": data.date,
        "status": {"$ne": "rejected"}
    })
    if existing:
        raise HTTPException(status_code=400, detail="Late request already exists for this date")
    
    req = LateRequest(
        employee_id=emp["id"],
        emp_name=emp["full_name"],
        team=emp["team"],
        department=emp["department"],
        date=data.date,
        expected_time=data.expected_time,
        actual_time=data.actual_time,
        reason=data.reason,
        applied_by_admin=is_admin,
        status="approved" if (data.auto_approve and is_admin) else "pending",
        is_lop=data.is_lop if (data.auto_approve and is_admin) else None,
        approved_by=current_user["id"] if (data.auto_approve and is_admin) else None
    )
    doc = req.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.late_requests.insert_one(doc.copy())
    return serialize_doc(doc)

@api_router.put("/late-requests/{request_id}/approve")
async def approve_late_request(request_id: str, data: Optional[RequestApproveBody] = None, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    update = {"status": "approved", "approved_by": current_user["id"], "approved_at": get_ist_now().isoformat()}
    if data and data.is_lop is not None:
        update["is_lop"] = data.is_lop
    if data and data.lop_remark:
        update["lop_remark"] = data.lop_remark
    result = await db.late_requests.update_one({"id": request_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    rec = await db.late_requests.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(rec)

@api_router.put("/late-requests/{request_id}/reject")
async def reject_late_request(request_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    result = await db.late_requests.update_one({"id": request_id}, {"$set": {"status": "rejected", "approved_by": current_user["id"]}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    rec = await db.late_requests.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(rec)

@api_router.put("/late-requests/{request_id}")
async def edit_late_request(request_id: str, data: LateRequestCreate, current_user: dict = Depends(get_current_user)):
    rec = await db.late_requests.find_one({"id": request_id}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Request not found")
    is_admin = current_user["role"] in ALL_ADMIN_ROLES
    # Employees may only edit their own pending requests; HR/Admin can edit any status
    if not is_admin and rec.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Can only edit pending requests")
    update = {
        "date": data.date,
        "reason": data.reason,
        "expected_time": data.expected_time,
        "actual_time": data.actual_time,
        "edited_by": current_user["id"],
        "edited_at": get_ist_now().isoformat(),
    }
    await db.late_requests.update_one({"id": request_id}, {"$set": update})
    await log_audit(current_user["id"], "edit", "late_request", request_id)
    updated = await db.late_requests.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(updated)

# ============== LATE REQUEST BULK IMPORT ==============

LR_IMPORT_COLUMNS = [
    "Email", "Request Date", "Actual Arrival", "Expected Arrival",
    "Reason", "Status", "Action Type", "Remark", "Applied On", "Approved By", "Approved On",
]

# Sheet-column aliases → canonical field names. Case-insensitive, whitespace-tolerant.
LR_COLUMN_ALIASES: dict[str, str] = {
    # email
    "Email": "email",
    "Emp Mail ID": "email",
    "Employee Email": "email",
    "Email ID": "email",
    "Mail": "email",
    # date
    "Request Date": "date",
    "Request_date": "date",
    "Date": "date",
    "Attendance Date": "date",
    "Late Date": "date",
    # actual time (when employee actually arrived)
    "Actual Arrival": "actual_time",
    "Actual Time": "actual_time",
    "Arrival Time": "actual_time",
    "Arrived At": "actual_time",
    "Expected_Arrival": "actual_time",  # per user: this column = actual arrival time
    "Expected Arrival": "actual_time",
    # request submission time (informational only — stored in extra_data)
    "Request_Time": "request_time",
    "Request Time": "request_time",
    # expected shift time (rarely present, but supported)
    "Expected Time": "expected_time",
    "Shift Start": "expected_time",
    "Scheduled Time": "expected_time",
    # reason
    "Reason": "reason",
    "Remarks": "reason",
    "Notes": "reason",
    # status
    "Status": "status",
    "Request Status": "status",
    # is_lop via Action_Type
    "Action_Type": "action_type",
    "Action Type": "action_type",
    "LOP Status": "action_type",
    # remark / lop_remark
    "Remark": "lop_remark",
    "LOP Remark": "lop_remark",
    "Approver Remark": "lop_remark",
    # applied_at
    "Applied On": "applied_at",
    "Applied_on": "applied_at",
    "Applied Date": "applied_at",
    "Date Applied": "applied_at",
    # approved_by
    "Approved By": "approved_by",
    "Approved_by": "approved_by",
    "Approver": "approved_by",
    # approved_at
    "Approved On": "approved_at",
    "Approved_on": "approved_at",
    "Approval Date": "approved_at",
    "Date Approved": "approved_at",
}


def _build_lr_alias_index() -> dict[str, str]:
    return {_normalize_header(k): v for k, v in LR_COLUMN_ALIASES.items()}


def _normalize_action_type(value) -> Optional[bool]:
    """Map Action_Type to is_lop. NO_LOP → False; LOP → True; NULL/blank → None."""
    if value in (None, ""):
        return None
    s = str(value).strip().upper().replace("-", "_").replace(" ", "_")
    if s in ("NO_LOP", "NOLOP", "NL"):
        return False
    if s in ("LOP",):
        return True
    if s in ("NULL", "NONE", "NA", "N/A", "-"):
        return None
    return None


@api_router.get("/late-requests/import-template")
async def download_lr_import_template(current_user: dict = Depends(get_current_user)):
    """Download styled .xlsx template for late-request bulk import."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Late Request Import"
    header_fill = openpyxl.styles.PatternFill(start_color="063c88", end_color="063c88", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(LR_IMPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    sample = [
        ["employee@blubridge.com", "2026-04-22", "10:15", "10:00", "Heavy traffic", "Approved", "NO_LOP", "Approved", "2026-04-22 10:18:00", "Admin", "2026-04-23 14:00:00"],
        ["another@blubridge.com", "22/04/2026", "10:45", "10:00", "Medical emergency", "Approved", "LOP", "Approved with LOP", "2026-04-22 10:50:00", "Admin", "2026-04-23 14:05:00"],
        ["third@blubridge.com", "23/04/2026", "11:00", "", "Late submission - vehicle breakdown", "Pending", "", "", "2026-04-23 11:05:00", "", ""],
    ]
    for r_idx, row_vals in enumerate(sample, start=2):
        for c_idx, v in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    widths = [30, 14, 12, 12, 30, 12, 12, 24, 20, 18, 20]
    for c_idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w

    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Sheet Column", "Required", "Maps To", "Notes"])
    ws2.append(["Email", "Yes", "employee_id (via email)", "Aliases: Emp Mail ID, Employee Email, Email ID, Mail"])
    ws2.append(["Request Date", "Yes", "date", "Aliases: Request_date, Date, Late Date. Accepts YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY, or Excel datetime"])
    ws2.append(["Actual Arrival", "Yes", "actual_time", "Aliases: Expected_Arrival, Actual Time, Arrival Time. Time when employee actually arrived. HH:MM (24h) or HH:MM AM/PM"])
    ws2.append(["Expected Arrival", "Optional", "expected_time", "Shift start time (e.g., 10:00). Leave blank if not tracked."])
    ws2.append(["Reason", "Yes", "reason", "Aliases: Remarks, Notes. Free text"])
    ws2.append(["Status", "No", "status", "Approved / Pending / Rejected (default: Pending)"])
    ws2.append(["Action Type", "No", "is_lop", "Aliases: Action_Type, LOP Status. NO_LOP → false; LOP → true; NULL/blank → null"])
    ws2.append(["Remark", "No", "lop_remark", "Free text approver note"])
    ws2.append(["Applied On", "No", "applied_at", "Aliases: Applied_on, Applied Date. Stored as ISO datetime in extra_data"])
    ws2.append(["Approved By", "No", "approved_by", "'Admin' or username/email/full name → resolved to user ID. Numeric IDs map to current admin."])
    ws2.append(["Approved On", "No", "approved_at", "Aliases: Approved_on, Approval Date. Stored when status = approved"])
    for col_idx in range(1, 5):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 36
    for c in range(1, 5):
        ws2.cell(1, c).font = openpyxl.styles.Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=late_requests_import_template.xlsx"}
    )


@api_router.post("/late-requests/import/preview")
async def preview_lr_import(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview detected columns + sample rows for late-request import."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")
    headers = [h for h in headers if h]

    alias_index = _build_lr_alias_index()
    column_mapping = {h: alias_index.get(_normalize_header(h)) for h in headers}
    mapped = {h: c for h, c in column_mapping.items() if c}
    ignored = [h for h, c in column_mapping.items() if not c]

    required_fields = ["email", "date", "actual_time", "reason"]
    detected_canonical = set(mapped.values())
    missing_required = [f for f in required_fields if f not in detected_canonical]

    sample = []
    for r in rows[:5]:
        sample.append({k: (v.isoformat() if isinstance(v, datetime) else (str(v) if v is not None else None)) for k, v in r.items()})

    return {
        "filename": file.filename,
        "total_rows": len(rows),
        "headers": headers,
        "column_mapping": mapped,
        "ignored_columns": ignored,
        "missing_required": missing_required,
        "ready_to_import": len(missing_required) == 0,
        "sample_rows": sample,
    }


@api_router.post("/late-requests/bulk-import")
async def bulk_import_late_requests(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk-import late-arrival requests from .xlsx or .csv. Admin-only.

    - Maps employee by email; skipped if not found.
    - Required columns: Email, Request Date, Actual Arrival, Reason.
    - Auto-handles combined Date+Time, in-batch dedupe, status mapping, LOP from Action_Type.
    """
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")

    alias_index = _build_lr_alias_index()
    headers_clean = [h for h in headers if h]
    sheet_to_canon = {h: alias_index.get(_normalize_header(h)) for h in headers_clean}
    detected_canonical = {c for c in sheet_to_canon.values() if c}
    ignored_columns = [h for h, c in sheet_to_canon.items() if not c]

    missing = []
    if "email" not in detected_canonical:
        missing.append("Email")
    if "date" not in detected_canonical:
        missing.append("Request Date")
    if "actual_time" not in detected_canonical:
        missing.append("Actual Arrival")
    if "reason" not in detected_canonical:
        missing.append("Reason")
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing mandatory column(s): {', '.join(missing)}")

    # Pre-fetch employees by email
    email_to_emp: dict = {}
    async for e in db.employees.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "official_email": 1, "full_name": 1, "team": 1, "department": 1}):
        email = (e.get("official_email") or "").strip().lower()
        if email:
            email_to_emp[email] = e

    # Pre-fetch users + employees for Approved By resolution
    approver_index: dict[str, str] = {}
    async for u in db.users.find({}, {"_id": 0, "id": 1, "username": 1, "email": 1}):
        if u.get("username"):
            approver_index[_normalize_header(u["username"])] = u["id"]
        if u.get("email"):
            approver_index[_normalize_header(u["email"])] = u["id"]
    async for e in db.employees.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "full_name": 1, "official_email": 1}):
        if e.get("full_name"):
            approver_index.setdefault(_normalize_header(e["full_name"]), e["id"])
        if e.get("official_email"):
            approver_index.setdefault(_normalize_header(e["official_email"]), e["id"])

    def _resolve_lr_approver(value) -> tuple[Optional[str], Optional[str]]:
        if value in (None, ""):
            return (None, None)
        s = str(value).strip()
        if not s:
            return (None, None)
        # Numeric IDs (e.g., 52) → map to current admin per user spec
        if s.isdigit() or (s.replace(".", "", 1).isdigit()):
            return (current_user["id"], s)
        norm = _normalize_header(s)
        if norm in ("admin", "administrator", "hr admin", "hr"):
            return (current_user["id"], None)
        if norm in approver_index:
            return (approver_index[norm], None)
        return (None, s)

    total = len(rows)
    success = 0
    failed = 0
    skipped_duplicates = 0
    errors: List[dict] = []
    inserts: List[dict] = []
    inbatch_keys: set = set()

    for idx, raw_row in enumerate(rows, start=2):
        row = _remap_row(raw_row, alias_index)

        email = (str(row.get("email") or "")).strip().lower()
        if not email:
            failed += 1
            errors.append({"row": idx, "email": "", "reason": "Email is blank"})
            continue
        emp = email_to_emp.get(email)
        if not emp:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "No employee found with this email"})
            continue

        # Date — date column may be a combined datetime; we only need the date portion
        date_iso, _ = _parse_import_datetime(row.get("date")) if row.get("date") not in (None, "") else (None, None)
        if not date_iso:
            date_iso = _parse_import_date(row.get("date"))
        if not date_iso:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": f"Invalid Request Date: '{row.get('date')}'"})
            continue

        # Actual arrival time
        _, actual_time = _parse_import_datetime(row.get("actual_time")) if row.get("actual_time") not in (None, "") else (None, None)
        if not actual_time:
            actual_time = _parse_import_time(row.get("actual_time"))
        if not actual_time:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": f"Invalid Actual Arrival: '{row.get('actual_time')}' (use HH:MM)"})
            continue

        # Optional expected_time
        expected_time = None
        if row.get("expected_time") not in (None, ""):
            _, expected_time = _parse_import_datetime(row.get("expected_time"))
            if not expected_time:
                expected_time = _parse_import_time(row.get("expected_time"))

        reason = str(row.get("reason")).strip() if row.get("reason") not in (None, "") else None
        if not reason:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "Reason is required"})
            continue

        status_norm = _normalize_status(row.get("status")) or "pending"
        is_lop = _normalize_action_type(row.get("action_type"))
        lop_remark = str(row.get("lop_remark")).strip() if row.get("lop_remark") not in (None, "") else None
        # Treat 'NULL'/'-' literal as blank
        if lop_remark and lop_remark.upper() in ("NULL", "NONE", "N/A", "NA", "-"):
            lop_remark = None

        approver_id, approver_unresolved = _resolve_lr_approver(row.get("approved_by"))
        applied_at_dt, applied_at_time = _parse_import_datetime(row.get("applied_at")) if row.get("applied_at") not in (None, "") else (None, None)
        approved_at_dt, approved_at_time = _parse_import_datetime(row.get("approved_at")) if row.get("approved_at") not in (None, "") else (None, None)
        applied_at_iso = f"{applied_at_dt}T{applied_at_time or '00:00'}" if applied_at_dt else None
        approved_at_iso = f"{approved_at_dt}T{approved_at_time or '00:00'}" if approved_at_dt else None

        # Request_Time — informational only
        request_time = None
        if row.get("request_time") not in (None, ""):
            _, request_time = _parse_import_datetime(row.get("request_time"))
            if not request_time:
                request_time = _parse_import_time(row.get("request_time"))

        # Duplicate guard: same employee + date, non-rejected
        existing = await db.late_requests.find_one({
            "employee_id": emp["id"],
            "date": date_iso,
            "status": {"$ne": "rejected"}
        }, {"_id": 0, "id": 1})
        if existing and status_norm != "rejected":
            skipped_duplicates += 1
            errors.append({"row": idx, "email": email, "reason": f"Late request already exists for {date_iso}"})
            continue

        batch_key = (emp["id"], date_iso, status_norm)
        if batch_key in inbatch_keys and status_norm != "rejected":
            skipped_duplicates += 1
            errors.append({"row": idx, "email": email, "reason": f"Duplicate of an earlier row for {date_iso}"})
            continue
        inbatch_keys.add(batch_key)

        final_approved_by = approver_id
        if status_norm == "approved" and not final_approved_by:
            final_approved_by = current_user["id"]

        extra = {
            k: v for k, v in {
                "applied_at": applied_at_iso,
                "approved_at": approved_at_iso,
                "approver_unresolved": approver_unresolved,
                "request_time": request_time,
            }.items() if v is not None
        } or None

        doc = {
            "id": str(uuid.uuid4()),
            "employee_id": emp["id"],
            "emp_name": emp.get("full_name", ""),
            "team": emp.get("team", ""),
            "department": emp.get("department", ""),
            "date": date_iso,
            "expected_time": expected_time,
            "actual_time": actual_time,
            "reason": reason,
            "status": status_norm,
            "is_lop": is_lop,
            "lop_remark": lop_remark,
            "approved_by": final_approved_by,
            "applied_by_admin": True,
            "extra_data": extra,
            "created_at": get_ist_now().isoformat(),
        }
        inserts.append(doc)
        success += 1

    if inserts:
        chunk = 500
        for i in range(0, len(inserts), chunk):
            await db.late_requests.insert_many([d.copy() for d in inserts[i:i + chunk]])

    return {
        "total": total,
        "success": success,
        "skipped_duplicates": skipped_duplicates,
        "failed": failed,
        "column_mapping": {h: c for h, c in sheet_to_canon.items() if c},
        "ignored_columns": ignored_columns,
        "errors": errors,
    }

# ============== EARLY OUT REQUEST ROUTES ==============

@api_router.get("/early-out-requests")
async def get_early_out_requests(
    status: Optional[str] = None,
    employee_name: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    is_admin = current_user["role"] in ALL_ADMIN_ROLES
    if not is_admin:
        query["employee_id"] = current_user.get("employee_id", "")
    if status and status != "All":
        query["status"] = status
    if employee_name:
        query["emp_name"] = {"$regex": employee_name, "$options": "i"}
    records = await db.early_out_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [serialize_doc(r) for r in records]

@api_router.post("/early-out-requests")
async def create_early_out_request(data: EarlyOutRequestCreate, current_user: dict = Depends(get_current_user)):
    emp, is_admin = await _resolve_employee(data.employee_id, current_user)
    
    # JOB 4: Prevent duplicate early out request
    existing = await db.early_out_requests.find_one({
        "employee_id": emp["id"],
        "date": data.date,
        "status": {"$ne": "rejected"}
    })
    if existing:
        raise HTTPException(status_code=400, detail="Early out request already exists for this date")
    
    req = EarlyOutRequest(
        employee_id=emp["id"],
        emp_name=emp["full_name"],
        team=emp["team"],
        department=emp["department"],
        date=data.date,
        expected_time=data.expected_time,
        actual_time=data.actual_time,
        reason=data.reason,
        applied_by_admin=is_admin,
        status="approved" if (data.auto_approve and is_admin) else "pending",
        is_lop=data.is_lop if (data.auto_approve and is_admin) else None,
        approved_by=current_user["id"] if (data.auto_approve and is_admin) else None
    )
    doc = req.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.early_out_requests.insert_one(doc.copy())
    return serialize_doc(doc)

@api_router.put("/early-out-requests/{request_id}/approve")
async def approve_early_out_request(request_id: str, data: Optional[RequestApproveBody] = None, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    update = {"status": "approved", "approved_by": current_user["id"], "approved_at": get_ist_now().isoformat()}
    if data and data.is_lop is not None:
        update["is_lop"] = data.is_lop
    if data and data.lop_remark:
        update["lop_remark"] = data.lop_remark
    result = await db.early_out_requests.update_one({"id": request_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    rec = await db.early_out_requests.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(rec)

@api_router.put("/early-out-requests/{request_id}/reject")
async def reject_early_out_request(request_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    result = await db.early_out_requests.update_one({"id": request_id}, {"$set": {"status": "rejected", "approved_by": current_user["id"]}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    rec = await db.early_out_requests.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(rec)

@api_router.put("/early-out-requests/{request_id}")
async def edit_early_out_request(request_id: str, data: EarlyOutRequestCreate, current_user: dict = Depends(get_current_user)):
    rec = await db.early_out_requests.find_one({"id": request_id}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Request not found")
    is_admin = current_user["role"] in ALL_ADMIN_ROLES
    if not is_admin and rec.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Can only edit pending requests")
    update = {
        "date": data.date,
        "reason": data.reason,
        "expected_time": data.expected_time,
        "actual_time": data.actual_time,
        "edited_by": current_user["id"],
        "edited_at": get_ist_now().isoformat(),
    }
    await db.early_out_requests.update_one({"id": request_id}, {"$set": update})
    await log_audit(current_user["id"], "edit", "early_out_request", request_id)
    updated = await db.early_out_requests.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(updated)

# ============== EARLY OUT BULK IMPORT ==============

EO_IMPORT_COLUMNS = [
    "Email", "Request Date", "Actual Leave Time", "Expected End Time",
    "Reason", "Status", "Action Type", "Remark", "Applied On", "Approved By", "Approved On",
]

# Sheet-column aliases → canonical field names. Case-insensitive, whitespace-tolerant.
EO_COLUMN_ALIASES: dict[str, str] = {
    # email
    "Email": "email",
    "Emp Mail ID": "email",
    "Employee Email": "email",
    "Email ID": "email",
    "Mail": "email",
    # date
    "Request Date": "date",
    "Request_date": "date",
    "Date": "date",
    "Attendance Date": "date",
    "Early Out Date": "date",
    # actual leave time (when employee actually left early)
    "Actual Leave Time": "actual_time",
    "Actual Time": "actual_time",
    "Left At": "actual_time",
    "Leave Time": "actual_time",
    # expected end time (shift end)
    "Expected End Time": "expected_time",
    "Expected Time": "expected_time",
    "Shift End": "expected_time",
    # reason
    "Reason": "reason",
    "Remarks": "reason",
    "Notes": "reason",
    # status
    "Status": "status",
    "Request Status": "status",
    # is_lop via Action_Type
    "Action_Type": "action_type",
    "Action Type": "action_type",
    "ACTION_Type": "action_type",
    "LOP Status": "action_type",
    # remark / lop_remark
    "Remark": "lop_remark",
    "LOP Remark": "lop_remark",
    "Approver Remark": "lop_remark",
    # applied_at
    "Applied On": "applied_at",
    "Applied_on": "applied_at",
    "Applied Date": "applied_at",
    "Date Applied": "applied_at",
    # approved_by
    "Approved By": "approved_by",
    "Approved_BY": "approved_by",
    "Approved_by": "approved_by",
    "Approver": "approved_by",
    # approved_at
    "Approved On": "approved_at",
    "Approved_on": "approved_at",
    "Approval Date": "approved_at",
    "Date Approved": "approved_at",
    # update timestamp (informational)
    "Update At": "updated_at",
    "Update_At": "updated_at",
    "Updated At": "updated_at",
    "Updated_at": "updated_at",
    "Last Updated": "updated_at",
}


def _build_eo_alias_index() -> dict[str, str]:
    return {_normalize_header(k): v for k, v in EO_COLUMN_ALIASES.items()}


@api_router.get("/early-out-requests/import-template")
async def download_eo_import_template(current_user: dict = Depends(get_current_user)):
    """Download styled .xlsx template for early-out bulk import."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Early Out Import"
    header_fill = openpyxl.styles.PatternFill(start_color="063c88", end_color="063c88", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(EO_IMPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    sample = [
        ["employee@blubridge.com", "2026-04-22", "16:30", "18:00", "Doctor appointment", "Approved", "NO_LOP", "Approved", "2026-04-22 15:00:00", "Admin", "2026-04-23 10:00:00"],
        ["another@blubridge.com", "22/04/2026", "", "", "Family emergency", "Pending", "", "", "2026-04-22 16:00:00", "", ""],
    ]
    for r_idx, row_vals in enumerate(sample, start=2):
        for c_idx, v in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    widths = [30, 14, 14, 14, 30, 12, 12, 24, 20, 18, 20]
    for c_idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w

    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Sheet Column", "Required", "Maps To", "Notes"])
    ws2.append(["Email", "Yes", "employee_id (via email)", "Aliases: Emp Mail ID, Employee Email, Email ID, Mail"])
    ws2.append(["Request Date", "Yes", "date", "Aliases: Request_date, Date. Accepts YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY, or Excel datetime"])
    ws2.append(["Actual Leave Time", "Optional", "actual_time", "Aliases: Actual Time, Left At, Leave Time. Time when employee actually left. HH:MM (24h) or HH:MM AM/PM"])
    ws2.append(["Expected End Time", "Optional", "expected_time", "Shift end time (e.g., 18:00). Leave blank if not tracked."])
    ws2.append(["Reason", "Yes", "reason", "Aliases: Remarks, Notes. Free text"])
    ws2.append(["Status", "No", "status", "Approved / Pending / Rejected (default: Pending)"])
    ws2.append(["Action Type", "No", "is_lop", "Aliases: Action_Type, ACTION_Type, LOP Status. NO_LOP → false; LOP → true; NULL/blank → null"])
    ws2.append(["Remark", "No", "lop_remark", "Free text approver note. 'NULL' is treated as blank."])
    ws2.append(["Applied On", "No", "applied_at", "Aliases: Applied_on, Applied Date. Stored in extra_data"])
    ws2.append(["Approved By", "No", "approved_by", "'Admin' or username/email/full name → resolved to user ID. Numeric IDs map to current admin."])
    ws2.append(["Approved On", "No", "approved_at", "Aliases: Approved_on, Approval Date. Stored when status = approved"])
    for col_idx in range(1, 5):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 36
    for c in range(1, 5):
        ws2.cell(1, c).font = openpyxl.styles.Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=early_out_import_template.xlsx"}
    )


@api_router.post("/early-out-requests/import/preview")
async def preview_eo_import(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview detected columns + sample rows for early-out import."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")
    headers = [h for h in headers if h]

    alias_index = _build_eo_alias_index()
    column_mapping = {h: alias_index.get(_normalize_header(h)) for h in headers}
    mapped = {h: c for h, c in column_mapping.items() if c}
    ignored = [h for h, c in column_mapping.items() if not c]

    required_fields = ["email", "date", "reason"]
    detected_canonical = set(mapped.values())
    missing_required = [f for f in required_fields if f not in detected_canonical]

    sample = []
    for r in rows[:5]:
        sample.append({k: (v.isoformat() if isinstance(v, datetime) else (str(v) if v is not None else None)) for k, v in r.items()})

    return {
        "filename": file.filename,
        "total_rows": len(rows),
        "headers": headers,
        "column_mapping": mapped,
        "ignored_columns": ignored,
        "missing_required": missing_required,
        "ready_to_import": len(missing_required) == 0,
        "sample_rows": sample,
    }


@api_router.post("/early-out-requests/bulk-import")
async def bulk_import_early_out(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk-import early-out requests from .xlsx or .csv. Admin-only.

    - Maps employee by email; skipped if not found.
    - Required columns: Email, Request Date, Reason.
    - Time columns optional (file may not contain them).
    """
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")

    alias_index = _build_eo_alias_index()
    headers_clean = [h for h in headers if h]
    sheet_to_canon = {h: alias_index.get(_normalize_header(h)) for h in headers_clean}
    detected_canonical = {c for c in sheet_to_canon.values() if c}
    ignored_columns = [h for h, c in sheet_to_canon.items() if not c]

    missing = []
    if "email" not in detected_canonical:
        missing.append("Email")
    if "date" not in detected_canonical:
        missing.append("Request Date")
    if "reason" not in detected_canonical:
        missing.append("Reason")
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing mandatory column(s): {', '.join(missing)}")

    # Pre-fetch employees by email
    email_to_emp: dict = {}
    async for e in db.employees.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "official_email": 1, "full_name": 1, "team": 1, "department": 1}):
        email = (e.get("official_email") or "").strip().lower()
        if email:
            email_to_emp[email] = e

    # Pre-fetch users + employees for Approved By resolution
    approver_index: dict[str, str] = {}
    async for u in db.users.find({}, {"_id": 0, "id": 1, "username": 1, "email": 1}):
        if u.get("username"):
            approver_index[_normalize_header(u["username"])] = u["id"]
        if u.get("email"):
            approver_index[_normalize_header(u["email"])] = u["id"]
    async for e in db.employees.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "full_name": 1, "official_email": 1}):
        if e.get("full_name"):
            approver_index.setdefault(_normalize_header(e["full_name"]), e["id"])
        if e.get("official_email"):
            approver_index.setdefault(_normalize_header(e["official_email"]), e["id"])

    def _resolve_eo_approver(value) -> tuple[Optional[str], Optional[str]]:
        if value in (None, ""):
            return (None, None)
        s = str(value).strip()
        if not s or s.upper() in ("NULL", "NONE", "N/A", "NA", "-"):
            return (None, None)
        if s.isdigit() or (s.replace(".", "", 1).isdigit()):
            return (current_user["id"], s)
        norm = _normalize_header(s)
        if norm in ("admin", "administrator", "hr admin", "hr"):
            return (current_user["id"], None)
        if norm in approver_index:
            return (approver_index[norm], None)
        return (None, s)

    def _clean_null(value) -> Optional[str]:
        if value in (None, ""):
            return None
        s = str(value).strip()
        if not s or s.upper() in ("NULL", "NONE", "N/A", "NA", "-"):
            return None
        return s

    total = len(rows)
    success = 0
    failed = 0
    skipped_duplicates = 0
    errors: List[dict] = []
    inserts: List[dict] = []
    inbatch_keys: set = set()

    for idx, raw_row in enumerate(rows, start=2):
        row = _remap_row(raw_row, alias_index)

        email = (str(row.get("email") or "")).strip().lower()
        if not email:
            failed += 1
            errors.append({"row": idx, "email": "", "reason": "Email is blank"})
            continue
        emp = email_to_emp.get(email)
        if not emp:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "No employee found with this email"})
            continue

        date_iso, _ = _parse_import_datetime(row.get("date")) if row.get("date") not in (None, "") else (None, None)
        if not date_iso:
            date_iso = _parse_import_date(row.get("date"))
        if not date_iso:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": f"Invalid Request Date: '{row.get('date')}'"})
            continue

        # Optional times
        actual_time = None
        if row.get("actual_time") not in (None, ""):
            _, actual_time = _parse_import_datetime(row.get("actual_time"))
            if not actual_time:
                actual_time = _parse_import_time(row.get("actual_time"))

        expected_time = None
        if row.get("expected_time") not in (None, ""):
            _, expected_time = _parse_import_datetime(row.get("expected_time"))
            if not expected_time:
                expected_time = _parse_import_time(row.get("expected_time"))

        reason = _clean_null(row.get("reason"))
        if not reason:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "Reason is required"})
            continue

        status_norm = _normalize_status(row.get("status")) or "pending"
        is_lop = _normalize_action_type(row.get("action_type"))
        lop_remark = _clean_null(row.get("lop_remark"))

        approver_id, approver_unresolved = _resolve_eo_approver(row.get("approved_by"))
        applied_at_dt, applied_at_time = _parse_import_datetime(row.get("applied_at")) if row.get("applied_at") not in (None, "") else (None, None)
        approved_at_dt, approved_at_time = _parse_import_datetime(row.get("approved_at")) if row.get("approved_at") not in (None, "") else (None, None)
        applied_at_iso = f"{applied_at_dt}T{applied_at_time or '00:00'}" if applied_at_dt else None
        approved_at_iso = f"{approved_at_dt}T{approved_at_time or '00:00'}" if approved_at_dt else None

        updated_at_iso = None
        if row.get("updated_at") not in (None, ""):
            ud_d, ud_t = _parse_import_datetime(row.get("updated_at"))
            if ud_d:
                updated_at_iso = f"{ud_d}T{ud_t or '00:00'}"

        # Duplicate guard: same employee + date, non-rejected
        existing = await db.early_out_requests.find_one({
            "employee_id": emp["id"],
            "date": date_iso,
            "status": {"$ne": "rejected"}
        }, {"_id": 0, "id": 1})
        if existing and status_norm != "rejected":
            skipped_duplicates += 1
            errors.append({"row": idx, "email": email, "reason": f"Early-out request already exists for {date_iso}"})
            continue

        batch_key = (emp["id"], date_iso, status_norm)
        if batch_key in inbatch_keys and status_norm != "rejected":
            skipped_duplicates += 1
            errors.append({"row": idx, "email": email, "reason": f"Duplicate of an earlier row for {date_iso}"})
            continue
        inbatch_keys.add(batch_key)

        final_approved_by = approver_id
        if status_norm == "approved" and not final_approved_by:
            final_approved_by = current_user["id"]

        extra = {
            k: v for k, v in {
                "applied_at": applied_at_iso,
                "approved_at": approved_at_iso,
                "updated_at": updated_at_iso,
                "approver_unresolved": approver_unresolved,
            }.items() if v is not None
        } or None

        doc = {
            "id": str(uuid.uuid4()),
            "employee_id": emp["id"],
            "emp_name": emp.get("full_name", ""),
            "team": emp.get("team", ""),
            "department": emp.get("department", ""),
            "date": date_iso,
            "expected_time": expected_time,
            "actual_time": actual_time,
            "reason": reason,
            "status": status_norm,
            "is_lop": is_lop,
            "lop_remark": lop_remark,
            "approved_by": final_approved_by,
            "applied_by_admin": True,
            "extra_data": extra,
            "created_at": get_ist_now().isoformat(),
        }
        inserts.append(doc)
        success += 1

    if inserts:
        chunk = 500
        for i in range(0, len(inserts), chunk):
            await db.early_out_requests.insert_many([d.copy() for d in inserts[i:i + chunk]])

    return {
        "total": total,
        "success": success,
        "skipped_duplicates": skipped_duplicates,
        "failed": failed,
        "column_mapping": {h: c for h, c in sheet_to_canon.items() if c},
        "ignored_columns": ignored_columns,
        "errors": errors,
    }

# ============== MISSED PUNCH REQUEST ROUTES ==============

# ============== ATTENDANCE UPDATE HELPER ==============

async def _update_attendance_from_missed_punch(rec):
    """Production Missed-Punch Approval Engine.

    Applies an APPROVED missed-punch correction to the FINAL attendance row
    for (employee_id, date) WITHOUT touching raw biometric logs.

    Behaviour (per business spec):
      • Type 'Check-in' → REPLACE check_in / check_in_24h only.
      • Type 'Check-out' → REPLACE check_out / check_out_24h only.
      • Type 'Both' → REPLACE both pairs (requires both times).
      • Hard replace — corrected times always win over biometric values
        (the request was approved by HR, that's the whole point).
      • If no attendance row exists for the date → INSERT one.
      • Recomputes total_hours (cross-midnight aware) and shift-based
        status (Present / Loss of Pay / Login / Worked on Sunday) via the
        existing `calculate_attendance_status` engine.
      • Sets `source = 'corrected'` and `missed_punch_corrected = True`.
      • Writes a full before/after audit row to `attendance_corrections`.
      • Idempotent: if the record already contains the exact target values
        attributed to the same request_id, no rewrite is performed.

    Raw biometric logs (`biometric_punch_logs`) are NEVER touched.
    """

    emp_id = rec.get("employee_id")
    date = rec.get("date")
    punch_type = (rec.get("punch_type") or "").strip()
    check_in_raw = rec.get("check_in_time")
    check_out_raw = rec.get("check_out_time")
    request_id = rec.get("id")

    if not emp_id or not date or not punch_type:
        return  # malformed request — abort safely

    # Normalize attendance date to DD-MM-YYYY — the canonical format used by
    # the rest of the attendance collection / API filters / frontend. Missed-
    # punch requests come in from the HTML date input as YYYY-MM-DD; writing
    # the raw value hides the corrected record from the attendance grid.
    def _to_ddmmyyyy(ds: str) -> str:
        if not ds:
            return ds
        s = str(ds).strip()
        if len(s) == 10 and s[4] == "-" and s[7] == "-":
            try:
                return datetime.strptime(s, "%Y-%m-%d").strftime("%d-%m-%Y")
            except ValueError:
                return s
        return s

    date = _to_ddmmyyyy(date)

    def _to_24h(raw: Optional[str]) -> Optional[str]:
        """Accept 'YYYY-MM-DDTHH:MM' / 'HH:MM' / 'HH:MM:SS' / 'HH:MM AM/PM'."""
        if not raw:
            return None
        s = str(raw).strip()
        if "T" in s:
            s = s.split("T")[-1]
        try:
            return datetime.strptime(s[:5], "%H:%M").strftime("%H:%M")
        except ValueError:
            pass
        try:  # AM/PM
            return datetime.strptime(s, "%I:%M %p").strftime("%H:%M")
        except ValueError:
            return None

    def _to_12h(t24: Optional[str]) -> Optional[str]:
        if not t24:
            return None
        try:
            return datetime.strptime(t24, "%H:%M").strftime("%I:%M %p")
        except ValueError:
            return None

    new_in_24h = _to_24h(check_in_raw) if punch_type in ("Check-in", "Both") else None
    new_out_24h = _to_24h(check_out_raw) if punch_type in ("Check-out", "Both") else None

    # Type-specific validation — refuse silently if the request is incoherent
    # (validation already enforced at create-time; this is belt-and-braces).
    if punch_type == "Check-in" and not new_in_24h:
        return
    if punch_type == "Check-out" and not new_out_24h:
        return
    if punch_type == "Both" and not (new_in_24h and new_out_24h):
        return

    existing = await db.attendance.find_one({"employee_id": emp_id, "date": date}, {"_id": 0})
    employee = await db.employees.find_one({"id": emp_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not employee:
        return

    old_in = existing.get("check_in") if existing else None
    old_in_24h = existing.get("check_in_24h") if existing else None
    old_out = existing.get("check_out") if existing else None
    old_out_24h = existing.get("check_out_24h") if existing else None
    old_status = existing.get("status") if existing else None

    # Compose the post-correction in/out (always REPLACE on type-targeted fields)
    final_in_24h = new_in_24h if new_in_24h is not None else old_in_24h
    final_out_24h = new_out_24h if new_out_24h is not None else old_out_24h

    # Idempotency — skip when the same request has already produced the same
    # state. We compare both raw strings AND the source-of-correction id.
    if (
        existing
        and existing.get("source") == "corrected"
        and existing.get("missed_punch_request_id") == request_id
        and ((new_in_24h is None) or existing.get("check_in_24h") == new_in_24h)
        and ((new_out_24h is None) or existing.get("check_out_24h") == new_out_24h)
    ):
        return

    # Recompute shift-aware status (handles cross-midnight, Sundays, LOP rules)
    shift_timings = get_shift_timings(employee) if employee else None
    status_result = (
        calculate_attendance_status(final_in_24h, final_out_24h, shift_timings or {}, attendance_date=date)
        if final_in_24h
        else {"status": AttendanceStatus.NOT_LOGGED, "is_lop": False, "lop_reason": None, "total_hours_decimal": 0.0, "expected_logout": None}
    )

    total_hours_decimal = float(status_result.get("total_hours_decimal") or 0.0)
    total_hours_str = calculate_total_hours_str(total_hours_decimal) if total_hours_decimal else None

    payload = {
        "status": status_result.get("status", AttendanceStatus.PRESENT),
        "is_lop": bool(status_result.get("is_lop")),
        "lop_reason": status_result.get("lop_reason"),
        "expected_logout": status_result.get("expected_logout"),
        "total_hours_decimal": round(total_hours_decimal, 2),
        "total_hours": total_hours_str,
        "source": "corrected",
        "missed_punch_corrected": True,
        "missed_punch_request_id": request_id,
        "missed_punch_corrected_at": get_ist_now().isoformat(),
        "missed_punch_corrected_by": rec.get("approved_by"),
    }
    if new_in_24h:
        payload["check_in_24h"] = new_in_24h
        payload["check_in"] = _to_12h(new_in_24h)
    if new_out_24h:
        payload["check_out_24h"] = new_out_24h
        payload["check_out"] = _to_12h(new_out_24h)

    if existing:
        await db.attendance.update_one(
            {"employee_id": emp_id, "date": date},
            {"$set": payload},
        )
    else:
        new_doc = {
            "id": str(uuid.uuid4()),
            "employee_id": emp_id,
            "emp_name": employee.get("full_name", ""),
            "team": employee.get("team", ""),
            "department": employee.get("department", ""),
            "date": date,
            "shift_type": employee.get("shift_type", "General"),
            "expected_login": (shift_timings or {}).get("login_time"),
            "created_at": get_ist_now().isoformat(),
            **payload,
        }
        await db.attendance.insert_one(new_doc.copy())

    # Audit trail — separate collection so it survives even if attendance is later edited
    try:
        audit_doc = {
            "id": str(uuid.uuid4()),
            "request_id": request_id,
            "employee_id": emp_id,
            "emp_name": employee.get("full_name", ""),
            "date": date,
            "punch_type": punch_type,
            "old_check_in": old_in,
            "old_check_in_24h": old_in_24h,
            "old_check_out": old_out,
            "old_check_out_24h": old_out_24h,
            "old_status": old_status,
            "new_check_in": payload.get("check_in", old_in),
            "new_check_in_24h": payload.get("check_in_24h", old_in_24h),
            "new_check_out": payload.get("check_out", old_out),
            "new_check_out_24h": payload.get("check_out_24h", old_out_24h),
            "new_status": payload["status"],
            "new_total_hours": payload.get("total_hours"),
            "approved_by": rec.get("approved_by"),
            "approved_at": rec.get("approved_at"),
            "source_before": (existing or {}).get("source", "biometric") if existing else "none",
            "source_after": "corrected",
            "created_at": get_ist_now().isoformat(),
        }
        await db.attendance_corrections.insert_one(audit_doc.copy())
    except Exception as e:  # never let audit failure roll back the correction itself
        logger.warning(f"Missed-punch audit insert failed for request {request_id}: {e}")

    logger.info(
        f"Missed punch applied: emp={emp_id} date={date} type={punch_type} "
        f"in {old_in_24h}->{payload.get('check_in_24h', old_in_24h)} "
        f"out {old_out_24h}->{payload.get('check_out_24h', old_out_24h)} "
        f"status {old_status}->{payload['status']} request={request_id}"
    )

    # Mark the source request as APPLIED — explicit flag the historical
    # backfill / payroll / dashboards can rely on without joining audit rows.
    if request_id:
        try:
            await db.missed_punches.update_one(
                {"id": request_id},
                {"$set": {
                    "correction_applied_at": get_ist_now().isoformat(),
                    "correction_applied_by": rec.get("approved_by") or "system",
                    "is_applied": True,
                }},
            )
        except Exception as e:
            logger.warning(f"Failed to stamp correction_applied_at on missed_punch {request_id}: {e}")

@api_router.get("/missed-punches")
async def get_missed_punches(
    status: Optional[str] = None,
    tab: Optional[str] = None,  # 'pending' | 'history' — filters and counts per-tab
    employee_name: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    current_user: dict = Depends(get_current_user)
):
    base_query: dict = {}
    is_admin = current_user["role"] in ALL_ADMIN_ROLES
    if not is_admin:
        base_query["employee_id"] = current_user.get("employee_id", "")
    if status and status != "All":
        base_query["status"] = status
    if employee_name:
        base_query["emp_name"] = {"$regex": employee_name, "$options": "i"}
    if from_date:
        base_query.setdefault("date", {})["$gte"] = from_date
    if to_date:
        base_query.setdefault("date", {})["$lte"] = to_date

    # Global per-tab counts (computed against base_query, ignore active tab)
    pending_count = await db.missed_punches.count_documents({**base_query, "status": "pending"})
    history_count = await db.missed_punches.count_documents({**base_query, "status": {"$ne": "pending"}})

    # Apply tab filter to the records & total returned
    query = dict(base_query)
    if tab == "pending":
        query["status"] = "pending"
    elif tab == "history":
        query["status"] = {"$ne": "pending"}

    total = await db.missed_punches.count_documents(query)
    skip = (page - 1) * per_page
    records = await db.missed_punches.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    return {
        "data": [serialize_doc(r) for r in records],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pending_count": pending_count,
        "history_count": history_count,
    }

@api_router.post("/missed-punches")
async def create_missed_punch(data: MissedPunchCreate, current_user: dict = Depends(get_current_user)):
    emp, is_hr = await _resolve_employee(data.employee_id, current_user)
    
    # JOB 4: Prevent duplicate missed punch (same employee + date + punch_type)
    existing = await db.missed_punches.find_one({
        "employee_id": emp["id"],
        "date": data.date,
        "punch_type": data.punch_type,
        "status": {"$ne": "rejected"}
    })
    if existing:
        raise HTTPException(status_code=400, detail="Request already exists for this date and punch type")
    
    req = MissedPunchRequest(
        employee_id=emp["id"],
        emp_name=emp["full_name"],
        team=emp["team"],
        department=emp["department"],
        date=data.date,
        punch_type=data.punch_type,
        check_in_time=data.check_in_time,
        check_out_time=data.check_out_time,
        reason=data.reason,
        applied_by_admin=is_hr,
        status="approved" if (data.auto_approve and is_hr) else "pending",
        approved_by=current_user["id"] if (data.auto_approve and is_hr) else None
    )
    doc = req.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.missed_punches.insert_one(doc.copy())
    
    # JOB 3: If auto-approved, update attendance
    if data.auto_approve and is_hr:
        await _update_attendance_from_missed_punch(doc)
    
    return serialize_doc(doc)

@api_router.put("/missed-punches/{request_id}/approve")
async def approve_missed_punch(request_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    rec = await db.missed_punches.find_one({"id": request_id}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Request not found")
    if rec.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Request is not pending")
    result = await db.missed_punches.update_one({"id": request_id}, {"$set": {"status": "approved", "approved_by": current_user["id"], "approved_at": get_ist_now().isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # JOB 3: Update attendance record on approval
    updated_rec = await db.missed_punches.find_one({"id": request_id}, {"_id": 0})
    await _update_attendance_from_missed_punch(updated_rec)
    
    return serialize_doc(updated_rec)

@api_router.put("/missed-punches/{request_id}/reject")
async def reject_missed_punch(request_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.HR]:
        raise HTTPException(status_code=403, detail="Permission denied")
    result = await db.missed_punches.update_one({"id": request_id}, {"$set": {"status": "rejected", "approved_by": current_user["id"]}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    rec = await db.missed_punches.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(rec)


class BackfillBody(BaseModel):
    """Optional knobs for the historical backfill endpoint."""
    dry_run: bool = False             # if True, don't actually mutate anything — return counts only
    batch_size: int = 500             # process up to N requests per call
    from_date: Optional[str] = None   # YYYY-MM-DD inclusive (filter on missed_punches.date)
    to_date: Optional[str] = None     # YYYY-MM-DD inclusive
    employee_id: Optional[str] = None # restrict to a single employee
    force: bool = False               # re-apply even if already marked applied (still idempotent in the engine)


@api_router.post("/missed-punches/backfill")
async def backfill_missed_punches(body: BackfillBody = BackfillBody(), current_user: dict = Depends(get_current_user)):
    """Historical Backfill Engine.

    Iterates over APPROVED missed-punch requests whose corrections were never
    propagated to the final attendance table and applies them via the same
    `_update_attendance_from_missed_punch` engine the live approval flow uses.

    Idempotency is enforced on three layers:
      1. Server-side filter: skip requests already stamped with
         `correction_applied_at` (unless `force=True`).
      2. Skip requests that already have an `attendance_corrections` row
         keyed by request_id.
      3. The engine itself is idempotent — re-applying the same target state
         from the same request_id is a no-op (no extra audit row written).

    Raw biometric logs are NEVER touched. Reports / payroll automatically
    pick up the corrected values because they read the same final
    `attendance` collection the engine writes to.
    """
    if current_user["role"] not in ALL_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")

    q = {"status": "approved"}
    if not body.force:
        q["$or"] = [
            {"correction_applied_at": {"$exists": False}},
            {"correction_applied_at": None},
        ]
    if body.employee_id:
        q["employee_id"] = body.employee_id
    if body.from_date or body.to_date:
        date_q = {}
        if body.from_date:
            date_q["$gte"] = body.from_date
        if body.to_date:
            date_q["$lte"] = body.to_date
        q["date"] = date_q

    candidates = await db.missed_punches.find(q, {"_id": 0}).limit(max(1, min(body.batch_size, 5000))).to_list(5000)

    stats = {
        "candidates": len(candidates),
        "applied": 0,
        "skipped_already_applied": 0,
        "skipped_invalid": 0,
        "errors": [],
        "dry_run": body.dry_run,
    }

    for rec in candidates:
        rid = rec.get("id")
        # Skip if audit row already exists for this request (defence-in-depth)
        if not body.force and rid:
            prior = await db.attendance_corrections.find_one({"request_id": rid}, {"_id": 0})
            if prior:
                # Heal the source flag if missing (legacy data)
                if not rec.get("correction_applied_at") and not body.dry_run:
                    await db.missed_punches.update_one(
                        {"id": rid},
                        {"$set": {
                            "correction_applied_at": prior.get("created_at") or get_ist_now().isoformat(),
                            "correction_applied_by": prior.get("approved_by") or "system_backfill",
                            "is_applied": True,
                        }},
                    )
                stats["skipped_already_applied"] += 1
                continue

        # Validate the request has the times required by its punch_type —
        # malformed historical rows must be flagged, not silently swallowed.
        ptype = (rec.get("punch_type") or "").strip()
        if ptype == "Check-in" and not rec.get("check_in_time"):
            stats["skipped_invalid"] += 1
            stats["errors"].append({"id": rid, "reason": "Check-in request missing check_in_time"})
            continue
        if ptype == "Check-out" and not rec.get("check_out_time"):
            stats["skipped_invalid"] += 1
            stats["errors"].append({"id": rid, "reason": "Check-out request missing check_out_time"})
            continue
        if ptype == "Both" and not (rec.get("check_in_time") and rec.get("check_out_time")):
            stats["skipped_invalid"] += 1
            stats["errors"].append({"id": rid, "reason": "Both request missing one or both times"})
            continue

        if body.dry_run:
            stats["applied"] += 1
            continue

        rec_for_engine = dict(rec)
        # Backfill-attribution for the audit row
        rec_for_engine["approved_by"] = rec.get("approved_by") or "system_backfill"
        rec_for_engine["approved_at"] = rec.get("approved_at") or get_ist_now().isoformat()

        try:
            await _update_attendance_from_missed_punch(rec_for_engine)
            stats["applied"] += 1
        except Exception as e:
            stats["errors"].append({"id": rid, "reason": str(e)})

    await log_audit(
        current_user["id"], "backfill", "missed_punches",
        f"applied={stats['applied']}, skipped={stats['skipped_already_applied']}, "
        f"invalid={stats['skipped_invalid']}, dry_run={body.dry_run}",
    )
    return stats


@api_router.get("/missed-punches/backfill/status")
async def backfill_status(current_user: dict = Depends(get_current_user)):
    """Quick health snapshot for the backfill engine — used by an admin UI
    so HR can see how much work is outstanding before kicking off a run."""
    if current_user["role"] not in ALL_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Permission denied")
    total_approved = await db.missed_punches.count_documents({"status": "approved"})
    pending_apply = await db.missed_punches.count_documents({
        "status": "approved",
        "$or": [
            {"correction_applied_at": {"$exists": False}},
            {"correction_applied_at": None},
        ],
    })
    audit_rows = await db.attendance_corrections.count_documents({})
    last_correction = await db.attendance_corrections.find_one(
        {}, {"_id": 0, "request_id": 1, "created_at": 1, "employee_id": 1, "date": 1},
        sort=[("created_at", -1)],
    )
    return {
        "total_approved_requests": total_approved,
        "pending_correction_apply": pending_apply,
        "applied_correction_audit_rows": audit_rows,
        "last_correction": last_correction,
    }



@api_router.put("/missed-punches/{request_id}")
async def edit_missed_punch(request_id: str, data: MissedPunchCreate, current_user: dict = Depends(get_current_user)):
    rec = await db.missed_punches.find_one({"id": request_id}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Request not found")
    is_admin = current_user["role"] in ALL_ADMIN_ROLES
    if not is_admin and rec.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Can only edit pending requests")
    update = {
        "date": data.date,
        "punch_type": data.punch_type,
        "check_in_time": data.check_in_time,
        "check_out_time": data.check_out_time,
        "reason": data.reason,
        "edited_by": current_user["id"],
        "edited_at": get_ist_now().isoformat(),
    }
    await db.missed_punches.update_one({"id": request_id}, {"$set": update})
    await log_audit(current_user["id"], "edit", "missed_punch", request_id)
    updated = await db.missed_punches.find_one({"id": request_id}, {"_id": 0})
    return serialize_doc(updated)

# ============== MISSED PUNCH BULK IMPORT ==============

MP_IMPORT_COLUMNS = [
    "Emp Mail ID", "Punch Date", "Punch Type", "In Time", "Out Time",
    "Reason", "Status", "Applied Date", "Approved By", "Approval Date", "Comments"
]

# Sheet-column aliases → canonical field names. Case-insensitive, whitespace-tolerant.
MP_COLUMN_ALIASES: dict[str, str] = {
    # email → employee
    "Emp Mail ID": "email",
    "Employee Email": "email",
    "Email": "email",
    "Email ID": "email",
    "Mail": "email",
    # date
    "Punch Date": "date",
    "Date": "date",
    "Attendance Date": "date",
    # punch type
    "Punch Type": "punch_type",
    "Type": "punch_type",
    "Punch": "punch_type",
    # in time
    "In Time": "check_in",
    "Punch In": "check_in",
    "Check In": "check_in",
    "Check-In": "check_in",
    "Login Time": "check_in",
    # out time
    "Out Time": "check_out",
    "Punch Out": "check_out",
    "Check Out": "check_out",
    "Check-Out": "check_out",
    "Logout Time": "check_out",
    # reason
    "Reason": "reason",
    "Remarks": "reason",
    "Remark": "reason",
    "Notes": "reason",
    # status
    "Status": "status",
    "Request Status": "status",
    # applied at
    "Applied Date": "applied_at",
    "Applied On": "applied_at",
    "Date Applied": "applied_at",
    # approved by
    "Approved By": "approved_by",
    "Approver": "approved_by",
    "Approving Authority": "approved_by",
    # approval date
    "Approval Date": "approved_at",
    "Approved On": "approved_at",
    "Date Approved": "approved_at",
    # comments
    "Comments": "comments",
    "Approver Notes": "comments",
}


def _normalize_punch_type(value) -> Optional[str]:
    """Normalize Punch Type to canonical 'Check-in', 'Check-out', or 'Both'.
    Returns None if blank, or 'INVALID' sentinel when unrecognized."""
    if value in (None, ""):
        return None
    s = str(value).strip().lower().replace("_", "-").replace(" ", "")
    if s in ("check-in", "checkin", "in", "punchin"):
        return "Check-in"
    if s in ("check-out", "checkout", "out", "punchout"):
        return "Check-out"
    if s in ("both", "checkin+checkout", "in+out", "in&out"):
        return "Both"
    return "INVALID"


def _build_mp_alias_index() -> dict[str, str]:
    return {_normalize_header(k): v for k, v in MP_COLUMN_ALIASES.items()}


_COMBINED_DT_FORMATS = (
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
    "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M",
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
    "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
    "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
    "%d-%m-%Y %I:%M %p", "%d-%m-%Y %I:%M:%S %p",
    "%d/%m/%Y %I:%M %p", "%d/%m/%Y %I:%M:%S %p",
    "%Y-%m-%d %I:%M %p", "%Y-%m-%d %I:%M:%S %p",
    "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M",
)


def _parse_import_datetime(value) -> tuple[Optional[str], Optional[str]]:
    """Parse a cell that may contain combined Date+Time, time-only, or date-only.
    Returns (date_iso or None, time_hhmm or None)."""
    if value is None:
        return (None, None)
    if isinstance(value, datetime):
        # If openpyxl returned a datetime, treat midnight as 'date-only'
        d = value.strftime("%Y-%m-%d")
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return (d, None)
        return (d, value.strftime("%H:%M"))
    # datetime.time
    if hasattr(value, "hour") and hasattr(value, "minute") and not hasattr(value, "year"):
        try:
            return (None, f"{value.hour:02d}:{value.minute:02d}")
        except Exception:
            return (None, None)
    s = str(value).strip()
    if not s:
        return (None, None)
    # Try combined Date+Time formats first
    for fmt in _COMBINED_DT_FORMATS:
        try:
            dt = datetime.strptime(s, fmt)
            return (dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M"))
        except ValueError:
            continue
    # Try date-only
    d = _parse_import_date(s)
    if d:
        return (d, None)
    # Try time-only
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p", "%I:%M%p"):
        try:
            return (None, datetime.strptime(s, fmt).strftime("%H:%M"))
        except ValueError:
            continue
    return (None, None)


def _parse_import_time(value) -> Optional[str]:
    """Parse a time cell to canonical 24h "HH:MM" string. Returns None if blank/invalid.
    Also accepts combined Date+Time cells - returns only the time part."""
    _, t = _parse_import_datetime(value)
    return t


@api_router.get("/missed-punches/import-template")
async def download_mp_import_template(current_user: dict = Depends(get_current_user)):
    """Download styled .xlsx template for missed-punch bulk import."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missed Punch Import"
    header_fill = openpyxl.styles.PatternFill(start_color="063c88", end_color="063c88", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(MP_IMPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    sample = [
        ["employee@blubridge.com", "2026-04-22", "Check-in", "09:00", "", "Forgot to punch in", "Approved", "2026-04-23", "Admin", "2026-04-23", "Verified by HR"],
        ["another@blubridge.com", "22/04/2026", "Check-out", "", "18:30", "Forgot to punch out", "Pending", "2026-04-23", "", "", ""],
        ["third@blubridge.com", "23/04/2026", "Both", "09:15", "18:45", "Biometric down all day", "Approved", "2026-04-24", "Admin", "2026-04-24", ""],
    ]
    for r_idx, row_vals in enumerate(sample, start=2):
        for c_idx, v in enumerate(row_vals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    widths = [30, 14, 14, 12, 12, 30, 12, 14, 18, 14, 30]
    for c_idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w

    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Sheet Column", "Required", "Maps To", "Notes"])
    ws2.append(["Emp Mail ID", "Yes", "employee_id (via email)", "Aliases: Employee Email, Email, Email ID, Mail"])
    ws2.append(["Punch Date", "Optional", "date", "Aliases: Date, Attendance Date. Accepts YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY. If omitted, date is auto-extracted from combined Date+Time in 'In Time'/'Out Time'."])
    ws2.append(["Punch Type", "Optional", "punch_type", "Aliases: Type, Punch. Accepted: 'Check-in', 'Check-out', 'Both'. If omitted, auto-detected: both times → Both, only In → Check-in, only Out → Check-out."])
    ws2.append(["In Time", "At least one of In/Out required", "check_in_time", "Aliases: Punch In, Check In, Login Time. Accepts HH:MM, HH:MM AM/PM, OR combined 'DD-MM-YYYY HH:MM' (date+time). For cross-midnight shifts, In Time's date is used as the anchor Punch Date."])
    ws2.append(["Out Time", "At least one of In/Out required", "check_out_time", "Aliases: Punch Out, Check Out, Logout Time. Accepts HH:MM, HH:MM AM/PM, OR combined 'DD-MM-YYYY HH:MM'."])
    ws2.append(["Reason", "Yes", "reason", "Aliases: Remarks, Notes. Free text"])
    ws2.append(["Status", "No", "status", "Approved / Pending / Rejected (default: Pending). Approved status auto-updates the attendance record."])
    ws2.append(["Applied Date", "No", "applied_at", "Stored as-is in extra_data"])
    ws2.append(["Approved By", "No", "approved_by", "'Admin' or admin's username/email/full name → resolved to user ID"])
    ws2.append(["Approval Date", "No", "approved_at", "Stored as-is when status = approved"])
    ws2.append(["Comments", "No", "comments", "Aliases: Approver Notes. Stored in extra_data"])
    for col_idx in range(1, 5):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 36
    for c in range(1, 5):
        ws2.cell(1, c).font = openpyxl.styles.Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=missed_punches_import_template.xlsx"}
    )


@api_router.post("/missed-punches/import/preview")
async def preview_mp_import(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview detected columns + sample rows for missed-punch import."""
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")
    headers = [h for h in headers if h]

    alias_index = _build_mp_alias_index()
    column_mapping = {h: alias_index.get(_normalize_header(h)) for h in headers}
    mapped = {h: c for h, c in column_mapping.items() if c}
    ignored = [h for h, c in column_mapping.items() if not c]

    required_fields = ["email"]
    detected_canonical = set(mapped.values())
    missing_required = [f for f in required_fields if f not in detected_canonical]
    # Need at least one of (check_in, check_out) - date & punch_type can be auto-derived from these
    has_time = "check_in" in detected_canonical or "check_out" in detected_canonical
    if not has_time:
        missing_required.append("In Time or Out Time")

    sample = []
    for r in rows[:5]:
        sample.append({k: (v.isoformat() if isinstance(v, datetime) else (str(v) if v is not None else None)) for k, v in r.items()})

    return {
        "filename": file.filename,
        "total_rows": len(rows),
        "headers": headers,
        "column_mapping": mapped,
        "ignored_columns": ignored,
        "missing_required": missing_required,
        "ready_to_import": len(missing_required) == 0,
        "sample_rows": sample,
    }


@api_router.post("/missed-punches/bulk-import")
async def bulk_import_missed_punches(
    file: UploadFile = FastAPIFile(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk-import missed-punch requests from .xlsx or .csv. Admin-only.

    - Maps employee by email; skipped if not found.
    - Requires at least one of In Time / Out Time per row.
    - Approved rows automatically update the attendance record (matching the
      manual approval flow via _update_attendance_from_missed_punch).
    """
    if current_user["role"] not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    headers, rows = _read_leave_import_rows(file_bytes, file.filename or "")

    alias_index = _build_mp_alias_index()
    headers_clean = [h for h in headers if h]
    sheet_to_canon = {h: alias_index.get(_normalize_header(h)) for h in headers_clean}
    detected_canonical = {c for c in sheet_to_canon.values() if c}
    ignored_columns = [h for h, c in sheet_to_canon.items() if not c]

    if "email" not in detected_canonical:
        raise HTTPException(status_code=400, detail="Missing mandatory column: Emp Mail ID")
    if "check_in" not in detected_canonical and "check_out" not in detected_canonical:
        raise HTTPException(status_code=400, detail="At least one of 'In Time' / 'Out Time' columns is required")

    # Pre-fetch employees by email
    email_to_emp: dict = {}
    async for e in db.employees.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "official_email": 1, "full_name": 1, "team": 1, "department": 1}):
        email = (e.get("official_email") or "").strip().lower()
        if email:
            email_to_emp[email] = e

    # Pre-fetch users + employees for Approved By resolution
    approver_index: dict[str, str] = {}
    async for u in db.users.find({}, {"_id": 0, "id": 1, "username": 1, "email": 1}):
        if u.get("username"):
            approver_index[_normalize_header(u["username"])] = u["id"]
        if u.get("email"):
            approver_index[_normalize_header(u["email"])] = u["id"]
    async for e in db.employees.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "full_name": 1, "official_email": 1}):
        if e.get("full_name"):
            approver_index.setdefault(_normalize_header(e["full_name"]), e["id"])
        if e.get("official_email"):
            approver_index.setdefault(_normalize_header(e["official_email"]), e["id"])

    def _resolve_mp_approver(value) -> tuple[Optional[str], Optional[str]]:
        if value in (None, ""):
            return (None, None)
        s = str(value).strip()
        if not s:
            return (None, None)
        norm = _normalize_header(s)
        if norm in ("admin", "administrator", "hr admin", "hr"):
            return (current_user["id"], None)
        if norm in approver_index:
            return (approver_index[norm], None)
        return (None, s)

    total = len(rows)
    success = 0
    failed = 0
    skipped_duplicates = 0
    errors: List[dict] = []
    inserts: List[dict] = []
    approved_inserts: List[dict] = []  # records to push through attendance update
    inbatch_keys: set = set()  # (employee_id, date, punch_type) seen earlier in this upload

    for idx, raw_row in enumerate(rows, start=2):
        row = _remap_row(raw_row, alias_index)

        email = (str(row.get("email") or "")).strip().lower()
        if not email:
            failed += 1
            errors.append({"row": idx, "email": "", "reason": "Employee Email is blank"})
            continue
        emp = email_to_emp.get(email)
        if not emp:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "No employee found with this email"})
            continue

        # Parse In Time / Out Time cells - they may contain combined Date+Time, time-only,
        # or date-only values. Extract date+time parts independently.
        in_raw = row.get("check_in")
        out_raw = row.get("check_out")
        in_date, in_time = _parse_import_datetime(in_raw) if in_raw not in (None, "") else (None, None)
        out_date, out_time = _parse_import_datetime(out_raw) if out_raw not in (None, "") else (None, None)

        # Validate: if a value was provided but couldn't be parsed at all → fail
        if in_raw not in (None, "") and in_date is None and in_time is None:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": f"Invalid In Time format '{in_raw}' (use HH:MM, DD-MM-YYYY HH:MM, etc.)"})
            continue
        if out_raw not in (None, "") and out_date is None and out_time is None:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": f"Invalid Out Time format '{out_raw}' (use HH:MM, DD-MM-YYYY HH:MM, etc.)"})
            continue

        # Derive Punch Date: prefer explicit Punch Date column → fallback to In Time's date → Out Time's date
        date_iso = _parse_import_date(row.get("date")) if row.get("date") not in (None, "") else None
        if not date_iso:
            date_iso = in_date or out_date
        if not date_iso:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "Punch Date is required (provide either 'Punch Date' column or combined Date+Time in 'In Time'/'Out Time')"})
            continue

        check_in_time = in_time
        check_out_time = out_time

        # Derive Punch Type: explicit value wins, else auto-detect from time presence
        raw_punch = row.get("punch_type")
        punch_norm = _normalize_punch_type(raw_punch)
        if punch_norm == "INVALID":
            failed += 1
            errors.append({"row": idx, "email": email, "reason": f"Invalid Punch Type '{raw_punch}' (use 'Check-in', 'Check-out', or 'Both')"})
            continue
        if punch_norm is None:
            # Auto-detect
            if check_in_time and check_out_time:
                punch_norm = "Both"
            elif check_in_time:
                punch_norm = "Check-in"
            elif check_out_time:
                punch_norm = "Check-out"
            else:
                failed += 1
                errors.append({"row": idx, "email": email, "reason": "Cannot determine Punch Type - In Time or Out Time is required"})
                continue
        punch_type = punch_norm

        # Use only the time(s) matching Punch Type; ignore the other.
        if punch_type == "Check-in":
            if not check_in_time:
                failed += 1
                errors.append({"row": idx, "email": email, "reason": "In Time is required when Punch Type is 'Check-in'"})
                continue
            check_out_time = None  # explicitly ignored
        elif punch_type == "Check-out":
            if not check_out_time:
                failed += 1
                errors.append({"row": idx, "email": email, "reason": "Out Time is required when Punch Type is 'Check-out'"})
                continue
            check_in_time = None  # explicitly ignored
        else:  # Both
            if not check_in_time or not check_out_time:
                failed += 1
                errors.append({"row": idx, "email": email, "reason": "Both punch type requires both In Time and Out Time"})
                continue

        reason = (str(row.get("reason")).strip() if row.get("reason") not in (None, "") else None)
        if not reason:
            failed += 1
            errors.append({"row": idx, "email": email, "reason": "Reason is required"})
            continue

        status_norm = _normalize_status(row.get("status")) or "pending"
        approver_id, approver_unresolved = _resolve_mp_approver(row.get("approved_by"))
        applied_at = _parse_import_date(row.get("applied_at"))
        approved_at = _parse_import_date(row.get("approved_at"))
        comments = (str(row.get("comments")).strip() if row.get("comments") not in (None, "") else None)

        # Duplicate guard (matches manual create rule): same employee + date + punch_type, non-rejected
        existing = await db.missed_punches.find_one({
            "employee_id": emp["id"],
            "date": date_iso,
            "punch_type": punch_type,
            "status": {"$ne": "rejected"}
        }, {"_id": 0, "id": 1})
        if existing:
            skipped_duplicates += 1
            errors.append({"row": idx, "email": email, "reason": f"Missed punch already exists for {date_iso} ({punch_type})"})
            continue

        # Also reject if an earlier row in THIS upload covers the same key
        batch_key = (emp["id"], date_iso, punch_type)
        if batch_key in inbatch_keys:
            skipped_duplicates += 1
            errors.append({"row": idx, "email": email, "reason": f"Duplicate of an earlier row for {date_iso} ({punch_type})"})
            continue
        inbatch_keys.add(batch_key)

        final_approved_by = approver_id
        if status_norm == "approved" and not final_approved_by:
            final_approved_by = current_user["id"]

        doc = {
            "id": str(uuid.uuid4()),
            "employee_id": emp["id"],
            "emp_name": emp.get("full_name", ""),
            "team": emp.get("team", ""),
            "department": emp.get("department", ""),
            "date": date_iso,
            "punch_type": punch_type,
            "check_in_time": check_in_time,
            "check_out_time": check_out_time,
            "reason": reason,
            "status": status_norm,
            "approved_by": final_approved_by,
            "applied_by_admin": True,
            "extra_data": {
                k: v for k, v in {
                    "applied_at": applied_at,
                    "approved_at": approved_at,
                    "approver_unresolved": approver_unresolved,
                    "comments": comments,
                }.items() if v is not None
            } or None,
            "created_at": get_ist_now().isoformat(),
        }
        inserts.append(doc)
        if status_norm == "approved":
            approved_inserts.append(doc)
        success += 1

    if inserts:
        chunk = 500
        for i in range(0, len(inserts), chunk):
            await db.missed_punches.insert_many([d.copy() for d in inserts[i:i + chunk]])

    # Apply attendance updates for approved rows (matches the manual approve flow)
    for doc in approved_inserts:
        try:
            await _update_attendance_from_missed_punch(doc)
        except Exception as e:
            errors.append({"row": "—", "email": doc.get("employee_id"), "reason": f"Attendance update warning: {str(e)[:120]}"})

    await log_audit(
        current_user["id"], "bulk_import", "missed_punches", None,
        f"Missed punch bulk import: total={total} success={success} duplicates={skipped_duplicates} failed={failed}"
    )

    return {
        "total": total,
        "success": success,
        "skipped_duplicates": skipped_duplicates,
        "failed": failed,
        "column_mapping": {h: c for h, c in sheet_to_canon.items() if c},
        "ignored_columns": ignored_columns,
        "errors": errors[:1000],
    }


# ---------------------------------------------------------------------------
# Help documentation — downloadable role-specific PDF
# ---------------------------------------------------------------------------
from help_docs import generate_help_pdf, generate_help_xlsx, ROLE_META  # noqa: E402


@api_router.get("/help/download")
async def download_help_guide(
    format: str = "pdf",
    current_user: dict = Depends(get_current_user),
):
    """Return a role-specific BluBridge HRMS User Guide (pdf or xlsx) for the logged-in user."""
    role = (current_user.get("role") or "employee").lower()
    user_name = current_user.get("name")
    fmt = (format or "pdf").lower()
    role_title = ROLE_META.get(role, ROLE_META["employee"])["title"]
    safe_role = role_title.replace(" ", "_")

    if fmt in ("xlsx", "excel", "xls"):
        data = generate_help_xlsx(role=role, user_name=user_name)
        filename = f"BluBridge_HRMS_{safe_role}_User_Guide.xlsx"
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        data = generate_help_pdf(role=role, user_name=user_name)
        filename = f"BluBridge_HRMS_{safe_role}_User_Guide.pdf"
        media = "application/pdf"

    return StreamingResponse(
        io.BytesIO(data),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Include router
# Register centralized Settings module routes
settings_services = settings_module.register(api_router, {
    "db": db,
    "get_current_user": get_current_user,
    "log_audit": log_audit,
    "get_ist_now": get_ist_now,
    "serialize_doc": serialize_doc,
    "parse_time_24h_to_minutes": parse_time_24h_to_minutes,
    "UserRole": UserRole,
    "ADMIN_ROLES": ADMIN_ROLES,
    "SYSTEM_ROLES": SYSTEM_ROLES,
    "EmployeeStatus": EmployeeStatus,
    "calculate_attendance_status": calculate_attendance_status,
    "get_shift_timings": get_shift_timings,
})


# ============== HRMS EMAIL MANUAL TRIGGER (HR-only, testing / on-demand) ==============
@api_router.post("/email-jobs/{job_name}/run")
async def run_email_job_now(job_name: str, current_user: dict = Depends(get_current_user)):
    """Manually trigger one of the scheduled HRMS email jobs.

    Run Now is admin-controlled and bypasses BOTH the enabled-toggle AND the
    dedup audit (so HR can re-send the daily summary on demand). The audit log
    still records each delivery — scoped uniquely with a timestamp suffix so
    the unique partial index never conflicts with prior sends.
    """
    if current_user.get("role") not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")
    handlers = _get_email_job_handlers()
    handler = handlers.get(job_name)
    if not handler:
        raise HTTPException(status_code=404, detail=f"Unknown job. Valid: {list(handlers)}")
    asyncio.create_task(handler(db, force=True))
    return {"status": "triggered", "job": job_name, "force": True}


@api_router.get("/email-jobs/audit")
async def email_jobs_audit(
    email_type: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user),
):
    if current_user.get("role") not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")
    q: dict = {}
    if email_type:
        q["email_type"] = email_type
    rows = await db.email_audit_logs.find(q, {"_id": 0}).sort("sent_at", -1).limit(max(1, min(limit, 500))).to_list(500)
    return {"count": len(rows), "rows": rows}


# ============== ADMIN-CONTROLLED CRON MANAGEMENT ==============
@api_router.get("/admin/cron-settings")
async def admin_get_cron_settings(current_user: dict = Depends(get_current_user)):
    """List all email cron jobs with metadata + admin toggle + last-run state.
    Admin-only (HR + system_admin)."""
    if current_user.get("role") not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")
    rows_by_name = {}
    async for doc in db.cron_settings.find({}, {"_id": 0}):
        rows_by_name[doc["job_name"]] = doc
    out = []
    for job_name, meta in _CRON_JOB_META.items():
        row = rows_by_name.get(job_name) or {"job_name": job_name, "enabled": True}
        out.append({
            "job_name": job_name,
            "label": meta["label"],
            "schedule": meta["schedule"],
            "scope": meta.get("scope"),
            "enabled": bool(row.get("enabled", True)),
            "cc_emails": list(row.get("cc_emails") or []),
            "last_run_at": row.get("last_run_at"),
            "last_result": row.get("last_result"),
            "last_error": row.get("last_error"),
        })
    return {"jobs": out}


@api_router.put("/admin/cron-settings/{job_name}")
async def admin_set_cron_enabled(
    job_name: str,
    payload: dict = Body(...),
    current_user: dict = Depends(get_current_user),
):
    """Toggle a cron's enabled flag. Admin-only.
    Body: { "enabled": true/false }
    """
    if current_user.get("role") not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")
    if job_name not in _CRON_JOB_META:
        raise HTTPException(status_code=404, detail=f"Unknown cron. Valid: {list(_CRON_JOB_META.keys())}")
    enabled = bool(payload.get("enabled"))
    await db.cron_settings.update_one(
        {"job_name": job_name},
        {
            "$set": {
                "enabled": enabled,
                "updated_at": datetime.utcnow().isoformat(),
                "updated_by": current_user.get("username") or current_user.get("id"),
            },
            "$setOnInsert": {"job_name": job_name},
        },
        upsert=True,
    )
    return {"job_name": job_name, "enabled": enabled, "status": "ok"}


_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")


@api_router.put("/admin/cron-settings/{job_name}/cc")
async def admin_set_cron_cc(
    job_name: str,
    payload: dict = Body(...),
    current_user: dict = Depends(get_current_user),
):
    """Set the CC recipient list for a cron job. Admin-only.
    Body: { "cc_emails": ["a@x.com", "b@x.com"] }
    Validates format, dedupes case-insensitively, accepts an empty list to clear.
    """
    if current_user.get("role") not in [UserRole.HR, UserRole.SYSTEM_ADMIN]:
        raise HTTPException(status_code=403, detail="Permission denied")
    if job_name not in _CRON_JOB_META:
        raise HTTPException(status_code=404, detail=f"Unknown cron. Valid: {list(_CRON_JOB_META.keys())}")

    raw = payload.get("cc_emails", [])
    if not isinstance(raw, list):
        raise HTTPException(status_code=400, detail="cc_emails must be a list")

    seen_lower: set = set()
    cleaned: list = []
    invalid: list = []
    for entry in raw:
        if entry is None:
            continue
        e = str(entry).strip()
        if not e:
            continue
        if not _EMAIL_RE.match(e):
            invalid.append(e)
            continue
        key = e.lower()
        if key in seen_lower:
            continue
        seen_lower.add(key)
        cleaned.append(e)

    if invalid:
        raise HTTPException(status_code=400, detail=f"Invalid email format: {', '.join(invalid)}")
    if len(cleaned) > 25:
        raise HTTPException(status_code=400, detail="Maximum 25 CC recipients per cron")

    await db.cron_settings.update_one(
        {"job_name": job_name},
        {
            "$set": {
                "cc_emails": cleaned,
                "cc_updated_at": datetime.utcnow().isoformat(),
                "cc_updated_by": current_user.get("username") or current_user.get("id"),
            },
            "$setOnInsert": {"job_name": job_name, "enabled": True},
        },
        upsert=True,
    )
    return {"job_name": job_name, "cc_emails": cleaned, "status": "ok"}


app.include_router(api_router)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def ensure_indexes():
    try:
        await db.attendance.create_index(
            [("employee_id", 1), ("date", 1)],
            unique=True,
            name="unique_employee_date"
        )
    except Exception:
        pass

    # Missed-punch correction audit trail — keyed lookups by request_id and
    # by employee+date for forensics. Non-unique (a single request can cause
    # multiple audit rows if re-applied via bulk-import edge cases).
    try:
        await db.attendance_corrections.create_index([("request_id", 1)])
        await db.attendance_corrections.create_index([("employee_id", 1), ("date", 1)])
        await db.attendance_corrections.create_index([("created_at", -1)])
    except Exception:
        pass
    
    # Holiday dedup: remove duplicate holidays and create unique index
    try:
        # Remove duplicates by keeping only the first document per (id, year)
        pipeline = [
            {"$group": {"_id": {"id": "$id", "year": "$year"}, "docs": {"$push": "$_id"}, "count": {"$sum": 1}}},
            {"$match": {"count": {"$gt": 1}}}
        ]
        async for group in db.holidays.aggregate(pipeline):
            # Keep first, delete rest
            ids_to_delete = group["docs"][1:]
            if ids_to_delete:
                await db.holidays.delete_many({"_id": {"$in": ids_to_delete}})
        
        await db.holidays.create_index(
            [("id", 1), ("year", 1)],
            unique=True,
            name="unique_holiday_id_year"
        )
    except Exception:
        pass
    
    # Seed default admin users if not exists + migrate old roles
    try:
        # Migrate old roles to new role system
        await db.users.update_many({"role": {"$in": ["super_admin", "admin", "hr_manager"]}}, {"$set": {"role": "hr"}})
        await db.users.update_many({"role": "team_lead"}, {"$set": {"role": "hr"}})
        
        existing_admin = await db.users.find_one({"username": "admin"})
        if not existing_admin:
            admin_user = User(
                username="admin",
                email="admin@blubridge.com",
                password_hash=hash_password("pass123"),
                name="HR Admin",
                role=UserRole.HR,
                is_first_login=False,
                onboarding_status="completed"
            )
            admin_doc = admin_user.model_dump()
            admin_doc['created_at'] = admin_doc['created_at'].isoformat()
            await db.users.insert_one(admin_doc.copy())
            print("Default HR admin user seeded successfully")
        else:
            # Update existing admin to HR role with new password
            await db.users.update_one(
                {"username": "admin"},
                {"$set": {"role": "hr", "name": "HR Admin", "password_hash": hash_password("pass123")}}
            )
            print("Admin user updated to HR role")
        
        # Create system_admin user if not exists
        existing_sysadmin = await db.users.find_one({"username": "sysadmin"})
        if not existing_sysadmin:
            sys_admin = User(
                username="sysadmin",
                email="sysadmin@blubridge.com",
                password_hash=hash_password("pass123"),
                name="System Admin",
                role=UserRole.SYSTEM_ADMIN,
                is_first_login=False,
                onboarding_status="completed"
            )
            sys_doc = sys_admin.model_dump()
            sys_doc['created_at'] = sys_doc['created_at'].isoformat()
            await db.users.insert_one(sys_doc.copy())
            print("System admin user seeded successfully")
        
        # Create office_admin user if not exists
        existing_offadmin = await db.users.find_one({"username": "offadmin"})
        if not existing_offadmin:
            off_admin = User(
                username="offadmin",
                email="offadmin@blubridge.com",
                password_hash=hash_password("pass123"),
                name="Office Admin",
                role=UserRole.OFFICE_ADMIN,
                is_first_login=False,
                onboarding_status="completed"
            )
            off_doc = off_admin.model_dump()
            off_doc['created_at'] = off_doc['created_at'].isoformat()
            await db.users.insert_one(off_doc.copy())
            print("Office admin user seeded successfully")
        
        # Create notifications index
        await db.notifications.create_index([("user_id", 1), ("read", 1)])
        await db.notifications.create_index([("created_at", -1)])
        
        # Create operational checklists index + backfill for existing employees
        await db.operational_checklists.create_index([("employee_id", 1)], unique=True)
        existing_emps = await db.employees.find({"is_deleted": {"$ne": True}}, {"_id": 0, "id": 1, "full_name": 1, "department": 1, "designation": 1}).to_list(1000)
        for emp in existing_emps:
            exists = await db.operational_checklists.find_one({"employee_id": emp["id"]})
            if not exists:
                cl = OperationalChecklist(
                    employee_id=emp["id"],
                    emp_name=emp.get("full_name", ""),
                    department=emp.get("department"),
                    designation=emp.get("designation"),
                    items=[item.copy() for item in OPERATIONAL_CHECKLIST_ITEMS]
                )
                cl_doc = cl.model_dump()
                cl_doc['created_at'] = cl_doc['created_at'].isoformat()
                cl_doc['updated_at'] = cl_doc['updated_at'].isoformat()
                try:
                    await db.operational_checklists.insert_one(cl_doc.copy())
                except Exception:
                    pass
        
        print("Admin users seeded/migrated successfully")
    except Exception as e:
        print(f"Admin seeding check: {e}")

    # HRMS automated email system — indexes + cron scheduler
    try:
        await _ensure_email_indexes(db)
        await _ensure_cron_settings_seed(db, list(_CRON_JOB_META.keys()))
        _start_email_scheduler(db)
    except Exception as e:
        print(f"Email scheduler startup failed: {e}")

    # Password reset tokens — unique on token + TTL on expires_at
    try:
        await _ensure_password_reset_token_indexes()
    except Exception as e:
        print(f"Password reset index setup failed: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
