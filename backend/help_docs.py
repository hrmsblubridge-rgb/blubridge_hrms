"""
Role-based Help Documentation generator for BluBridge HRMS.
Produces a downloadable PDF User Guide tailored to the logged-in user's role.
"""
from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle,
    ListFlowable, ListItem, KeepTogether
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Role metadata
# ---------------------------------------------------------------------------
ROLE_META = {
    "hr": {
        "title": "HR Team",
        "subtitle": "Complete User Guide for HR Administrators",
        "intro": (
            "As an HR Team member, you have the widest access in BluBridge HRMS. "
            "You can manage employees end-to-end, approve attendance & leave, run payroll, "
            "configure policies, holidays, star rewards, verification, and review audit logs."
        ),
    },
    "system_admin": {
        "title": "System Admin",
        "subtitle": "Complete User Guide for System Administrators",
        "intro": (
            "As a System Admin, you focus on platform governance: managing user roles, "
            "reviewing audit logs, and supervising attendance & leave records across the organisation. "
            "You cannot create or modify salary/payroll data."
        ),
    },
    "office_admin": {
        "title": "Office Admin",
        "subtitle": "Complete User Guide for Office Administrators",
        "intro": (
            "As an Office Admin, you handle day-to-day operational setup: the Operational "
            "Checklist, holidays, and monitoring attendance, leave, late requests, early-outs, and "
            "missed punches for the office."
        ),
    },
    "employee": {
        "title": "Employee",
        "subtitle": "Complete User Guide for Employees",
        "intro": (
            "As an Employee, this portal is your self-service hub. Mark attendance, apply for leave, "
            "raise late/early-out/missed-punch requests, view payslips, download documents, and "
            "manage your profile."
        ),
    },
}

# ---------------------------------------------------------------------------
# Content blocks — step-by-step guides
# Each module is: (title, description, [step1, step2, ...], tips[])
# ---------------------------------------------------------------------------
COMMON_LOGIN = (
    "Logging In",
    "How to sign in to BluBridge HRMS.",
    [
        "Open the BluBridge HRMS URL shared by your administrator in any modern browser.",
        "On the Login screen, enter your Username (or Email) and Password.",
        "Click the 'Sign In' button.",
        "On first login, you may be prompted to change your password — follow the on-screen steps.",
        "Once authenticated, you will be redirected to your role-specific Dashboard.",
    ],
    [
        "If you forgot your password, contact your HR / System Admin to reset it.",
        "Always sign out from the profile dropdown when using a shared device.",
    ],
)

COMMON_PROFILE = (
    "Profile & Change Password",
    "Keep your personal details and password up to date.",
    [
        "Click your avatar in the top-right corner of the page.",
        "Select 'Profile' to view and edit your name, contact details, and other personal info.",
        "Click 'Save' to persist your changes.",
        "To change your password, open the avatar menu again and choose 'Change Password'.",
        "Enter your Current Password, then New Password, confirm it, and click 'Update Password'.",
    ],
    [
        "Passwords are case-sensitive. Use a mix of letters, numbers, and symbols.",
        "After changing password, you will remain logged in — no need to sign in again.",
    ],
)

COMMON_NOTIFICATIONS = (
    "Notifications",
    "Stay informed about approvals, requests, and announcements.",
    [
        "Click the bell icon in the top header to open the Notifications panel.",
        "Unread notifications are highlighted. Click one to jump to the related record.",
        "Use 'Mark all as read' to clear the unread badge.",
    ],
    ["The bell auto-refreshes every minute — no manual reload required."],
)

# ---------- HR modules ----------
HR_MODULES = [
    COMMON_LOGIN,
    (
        "Dashboard (HR)",
        "Your command centre for org-wide HR metrics.",
        [
            "Click 'Dashboard' in the left sidebar.",
            "Review top KPI cards: Total Employees, Present Today, On Leave, Pending Approvals.",
            "Inspect charts for attendance trends and department-wise breakdowns.",
            "Click any KPI card to drill into the underlying module (e.g., Pending Approvals → Leave).",
        ],
        ["Use the date filters where available to compare periods."],
    ),
    (
        "Employees Module",
        "Create, update, deactivate, reactivate, and bulk-import employees.",
        [
            "Open 'Employees' from the sidebar to see the full list with search and filters.",
            "Use the autocomplete search at the top to quickly locate an employee by Name / ID / Email.",
            "Click '+ Add Employee' to create a new record. Fill personal, job, salary, and access details.",
            "Click an employee row to open their detailed profile tabs (Personal, Job, Salary, Documents, Education, Attendance).",
            "Click 'Edit' to update any field; click 'Save' to persist.",
            "To deactivate, click 'Deactivate' on the profile, select Inactive Type (Resigned / Relieved / Terminated), set Inactive Date, reason, and Last Day Payable date.",
            "To reactivate a previously inactive employee, open their profile and click 'Reactivate' — this restores their login access.",
            "Use 'Bulk Import' to upload a CSV/Excel. Download the template first from 'Import Template'.",
            "Use 'Delete' (soft-delete) only when absolutely necessary — prefer Deactivate.",
        ],
        [
            "Inactive employees' payroll is strictly clamped at their Inactive Date.",
            "Always upload a profile photo and government IDs for compliance.",
        ],
    ),
    (
        "Verification",
        "Verify newly onboarded employees before granting full access.",
        [
            "Open 'Verification' from the sidebar. A red badge shows pending count.",
            "Review each employee's submitted documents and self-declared info.",
            "Click 'Verify' to approve or 'Reject' with a reason if details are incorrect.",
            "Verified employees receive a verification badge on their profile.",
        ],
        ["Check the badge counter in the sidebar — it updates every 60 seconds."],
    ),
    (
        "Operational Setup (Checklist)",
        "Track office setup tasks for every new joiner.",
        [
            "Open 'Operational Setup' from the sidebar.",
            "Select an employee; the checklist (laptop, ID card, email, seating, etc.) opens.",
            "Tick each item as it is completed and add remarks.",
            "Click 'Save Progress' to persist. The sidebar badge auto-updates.",
        ],
        ["Completion is visible on the employee's Onboarding view."],
    ),
    (
        "Attendance",
        "Monitor and regularise daily attendance across the organisation.",
        [
            "Open 'Attendance'. Use date filters and employee autocomplete to narrow down.",
            "Check status codes: P (Present), A (Absent), L (Leave), H (Holiday), WO (Weekoff), LOP (Loss Of Pay).",
            "Click a cell to view check-in / check-out times and correct missed punches.",
            "Approve/Reject attendance regularisation requests submitted by employees.",
            "Export attendance to Excel via the 'Export' button when available.",
        ],
        [
            "Sundays are treated as Weekoffs; holidays follow the Holidays module.",
            "Biometric data auto-syncs; manual edits are logged in Audit Logs.",
        ],
    ),
    (
        "Leave",
        "Approve, reject, and manage employee leave.",
        [
            "Open 'Leave' from the sidebar.",
            "Filter by status (Pending / Approved / Rejected) or employee.",
            "Click a request to view dates, leave type, reason, and attached proof.",
            "Click 'Approve' or 'Reject'; add remarks before confirming.",
            "Use 'Mark as LOP' for unapproved or unpaid days.",
        ],
        ["LOP days feed directly into the payroll calculation."],
    ),
    (
        "Late Requests / Early Out / Missed Punch",
        "Handle day-to-day attendance exception requests.",
        [
            "Open the respective module from the sidebar (Late Requests / Early Out / Missed Punch).",
            "Review the reason, date, and any attached proof from the employee.",
            "Click 'Approve' or 'Reject' with remarks.",
            "Approved missed punches are auto-applied to the attendance record.",
        ],
        ["Repeated late/early-out requests from the same employee trigger a warning icon."],
    ),
    (
        "Holidays",
        "Configure the annual holiday calendar.",
        [
            "Open 'Holidays' from the sidebar.",
            "Click '+ Add Holiday', pick a date, enter the name, and choose type (National / Optional / Regional).",
            "Save. The holiday is visible to all employees instantly and influences payroll.",
            "Edit or delete an existing holiday from the list actions.",
        ],
        ["Optional holidays do not count as Weekoff Pay; only Sundays do."],
    ),
    (
        "Policies",
        "Publish company policies and acknowledge reads.",
        [
            "Open 'Policies' from the sidebar.",
            "Click '+ New Policy', enter title, effective date, category, and upload the PDF.",
            "Publish — all employees receive a notification.",
            "Track who has read each policy on the Policy detail page.",
        ],
        ["Replace an existing policy by uploading a new version; old versions stay archived."],
    ),
    (
        "Star Reward",
        "Recognise employees with star points.",
        [
            "Open 'Star Reward' from the sidebar.",
            "Select an employee, enter stars to award, and a reason.",
            "Click 'Award Stars'. The employee sees the award on their dashboard.",
            "Review the leaderboard for top performers.",
        ],
        ["Use this to reinforce values — feedback is stored with the award."],
    ),
    (
        "Team",
        "View teams and reporting hierarchy.",
        [
            "Open 'Team' from the sidebar.",
            "Browse departments and drill into team members.",
            "Reassign reporting managers by editing the employee's Job tab.",
        ],
        ["Team hierarchy drives leave approval routing."],
    ),
    (
        "Payroll",
        "Run payroll, configure salary, and generate payslips.",
        [
            "Open 'Payroll' from the sidebar. Select the month and year.",
            "Switch between 'Payroll View' (earnings & deductions) and 'Attendance View' (compliance sheet).",
            "In Attendance View, review Total Days, Working Days, Weekoff Pay (Sundays), Extra Pay, LOP, and Final Payable.",
            "Use the employee autocomplete to filter specific records.",
            "Click 'Regenerate' if you update attendance mid-month — the engine recalculates in bulk.",
            "Click 'Download Payslip' to get a PDF for any employee.",
            "Use 'Export' to download the payroll sheet as Excel.",
        ],
        [
            "The payroll engine stops calculations at the Inactive Date for relieved employees.",
            "Weekoffs are Sundays only; worked Holidays add to Extra Pay.",
        ],
    ),
    (
        "Issue Tickets",
        "Manage support tickets raised by employees.",
        [
            "Open 'Issue Tickets' from the sidebar.",
            "Filter by status (Open / In Progress / Resolved / Closed).",
            "Click a ticket to see description, attachments, and the conversation thread.",
            "Reply, change priority, assign to a team, or mark resolved.",
        ],
        ["Always update status and add closing notes when resolved."],
    ),
    (
        "Reports",
        "Download attendance, leave, payroll, and custom HR reports.",
        [
            "Open 'Reports' from the sidebar.",
            "Select the report type (Attendance / Leave / Payroll / Headcount).",
            "Pick the date range and filters.",
            "Click 'Generate' and then 'Download' to export Excel/PDF.",
        ],
        ["Generated reports reflect data at the moment of generation."],
    ),
    (
        "Role Management",
        "Assign and change user roles.",
        [
            "Open 'Role Management' from the sidebar.",
            "Search the user, click the role dropdown, and choose a new role (HR / System Admin / Office Admin / Employee).",
            "Click 'Update'. The change is audit-logged instantly.",
        ],
        ["Role changes take effect on the user's next login."],
    ),
    (
        "Audit Logs",
        "Review every security-sensitive action across the system.",
        [
            "Open 'Audit Logs' from the sidebar.",
            "Filter by user, action type, module, and date range.",
            "Click a row to view the before/after payload.",
        ],
        ["Audit logs are immutable — used for compliance and security reviews."],
    ),
    COMMON_PROFILE,
    COMMON_NOTIFICATIONS,
]

# ---------- System Admin modules ----------
SYSADMIN_MODULES = [
    COMMON_LOGIN,
    (
        "Dashboard (System Admin)",
        "High-level view of users, roles, and platform health.",
        [
            "Click 'Dashboard' in the sidebar to see KPI cards.",
            "Focus on user counts per role and recent activity.",
        ],
        [],
    ),
    (
        "Employees",
        "Search and review employee master data (read/edit access varies by setup).",
        [
            "Open 'Employees' from the sidebar.",
            "Use autocomplete search by Name / ID / Email.",
            "Click a row to open the employee profile.",
        ],
        ["Payroll and salary are managed by HR, not System Admin."],
    ),
    (
        "Attendance, Leave, Late Requests, Early Out, Missed Punch",
        "Supervise operational attendance records across the org.",
        [
            "Open each module from the sidebar.",
            "Review status, approve/reject pending requests if your RBAC permits.",
            "Use filters to narrow by department, date, or employee.",
        ],
        ["All approvals here are captured in the Audit Logs."],
    ),
    (
        "Role Management",
        "Core responsibility: assign user roles.",
        [
            "Open 'Role Management' from the sidebar.",
            "Search the user, open the role dropdown.",
            "Select HR / System Admin / Office Admin / Employee and click 'Update'.",
            "Confirm the change; it is logged automatically.",
        ],
        [
            "Promoting someone to System Admin grants platform-wide audit access — do this cautiously.",
            "Role changes log out the affected user only on their next request.",
        ],
    ),
    (
        "Audit Logs",
        "Investigate user actions for compliance or incident response.",
        [
            "Open 'Audit Logs' from the sidebar.",
            "Filter by user, module, action (Create / Update / Delete / Login), or date range.",
            "Click a row to expand the full JSON payload and metadata.",
            "Export a filtered view if required for an investigation.",
        ],
        ["Audit log entries cannot be edited or deleted — by design."],
    ),
    COMMON_PROFILE,
    COMMON_NOTIFICATIONS,
]

# ---------- Office Admin modules ----------
OFFICEADMIN_MODULES = [
    COMMON_LOGIN,
    (
        "Dashboard (Office Admin)",
        "Office-level operational snapshot.",
        [
            "Click 'Dashboard' to see today's Present / On Leave / Pending Checklist counts.",
        ],
        [],
    ),
    (
        "Employees",
        "Look up employees for operational actions.",
        [
            "Open 'Employees' from the sidebar.",
            "Use autocomplete to find an employee quickly.",
            "Click to view profile, then jump into Operational Checklist or Attendance for them.",
        ],
        [],
    ),
    (
        "Operational Setup (Checklist)",
        "The main office-admin module — onboarding task tracking.",
        [
            "Open 'Operational Setup' from the sidebar (sidebar badge shows pending items).",
            "Select a new joiner from the list.",
            "Tick each onboarding task (Laptop, Email, ID Card, Seating, ID docs, etc.).",
            "Add remarks / serial numbers where relevant.",
            "Click 'Save Progress'. The employee's onboarding status updates live.",
        ],
        ["Complete checklists before the employee's first payroll cycle to avoid delays."],
    ),
    (
        "Attendance, Leave, Late Requests, Early Out, Missed Punch",
        "Monitor day-to-day attendance operations.",
        [
            "Open each module from the sidebar.",
            "Filter by date or department.",
            "Review / approve / reject requests based on your approval rights.",
        ],
        [],
    ),
    (
        "Holidays",
        "Maintain the office holiday calendar.",
        [
            "Open 'Holidays'. Click '+ Add Holiday'.",
            "Set Date, Name, Type (National / Optional / Regional). Save.",
            "Edit or delete via row actions.",
        ],
        ["Office holidays reflect instantly for every employee."],
    ),
    COMMON_PROFILE,
    COMMON_NOTIFICATIONS,
]

# ---------- Employee modules ----------
EMPLOYEE_MODULES = [
    COMMON_LOGIN,
    (
        "Employee Dashboard",
        "Your personal HR snapshot.",
        [
            "After login you land on the Dashboard.",
            "See today's attendance status, upcoming holidays, leave balance, and latest announcements.",
            "Use quick-action tiles to jump into Attendance, Leave, or Salary.",
        ],
        [],
    ),
    (
        "My Attendance",
        "Mark attendance and review your history.",
        [
            "Open 'My Attendance' from the sidebar.",
            "Click 'Check In' when you start work; 'Check Out' when you end.",
            "Review the calendar — P, A, L, H, WO are highlighted.",
            "Click any date to see check-in/out times and working hours.",
        ],
        [
            "Biometric punches sync automatically; you do not need to re-punch in the app.",
            "If a punch is missing, raise a 'Missed Punch' request.",
        ],
    ),
    (
        "Leave",
        "Apply for and track your leaves.",
        [
            "Open 'Leave' from the sidebar and click 'Apply Leave'.",
            "Pick Leave Type (Casual / Sick / Earned / LOP), From-To dates, and enter a reason.",
            "Attach supporting proof if required (e.g., medical certificate) and submit.",
            "Track status in the 'My Leaves' list (Pending / Approved / Rejected).",
            "Click a leave to cancel it while still Pending.",
        ],
        ["Plan leaves in advance; emergency leaves should be followed up with a mail/call to your manager."],
    ),
    (
        "Late Request",
        "Request permission when you arrive late.",
        [
            "Open 'Late Request' from the sidebar.",
            "Click 'New Request'. Enter Date, expected In-time, and Reason.",
            "Submit. Track Pending / Approved / Rejected status in the list.",
        ],
        ["Frequent late requests trigger an HR review."],
    ),
    (
        "Early Out",
        "Request early departure from work.",
        [
            "Open 'Early Out' from the sidebar.",
            "Click 'New Request'. Enter Date, expected Out-time, and Reason. Submit.",
            "Track status in the list.",
        ],
        [],
    ),
    (
        "Missed Punch",
        "Correct a missing check-in or check-out.",
        [
            "Open 'Missed Punch' from the sidebar.",
            "Click 'New Request'. Select Date, indicate missed In or Out, enter actual time and reason.",
            "Submit. Once approved, your attendance will be automatically corrected.",
        ],
        ["Attach a screenshot or proof if possible."],
    ),
    (
        "My Salary",
        "View and download your payslips.",
        [
            "Open 'My Salary' from the sidebar.",
            "Pick the month — you see earnings, deductions, and net payable.",
            "Click 'Download Payslip' to get a PDF for any approved month.",
        ],
        [
            "Raise an Issue Ticket if you notice a discrepancy.",
            "Payslips are available once HR finalises payroll for the month.",
        ],
    ),
    (
        "Holidays",
        "See the official holiday list.",
        [
            "Open 'Holidays' from the sidebar.",
            "Review the calendar or list view. Filter by month or type.",
        ],
        [],
    ),
    (
        "Education & Experience",
        "Maintain your academic and work history.",
        [
            "Open 'Education & Experience' from the sidebar.",
            "Click '+ Add' under Education — enter Degree, Institution, Year, Grade.",
            "Click '+ Add' under Experience — enter Company, Role, From-To, and responsibilities.",
            "Upload certificates where applicable.",
        ],
        ["Keep this up to date — HR uses it for appraisals and compliance."],
    ),
    (
        "Policies",
        "Read and acknowledge company policies.",
        [
            "Open 'Policies' from the sidebar.",
            "Click a policy to read or download the PDF.",
            "Click 'Mark as Read' to record your acknowledgement.",
        ],
        ["HR can audit who has acknowledged each policy."],
    ),
    (
        "My Documents",
        "Access employer-issued documents.",
        [
            "Open 'My Documents' from the sidebar.",
            "Browse by category (Offer Letter / Experience / Address Proof / Other).",
            "Click 'Download' to save a PDF locally.",
        ],
        [
            "If you cannot open a PDF in the browser, use the Download option.",
            "Missing a document? Raise an Issue Ticket tagged to HR.",
        ],
    ),
    (
        "Support Tickets",
        "Raise and track issues with HR or IT.",
        [
            "Open 'Support Tickets' from the sidebar.",
            "Click '+ New Ticket'. Choose Category (HR / IT / Payroll / Facilities).",
            "Enter subject, detailed description, and attach files if needed. Submit.",
            "Track replies in the ticket thread; add comments as the conversation progresses.",
        ],
        ["Keep one topic per ticket for clarity."],
    ),
    (
        "Profile",
        "Update your personal info and profile photo.",
        [
            "Open 'Profile' from the sidebar or profile avatar.",
            "Edit contact number, address, emergency contact, bank details.",
            "Upload/replace your profile photo. Save.",
        ],
        ["Some fields (name, DOB, joining date) can only be changed by HR."],
    ),
    COMMON_PROFILE,
    COMMON_NOTIFICATIONS,
]

ROLE_CONTENT = {
    "hr": HR_MODULES,
    "system_admin": SYSADMIN_MODULES,
    "office_admin": OFFICEADMIN_MODULES,
    "employee": EMPLOYEE_MODULES,
}


# ---------------------------------------------------------------------------
# PDF builder
# ---------------------------------------------------------------------------
def _build_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="HGTitle", fontName="Helvetica-Bold", fontSize=28,
        leading=34, alignment=TA_LEFT, textColor=colors.HexColor("#063c88"),
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="HGSubtitle", fontName="Helvetica", fontSize=13,
        leading=18, alignment=TA_LEFT, textColor=colors.HexColor("#475569"),
        spaceAfter=16,
    ))
    styles.add(ParagraphStyle(
        name="HGSection", fontName="Helvetica-Bold", fontSize=16,
        leading=22, alignment=TA_LEFT, textColor=colors.HexColor("#063c88"),
        spaceBefore=14, spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="HGDesc", fontName="Helvetica-Oblique", fontSize=10.5,
        leading=15, alignment=TA_JUSTIFY, textColor=colors.HexColor("#334155"),
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="HGStep", fontName="Helvetica", fontSize=10.5,
        leading=15, alignment=TA_JUSTIFY, textColor=colors.HexColor("#1e293b"),
    ))
    styles.add(ParagraphStyle(
        name="HGTip", fontName="Helvetica-Oblique", fontSize=9.5,
        leading=13, alignment=TA_LEFT, textColor=colors.HexColor("#0a5cba"),
    ))
    styles.add(ParagraphStyle(
        name="HGLabel", fontName="Helvetica-Bold", fontSize=10,
        leading=13, alignment=TA_LEFT, textColor=colors.HexColor("#334155"),
        spaceBefore=8, spaceAfter=2,
    ))
    styles.add(ParagraphStyle(
        name="HGIntro", fontName="Helvetica", fontSize=11,
        leading=16, alignment=TA_JUSTIFY, textColor=colors.HexColor("#1e293b"),
        spaceAfter=12,
    ))
    styles.add(ParagraphStyle(
        name="HGTOC", fontName="Helvetica", fontSize=11,
        leading=17, alignment=TA_LEFT, textColor=colors.HexColor("#1e293b"),
    ))
    return styles


def _footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#64748b"))
    canvas.drawString(15 * mm, 10 * mm, "BluBridge HRMS User Guide")
    canvas.drawRightString(
        A4[0] - 15 * mm, 10 * mm, f"Page {doc.page}"
    )
    canvas.restoreState()


def generate_help_pdf(role: str, user_name: str | None = None) -> bytes:
    """Return the PDF bytes for the given role."""
    role = (role or "employee").lower()
    if role not in ROLE_CONTENT:
        role = "employee"

    meta = ROLE_META[role]
    modules = ROLE_CONTENT[role]

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title=f"BluBridge HRMS — {meta['title']} User Guide",
        author="BluBridge HRMS",
    )
    styles = _build_styles()
    story = []

    # --- Cover ---
    story.append(Spacer(1, 40 * mm))
    story.append(Paragraph("BluBridge HRMS", styles["HGTitle"]))
    story.append(Paragraph(meta["subtitle"], styles["HGSubtitle"]))
    story.append(Spacer(1, 4 * mm))
    info_rows = [
        ["Role", meta["title"]],
        ["Generated On", datetime.utcnow().strftime("%d %b %Y")],
    ]
    if user_name:
        info_rows.insert(0, ["Prepared For", user_name])
    info_tbl = Table(info_rows, colWidths=[45 * mm, 100 * mm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10.5),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#475569")),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor("#0f172a")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 10 * mm))
    story.append(Paragraph(meta["intro"], styles["HGIntro"]))
    story.append(PageBreak())

    # --- Table of Contents ---
    story.append(Paragraph("Table of Contents", styles["HGSection"]))
    story.append(Spacer(1, 4))
    for i, (title, *_rest) in enumerate(modules, start=1):
        story.append(Paragraph(f"{i}. {title}", styles["HGTOC"]))
    story.append(PageBreak())

    # --- Module sections ---
    for i, (title, desc, steps, tips) in enumerate(modules, start=1):
        block = [
            Paragraph(f"{i}. {title}", styles["HGSection"]),
            Paragraph(desc, styles["HGDesc"]),
            Paragraph("Step-by-step:", styles["HGLabel"]),
            ListFlowable(
                [ListItem(Paragraph(s, styles["HGStep"]), leftIndent=6)
                 for s in steps],
                bulletType="1", start="1", leftIndent=14,
                bulletFormat="%s.", bulletFontSize=10.5,
            ),
        ]
        if tips:
            block.append(Paragraph("Tips:", styles["HGLabel"]))
            block.append(ListFlowable(
                [ListItem(Paragraph(t, styles["HGTip"]), leftIndent=6)
                 for t in tips],
                bulletType="bullet", leftIndent=14,
            ))
        story.append(KeepTogether(block))
        story.append(Spacer(1, 6))

    # --- Closing ---
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "If any step is unclear or you need further help, raise a Support Ticket "
        "from the sidebar or contact your HR team directly.",
        styles["HGIntro"],
    ))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------
def generate_help_xlsx(role: str, user_name: str | None = None) -> bytes:
    """Return an XLSX workbook (bytes) for the given role."""
    role = (role or "employee").lower()
    if role not in ROLE_CONTENT:
        role = "employee"

    meta = ROLE_META[role]
    modules = ROLE_CONTENT[role]

    wb = openpyxl.Workbook()

    # --- Colours & styles ---
    navy = "FF063C88"
    navy_soft = "FFE7EEF7"
    grey = "FF475569"
    slate = "FF1E293B"
    thin = Side(border_style="thin", color="FFCBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(name="Calibri", size=18, bold=True, color="FFFFFFFF")
    title_fill = PatternFill("solid", fgColor=navy)
    h2_font = Font(name="Calibri", size=13, bold=True, color="FFFFFFFF")
    h2_fill = PatternFill("solid", fgColor=navy)
    label_font = Font(name="Calibri", size=11, bold=True, color=slate)
    label_fill = PatternFill("solid", fgColor=navy_soft)
    body_font = Font(name="Calibri", size=11, color=slate)
    muted_font = Font(name="Calibri", size=10, italic=True, color=grey)
    tip_font = Font(name="Calibri", size=10, italic=True, color=navy)
    step_num_font = Font(name="Calibri", size=11, bold=True, color=navy)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Overview sheet ---
    ws = wb.active
    ws.title = "Overview"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = f"BluBridge HRMS — {meta['title']} User Guide"
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 34

    rows = [
        ("Role", meta["title"]),
        ("Generated On", datetime.utcnow().strftime("%d %b %Y")),
    ]
    if user_name:
        rows.insert(0, ("Prepared For", user_name))
    rows.append(("Overview", meta["intro"]))

    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=1).fill = label_fill
        ws.cell(row=r, column=1).alignment = wrap_top
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=value).font = body_font
        ws.cell(row=r, column=2).alignment = wrap_top
        ws.cell(row=r, column=2).border = border
        if label == "Overview":
            ws.row_dimensions[r].height = 80
        else:
            ws.row_dimensions[r].height = 22
        r += 1

    # Table of Contents
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    toc_title = ws.cell(row=r, column=1, value="Table of Contents")
    toc_title.font = h2_font
    toc_title.fill = h2_fill
    toc_title.alignment = center
    ws.row_dimensions[r].height = 26
    r += 1
    for i, (title, *_rest) in enumerate(modules, start=1):
        ws.cell(row=r, column=1, value=f"{i}.").font = step_num_font
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=title).font = body_font
        r += 1

    # --- One sheet per module ---
    used_titles = set()
    for idx, (title, desc, steps, tips) in enumerate(modules, start=1):
        # Excel tab names: max 31 chars, no : \ / ? * [ ]
        clean = "".join(ch for ch in title if ch not in ":\\/?*[]")
        base = f"{idx:02d}. {clean}"[:31]
        sheet_name = base
        n = 2
        while sheet_name in used_titles:
            sheet_name = f"{base[:28]}_{n}"
            n += 1
        used_titles.add(sheet_name)
        s = wb.create_sheet(sheet_name)
        s.column_dimensions["A"].width = 6
        s.column_dimensions["B"].width = 95

        # Title
        s.merge_cells("A1:B1")
        t = s["A1"]
        t.value = f"{idx}. {title}"
        t.font = title_font
        t.fill = title_fill
        t.alignment = center
        s.row_dimensions[1].height = 30

        # Description
        s.merge_cells("A3:B3")
        d = s["A3"]
        d.value = desc
        d.font = muted_font
        d.alignment = wrap_top
        s.row_dimensions[3].height = 32

        # Steps header
        s.merge_cells("A5:B5")
        sh = s["A5"]
        sh.value = "Step-by-step"
        sh.font = label_font
        sh.fill = label_fill
        sh.alignment = Alignment(vertical="center")
        s.row_dimensions[5].height = 22

        row = 6
        for i, step in enumerate(steps, start=1):
            sc = s.cell(row=row, column=1, value=i)
            sc.font = step_num_font
            sc.alignment = Alignment(horizontal="center", vertical="top")
            sc.border = border
            bc = s.cell(row=row, column=2, value=step)
            bc.font = body_font
            bc.alignment = wrap_top
            bc.border = border
            # rough height by length
            s.row_dimensions[row].height = max(22, 16 * (1 + len(step) // 90))
            row += 1

        if tips:
            row += 1
            s.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            th = s.cell(row=row, column=1, value="Tips")
            th.font = label_font
            th.fill = label_fill
            th.alignment = Alignment(vertical="center")
            s.row_dimensions[row].height = 22
            row += 1
            for tip in tips:
                s.cell(row=row, column=1, value="•").alignment = Alignment(horizontal="center")
                s.cell(row=row, column=1).font = tip_font
                tc = s.cell(row=row, column=2, value=tip)
                tc.font = tip_font
                tc.alignment = wrap_top
                s.row_dimensions[row].height = max(20, 16 * (1 + len(tip) // 90))
                row += 1

        # Freeze the title row
        s.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

