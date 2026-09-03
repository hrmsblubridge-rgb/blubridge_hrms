"""BRSF Excel / CSV export + import.

The exported XLSX mirrors the HR reference sheet: two header rows with merged
"Positive Stars" (E:J) and "Negative Stars" (K:R) bands, one employee per row
and live Excel formulas for the three total columns. Import reuses the existing
BRSF manual-entry / manual-override layers and validation — system values,
weekly children and per-instance records are never touched or fabricated.
"""
import csv
import io
import uuid
from datetime import datetime, timezone
from typing import Optional

from fastapi import Body, Depends, File as FastAPIFile, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from server import api_router, db, get_current_user, get_ist_now
from brsf_stars import (
    _require_star_admin,
    _save_line,
    _utc_now_iso,
    eligible_employees,
)
from brsf_validation import (
    as_star_int,
    validate_monthly_entry,
    validate_override,
)

MONTH_NAMES = ["January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December"]

# visible column -> (brsf code, xlsx header, csv header, static excel range or None)
SHEET_COLUMNS = [
    ("E", "P01", "Full Present 26/26 (+2 star per month)", "Full Attendance", (0, 2)),
    ("F", "P03", "Innovation Per Month (Max +3 stars)", "Innovation", (0, 3)),
    ("G", "P02", "Performance per month (Max +5 stars)", "Performance", (0, 5)),
    ("H", "P04", "Learning per month (Max +5 stars)", "Learning", (0, 5)),
    ("I", "P06", "Extra Effort Sunday / Holiday (+1 Per instance)", "Extra Effort", None),
    ("J", "P05", "Attendance Per Month (10+ hours average research engagement) (Max +5)",
     "Research Attendance", (0, 5)),
    ("K", "N01", "Invalid Leave Request Late Notification (-1 star Per In)", "Invalid Leave", None),
    ("L", "N03", "Frequent Emergencies (More than 2 Per month -3)", "Frequent Emergencies", (-3, 0)),
    ("M", "N04", "Short Research Duration (Average <9.5 hrs -1 star Per Week)", "Short Research", None),
    ("N", "N05", "No Proof / Verification of claimed Leave (-3 star Per Sequence)", "No Proof", None),
    ("O", "N02", "Emergency Leave Violation without proof (-2 Stars Per Instance)",
     "Emergency Violation", None),
    ("P", "N06", "Frequent Absences More than 4 Leave/Absence (-3 Monthly)", "Frequent Absences", (-3, 0)),
    ("Q", "N07", "No Show / Unreachable (-3 Per Instance)", "No Show", None),
    ("R", "N08", "Behavior / UnSafe Conduct (-4 star) per instance", "Unsafe Conduct", None),
]
CODE_BY_COLUMN = {c: code for c, code, _, _, _ in SHEET_COLUMNS}
MONTHLY_MANUAL_CODES = {"P02", "P03", "P04"}   # imported value = manual monthly entry
FIRST_DATA_ROW = 3
META_SHEET = "_BRSF_Metadata"

THIN = Side(style="thin", color="FF999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="FFF2F2F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFEDF3FF")


def _month_label(month: str) -> str:
    return f"{MONTH_NAMES[int(month[5:7]) - 1]} {month[:4]}"


def _fmt_date(iso) -> str:
    if not iso or len(str(iso)) < 10:
        return ""
    y, m, d = str(iso)[:10].split("-")
    return f"{d}-{m}-{y}"


async def _month_rows(month: str) -> list:
    """Eligible employees for the month + their currently applied final stars."""
    year, mon = int(month[:4]), int(month[5:7])
    emps = await eligible_employees(month)
    finals, has_lines = {}, set()
    async for l in db.brsf_star_lines.find(
        {"year": year, "month": mon, "employee_id": {"$in": [e["id"] for e in emps]}},
        {"_id": 0, "employee_id": 1, "code": 1, "final_value": 1},
    ):
        finals.setdefault(l["employee_id"], {})[l["code"]] = l.get("final_value") or 0
        has_lines.add(l["employee_id"])
    rows = []
    for e in emps:
        vals = finals.get(e["id"], {})
        rows.append({
            "employee": e,
            "values": {code: int(round(vals.get(code, 0) or 0)) for _, code, _, _, _ in SHEET_COLUMNS},
            "calculated": e["id"] in has_lines,
        })
    return rows


def _build_workbook(month: str, rows: list, batch_id: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"BRSF {_month_label(month)}"[:31]

    # ---- header rows
    ws["E1"] = "Positive Stars"
    ws.merge_cells("E1:J1")
    ws["K1"] = "Negative Stars"
    ws.merge_cells("K1:R1")
    for col, label in (("A", "Employee Name"), ("B", "Team Name"), ("C", "Date Of Joining"),
                       ("D", "Date Of Employee Confirmation"), ("S", "Total Positive Stars"),
                       ("T", "Total Negative Stars"), ("U", "Total Stars")):
        ws[f"{col}1"] = label
        ws.merge_cells(f"{col}1:{col}2")
    for col, _code, header, _csv, _rng in SHEET_COLUMNS:
        ws[f"{col}2"] = header

    for row in (1, 2):
        for col in range(1, 22):
            c = ws.cell(row=row, column=col)
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
            c.fill = HEAD_FILL
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 78
    widths = {"A": 20, "B": 20, "C": 14, "D": 18, "S": 12, "T": 12, "U": 11}
    for col in range(1, 22):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(letter, 16)
    ws.freeze_panes = "A3"

    # ---- data rows
    for i, r in enumerate(rows):
        excel_row = FIRST_DATA_ROW + i
        emp = r["employee"]
        ws.cell(row=excel_row, column=1, value=emp.get("full_name"))
        ws.cell(row=excel_row, column=2, value=emp.get("team") or "")
        ws.cell(row=excel_row, column=3, value=_fmt_date(emp.get("date_of_joining")))
        ws.cell(row=excel_row, column=4, value=_fmt_date(emp.get("confirmation_date")))
        for col, code, _h, _c, _rng in SHEET_COLUMNS:
            cell = ws[f"{col}{excel_row}"]
            cell.value = r["values"].get(code, 0)
            cell.alignment = Alignment(horizontal="center")
            cell.protection = Protection(locked=False)
        ws[f"S{excel_row}"] = f"=SUM(E{excel_row}:J{excel_row})"
        ws[f"T{excel_row}"] = f"=SUM(K{excel_row}:R{excel_row})"
        ws[f"U{excel_row}"] = f"=S{excel_row}+T{excel_row}"
        for col in range(1, 22):
            c = ws.cell(row=excel_row, column=col)
            c.border = BORDER
            if col <= 4:
                c.alignment = Alignment(vertical="center", wrap_text=True)
            if col >= 19:
                c.fill = TOTAL_FILL
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[excel_row].height = 30

    last_row = FIRST_DATA_ROW + max(len(rows), 1) - 1

    # ---- excel-side convenience validation (backend validation is authoritative)
    for col, _code, _h, _c, rng in SHEET_COLUMNS:
        if not rng:
            continue
        lo, hi = rng
        dv = DataValidation(type="whole", operator="between", formula1=str(lo), formula2=str(hi),
                            allow_blank=True, showErrorMessage=True,
                            errorTitle="Invalid star value",
                            error=f"Allowed whole numbers: {lo} to {hi}.")
        ws.add_data_validation(dv)
        dv.add(f"{col}{FIRST_DATA_ROW}:{col}{last_row}")

    # totals keep their formulas: sheet is protected, only E:R is unlocked
    ws.protection.sheet = True
    ws.protection.enable()

    # ---- hidden metadata (employee identity + month) — never edited by HR
    meta = wb.create_sheet(META_SHEET)
    meta.append(["Sheet Row", "Employee ID", "Employee Email", "BRSF Month", "Export Batch ID"])
    for i, r in enumerate(rows):
        meta.append([FIRST_DATA_ROW + i, r["employee"]["id"],
                     r["employee"].get("email") or "", month, batch_id])
    meta.sheet_state = "hidden"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@api_router.get("/brsf/export")
async def brsf_export(month: str, format: str = "xlsx",
                      current_user: dict = Depends(get_current_user)):
    """Screenshot-style XLSX (merged bands + formulas) or the flat CSV variant."""
    _require_star_admin(current_user)
    if not month or len(month) < 7:
        raise HTTPException(status_code=400, detail="A month (YYYY-MM) is required for export")
    rows = await _month_rows(month)
    stem = f"BRSF_Star_Reward_{_month_label(month).replace(' ', '_')}"

    if format == "csv":
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["Employee Name", "Team Name", "Date Of Joining", "Date Of Employee Confirmation",
                    *[c for _col, _code, _h, c, _r in SHEET_COLUMNS],
                    "Total Positive Stars", "Total Negative Stars", "Total Stars"])
        for r in rows:
            vals = [r["values"].get(code, 0) for _col, code, _h, _c, _r in SHEET_COLUMNS]
            pos = sum(v for v in vals[:6])
            neg = sum(v for v in vals[6:])
            e = r["employee"]
            w.writerow([e.get("full_name"), e.get("team") or "", _fmt_date(e.get("date_of_joining")),
                        _fmt_date(e.get("confirmation_date")), *vals, pos, neg, pos + neg])
        data = io.BytesIO(out.getvalue().encode("utf-8-sig"))
        return StreamingResponse(data, media_type="text/csv", headers={
            "Content-Disposition": f'attachment; filename="{stem}.csv"'})

    buf = _build_workbook(month, rows, str(uuid.uuid4()))
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'})


# ------------------------------------------------------------------ import
def _parse_cell(raw) -> Optional[int]:
    """Blank -> None (no change). Anything else must be a whole number."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    return as_star_int(raw, "Star value")


def _read_xlsx(content: bytes) -> dict:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    meta_month, meta_by_row = None, {}
    if META_SHEET in wb.sheetnames:
        for r in wb[META_SHEET].iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            meta_by_row[int(r[0])] = {"employee_id": r[1], "email": r[2]}
            meta_month = meta_month or (str(r[3]) if r[3] else None)
    rows = []
    for excel_row in range(FIRST_DATA_ROW, ws.max_row + 1):
        name = ws[f"A{excel_row}"].value
        cells = {col: ws[f"{col}{excel_row}"].value for col, *_ in SHEET_COLUMNS}
        if not name and all(v is None for v in cells.values()):
            continue
        rows.append({"row": excel_row, "name": str(name or "").strip(),
                     "employee_id": (meta_by_row.get(excel_row) or {}).get("employee_id"),
                     "cells": cells})
    return {"file_month": meta_month, "rows": rows}


def _read_csv(content: bytes) -> dict:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not all_rows:
        return {"file_month": None, "rows": []}
    rows = []
    for i, r in enumerate(all_rows[1:], start=2):
        padded = r + [""] * (21 - len(r))
        cells = {col: padded[4 + idx] for idx, (col, *_) in enumerate(SHEET_COLUMNS)}
        rows.append({"row": i, "name": (padded[0] or "").strip(),
                     "employee_id": None, "cells": cells})
    return {"file_month": None, "rows": rows}


@api_router.post("/brsf/import/preview")
async def brsf_import_preview(month: str = Form(...), file: UploadFile = FastAPIFile(...),
                              current_user: dict = Depends(get_current_user)):
    """Parse + validate the uploaded sheet. Nothing is written to the star lines."""
    _require_star_admin(current_user)
    content = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        parsed, source = _read_xlsx(content), "Excel Import"
    elif name.endswith(".csv"):
        parsed, source = _read_csv(content), "CSV Import"
    else:
        raise HTTPException(status_code=400, detail="Upload an .xlsx or .csv file exported from this module")

    if parsed["file_month"] and parsed["file_month"] != month:
        raise HTTPException(
            status_code=400,
            detail=f"This file was exported for {_month_label(parsed['file_month'])}. "
                   f"Select {_month_label(parsed['file_month'])} before importing.")

    year, mon = int(month[:4]), int(month[5:7])
    emps = await eligible_employees(month)
    by_id = {e["id"]: e for e in emps}
    by_name = {(e.get("full_name") or "").strip().lower(): e for e in emps}

    changes, errors, skipped = [], [], []
    no_change = 0
    matched_ids = set()
    for r in parsed["rows"]:
        emp = by_id.get(r["employee_id"]) or by_name.get(r["name"].lower())
        if not emp:
            skipped.append({"row": r["row"], "employee": r["name"] or "(blank)",
                            "reason": "Not an eligible employee for this month — row skipped"})
            continue
        matched_ids.add(emp["id"])
        lines = {l["code"]: l async for l in db.brsf_star_lines.find(
            {"employee_id": emp["id"], "year": year, "month": mon}, {"_id": 0})}
        if not lines:
            skipped.append({"row": r["row"], "employee": emp["full_name"],
                            "reason": "No BRSF records for this month yet — run Auto Calculate first"})
            continue
        for col, code, _h, _c, _rng in SHEET_COLUMNS:
            line = lines.get(code)
            if not line:
                continue
            try:
                value = _parse_cell(r["cells"].get(col))
            except HTTPException as exc:
                errors.append({"row": r["row"], "employee": emp["full_name"], "code": code,
                               "criteria": line["name"], "existing": line.get("final_value"),
                               "imported": r["cells"].get(col), "status": "Invalid",
                               "message": exc.detail})
                continue
            if value is None:
                continue  # blank cell = do not touch this criterion
            existing = int(round(line.get("final_value") or 0))
            if value == existing:
                no_change += 1
                continue
            try:
                if code in MONTHLY_MANUAL_CODES:
                    validate_monthly_entry(line, value)
                else:
                    validate_override(line, value)
            except HTTPException as exc:
                errors.append({"row": r["row"], "employee": emp["full_name"], "code": code,
                               "criteria": line["name"], "existing": existing, "imported": value,
                               "status": "Invalid", "message": exc.detail})
                continue
            changes.append({"row": r["row"], "employee_id": emp["id"],
                            "employee": emp["full_name"], "code": code, "criteria": line["name"],
                            "line_id": line["id"], "existing": existing, "imported": value,
                            "system_value": line.get("system_value"),
                            "target": "manual" if code in MONTHLY_MANUAL_CODES else "override",
                            "status": "Valid"})

    batch = {
        "id": str(uuid.uuid4()), "month": month, "source": source,
        "filename": file.filename, "changes": changes,
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("full_name") or current_user.get("username"),
        "created_at": _utc_now_iso(), "applied": False,
    }
    doc = dict(batch)
    doc["created_dt"] = datetime.now(timezone.utc)   # TTL housekeeping
    await db.brsf_import_batches.insert_one(doc)
    batch.pop("_id", None)
    return {
        "batch_id": batch["id"], "month": month, "source": source,
        "summary": {
            "employees_in_file": len(parsed["rows"]),
            "valid_employees": len(matched_ids),
            "changed_values": len(changes),
            "no_change": no_change,
            "errors": len(errors),
            "skipped": len(skipped),
        },
        "changes": changes, "errors": errors, "skipped": skipped,
    }


@api_router.post("/brsf/import/confirm")
async def brsf_import_confirm(payload: dict = Body(...),
                              current_user: dict = Depends(get_current_user)):
    """Apply a previewed batch: manual criteria -> manual value, automated -> override."""
    _require_star_admin(current_user)
    batch = await db.brsf_import_batches.find_one({"id": payload.get("batch_id")}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Import preview expired — upload the file again")
    if batch.get("applied"):
        raise HTTPException(status_code=400, detail="This import batch has already been applied")

    applied, failed = 0, []
    for ch in batch["changes"]:
        line = await db.brsf_star_lines.find_one({"id": ch["line_id"]}, {"_id": 0})
        if not line:
            failed.append({**ch, "message": "Star line no longer exists"})
            continue
        prev_final = line.get("final_value")
        try:
            if ch["target"] == "manual":
                line["manual_value"] = validate_monthly_entry(line, ch["imported"])
                line["entry_mode"] = "monthly"
            else:
                line["override_value"] = validate_override(line, ch["imported"])
                line["override_reason"] = f"{batch['source']} — {batch.get('filename') or 'sheet'}"
        except HTTPException as exc:
            failed.append({**ch, "message": exc.detail})
            continue
        line["changed_by"] = current_user.get("full_name") or current_user.get("username")
        line["changed_at"] = _utc_now_iso()
        saved = await _save_line(line)
        await db.brsf_star_audit.insert_one({
            "id": str(uuid.uuid4()),
            "employee_id": line["employee_id"], "year": line["year"], "month": line["month"],
            "code": line["code"], "criteria": line["name"], "action": batch["source"],
            "previous_value": prev_final, "new_value": saved["final_value"],
            "imported_value": ch["imported"],
            "system_calculated_value": line.get("system_value"),
            "reason": f"{batch['source']} from {batch.get('filename') or 'sheet'}",
            "source": batch["source"], "import_batch_id": batch["id"],
            "updated_by": current_user.get("id"),
            "updated_by_name": current_user.get("full_name") or current_user.get("username"),
            "updated_at": get_ist_now().isoformat(),
        })
        applied += 1

    await db.brsf_import_batches.update_one(
        {"id": batch["id"]},
        {"$set": {"applied": True, "applied_at": _utc_now_iso(), "applied_count": applied,
                  "failed": failed}})
    return {"success": True, "applied": applied, "failed": failed,
            "message": f"{applied} star value(s) imported for {_month_label(batch['month'])}. "
                       "System values and child records untouched."}
