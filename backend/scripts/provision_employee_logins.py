"""
Provision login credentials for every active employee and export them to Excel.

Policy (confirmed by user):
  * Scope: ALL active employees (create-or-reset for role=employee users only).
  * Username = email prefix (left of @).
  * Password = {first 4 letters of full_name, lowercase, no spaces} + '@' +
               {last 4 digits of phone}
               Fallbacks: if name <4 chars → pad with 'x'; if phone missing
               or <4 digits → use last 4 chars of emp_id (or uuid).
  * is_first_login = True so user is forced to change on first login.
  * Admin accounts (role != employee) are LEFT UNTOUCHED.
  * No emails are sent.

Output: /app/backend/scripts/Employee_Credentials_YYYY-MM-DD.xlsx
"""

from __future__ import annotations

import asyncio
import hashlib
import os
import re
import sys
import uuid
from datetime import datetime, timezone, timedelta
from pathlib import Path

from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---- Load env from backend/.env ---------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
ENV_PATH = ROOT / ".env"
if ENV_PATH.exists():
    for ln in ENV_PATH.read_text().splitlines():
        ln = ln.strip()
        if not ln or ln.startswith("#") or "=" not in ln:
            continue
        k, v = ln.split("=", 1)
        v = v.strip().strip('"').strip("'")
        os.environ.setdefault(k.strip(), v)

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]

IST = timezone(timedelta(hours=5, minutes=30))


def hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()


def sanitize_name_part(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z]", "", name or "").lower()
    if len(cleaned) >= 4:
        return cleaned[:4]
    return (cleaned + "xxxx")[:4]


def sanitize_phone_part(phone, fallback: str) -> str:
    digits = re.sub(r"\D", "", str(phone or ""))
    if len(digits) >= 4:
        return digits[-4:]
    # fallback: last 4 chars of emp_id/uuid, digits-only
    digs = re.sub(r"\D", "", fallback or "")
    if len(digs) >= 4:
        return digs[-4:]
    return (digs + "0000")[:4]


def build_credentials(emp: dict) -> tuple[str, str]:
    email = (emp.get("official_email") or emp.get("email") or "").strip()
    username = email.split("@")[0] if email else re.sub(r"\W", "", emp.get("full_name", ""))[:16].lower()
    name_part = sanitize_name_part(emp.get("full_name", ""))
    phone_part = sanitize_phone_part(emp.get("phone"), emp.get("emp_id") or emp.get("id"))
    password = f"{name_part}@{phone_part}"
    return username, password


async def main():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    emps = await db.employees.find(
        {"is_deleted": {"$ne": True}, "employee_status": "Active"},
        {"_id": 0},
    ).to_list(2000)
    emps.sort(key=lambda e: e.get("emp_id", ""))
    print(f"Active employees found: {len(emps)}")

    rows = []
    created = reset = skipped_admin = skipped_no_email = 0
    admin_usernames_preview = []

    for emp in emps:
        email = (emp.get("official_email") or emp.get("email") or "").strip()
        if not email:
            skipped_no_email += 1
            rows.append({
                "Emp ID": emp.get("emp_id") or "-",
                "Full Name": emp.get("full_name") or "-",
                "Email": "-",
                "Username": "-",
                "Password": "<skipped: no email>",
                "Role": "-",
                "Department": emp.get("department") or "-",
                "Team": emp.get("team") or "-",
                "Status": "SKIPPED (no email)",
            })
            continue

        username, password = build_credentials(emp)
        existing = await db.users.find_one({"username": username}, {"_id": 0})

        # Respect admin accounts — never touch them
        if existing and existing.get("role") in ("hr", "system_admin", "office_admin"):
            admin_usernames_preview.append(f"{username}({existing.get('role')})")
            skipped_admin += 1
            rows.append({
                "Emp ID": emp.get("emp_id") or "-",
                "Full Name": emp.get("full_name") or "-",
                "Email": email,
                "Username": username,
                "Password": "<unchanged — admin account>",
                "Role": existing.get("role"),
                "Department": emp.get("department") or "-",
                "Team": emp.get("team") or "-",
                "Status": "SKIPPED (admin)",
            })
            continue

        now_ist = datetime.now(IST).isoformat()

        if existing:
            await db.users.update_one(
                {"username": username},
                {"$set": {
                    "password_hash": hash_password(password),
                    "is_first_login": True,
                    "is_active": True,
                    "role": "employee",
                    "employee_id": emp.get("id"),
                    "department": emp.get("department"),
                    "team": emp.get("team"),
                    "email": email,
                    "name": emp.get("full_name", ""),
                    "updated_at": now_ist,
                }},
            )
            reset += 1
            status = "RESET"
        else:
            await db.users.insert_one({
                "id": str(uuid.uuid4()),
                "username": username,
                "email": email,
                "password_hash": hash_password(password),
                "name": emp.get("full_name", ""),
                "role": "employee",
                "employee_id": emp.get("id"),
                "department": emp.get("department"),
                "team": emp.get("team"),
                "onboarding_status": emp.get("onboarding_status", "pending"),
                "is_first_login": True,
                "is_active": True,
                "created_at": now_ist,
            })
            created += 1
            status = "CREATED"

        rows.append({
            "Emp ID": emp.get("emp_id") or "-",
            "Full Name": emp.get("full_name") or "-",
            "Email": email,
            "Username": username,
            "Password": password,
            "Role": "employee",
            "Department": emp.get("department") or "-",
            "Team": emp.get("team") or "-",
            "Status": status,
        })

    # ---- Build Excel --------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Credentials"

    headers = [
        "Emp ID", "Full Name", "Email", "Username", "Password",
        "Role", "Department", "Team", "Status",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="063C88")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    widths = {
        "Emp ID": 10, "Full Name": 28, "Email": 38, "Username": 24,
        "Password": 18, "Role": 12, "Department": 20, "Team": 28, "Status": 20,
    }
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = widths.get(h, 16)

    mono = Font(name="Consolas")
    for r in rows:
        ws.append([r[h] for h in headers])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=headers.index("Username") + 1).font = mono
        ws.cell(row=row_idx, column=headers.index("Password") + 1).font = mono

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary.append(["Generated at (IST)", datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")])
    summary.append(["Total active employees", len(emps)])
    summary.append(["Accounts CREATED", created])
    summary.append(["Accounts RESET", reset])
    summary.append(["Admin accounts SKIPPED", skipped_admin])
    summary.append(["Rows with no email (skipped)", skipped_no_email])
    summary.append([])
    summary.append(["Password pattern", "{first4letters_of_name}@{last4digits_of_phone}"])
    summary.append(["Force change on first login", "Yes (is_first_login=true)"])
    summary.append(["Emails sent", "No"])
    for col in ("A", "B"):
        summary.column_dimensions[col].width = 40
    for r in (1, 2, 3, 4, 5, 6):
        summary.cell(row=r, column=1).font = Font(bold=True)

    stamp = datetime.now(IST).strftime("%Y-%m-%d")
    out_path = ROOT / "scripts" / f"Employee_Credentials_{stamp}.xlsx"
    wb.save(out_path)

    print("\n=== Summary ===")
    print(f"  Created : {created}")
    print(f"  Reset   : {reset}")
    print(f"  Admin skipped : {skipped_admin}  ({admin_usernames_preview[:5]}{'...' if len(admin_usernames_preview) > 5 else ''})")
    print(f"  No-email skipped : {skipped_no_email}")
    print(f"\nFile: {out_path}")
    client.close()


if __name__ == "__main__":
    asyncio.run(main())
