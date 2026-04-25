"""
One-time migration: replace existing employee data with master list from Excel.

USER-CONFIRMED PLAN:
1) Hard-delete all employees + all linked transactional records
2) Keep system admin/HR/office_admin login users (admin / sysadmin / offadmin)
3) Delete employee login user accounts
4) Replace `departments` and `teams` with new master lists
5) Insert 90 - 6 duplicates - 1 NULL row = 83 employees from Excel
6) Normalize typo: 'Administation' -> 'Administration'
7) Auto-generate custom_employee_id (EMP100+) for rows missing Employee ID
8) Skip row where Department=NULL or Team=NULL
9) Dedupe by email — keep first occurrence

Run:
    cd /app/backend && python3 scripts/seed_employees_from_excel.py
"""
import asyncio
import os
import sys
import uuid
from datetime import datetime, timezone, timedelta

import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient

# Make the parent backend dir importable for shared constants if needed
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

EXCEL_PATH = os.environ.get("MASTER_EXCEL_PATH", "/tmp/employees_master.xlsx")
IST = timezone(timedelta(hours=5, minutes=30))

# Logins to PRESERVE (system / admin accounts, not employee logins)
PRESERVED_USERNAMES = {"admin", "sysadmin", "offadmin"}

# Master department -> teams mapping derived from the Excel
DEPARTMENTS = ["Business & Product", "Research Unit", "Support Staff", "System Engineer"]

TEAM_DEPARTMENT = {
    "Accounts": "Support Staff",
    "Administration": "Support Staff",
    "Business Analyst": "Business & Product",
    "Checkpointing": "Research Unit",
    "Compiler - Auto Differentiation": "Research Unit",
    "Data": "Research Unit",
    "Framework - Graph & Auto-differentiation": "Research Unit",
    "Framework - Parallelism": "Research Unit",
    "Framework - Quantz": "Research Unit",
    "Framework - Tensor & Ops": "Research Unit",
    "Hardware & Systems": "System Engineer",
    "Hr Intern": "Support Staff",
    "Marketing & Growth": "Support Staff",
    "Process & Operation": "Support Staff",
    "Product Team": "Business & Product",
    "Quantization": "Research Unit",
    "Tensor & Ops": "Research Unit",
    "Tokenizer": "Research Unit",
    "Unassigned": "Research Unit",
    "Vigilance": "Support Staff",
}

# Collections to wipe entirely (transactional / per-employee data)
WIPE_COLLECTIONS = [
    "employees",
    "attendance",
    "leaves",
    "late_requests",
    "early_out_requests",
    "missed_punches",
    "notifications",
    "payroll",                 # if exists
    "biometric_punch_logs",
    "audit_logs",
    "onboarding",
    "onboarding_documents",
    "salary_structures",
    "operational_checklists",
]


def fmt_date(value) -> str | None:
    """Excel cells can be datetime, date, or string. Normalize to YYYY-MM-DD."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s.upper() == "NULL":
        return None
    return s


def fmt_text(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.upper() == "NULL":
        return None
    return s


def fmt_phone(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None


def fmt_biometric(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None


def normalize_team(team_raw: str | None) -> str | None:
    """Fix the 'Administation' -> 'Administration' typo."""
    if not team_raw:
        return None
    t = team_raw.strip()
    if t.lower() == "administation":
        return "Administration"
    return t


def load_excel_rows(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return rows


def build_employee_doc(row: dict, custom_id_counter: list[int]) -> dict | None:
    """Convert one Excel row into an Employee document. Returns None if row should be skipped."""
    full_name = fmt_text(row.get("Employee Name"))
    email = fmt_text(row.get("Email"))
    department = fmt_text(row.get("Department"))
    team = normalize_team(fmt_text(row.get("Team")))

    # Skip rows that have no name or no email (cannot be a valid employee)
    if not full_name or not email:
        return None
    # Skip per spec: row where Department or Team is NULL
    if not department or not team:
        return None

    custom_id_raw = fmt_text(row.get("Employee ID"))
    if not custom_id_raw:
        # Auto-generate per user choice (EMP100, EMP101, ...)
        custom_id_raw = f"EMP{custom_id_counter[0]}"
        custom_id_counter[0] += 1

    # Map a few values to existing system enums where possible
    employment_type = fmt_text(row.get("Employment Type")) or "Full-time"
    tier_level = fmt_text(row.get("Tier Level")) or "Mid"
    work_location = fmt_text(row.get("Work Location")) or "Office"
    shift_type = fmt_text(row.get("Shift Type")) or "General"
    user_role = fmt_text(row.get("User Role")) or "employee"
    gender = fmt_text(row.get("Gender"))

    monthly_salary = row.get("Monthly Salary")
    try:
        monthly_salary = float(monthly_salary) if monthly_salary not in (None, "") else 0.0
    except (TypeError, ValueError):
        monthly_salary = 0.0

    now = datetime.now(IST)

    return {
        "id": str(uuid.uuid4()),
        # Will fill emp_id after sequence is decided
        "emp_id": None,
        "full_name": full_name,
        "official_email": email,
        "phone_number": fmt_phone(row.get("Phone")),
        "gender": gender,
        "date_of_birth": fmt_date(row.get("Date of Birth")),
        "custom_employee_id": custom_id_raw,
        "date_of_joining": fmt_date(row.get("Date of Joining")) or now.strftime("%Y-%m-%d"),
        "employment_type": employment_type,
        "employee_status": "Active",
        "designation": fmt_text(row.get("Designation")) or "Employee",
        "tier_level": tier_level,
        "reporting_manager_id": None,
        "department": department,
        "team": team,
        "work_location": work_location,
        "leave_policy": "Standard",
        "shift_type": shift_type,
        "attendance_tracking_enabled": True,
        "custom_login_time": None,
        "custom_logout_time": None,
        "custom_total_hours": None,
        "monthly_salary": monthly_salary,
        "user_role": user_role,
        "login_enabled": True,
        "biometric_id": fmt_biometric(row.get("Biometric ID")),
        "avatar": None,
        "stars": 0,
        "unsafe_count": 0,
        "onboarding_status": "pending",
        "onboarding_completed_at": None,
        "is_deleted": False,
        "deleted_at": None,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
    }


async def main():
    mongo_url = os.environ["MONGO_URL"]
    db_name = os.environ["DB_NAME"]
    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]

    print(f"Connected to {db_name}\n")

    # ---------- 1. WIPE TRANSACTIONAL COLLECTIONS ----------
    existing = set(await db.list_collection_names())
    for coll in WIPE_COLLECTIONS:
        if coll in existing:
            r = await db[coll].delete_many({})
            print(f"  wiped {coll}: deleted {r.deleted_count}")
        else:
            print(f"  skipped {coll} (does not exist)")

    # ---------- 2. WIPE EMPLOYEE LOGIN USERS (keep admin/sysadmin/offadmin) ----------
    r = await db.users.delete_many({"username": {"$nin": list(PRESERVED_USERNAMES)}})
    print(f"  deleted employee user accounts: {r.deleted_count}")
    remaining = await db.users.count_documents({})
    print(f"  remaining users (admin/sysadmin/offadmin): {remaining}")

    # ---------- 3. REPLACE DEPARTMENTS ----------
    await db.departments.delete_many({})
    now = datetime.now(IST).isoformat()
    dept_docs = [
        {"id": str(uuid.uuid4()), "name": d, "created_at": now, "updated_at": now}
        for d in DEPARTMENTS
    ]
    await db.departments.insert_many([d.copy() for d in dept_docs])
    print(f"  inserted {len(dept_docs)} departments: {[d['name'] for d in dept_docs]}")

    # ---------- 4. REPLACE TEAMS ----------
    await db.teams.delete_many({})
    team_docs = [
        {"id": str(uuid.uuid4()), "name": team, "department": dept, "created_at": now, "updated_at": now}
        for team, dept in TEAM_DEPARTMENT.items()
    ]
    await db.teams.insert_many([t.copy() for t in team_docs])
    print(f"  inserted {len(team_docs)} teams")

    # ---------- 5. LOAD + DEDUPE EXCEL ROWS ----------
    rows = load_excel_rows(EXCEL_PATH)
    print(f"\nLoaded {len(rows)} rows from Excel: {EXCEL_PATH}")

    # Auto-generated custom_employee_id counter (EMP100+) for rows missing Employee ID
    custom_id_counter = [100]

    seen_emails: set[str] = set()
    docs: list[dict] = []
    skipped = []
    for idx, row in enumerate(rows, start=1):
        doc = build_employee_doc(row, custom_id_counter)
        if doc is None:
            skipped.append((idx, row.get("Employee Name"), "missing data / NULL dept|team"))
            continue
        email_lower = doc["official_email"].strip().lower()
        if email_lower in seen_emails:
            skipped.append((idx, doc["full_name"], f"duplicate email {email_lower}"))
            continue
        seen_emails.add(email_lower)
        docs.append(doc)

    # Assign sequential emp_id (system-generated) like EMP0001
    for i, doc in enumerate(docs, start=1):
        doc["emp_id"] = f"EMP{str(i).zfill(4)}"

    # ---------- 6. INSERT EMPLOYEES ----------
    if docs:
        await db.employees.insert_many([d.copy() for d in docs])
    print(f"\nInserted {len(docs)} employees")
    print(f"Skipped {len(skipped)} rows:")
    for idx, name, reason in skipped:
        print(f"  - row {idx} ({name}): {reason}")

    # Summary
    print("\n=== Summary ===")
    print(f"Employees: {await db.employees.count_documents({})}")
    print(f"Departments: {await db.departments.count_documents({})}")
    print(f"Teams: {await db.teams.count_documents({})}")
    print(f"Users (logins kept): {await db.users.count_documents({})}")

    client.close()


if __name__ == "__main__":
    asyncio.run(main())
