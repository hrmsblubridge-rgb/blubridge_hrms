"""
Vigilance service layer — pure(ish) helpers operating on the injected Motor `db`.

Sections:
  * Formatting / validation (dates, clock times, durations)
  * Data sources (date-aware active employees, attendance snapshot) — READ ONLY
  * Template generation (streaming xlsx)
  * Upload parsing (dynamic break detection + strict validation)
  * Repository (upsert/CRUD on `vigilance_entries`)
  * Aggregation (vigilance own-view, admin merged-view, attendance integration)

Storage model — collection `vigilance_entries`, one doc per
(target_employee_id + date + uploaded_by_employee_id):
  {
    id, target_employee_id, target_employee_name, target_email,
    target_team, target_department, target_designation,
    date (ISO yyyy-mm-dd),
    uploaded_by_employee_id, uploaded_by_name,
    system_login, system_logout,            # editable clock times "HH:MM AM/PM"
    total_research_hours, total_break_hours, # editable durations "HH:MM"
    breaks: [ {label, from, to, total} ],    # DYNAMIC, ordered, unlimited
    created_at, updated_at
  }
"""
import io
import re
import uuid
from datetime import datetime, date, timezone, timedelta, time as dtime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

FIXED_COLS = [
    "Name", "Email-id", "Team", "Date", "Punch-In", "Punch-Out", "Total Hours",
    "System Login", "System Logout", "Total Research Hours", "Total Break Hours",
]
N_FIXED = len(FIXED_COLS)                       # 11
ATT_SYSTEM_COLS = ["Name", "Email-id", "Team", "Date", "Punch-In", "Punch-Out", "Total Hours"]  # non-editable, attendance-derived
DEFAULT_BREAK_LABELS = ["Morning Break", "Lunch Break", "Evening Break", "Extra-Break1", "Extra-Break2"]
BREAK_SUBS = ["From", "To", "Total"]

VIGILANCE_DESIGNATION = "vigilance"

CLOCK_RE = re.compile(r"^(\d{1,2}):([0-5]\d)\s*(AM|PM)$", re.IGNORECASE)
TWENTYFOUR_RE = re.compile(r"^([01]?\d|2[0-3]):([0-5]\d)$")
DUR_RE = re.compile(r"^(\d{1,3}):([0-5]\d)(?::([0-5]\d))?$")


# ----------------------------------------------------------------------------
# Formatting / validation
# ----------------------------------------------------------------------------
def to_iso(value):
    """Normalise any supported date representation to ISO 'YYYY-MM-DD' or None.

    Accepts: ISO, DD-MM-YYYY (attendance store), DD-MMM-YYYY (display), datetime/date.
    """
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", s)            # DD-MM-YYYY
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    try:
        return datetime.strptime(s, "%d-%b-%Y").strftime("%Y-%m-%d")  # DD-MMM-YYYY
    except ValueError:
        return None


def iso_to_display(iso):
    """ISO 'YYYY-MM-DD' -> 'DD-MMM-YYYY' (universal HRMS format)."""
    if not iso:
        return ""
    try:
        d = datetime.strptime(str(iso)[:10], "%Y-%m-%d")
        return f"{d.day:02d}-{MONTHS[d.month - 1]}-{d.year}"
    except ValueError:
        return ""


def parse_display_date_strict(value):
    """Strict 'DD-MMM-YYYY' -> ISO. Returns None on any deviation (for validation)."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    try:
        return datetime.strptime(str(value).strip(), "%d-%b-%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def _hm_to_12h(h, mm):
    ampm = "AM" if h < 12 else "PM"
    h12 = h % 12 or 12
    return f"{h12:02d}:{mm} {ampm}"


def norm_clock(value):
    """Validate a clock time entered as EITHER 24-hour (13:45) OR 12-hour
    (01:45 PM) and normalise to canonical 12-hour 'HH:MM AM/PM' for storage.
    Returns (ok, normalised|''). Blank allowed.
    """
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return True, ""
    if isinstance(value, dtime):
        return True, _hm_to_12h(value.hour, f"{value.minute:02d}")
    if isinstance(value, datetime):
        return True, _hm_to_12h(value.hour, f"{value.minute:02d}")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel stores a clock time as a fraction of a day (0.0 == midnight).
        # 00:00 read as the number 0 must be accepted, not rejected.
        total = round((float(value) % 1) * 24 * 3600)
        h = (total // 3600) % 24
        mm = (total % 3600) // 60
        return True, _hm_to_12h(h, f"{mm:02d}")
    s = re.sub(r"\s+", " ", str(value).strip())
    m = CLOCK_RE.match(s)               # 12-hour with AM/PM
    if m:
        h = int(m.group(1))
        if h < 1 or h > 12:
            return False, None
        return True, f"{h:02d}:{m.group(2)} {m.group(3).upper()}"
    m = TWENTYFOUR_RE.match(s)          # 24-hour
    if m:
        return True, _hm_to_12h(int(m.group(1)), m.group(2))
    return False, None


def to_24h(value):
    """Convert a stored 12-hour (or already-24h) clock string to 24-hour 'HH:MM'.
    Blank stays blank. Used for the Vigilance-user download/export format.
    """
    if value is None or str(value).strip() == "":
        return ""
    if isinstance(value, dtime):
        return f"{value.hour:02d}:{value.minute:02d}"
    s = re.sub(r"\s+", " ", str(value).strip())
    m = CLOCK_RE.match(s)
    if m:
        h = int(m.group(1)) % 12
        if m.group(3).upper() == "PM":
            h += 12
        return f"{h:02d}:{m.group(2)}"
    m = TWENTYFOUR_RE.match(s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return str(value).strip()


def _clock_to_minutes(value):
    """Normalised clock ('HH:MM AM/PM' or 24h 'HH:MM') -> minutes of day, or None."""
    h24 = to_24h(value)
    m = TWENTYFOUR_RE.match(h24) if h24 else None
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


def compute_break_total(from_clock, to_clock):
    """Auto-calculate a break Total = (To − From). Overnight spans are allowed
    (a To earlier than From wraps past midnight, e.g. 23:50 → 00:10 = 00:20).
    Returns canonical 'HH:MM:SS' (seconds always 00 → renders as 'HH:MM'), or ''
    when either endpoint is missing/blank. Any user-supplied Total is IGNORED by
    callers — the Total column is always derived from From & To.
    """
    fm = _clock_to_minutes(from_clock)
    tm = _clock_to_minutes(to_clock)
    if fm is None or tm is None:
        return ""
    diff = (tm - fm) % (24 * 60)
    h, mm = divmod(diff, 60)
    return f"{h:02d}:{mm:02d}:00"


def norm_duration(value):
    """Validate/normalise a DURATION (no AM/PM) entered as EITHER 'HH:MM' OR
    'HH:MM:SS' (and Excel time/number cells) and store canonically as 'HH:MM:SS'.
    Blank allowed. ZERO durations ('00:00' / '00:00:00' / numeric 0) are VALID
    business values (no break / no research) and are NEVER rejected. Only a
    genuinely malformed value (e.g. 25:99, AA:BB, 12:70, 10::30, 10-30) fails.
    """
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return True, ""
    if isinstance(value, dtime):
        return True, f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}"
    if isinstance(value, datetime):
        return True, f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel time/duration serial (1.0 == 24h). 0 -> 00:00:00 (a VALID value).
        total = round(float(value) * 24 * 3600)
        if total < 0:
            return False, None
        h, rem = divmod(total, 3600)
        mm, ss = divmod(rem, 60)
        return True, f"{h:02d}:{mm:02d}:{ss:02d}"
    s = str(value).strip()
    m = DUR_RE.match(s)
    if not m:
        return False, None
    h, mm = int(m.group(1)), int(m.group(2))
    ss = int(m.group(3)) if m.group(3) is not None else 0
    return True, f"{h:02d}:{mm:02d}:{ss:02d}"


def display_duration(value):
    """Render a stored duration for UI / export: 'HH:MM' when the seconds part
    is :00 (or absent for legacy 'HH:MM' data), 'HH:MM:SS' when seconds are
    present and non-zero. No truncation; backward-compatible with legacy data.
    """
    if value is None or str(value).strip() == "":
        return ""
    s = str(value).strip()
    m = DUR_RE.match(s)
    if not m:
        return s
    h, mm = int(m.group(1)), int(m.group(2))
    ss = int(m.group(3)) if m.group(3) is not None else 0
    return f"{h:02d}:{mm:02d}:{ss:02d}" if ss else f"{h:02d}:{mm:02d}"


def _display_breaks(breaks):
    """Return breaks with their duration 'total' rendered for UI/export."""
    return [{**b, "total": display_duration(b.get("total", ""))} for b in (breaks or [])]


def att_total_hours_to_duration(att):
    """Convert attendance total_hours ('11h 36m' / decimal) to 'HH:MM' duration."""
    th = att.get("total_hours")
    if th:
        m = re.match(r"(\d+)\s*h\s*(\d+)\s*m", str(th))
        if m:
            return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
    d = att.get("total_hours_decimal")
    if d is not None:
        try:
            h = int(d)
            mn = round((float(d) - h) * 60)
            if mn == 60:
                h, mn = h + 1, 0
            return f"{h:02d}:{mn:02d}"
        except (ValueError, TypeError):
            return ""
    return ""


# ----------------------------------------------------------------------------
# Data sources (READ ONLY — never mutate attendance / employees)
# ----------------------------------------------------------------------------
async def get_active_employees(db, iso_from, iso_to, *, employee_ids=None):
    """Return employees that were active at any point within [iso_from, iso_to].

    Date-aware: an employee is active on day D when joining<=D and
    (no inactive_date OR D<=inactive_date). ISO date strings compare lexically.
    """
    query = {}
    if employee_ids is not None:
        query["id"] = {"$in": employee_ids}
    proj = {
        "_id": 0, "id": 1, "full_name": 1, "official_email": 1, "team": 1,
        "department": 1, "designation": 1, "date_of_joining": 1, "inactive_date": 1,
        "employee_status": 1,
    }
    emps = await db.employees.find(query, proj).to_list(100000)
    out = []
    for e in emps:
        doj = to_iso(e.get("date_of_joining"))
        if not doj or doj > iso_to:
            continue
        inactive = to_iso(e.get("inactive_date"))
        if inactive and inactive < iso_from:
            continue
        e["_doj_iso"] = doj
        e["_inactive_iso"] = inactive
        out.append(e)
    out.sort(key=lambda x: (x.get("full_name") or "").lower())
    return out


def is_active_on(emp, iso_day):
    if emp["_doj_iso"] > iso_day:
        return False
    if emp["_inactive_iso"] and iso_day > emp["_inactive_iso"]:
        return False
    return True


async def get_attendance_map(db, employee_ids, iso_from, iso_to):
    """Map (employee_id, iso_date) -> {punch_in, punch_out, total_hours} from attendance.

    Attendance `date` is stored as 'DD-MM-YYYY'; we normalise to ISO for keys.
    """
    cur = db.attendance.find(
        {"employee_id": {"$in": employee_ids}},
        {"_id": 0, "employee_id": 1, "date": 1, "check_in": 1, "check_out": 1,
         "total_hours": 1, "total_hours_decimal": 1},
    )
    out = {}
    async for a in cur:
        iso = to_iso(a.get("date"))
        if iso and iso_from <= iso <= iso_to:
            out[(a["employee_id"], iso)] = {
                "punch_in": a.get("check_in") or "",
                "punch_out": a.get("check_out") or "",
                "total_hours": att_total_hours_to_duration(a),
            }
    return out


def daterange_iso(iso_from, iso_to):
    d0 = datetime.strptime(iso_from, "%Y-%m-%d").date()
    d1 = datetime.strptime(iso_to, "%Y-%m-%d").date()
    days = []
    cur = d0
    while cur <= d1:
        days.append(cur.strftime("%Y-%m-%d"))
        cur = date.fromordinal(cur.toordinal() + 1)
    return days


# ----------------------------------------------------------------------------
# Template / export generation (premium merged 2-row headers)
# ----------------------------------------------------------------------------
_HDR_FONT = Font(bold=True, size=11, color="1F2937")
_SYS_FILL = PatternFill("solid", fgColor="DCE6F1")        # attendance/system (do not edit)
_EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")       # editable scalar fields
_BREAK_FILL = PatternFill("solid", fgColor="E2EFDA")      # break group columns
_SUB_FILL = PatternFill("solid", fgColor="F2F2F2")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="B7C2D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_hdr(cell, fill):
    cell.font = _HDR_FONT
    cell.fill = fill
    cell.alignment = _CENTER
    cell.border = _BORDER


def _write_grouped_header(ws, specs):
    """Write a premium 2-row header from an ordered list of column specs.

    specs: list of (kind, label, fill) where kind is:
      * 'scalar' -> ONE column whose label is merged VERTICALLY across both header
                    rows (single clean block).
      * 'group'  -> THREE columns (From/To/Total); the parent `label` is merged
                    HORIZONTALLY across all three (one clean block), with the
                    sub-labels in the 2nd header row.
    Styles are applied BEFORE merging (MergedCell objects are read-only).
    Returns the total column count.
    """
    col = 1
    for kind, label, fill in specs:
        if kind == "scalar":
            c1 = ws.cell(row=1, column=col, value=label)
            c2 = ws.cell(row=2, column=col)
            _style_hdr(c1, fill)
            _style_hdr(c2, fill)
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            col += 1
        else:  # group
            for j, sub in enumerate(BREAK_SUBS):
                pc = ws.cell(row=1, column=col + j, value=label if j == 0 else None)
                _style_hdr(pc, fill)
                sc = ws.cell(row=2, column=col + j, value=sub)
                _style_hdr(sc, _SUB_FILL)
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(BREAK_SUBS) - 1)
            col += len(BREAK_SUBS)
    return col - 1


def _apply_widths(ws, total_cols, n_scalar):
    widths = ([22, 28, 16, 14, 13, 13, 13, 14, 14, 18, 16][:n_scalar]
              + [13] * max(0, n_scalar - 11)
              + [12] * (total_cols - n_scalar))
    for i, w in enumerate(widths[:total_cols], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18


def build_template_workbook(prefill_rows, break_labels=None):
    """Return BytesIO of a styled .xlsx with a premium merged 2-row header.

    `prefill_rows`: list of dicts with keys name,email,team,date_display,
    punch_in,punch_out,total_hours (system columns prefilled).
    """
    break_labels = break_labels or list(DEFAULT_BREAK_LABELS)
    wb = Workbook()
    ws = wb.active
    ws.title = "Vigilance"

    specs = [("scalar", label, _SYS_FILL if label in ATT_SYSTEM_COLS else _EDIT_FILL)
             for label in FIXED_COLS]
    specs += [("group", bl, _BREAK_FILL) for bl in break_labels]
    total_cols = _write_grouped_header(ws, specs)
    _apply_widths(ws, total_cols, N_FIXED)
    ws.freeze_panes = "E3"   # lock 2 header rows + Name/Email/Team/Date columns

    n_break_cells = len(break_labels) * len(BREAK_SUBS)
    for r in prefill_rows:
        ws.append([
            r.get("name", ""), r.get("email", ""), r.get("team", ""),
            r.get("date_display", ""), r.get("punch_in", ""), r.get("punch_out", ""),
            r.get("total_hours", ""),
            "", "", "", "",                      # System Login/Logout, Research, Break Hours (editable)
        ] + [""] * n_break_cells)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ----------------------------------------------------------------------------
# Upload parsing (dynamic break detection + strict validation)
# ----------------------------------------------------------------------------
def _norm_header(v):
    return re.sub(r"\s+", " ", str(v).strip()).lower() if v is not None else ""


def _nkey(v):
    """Aggressive normalisation for fuzzy header matching: lowercase alphanumerics only.
    'Email-id' / 'Email id' / 'E-mail' all collapse to comparable keys."""
    return re.sub(r"[^a-z0-9]", "", str(v).strip().lower()) if v is not None else ""


# Header aliases for locating editable scalar + identity columns by NAME (not by
# position) so the upload survives reordering / slight renaming of columns.
_EMAIL_KEYS = {"emailid", "email", "emailaddress", "officialemail", "mail", "employeeemail", "empemail"}
_DATE_KEYS = {"date", "day", "workdate", "attendancedate"}
_NAME_KEYS = {"name", "employeename", "empname", "fullname", "employee"}
_SYS_LOGIN_KEYS = {"systemlogin", "syslogin", "login"}
_SYS_LOGOUT_KEYS = {"systemlogout", "syslogout", "logout"}
_RESEARCH_KEYS = {"totalresearchhours", "researchhours", "research", "totalresearch", "researchhrs"}
_BREAKHRS_KEYS = {"totalbreakhours", "breakhours", "totalbreak", "breakhrs"}
# Reserved scalar/system header keys — excluded from break-group detection.
_RESERVED_KEYS = (_EMAIL_KEYS | _DATE_KEYS | _NAME_KEYS | _SYS_LOGIN_KEYS | _SYS_LOGOUT_KEYS
                  | _RESEARCH_KEYS | _BREAKHRS_KEYS
                  | {"team", "punchin", "punchout", "totalhours"})
_SUB_KEYS = {"from": "From", "to": "To", "total": "Total"}


def parse_upload(file_bytes, employees_by_email, employees_by_name, uploaded_by):
    """Parse an uploaded .xlsx (resilient, name-based). Returns (entries, errors).

    `employees_by_email`: {lower_email: employee_dict}
    `employees_by_name` : {lower_full_name: employee_dict}
    `uploaded_by`       : {'employee_id','name'}
    All-or-nothing: caller rejects the whole upload if errors is non-empty.

    IMMUTABLE HRMS REFERENCE COLUMNS (Name, Email-id, Team, Date, Punch-In,
    Punch-Out, Total Hours) are NEVER validated for format and their uploaded
    values are NEVER written to the DB — they are reconstructed from the Employee
    Master + Attendance (the single source of truth). Renaming, reordering,
    blanking, hiding or tampering with these columns must NOT fail the upload.
    Email-id (or, as a fallback, Name) + Date are used ONLY to MAP each row to an
    employee/day. Only vigilance-EDITABLE fields are validated and imported.
    """
    errors = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return [], [{"row": 0, "message": "File is not a valid .xlsx workbook."}]
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], [{"row": 0, "message": "Template is empty or missing header rows."}]
    header1 = list(rows[0])
    header2 = list(rows[1])
    ncols = max(len(header1), len(header2))

    # Locate identity + editable scalar columns by fuzzy NAME (first match wins).
    def find_col(keyset):
        for idx in range(len(header1)):
            if _nkey(header1[idx]) in keyset:
                return idx
        return None

    email_idx = find_col(_EMAIL_KEYS)
    name_idx = find_col(_NAME_KEYS)
    date_idx = find_col(_DATE_KEYS)
    login_idx = find_col(_SYS_LOGIN_KEYS)
    logout_idx = find_col(_SYS_LOGOUT_KEYS)
    research_idx = find_col(_RESEARCH_KEYS)
    breakhrs_idx = find_col(_BREAKHRS_KEYS)

    # Detect break groups by NAME: a header-1 label that is NOT a reserved
    # scalar/system column and whose columns carry From/To/Total sub-headers.
    groups = []
    current = None
    for i in range(ncols):
        label = header1[i] if i < len(header1) else None
        sub = _nkey(header2[i]) if i < len(header2) else ""
        lkey = _nkey(label)
        if label not in (None, "") and lkey not in _RESERVED_KEYS and sub in _SUB_KEYS:
            lbl = str(label).strip()
            if current is None or current["label"] != lbl or current[_SUB_KEYS[sub]] is not None:
                current = {"label": lbl, "From": None, "To": None, "Total": None}
                groups.append(current)
            current[_SUB_KEYS[sub]] = i
        elif current is not None and lkey == "" and sub in _SUB_KEYS:
            current[_SUB_KEYS[sub]] = i
        elif label not in (None, "") and (lkey in _RESERVED_KEYS or sub not in _SUB_KEYS):
            current = None
    groups = [g for g in groups if g["From"] is not None or g["To"] is not None or g["Total"] is not None]

    def gv(row, idx):
        return row[idx] if (idx is not None and idx < len(row)) else None

    def nonempty(x):
        return x is not None and str(x).strip() != ""

    entries = []
    seen = set()
    for rnum in range(2, len(rows)):          # data rows (excel row == rnum + 1)
        excel_row = rnum + 1
        row = list(rows[rnum])

        sys_login_raw = gv(row, login_idx)
        sys_logout_raw = gv(row, logout_idx)
        research_raw = gv(row, research_idx)
        break_hours_raw = gv(row, breakhrs_idx)

        break_vals = [(g["label"], gv(row, g["From"]), gv(row, g["To"]), gv(row, g["Total"])) for g in groups]

        # Skip a fully-empty row (vigilance user entered nothing at all). A value
        # of 0 / 00:00 counts as entered (zero is a legitimate logged value).
        has_any_editable = any(nonempty(x) for x in (sys_login_raw, sys_logout_raw, research_raw, break_hours_raw))
        has_any_break = any(nonempty(v) for (_, f, t, tot) in break_vals for v in (f, t, tot))
        if not has_any_editable and not has_any_break:
            continue

        # Map to employee — Email-id first, then Name fallback. System reference
        # columns never block: if both are tampered/blank we simply cannot map.
        emp = None
        email_raw = gv(row, email_idx)
        if nonempty(email_raw):
            emp = employees_by_email.get(str(email_raw).strip().lower())
        if emp is None:
            name_raw = gv(row, name_idx)
            if nonempty(name_raw):
                emp = employees_by_name.get(str(name_raw).strip().lower())
        if emp is None:
            ident = email_raw if nonempty(email_raw) else gv(row, name_idx)
            errors.append({"row": excel_row, "message": f"Could not match this row to an employee — check Email-id or Name (got '{ident}')."})
            continue

        # Map to the day (record key). Lenient: Excel date / DD-MMM-YYYY / ISO / DD-MM-YYYY.
        date_raw = gv(row, date_idx)
        iso = to_iso(date_raw)
        if not iso:
            errors.append({"row": excel_row, "message": f"Could not read the Date for {emp.get('full_name')} (got '{date_raw}'). Keep the Date column from the template."})
            continue

        key = (emp["id"], iso, uploaded_by["employee_id"])
        if key in seen:
            errors.append({"row": excel_row, "message": f"Duplicate row for {emp.get('full_name')} on {iso_to_display(iso)} within the file."})
            continue

        # Validate ONLY the vigilance-editable fields.
        ok, sys_login = norm_clock(sys_login_raw)
        if not ok:
            errors.append({"row": excel_row, "message": f"System Login '{sys_login_raw}' must be a time like 09:30 (24h) or 09:30 AM."})
            continue
        ok, sys_logout = norm_clock(sys_logout_raw)
        if not ok:
            errors.append({"row": excel_row, "message": f"System Logout '{sys_logout_raw}' must be a time like 18:00 (24h) or 06:00 PM."})
            continue
        ok, research = norm_duration(research_raw)
        if not ok:
            errors.append({"row": excel_row, "message": f"Total Research Hours '{research_raw}' — use HH:MM or HH:MM:SS (00:00 is allowed)."})
            continue
        ok, break_hours = norm_duration(break_hours_raw)
        if not ok:
            errors.append({"row": excel_row, "message": f"Total Break Hours '{break_hours_raw}' — use HH:MM or HH:MM:SS (00:00 is allowed)."})
            continue

        breaks = []
        row_has_break_error = False
        for (label, f, t, tot) in break_vals:
            okf, nf = norm_clock(f)
            okt, nt = norm_clock(t)
            if not okf:
                errors.append({"row": excel_row, "message": f"'{label} From' = '{f}' must be a time like 09:30."})
                row_has_break_error = True
            if not okt:
                errors.append({"row": excel_row, "message": f"'{label} To' = '{t}' must be a time like 10:00."})
                row_has_break_error = True
            # Total is AUTO-CALCULATED (To − From, overnight allowed); the uploaded
            # Total cell is ignored entirely so manual totals can never be wrong.
            nv = compute_break_total(nf, nt) if (okf and okt) else ""
            if okf and okt and (nf or nt or nv):
                breaks.append({"label": label, "from": nf, "to": nt, "total": nv})
        if row_has_break_error:
            continue

        seen.add(key)
        entries.append({
            "target_employee_id": emp["id"],
            "target_employee_name": emp.get("full_name"),
            "target_email": emp.get("official_email"),
            "target_team": emp.get("team"),
            "target_department": emp.get("department"),
            "target_designation": emp.get("designation"),
            "date": iso,
            "system_login": sys_login,
            "system_logout": sys_logout,
            "total_research_hours": research,
            "total_break_hours": break_hours,
            "breaks": breaks,
        })

    return entries, errors


# ----------------------------------------------------------------------------
# Repository
# ----------------------------------------------------------------------------
def now_iso():
    return datetime.now(timezone.utc).isoformat()


async def upsert_entry(db, entry, uploaded_by):
    """Upsert by (target_employee_id, date, uploaded_by_employee_id)."""
    key = {
        "target_employee_id": entry["target_employee_id"],
        "date": entry["date"],
        "uploaded_by_employee_id": uploaded_by["employee_id"],
    }
    set_fields = {
        **entry,
        "uploaded_by_employee_id": uploaded_by["employee_id"],
        "uploaded_by_name": uploaded_by["name"],
        "updated_at": now_iso(),
    }
    res = await db.vigilance_entries.update_one(
        key,
        {"$set": set_fields, "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now_iso()}},
        upsert=True,
    )
    return res


def _clean(doc):
    doc.pop("_id", None)
    return doc


# ----------------------------------------------------------------------------
# Aggregation
# ----------------------------------------------------------------------------
def today_iso():
    """Today's date (IST) as ISO 'YYYY-MM-DD' — used as the default range."""
    return (datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)).strftime("%Y-%m-%d")


def _resolve_range(filters):
    """Return (iso_from, iso_to) defaulting to Today→Today when not supplied."""
    iso_from = filters.get("from_iso") or today_iso()
    iso_to = filters.get("to_iso") or iso_from
    if iso_to < iso_from:
        iso_to = iso_from
    return iso_from, iso_to


def _emp_passes(e, filters):
    name = (filters.get("employee_name") or "").strip().lower()
    if name and name not in (e.get("full_name") or "").lower():
        return False
    dept = filters.get("department")
    if dept and dept != "All" and e.get("department") != dept:
        return False
    desig = filters.get("designation")
    if desig and desig != "All" and e.get("designation") != desig:
        return False
    team = filters.get("team")
    if team and team != "All" and e.get("team") != team:
        return False
    return True


# Public alias — used by the template endpoint to honour the full filter state.
emp_passes_filters = _emp_passes


def _doc_passes(d, filters):
    """Filter a vigilance doc by the same criteria (for fallback rows of
    employees that fall outside the active-employee base grid)."""
    name = (filters.get("employee_name") or "").strip().lower()
    if name and name not in (d.get("target_employee_name") or "").lower():
        return False
    dept = filters.get("department")
    if dept and dept != "All" and d.get("target_department") != dept:
        return False
    desig = filters.get("designation")
    if desig and desig != "All" and d.get("target_designation") != desig:
        return False
    team = filters.get("team")
    if team and team != "All" and d.get("target_team") != team:
        return False
    return True


async def _base_grid(db, filters):
    """Build the SYSTEM-prefilled base grid: every active employee × every day
    in range, with attendance-derived system columns. Always non-empty when at
    least one employee is active in the range. Returns (base, iso_from, iso_to).
    """
    iso_from, iso_to = _resolve_range(filters)
    emps = [e for e in await get_active_employees(db, iso_from, iso_to) if _emp_passes(e, filters)]
    emp_ids = [e["id"] for e in emps]
    att = await get_attendance_map(db, emp_ids, iso_from, iso_to) if emp_ids else {}
    days = daterange_iso(iso_from, iso_to)
    base = {}
    for day in days:
        day_display = iso_to_display(day)
        for e in emps:
            if not is_active_on(e, day):
                continue
            gkey = (e["id"], day)
            a = att.get(gkey, {})
            base[gkey] = {
                "key": f"{e['id']}__{day}",
                "target_employee_id": e["id"],
                "target_employee_name": e.get("full_name"),
                "target_email": e.get("official_email"),
                "target_team": e.get("team"),
                "target_department": e.get("department"),
                "target_designation": e.get("designation"),
                "date": day,
                "date_display": day_display,
                "punch_in": a.get("punch_in", ""),
                "punch_out": a.get("punch_out", ""),
                "total_hours": a.get("total_hours", ""),
            }
    return base, iso_from, iso_to


def _fallback_row(d, att):
    """Build a base row from a vigilance doc (for an employee outside the
    active grid) so previously-uploaded data is never lost."""
    a = att.get((d["target_employee_id"], d["date"]), {})
    return {
        "key": f"{d['target_employee_id']}__{d['date']}",
        "target_employee_id": d["target_employee_id"],
        "target_employee_name": d.get("target_employee_name"),
        "target_email": d.get("target_email"),
        "target_team": d.get("target_team"),
        "target_department": d.get("target_department"),
        "target_designation": d.get("target_designation"),
        "date": d["date"],
        "date_display": iso_to_display(d["date"]),
        "punch_in": a.get("punch_in", ""),
        "punch_out": a.get("punch_out", ""),
        "total_hours": a.get("total_hours", ""),
    }


async def list_own_rows(db, uploaded_by_employee_id, filters):
    """Vigilance user view: ALWAYS the system base grid (active employees × days,
    attendance-prefilled). The user's OWN vigilance entries are merged in by
    (employee + date); other vigilance members' data stays hidden. System columns
    render for every row even when no vigilance data has been entered yet.
    """
    base, iso_from, iso_to = await _base_grid(db, filters)
    for r in base.values():
        r["id"] = None
        r["system_login"] = ""
        r["system_logout"] = ""
        r["total_research_hours"] = ""
        r["total_break_hours"] = ""
        r["breaks"] = []

    q = {"date": {"$gte": iso_from, "$lte": iso_to}, "uploaded_by_employee_id": uploaded_by_employee_id}
    docs = [_clean(d) async for d in db.vigilance_entries.find(q)]
    fb_ids = list({d["target_employee_id"] for d in docs if (d["target_employee_id"], d["date"]) not in base})
    fb_att = await get_attendance_map(db, fb_ids, iso_from, iso_to) if fb_ids else {}

    break_labels = []
    for d in docs:
        gkey = (d["target_employee_id"], d["date"])
        if gkey not in base:
            if not _doc_passes(d, filters):
                continue
            r = _fallback_row(d, fb_att)
            r.update({"id": None, "system_login": "", "system_logout": "",
                      "total_research_hours": "", "total_break_hours": "", "breaks": []})
            base[gkey] = r
        r = base[gkey]
        r["id"] = d["id"]
        r["system_login"] = d.get("system_login", "")
        r["system_logout"] = d.get("system_logout", "")
        r["total_research_hours"] = display_duration(d.get("total_research_hours", ""))
        r["total_break_hours"] = display_duration(d.get("total_break_hours", ""))
        r["breaks"] = _display_breaks(d.get("breaks", []))
        for b in d.get("breaks", []):
            if b["label"] not in break_labels:
                break_labels.append(b["label"])

    rows = sorted(base.values(), key=lambda r: (r["date"], (r["target_employee_name"] or "").lower()))
    return {"rows": rows, "break_labels": _ordered_break_labels(break_labels)}


async def list_admin_merged(db, filters):
    """Admin view: ALWAYS the system base grid (active employees × days), with
    ALL vigilance uploaders' submissions merged per (employee, date). One row per
    employee/day; system columns always render even with zero vigilance uploads.
    """
    base, iso_from, iso_to = await _base_grid(db, filters)
    for r in base.values():
        r["submissions"] = []

    q = {"date": {"$gte": iso_from, "$lte": iso_to}}
    docs = [_clean(d) async for d in db.vigilance_entries.find(q)]
    fb_ids = list({d["target_employee_id"] for d in docs if (d["target_employee_id"], d["date"]) not in base})
    fb_att = await get_attendance_map(db, fb_ids, iso_from, iso_to) if fb_ids else {}

    uploaders = {}
    break_labels = []
    for d in docs:
        gkey = (d["target_employee_id"], d["date"])
        if gkey not in base:
            if not _doc_passes(d, filters):
                continue
            r = _fallback_row(d, fb_att)
            r["submissions"] = []
            base[gkey] = r
        base[gkey]["submissions"].append({
            "id": d["id"],
            "uploaded_by_employee_id": d["uploaded_by_employee_id"],
            "uploaded_by_name": d.get("uploaded_by_name"),
            "system_login": d.get("system_login", ""),
            "system_logout": d.get("system_logout", ""),
            "total_research_hours": display_duration(d.get("total_research_hours", "")),
            "total_break_hours": display_duration(d.get("total_break_hours", "")),
            "breaks": _display_breaks(d.get("breaks", [])),
        })
        uploaders[d["uploaded_by_employee_id"]] = d.get("uploaded_by_name")
        for b in d.get("breaks", []):
            if b["label"] not in break_labels:
                break_labels.append(b["label"])

    rows = sorted(base.values(), key=lambda r: (r["date"], (r["target_employee_name"] or "").lower()))
    uploader_list = [{"employee_id": k, "name": v} for k, v in uploaders.items()]
    uploader_list.sort(key=lambda u: (u["name"] or "").lower())
    return {
        "rows": rows,
        "uploaders": uploader_list,
        "break_labels": _ordered_break_labels(break_labels),
    }


def _ordered_break_labels(labels):
    """Default breaks first (in canonical order), then any dynamic ones sorted."""
    ordered = [lbl for lbl in DEFAULT_BREAK_LABELS if lbl in labels]
    extras = sorted([lbl for lbl in labels if lbl not in DEFAULT_BREAK_LABELS])
    return ordered + extras


async def attendance_integration_map(db, employee_ids, iso_from, iso_to):
    """For admin Attendance module: (emp_id,iso_date) -> [{uploaded_by_employee_id,
    uploaded_by_name, research, break}].

    Real-time read of vigilance_entries (no cache). Frontend-only consumer.
    """
    q = {"date": {"$gte": iso_from, "$lte": iso_to}}
    if employee_ids:
        q["target_employee_id"] = {"$in": employee_ids}
    out = {}
    async for d in db.vigilance_entries.find(q, {"_id": 0}):
        k = f"{d['target_employee_id']}__{d['date']}"
        out.setdefault(k, []).append({
            "uploaded_by_employee_id": d.get("uploaded_by_employee_id"),
            "uploaded_by_name": d.get("uploaded_by_name"),
            "total_research_hours": display_duration(d.get("total_research_hours", "")),
            "total_break_hours": display_duration(d.get("total_break_hours", "")),
        })
    return out


async def list_vigilance_members(db):
    """All employees whose designation is 'Vigilance' — the stable column set for
    the Attendance integration (columns always render, even with no data yet)."""
    emps = await db.employees.find(
        {}, {"_id": 0, "id": 1, "full_name": 1, "designation": 1,
             "inactive_date": 1, "employee_status": 1}
    ).to_list(100000)
    members = [
        {"employee_id": e["id"], "name": e.get("full_name")}
        for e in emps
        if (e.get("designation") or "").strip().lower() == VIGILANCE_DESIGNATION
    ]
    members.sort(key=lambda m: (m["name"] or "").lower())
    return members


# ----------------------------------------------------------------------------
# Export
# ----------------------------------------------------------------------------
def build_export_workbook(rows, break_labels, *, admin_mode, clock_24h=False):
    """Build a filtered-results export with premium merged 2-row headers.

    clock_24h: when True (Vigilance-user download), editable clock fields
    (System Login/Logout + break From/To) are rendered in 24-hour 'HH:MM'.
    Punch-In/Out (attendance-derived) ALWAYS stay 12-hour; durations stay clean.
    """
    fc = (lambda v: to_24h(v)) if clock_24h else (lambda v: v or "")
    wb = Workbook()
    ws = wb.active
    ws.title = "Vigilance Report"

    if admin_mode:
        uploaders = []
        for r in rows:
            for s in r.get("submissions", []):
                if s.get("uploaded_by_name") not in uploaders:
                    uploaders.append(s.get("uploaded_by_name"))
        base = ["Name", "Email-id", "Team", "Department", "Date", "Punch-In", "Punch-Out", "Total Hours"]
        per_uploader_fields = ["System Login", "System Logout", "Total Research Hours", "Total Break Hours"]
        specs = [("scalar", b, _SYS_FILL) for b in base]
        for up in uploaders:
            specs += [("scalar", f"{f} ({up})", _EDIT_FILL) for f in per_uploader_fields]
            specs += [("group", f"{bl} ({up})", _BREAK_FILL) for bl in break_labels]
        total_cols = _write_grouped_header(ws, specs)
        _apply_widths(ws, total_cols, len(base))
        ws.freeze_panes = "F3"
        for r in rows:
            sub_by_up = {s.get("uploaded_by_name"): s for s in r.get("submissions", [])}
            vals = [
                r.get("target_employee_name", ""), r.get("target_email", ""),
                r.get("target_team", ""), r.get("target_department", ""),
                r.get("date_display", ""), r.get("punch_in", ""),
                r.get("punch_out", ""), r.get("total_hours", ""),
            ]
            for up in uploaders:
                s = sub_by_up.get(up, {})
                vals += [s.get("system_login", ""), s.get("system_logout", ""),
                         s.get("total_research_hours", ""), s.get("total_break_hours", "")]
                bmap = {b["label"]: b for b in s.get("breaks", [])}
                for bl in break_labels:
                    b = bmap.get(bl, {})
                    vals += [b.get("from", ""), b.get("to", ""), b.get("total", "")]
            ws.append(vals)
    else:
        base = ["Name", "Email-id", "Team", "Date", "Punch-In", "Punch-Out", "Total Hours",
                "System Login", "System Logout", "Total Research Hours", "Total Break Hours"]
        specs = [("scalar", b, _SYS_FILL if b in ATT_SYSTEM_COLS else _EDIT_FILL) for b in base]
        specs += [("group", bl, _BREAK_FILL) for bl in break_labels]
        total_cols = _write_grouped_header(ws, specs)
        _apply_widths(ws, total_cols, len(base))
        ws.freeze_panes = "E3"
        for r in rows:
            vals = [
                r.get("target_employee_name", ""), r.get("target_email", ""),
                r.get("target_team", ""), r.get("date_display", ""),
                r.get("punch_in", ""), r.get("punch_out", ""), r.get("total_hours", ""),
                fc(r.get("system_login", "")), fc(r.get("system_logout", "")),
                r.get("total_research_hours", ""), r.get("total_break_hours", ""),
            ]
            bmap = {b["label"]: b for b in r.get("breaks", [])}
            for bl in break_labels:
                b = bmap.get(bl, {})
                vals += [fc(b.get("from", "")), fc(b.get("to", "")), b.get("total", "")]
            ws.append(vals)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
