"""Export -> edit -> import -> re-export round-trip probe (run locally)."""
import os
import sys

import requests
from openpyxl import load_workbook

API = os.environ["API"]
MONTH = os.environ.get("MONTH", "2026-09")
tok = requests.post(f"{API}/api/auth/login",
                    json={"username": "admin", "password": "HrAdmin786$"}).json()["token"]
H = {"Authorization": f"Bearer {tok}"}


def export(fmt="xlsx", month=MONTH):
    r = requests.get(f"{API}/api/brsf/export", params={"month": month, "format": fmt}, headers=H)
    r.raise_for_status()
    path = f"/tmp/rt.{fmt}"
    open(path, "wb").write(r.content)
    return path


def preview(path, month=MONTH):
    with open(path, "rb") as f:
        return requests.post(f"{API}/api/brsf/import/preview", data={"month": month},
                             files={"file": (os.path.basename(path), f)}, headers=H)


p1 = export()
wb = load_workbook(p1)
ws = wb[wb.sheetnames[0]]
print("row3 before:", ws["A3"].value, "G3(Perf)", ws["G3"].value, "F3(Innov)", ws["F3"].value,
      "E3(P01)", ws["E3"].value, "L3(N03)", ws["L3"].value)

# valid edits + invalid edits + blank + unchanged
ws["G3"] = 4          # Performance -> manual monthly 4 (valid)
ws["F4"] = 5          # Innovation row4 -> invalid (max +3)
ws["E3"] = None       # blank -> must NOT change P01
ws["L5"] = -2         # N03 -> invalid (only 0 or -3)
ws["H3"] = 0.5        # decimal -> invalid
ws["S3"] = 999        # tampered total -> must be ignored
wb.save("/tmp/rt_edited.xlsx")

r = preview("/tmp/rt_edited.xlsx")
print("preview status", r.status_code)
data = r.json()
print("summary", data["summary"])
for c in data["changes"]:
    print("  CHANGE", c["employee"], c["code"], c["existing"], "->", c["imported"], c["target"])
for e in data["errors"]:
    print("  ERROR", e["employee"], e["code"], e["imported"], e["message"])
for s in data["skipped"][:3]:
    print("  SKIP", s["employee"], s["reason"])

# wrong-month guard
r2 = preview("/tmp/rt_edited.xlsx", month="2026-08")
print("wrong month ->", r2.status_code, r2.json().get("detail"))

conf = requests.post(f"{API}/api/brsf/import/confirm",
                     json={"batch_id": data["batch_id"]}, headers=H)
print("confirm", conf.status_code, conf.json().get("message"), "failed", conf.json().get("failed"))
print("re-confirm ->", requests.post(f"{API}/api/brsf/import/confirm",
                                     json={"batch_id": data["batch_id"]}, headers=H).status_code)

p2 = export()
ws2 = load_workbook(p2)[load_workbook(p2).sheetnames[0]]
print("row3 after: G3(Perf)", ws2["G3"].value, "E3(P01 unchanged)", ws2["E3"].value,
      "S3 formula", load_workbook(p2)[load_workbook(p2).sheetnames[0]]["S3"].value)

# csv round trip
pc = export("csv")
print("csv head:", open(pc).readline().strip()[:120])
rc = preview(pc)
print("csv preview", rc.status_code, rc.json()["summary"])
sys.exit(0)
