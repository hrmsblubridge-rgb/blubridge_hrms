"""Historical employee-count accuracy — regression tests.

Spec (user, Jun-2026): attendance reports must be date-aware, never filtered
by the employee's CURRENT Active/Inactive status.

  Include an employee on a date when:
      date >= date_of_joining
  AND date <= last working day (last_day_payable, fallback inactive_date);
      no exit date recorded → treat as still employed.

Covered surfaces:
  1. /reports/office-attendance      — inactive WITHOUT exit date stays counted
  2. /reports/attendance/export      — blank "-" cells outside employment,
                                       full exclusion when no window overlap
  3. /attendance (gap-fill stubs)    — Resigned employees stop generating
                                       Absent stubs after their last day
  4. /attendance/stats               — window totals include ex-employees who
                                       were employed during the window
"""
import io
import os
import sys
import uuid

import httpx
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

API = "http://localhost:8001/api"
HTTP_TIMEOUT = 60.0


def _login(username, password):
    r = httpx.post(
        f"{API}/auth/login",
        json={"username": username, "password": password},
        timeout=HTTP_TIMEOUT,
    )
    if r.status_code != 200:
        return None
    body = r.json()
    return body.get("token") or body.get("access_token")


@pytest.fixture(scope="module")
def admin_token():
    t = _login("sysadmin", "pass123")
    assert t, "sysadmin login failed"
    return t


def _h(token):
    return {"Authorization": f"Bearer {token}"}


def _base_emp(emp_id, name, **overrides):
    doc = {
        "id": emp_id,
        "emp_id": f"HIST-{emp_id[:8]}",
        "full_name": name,
        "department": "Test",
        "team": "Test",
        "designation": "Test",
        "date_of_joining": "2024-01-01",
        "employee_status": "Active",
        "attendance_tracking_enabled": True,
        "is_deleted": False,
    }
    doc.update(overrides)
    return doc


# ---------------------------------------------------------------------------
# 1) Office report — inactive WITHOUT exit date = still employed (spec rule)
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_office_report_inactive_without_exit_date_still_counted(admin_token):
    from server import db  # noqa

    emp_id = str(uuid.uuid4())
    office = "HistCount-Test-Office"
    loc_id = None
    try:
        r = httpx.post(
            f"{API}/settings/office-locations",
            json={"name": office},
            headers=_h(admin_token),
            timeout=HTTP_TIMEOUT,
        )
        assert r.status_code == 200, r.text
        loc_id = r.json()["id"]

        await db.employees.insert_one(_base_emp(
            emp_id, "__HIST_NOEXIT__",
            office_location=office,
            employee_status="Inactive",   # no inactive_date / last_day_payable
        ))

        r = httpx.get(
            f"{API}/reports/office-attendance",
            headers=_h(admin_token),
            params={"from_date": "02-03-2026", "to_date": "04-03-2026"},
            timeout=HTTP_TIMEOUT,
        )
        assert r.status_code == 200, r.text
        mine = [row for row in r.json() if row["office_location"] == office]
        assert len(mine) == 3
        for row in mine:
            assert row["total_employees"] == 1, (
                f"Inactive employee with NO exit date must still be counted: {row}"
            )
    finally:
        await db.employees.delete_one({"id": emp_id})
        if loc_id:
            httpx.delete(f"{API}/settings/office-locations/{loc_id}",
                         headers=_h(admin_token), timeout=HTTP_TIMEOUT)


# ---------------------------------------------------------------------------
# 2) Attendance report XLSX export — per-date employment window
# ---------------------------------------------------------------------------

def _export_status_by_date(token, name, frm, to):
    """Download the pivot XLSX filtered to one employee and return
    {date_iso: status_cell} from the single data row (or None if no row)."""
    from openpyxl import load_workbook
    r = httpx.get(
        f"{API}/reports/attendance/export",
        headers=_h(token),
        params={"from_date": frm, "to_date": to, "employee_name": name},
        timeout=HTTP_TIMEOUT,
    )
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    if ws.max_row < 3:
        return None
    out = {}
    col = 5
    while True:
        date_hdr = ws.cell(row=1, column=col).value
        if not date_hdr:
            break
        out[str(date_hdr)] = ws.cell(row=3, column=col + 5).value  # Status sub-col
        col += 6
    return out


@pytest.mark.asyncio
async def test_export_blanks_outside_employment_window(admin_token):
    from server import db  # noqa

    emp_id = str(uuid.uuid4())
    name = "__HIST_EXPORT__"
    try:
        await db.employees.insert_one(_base_emp(
            emp_id, name,
            date_of_joining="2026-03-05",
            employee_status="Inactive",
            inactive_date="2026-03-15",
            last_day_payable="2026-03-15",
        ))
        statuses = _export_status_by_date(admin_token, name, "01-03-2026", "20-03-2026")
        assert statuses is not None, "Employee must be included (window overlaps employment)"

        for iso in ["2026-03-01", "2026-03-02", "2026-03-03", "2026-03-04"]:
            assert statuses[iso] == "-", f"{iso} is before DOJ, expected '-': {statuses[iso]}"
        for iso in ["2026-03-16", "2026-03-17", "2026-03-20"]:
            assert statuses[iso] == "-", f"{iso} is after exit, expected '-': {statuses[iso]}"
        # Mid-employment (no punches) → Not Login variants, never "-"
        for iso in ["2026-03-05", "2026-03-10", "2026-03-15"]:
            assert "Not Login" in str(statuses[iso]), (
                f"{iso} is inside employment, expected Not Login: {statuses[iso]}"
            )
    finally:
        await db.employees.delete_one({"id": emp_id})


@pytest.mark.asyncio
async def test_export_excludes_employee_with_no_window_overlap(admin_token):
    from server import db  # noqa

    emp_id = str(uuid.uuid4())
    name = "__HIST_NOOVERLAP__"
    try:
        await db.employees.insert_one(_base_emp(
            emp_id, name,
            date_of_joining="2026-03-05",
            employee_status="Inactive",
            inactive_date="2026-03-15",
            last_day_payable="2026-03-15",
        ))
        # Window entirely AFTER exit → excluded
        assert _export_status_by_date(admin_token, name, "16-03-2026", "20-03-2026") is None
        # Window entirely BEFORE joining → excluded
        assert _export_status_by_date(admin_token, name, "01-03-2026", "04-03-2026") is None
    finally:
        await db.employees.delete_one({"id": emp_id})


# ---------------------------------------------------------------------------
# 3) /attendance gap-fill — Resigned employees stop at last working day
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_attendance_gapfill_resigned_stops_at_last_day(admin_token):
    from server import db  # noqa

    emp_id = str(uuid.uuid4())
    name = "__HIST_RESIGNED__"
    try:
        await db.employees.insert_one(_base_emp(
            emp_id, name,
            employee_status="Resigned",
            inactive_date="2026-03-15",
        ))
        r = httpx.get(
            f"{API}/attendance",
            headers=_h(admin_token),
            params={"employee_name": name,
                    "from_date": "10-03-2026", "to_date": "20-03-2026"},
            timeout=HTTP_TIMEOUT,
        )
        assert r.status_code == 200, r.text
        rows = [x for x in r.json() if x.get("employee_id") == emp_id]
        dates = sorted({x["date"] for x in rows})
        assert dates, "Expected Absent stubs during employment"
        expected = [f"{d:02d}-03-2026" for d in range(10, 16)]
        assert dates == expected, (
            f"Resigned employee must have stubs ONLY 10..15-Mar, got {dates}"
        )
    finally:
        await db.employees.delete_one({"id": emp_id})


# ---------------------------------------------------------------------------
# 4) /attendance/stats — window totals are date-aware
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_stats_totals_include_ex_employee_for_historical_window(admin_token):
    from server import db  # noqa

    def _total(frm, to):
        r = httpx.get(
            f"{API}/attendance/stats",
            headers=_h(admin_token),
            params={"from_date": frm, "to_date": to},
            timeout=HTTP_TIMEOUT,
        )
        assert r.status_code == 200, r.text
        return r.json()["total_employees"]

    emp_id = str(uuid.uuid4())
    try:
        before_overlap = _total("01-03-2026", "31-03-2026")
        before_after = _total("01-04-2026", "30-04-2026")

        await db.employees.insert_one(_base_emp(
            emp_id, "__HIST_STATS__",
            date_of_joining="2026-03-01",
            employee_status="Inactive",
            inactive_date="2026-03-15",
            last_day_payable="2026-03-15",
        ))

        assert _total("01-03-2026", "31-03-2026") == before_overlap + 1, (
            "March window must include the employee who worked 01..15-Mar"
        )
        assert _total("01-04-2026", "30-04-2026") == before_after, (
            "April window must NOT include an employee who exited 15-Mar"
        )
    finally:
        await db.employees.delete_one({"id": emp_id})
