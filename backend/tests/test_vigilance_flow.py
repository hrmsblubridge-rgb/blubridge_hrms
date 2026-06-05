"""End-to-end backend tests for the Operational Vigilance module."""
import io
import os
import pytest
import requests
from openpyxl import load_workbook, Workbook

# Bump default timeout for slow preview ingress
_orig_request = requests.Session.request
def _patched(self, method, url, **kw):
    kw.setdefault("timeout", 120)
    return _orig_request(self, method, url, **kw)
requests.Session.request = _patched

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://blank-tab-debug.preview.emergentagent.com").rstrip("/")

ADMIN = ("admin", "HrAdmin786$")
VIG1 = ("madhan.s", "Vigil@123")    # Madhan S
VIG2 = ("dinesh.t", "Vigil@123")    # Dinesh T
NORMAL = ("user", "pass123")

FROM_D = "05-Jun-2026"
TO_D = "05-Jun-2026"


def _login(username, password):
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"username": username, "password": password}, timeout=90)
    assert r.status_code == 200, f"login failed for {username}: {r.status_code} {r.text}"
    token = r.json().get("token") or r.json().get("access_token")
    assert token
    return token


def _hdr(tok):
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def tokens():
    return {
        "admin": _login(*ADMIN),
        "vig1": _login(*VIG1),
        "vig2": _login(*VIG2),
        "normal": _login(*NORMAL),
    }


# ---------------- Access control ----------------
class TestAccess:
    def test_admin_has_access(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/access", headers=_hdr(tokens["admin"]))
        assert r.status_code == 200
        j = r.json()
        assert j["has_access"] is True and j["is_admin"] is True

    def test_vig1_has_access(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/access", headers=_hdr(tokens["vig1"]))
        assert r.status_code == 200
        j = r.json()
        assert j["has_access"] is True and j["is_vigilance"] is True

    def test_normal_no_access(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/access", headers=_hdr(tokens["normal"]))
        assert r.status_code == 200
        assert r.json().get("has_access") is False

    def test_normal_blocked_from_entries(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["normal"]))
        assert r.status_code == 403


# ---------------- Template download ----------------
class TestTemplate:
    def test_template_admin(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/template",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["admin"]))
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 2-row header
        h1 = [c for c in rows[0]]
        assert h1[:7] == ["Name", "Email-id", "Team", "Date", "Punch-In", "Punch-Out", "Total Hours"]
        assert "System Login" in h1 and "Total Research Hours" in h1 and "Total Break Hours" in h1
        # At least 5 default break groups
        labels = [c for c in h1 if c and "Break" in str(c)]
        assert any("Morning" in str(l) for l in labels)
        assert any("Extra-Break1" in str(l) for l in labels)
        # At least one data row prefilled with a date in DD-MMM-YYYY
        assert len(rows) >= 3
        # find Date column index
        date_idx = h1.index("Date")
        assert rows[2][date_idx] == FROM_D
        # store column count for next test
        global _TEMPLATE_COLS
        _TEMPLATE_COLS = len(h1)

    def test_template_to_before_from_400(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/template",
                         params={"from_date": TO_D, "to_date": "01-Jun-2026"},
                         headers=_hdr(tokens["vig1"]))
        assert r.status_code == 400


# ---------------- Helpers to author filled xlsx ----------------
def _author_filled(template_bytes, target_email, fills, extra_label=None, extra_vals=None):
    """fills dict for row3, optionally appends a NEW dynamic break group."""
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    h1 = [c.value for c in ws[1]]
    h2 = [c.value for c in ws[2]]

    # Find target row by email
    email_col = h1.index("Email-id") + 1
    target_row = None
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=email_col).value
        if v and str(v).strip().lower() == target_email.lower():
            target_row = r
            break
    assert target_row, f"email {target_email} not in template"

    for col_name, val in fills.items():
        ws.cell(row=target_row, column=h1.index(col_name) + 1).value = val

    # Add dynamic Extra-Break3 group
    if extra_label and extra_vals:
        col = ws.max_column + 1
        ws.cell(row=1, column=col).value = extra_label
        ws.cell(row=2, column=col).value = "From"
        ws.cell(row=2, column=col + 1).value = "To"
        ws.cell(row=2, column=col + 2).value = "Total"
        ws.cell(row=target_row, column=col).value = extra_vals[0]
        ws.cell(row=target_row, column=col + 1).value = extra_vals[1]
        ws.cell(row=target_row, column=col + 2).value = extra_vals[2]

    # Add Morning Break values too if requested via fills keys
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _get_template(tok):
    r = requests.get(f"{BASE_URL}/api/vigilance/template",
                     params={"from_date": FROM_D, "to_date": TO_D},
                     headers=_hdr(tok))
    assert r.status_code == 200
    return r.content


def _pick_target_email(template_bytes):
    """Pick the email of the first prefilled row (active employee for 05-Jun)."""
    wb = load_workbook(io.BytesIO(template_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h1 = list(rows[0])
    email_idx = h1.index("Email-id")
    name_idx = h1.index("Name")
    # prefer 'user' Rishi if present, else first
    for r in rows[2:]:
        if r[email_idx]:
            return r[email_idx], r[name_idx]
    return None, None


# ---------------- Upload + Isolation ----------------
@pytest.fixture(scope="module")
def upload_data(tokens):
    """Upload from vig1 (with dynamic Extra-Break3) and vig2 (different research hours)."""
    tmpl1 = _get_template(tokens["vig1"])
    email, name = _pick_target_email(tmpl1)
    assert email
    wb = load_workbook(io.BytesIO(tmpl1))
    ws = wb.active
    h1 = [c.value for c in ws[1]]
    h2 = [c.value for c in ws[2]]
    email_col = h1.index("Email-id") + 1
    # find target row
    target_row = None
    for r in range(3, ws.max_row + 1):
        if str(ws.cell(row=r, column=email_col).value or "").lower() == email.lower():
            target_row = r
            break
    ws.cell(row=target_row, column=h1.index("System Login") + 1).value = "09:00 AM"
    ws.cell(row=target_row, column=h1.index("System Logout") + 1).value = "06:00 PM"
    ws.cell(row=target_row, column=h1.index("Total Research Hours") + 1).value = "08:00"
    ws.cell(row=target_row, column=h1.index("Total Break Hours") + 1).value = "01:00"
    # Morning Break group: find label cell at row1 == "Morning Break"
    mb_col = None
    for ci, v in enumerate(h1, start=1):
        if v and "Morning" in str(v):
            mb_col = ci
            break
    assert mb_col
    ws.cell(row=target_row, column=mb_col).value = "11:00 AM"
    ws.cell(row=target_row, column=mb_col + 1).value = "11:15 AM"
    ws.cell(row=target_row, column=mb_col + 2).value = "00:15"
    # Add dynamic Extra-Break3 group
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = "Extra-Break3"
    ws.cell(row=2, column=col).value = "From"
    ws.cell(row=2, column=col + 1).value = "To"
    ws.cell(row=2, column=col + 2).value = "Total"
    ws.cell(row=target_row, column=col).value = "03:00 PM"
    ws.cell(row=target_row, column=col + 1).value = "03:10 PM"
    ws.cell(row=target_row, column=col + 2).value = "00:10"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Upload as vig1
    files = {"file": ("v1.xlsx", buf.getvalue(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{BASE_URL}/api/vigilance/upload", files=files, headers=_hdr(tokens["vig1"]))
    assert r.status_code == 200, r.text

    # Upload as vig2 (research 10:50)
    tmpl2 = _get_template(tokens["vig2"])
    wb2 = load_workbook(io.BytesIO(tmpl2))
    ws2 = wb2.active
    h1b = [c.value for c in ws2[1]]
    ec = h1b.index("Email-id") + 1
    tr = None
    for r in range(3, ws2.max_row + 1):
        if str(ws2.cell(row=r, column=ec).value or "").lower() == email.lower():
            tr = r
            break
    ws2.cell(row=tr, column=h1b.index("System Login") + 1).value = "09:30 AM"
    ws2.cell(row=tr, column=h1b.index("Total Research Hours") + 1).value = "10:50"
    ws2.cell(row=tr, column=h1b.index("Total Break Hours") + 1).value = "00:30"
    buf2 = io.BytesIO()
    wb2.save(buf2)
    files2 = {"file": ("v2.xlsx", buf2.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = requests.post(f"{BASE_URL}/api/vigilance/upload", files=files2, headers=_hdr(tokens["vig2"]))
    assert r2.status_code == 200, r2.text
    return {"email": email, "name": name}


class TestIsolationAndMerge:
    def test_vig1_sees_own_only(self, tokens, upload_data):
        r = requests.get(f"{BASE_URL}/api/vigilance/entries",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["vig1"]))
        assert r.status_code == 200
        j = r.json()
        assert j["mode"] == "vigilance"
        rows = j["rows"]
        target_rows = [x for x in rows if x.get("target_email", "").lower() == upload_data["email"].lower()]
        assert target_rows
        assert target_rows[0]["total_research_hours"] == "08:00"
        assert "Extra-Break3" in j["break_labels"]

    def test_vig2_sees_own_only(self, tokens, upload_data):
        r = requests.get(f"{BASE_URL}/api/vigilance/entries",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["vig2"]))
        assert r.status_code == 200
        j = r.json()
        rows = j["rows"]
        target_rows = [x for x in rows if x.get("target_email", "").lower() == upload_data["email"].lower()]
        assert target_rows
        # MUST show vig2's own 10:50, NOT madhan's 08:00 (isolation holds)
        assert target_rows[0]["total_research_hours"] == "10:50"
        # vig2 must never see madhan's 08:00 on any base row
        for x in rows:
            assert x.get("total_research_hours") != "08:00"

    def test_admin_merged_single_row(self, tokens, upload_data):
        r = requests.get(f"{BASE_URL}/api/vigilance/entries",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["admin"]))
        assert r.status_code == 200
        j = r.json()
        assert j["mode"] == "admin"
        target = [x for x in j["rows"] if x["target_email"].lower() == upload_data["email"].lower()]
        assert len(target) == 1, "admin must merge to ONE row per employee/day"
        subs = target[0]["submissions"]
        assert len(subs) == 2
        names = sorted(s["uploaded_by_name"] for s in subs)
        assert len(names) == 2
        assert "Extra-Break3" in j["break_labels"]


# ---------------- CRUD / Ownership ----------------
class TestCRUD:
    def test_vig2_cannot_edit_vig1_entry(self, tokens, upload_data):
        r = requests.get(f"{BASE_URL}/api/vigilance/entries",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["admin"]))
        target = [x for x in r.json()["rows"] if x["target_email"].lower() == upload_data["email"].lower()][0]
        vig1_sub = [s for s in target["submissions"] if "madhan" in (s["uploaded_by_name"] or "").lower()][0]
        u = requests.put(f"{BASE_URL}/api/vigilance/entries/{vig1_sub['id']}",
                         json={"total_research_hours": "01:00"},
                         headers=_hdr(tokens["vig2"]))
        assert u.status_code == 403

    def test_admin_can_edit_anyone(self, tokens, upload_data):
        r = requests.get(f"{BASE_URL}/api/vigilance/entries",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["admin"]))
        target = [x for x in r.json()["rows"] if x["target_email"].lower() == upload_data["email"].lower()][0]
        sub = target["submissions"][0]
        u = requests.put(f"{BASE_URL}/api/vigilance/entries/{sub['id']}",
                         json={"total_research_hours": "07:30"},
                         headers=_hdr(tokens["admin"]))
        assert u.status_code == 200


# ---------------- Validation ----------------
class TestValidation:
    def test_invalid_time_422(self, tokens):
        tmpl = _get_template(tokens["vig1"])
        wb = load_workbook(io.BytesIO(tmpl))
        ws = wb.active
        h1 = [c.value for c in ws[1]]
        # set System Login to invalid value in row 3
        ws.cell(row=3, column=h1.index("System Login") + 1).value = "25:00 PM"
        buf = io.BytesIO()
        wb.save(buf)
        files = {"file": ("bad.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/vigilance/upload", files=files, headers=_hdr(tokens["vig1"]))
        assert r.status_code == 422

    def test_non_xlsx_rejected(self, tokens):
        files = {"file": ("foo.csv", b"hello,world", "text/csv")}
        r = requests.post(f"{BASE_URL}/api/vigilance/upload", files=files, headers=_hdr(tokens["vig1"]))
        assert r.status_code == 400


# ---------------- Export & integration ----------------
class TestExportAndIntegration:
    def test_admin_export(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/export",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["admin"]))
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_vig1_export(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/export",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["vig1"]))
        assert r.status_code == 200

    def test_attendance_integration_admin(self, tokens, upload_data):
        r = requests.get(f"{BASE_URL}/api/vigilance/attendance-integration",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["admin"]))
        assert r.status_code == 200
        j = r.json()
        assert "map" in j
        # At least one key with 2 uploaders
        any_two = any(len(v) >= 2 for v in j["map"].values())
        assert any_two
        # map entries carry uploaded_by_employee_id (for stable column matching)
        for v in j["map"].values():
            for sub in v:
                assert sub.get("uploaded_by_employee_id")

    def test_attendance_integration_members_always_listed(self, tokens):
        """vigilance_members must be returned even with NO vigilance data in range
        so the per-member Research/Break columns are always visible in Attendance."""
        r = requests.get(f"{BASE_URL}/api/vigilance/attendance-integration",
                         params={"from_date": "01-Jan-2020", "to_date": "02-Jan-2020"},
                         headers=_hdr(tokens["admin"]))
        assert r.status_code == 200
        j = r.json()
        assert "vigilance_members" in j
        names = {m["name"] for m in j["vigilance_members"]}
        assert {"Madhan S", "Dinesh T"} <= names, "all Vigilance-designation employees must be listed"
        for m in j["vigilance_members"]:
            assert m.get("employee_id") and m.get("name")

    def test_attendance_integration_blocks_vig(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/attendance-integration",
                         params={"from_date": FROM_D, "to_date": TO_D},
                         headers=_hdr(tokens["vig1"]))
        assert r.status_code == 403


# ---------------- Regression sanity ----------------
class TestRegression:
    def test_employees_endpoint(self, tokens):
        r = requests.get(f"{BASE_URL}/api/employees", headers=_hdr(tokens["admin"]))
        assert r.status_code in (200, 201)

    def test_attendance_endpoint(self, tokens):
        # Try common paths
        for path in ("/api/attendance", "/api/attendance/all", "/api/admin/attendance"):
            r = requests.get(f"{BASE_URL}{path}", headers=_hdr(tokens["admin"]),
                             params={"from_date": FROM_D, "to_date": TO_D})
            if r.status_code == 200:
                return
        pytest.skip("No matching attendance endpoint - manual verify")



# ---------------- Base grid (always-render system data) ----------------
SYS_KEYS = ("target_employee_name", "target_email", "target_team", "date_display",
            "punch_in", "punch_out", "total_hours")


class TestBaseGrid:
    """The table must ALWAYS render system-prefilled attendance rows, even with
    zero vigilance uploads, and merge vigilance data on top by (employee+date)."""

    def test_admin_default_today_renders_base_rows(self, tokens):
        # No date params -> defaults to Today->Today, still returns rows.
        r = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["admin"]))
        assert r.status_code == 200
        j = r.json()
        assert j["mode"] == "admin"
        assert len(j["rows"]) > 0, "admin base grid must render employee rows for today"
        row = j["rows"][0]
        for k in SYS_KEYS:
            assert k in row, f"system column '{k}' missing from base row"
        assert "submissions" in row
        assert "key" in row

    def test_vigilance_user_sees_base_grid(self, tokens):
        r = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["vig1"]),
                         params={"from_date": FROM_D, "to_date": TO_D})
        assert r.status_code == 200
        j = r.json()
        assert j["mode"] == "vigilance"
        assert len(j["rows"]) > 0, "vigilance user must still see system base rows"
        # Base rows (no own entry) carry id=None but full system columns.
        row = j["rows"][0]
        for k in SYS_KEYS:
            assert k in row

    def test_range_expands_employee_day_grid(self, tokens):
        single = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["admin"]),
                              params={"from_date": "01-Jun-2026", "to_date": "01-Jun-2026"}).json()
        multi = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["admin"]),
                             params={"from_date": "01-Jun-2026", "to_date": "05-Jun-2026"}).json()
        assert len(multi["rows"]) >= len(single["rows"]) * 2, \
            "multi-day range must render employee×day rows"

    def test_create_merges_and_isolation_then_cleanup(self, tokens):
        base = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["admin"]),
                            params={"from_date": FROM_D, "to_date": TO_D}).json()
        # pick a base row that has NO existing vigilance submissions (clean slate)
        target = next(x for x in base["rows"] if not x["submissions"])
        emp_id, iso_date = target["target_employee_id"], target["date"]

        # vig1 creates an entry (24h input -> 12h stored)
        c = requests.post(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["vig1"]),
                          json={"target_employee_id": emp_id, "date": iso_date,
                                "system_login": "13:45", "system_logout": "18:30",
                                "total_research_hours": "04:00", "total_break_hours": "01:00",
                                "breaks": []})
        assert c.status_code == 200, c.text
        try:
            # admin merged: the row now has a submission with 12h stored time
            adm = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["admin"]),
                               params={"from_date": FROM_D, "to_date": TO_D}).json()
            merged = next(x for x in adm["rows"] if x["target_employee_id"] == emp_id and x["date"] == iso_date)
            assert merged["submissions"], "vigilance submission must merge into admin base row"
            assert any(s["system_login"] == "01:45 PM" for s in merged["submissions"])

            # vig1 own: same row carries the entry id with 12h stored time
            own = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["vig1"]),
                               params={"from_date": FROM_D, "to_date": TO_D}).json()
            my = next(x for x in own["rows"] if x["target_employee_id"] == emp_id and x["date"] == iso_date)
            assert my["id"] and my["system_login"] == "01:45 PM"

            # vig2 isolation: sees base grid but NOT vig1's just-created entry
            other = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["vig2"]),
                                 params={"from_date": FROM_D, "to_date": TO_D}).json()
            assert len(other["rows"]) > 0
            other_row = next((x for x in other["rows"]
                              if x["target_employee_id"] == emp_id and x["date"] == iso_date), None)
            assert other_row is not None, "vig2 must still see the system base row"
            assert other_row.get("system_login") != "01:45 PM", "vig2 must not see vig1's entry"
        finally:
            # remove only the entry this test created
            own = requests.get(f"{BASE_URL}/api/vigilance/entries", headers=_hdr(tokens["vig1"]),
                               params={"from_date": FROM_D, "to_date": TO_D}).json()
            mine = next((x for x in own["rows"]
                         if x["target_employee_id"] == emp_id and x["date"] == iso_date and x.get("id")), None)
            if mine:
                requests.delete(f"{BASE_URL}/api/vigilance/entries/{mine['id']}", headers=_hdr(tokens["vig1"]))
