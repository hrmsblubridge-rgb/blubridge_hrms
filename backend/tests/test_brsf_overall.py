"""Iteration 74 — BRSF Overall Star tab (report + export + RBAC + no side effects).

Runs against http://localhost:8001 (preview URL Cloudflare-blocks python-requests).
Covers: reward band boundaries, month range + validation, per-cell state,
month-effective eligibility, cash_total = sum(cash cells only), stars match
Employees tab finals, Employee/Team filters, RBAC 403, xlsx/csv export shape
with cash-only SUM formula, and read-only invariance of brsf_star_lines.
"""
import io
import os
import csv

import sys
sys.path.insert(0, "/app/backend")

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("BRSF_API") or os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8001")
BASE_URL = BASE_URL.rstrip("/")

ADMIN = ("admin", "HrAdmin786$")
EMP = ("user", "pass123")


def _login(u, p):
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"username": u, "password": p}, timeout=90)
    assert r.status_code == 200, f"login {u} -> {r.status_code} {r.text[:200]}"
    j = r.json()
    return j.get("access_token") or j.get("token")


@pytest.fixture(scope="module")
def admin_h():
    return {"Authorization": f"Bearer {_login(*ADMIN)}"}


@pytest.fixture(scope="module")
def emp_h():
    try:
        return {"Authorization": f"Bearer {_login(*EMP)}"}
    except AssertionError:
        pytest.skip("Employee login unavailable")


# ---------------------- REWARD BAND BOUNDARIES ----------------------
def test_reward_band_boundaries():
    """cash_reward mapping direct import — 18 boundary values."""
    from brsf_overall import cash_reward
    cases = [
        (30, "Research Premier", 11000), (25, "Research Premier", 11000),
        (24, "Research Champion", 9000), (22, "Research Champion", 9000),
        (21, "Research Outstanding", 7000), (19, "Research Outstanding", 7000),
        (18, "Research Excellence", 5000), (16, "Research Excellence", 5000),
        (15, "Research Elite", 3000), (9, "Research Elite", 3000),
        (8, "Research Consistent", 2000), (6, "Research Consistent", 2000),
        (5, "Satisfactory", 1000), (4, "Satisfactory", 1000),
        (3, "Needs Improvement", 0), (1, "Needs Improvement", 0),
        (0, "Unsafe Behavior", 0), (-1, "Unsafe Behavior", 0), (-5, "Unsafe Behavior", 0),
    ]
    for stars, cat, cash in cases:
        r = cash_reward(stars)
        assert r["category"] == cat, f"stars={stars} → {r}"
        assert r["cash"] == cash, f"stars={stars} → {r}"


# ---------------------- BASIC REPORT SHAPE ----------------------
@pytest.fixture(scope="module")
def report_apr_sep(admin_h):
    r = requests.get(f"{BASE_URL}/api/brsf/overall",
                     params={"from_month": "2026-04", "to_month": "2026-09"},
                     headers=admin_h, timeout=60)
    assert r.status_code == 200, r.text[:300]
    return r.json()


def test_report_months_and_skipped(report_apr_sep):
    keys = [m["key"] for m in report_apr_sep["months"]]
    assert keys == ["2026-04", "2026-05", "2026-06", "2026-07", "2026-08"]
    assert report_apr_sep["skipped_months"] == ["2026-09"]
    assert isinstance(report_apr_sep.get("teams"), list)
    assert len(report_apr_sep.get("teams")) >= 1


def test_reward_bands_payload_labels(report_apr_sep):
    labels = [b["stars"] for b in report_apr_sep["reward_bands"]]
    assert labels == ["25+", "22-24", "19-21", "16-18", "9-15", "6-8", "4-5", "1-3"]


def test_rows_shape_and_cells(report_apr_sep):
    rows = report_apr_sep["rows"]
    assert len(rows) >= 1
    for r in rows[:3]:
        assert set(["full_name", "team", "cells", "cash_total"]).issubset(r.keys())
        assert len(r["cells"]) == 5
        for c in r["cells"]:
            assert c["state"] in ("value", "not_eligible", "not_calculated")


def test_cash_total_equals_sum_of_value_cells(report_apr_sep):
    """NC and '-' contribute nothing; stars are never summed."""
    for r in report_apr_sep["rows"]:
        expected = sum(c["cash"] for c in r["cells"] if c["state"] == "value")
        assert r["cash_total"] == expected, f"{r['full_name']} cash_total={r['cash_total']} vs {expected}"


# ---------------------- MONTH-EFFECTIVE ELIGIBILITY ----------------------
def _find_row(rep, name_substr):
    for r in rep["rows"]:
        if (r["full_name"] or "").lower().find(name_substr.lower()) >= 0:
            return r
    return None


def test_eligibility_inactive_anuj(report_apr_sep):
    """Anuj Kumar inactive 2026-08-07 → value/NC Apr-Jul, not_eligible Aug."""
    r = _find_row(report_apr_sep, "Anuj Kumar")
    assert r, "Anuj Kumar not found — must appear because eligible in earlier months"
    by = {c["month"]: c for c in r["cells"]}
    for m in ["2026-04", "2026-05", "2026-06", "2026-07"]:
        assert by[m]["state"] in ("value", "not_calculated"), f"{m} → {by[m]}"
    assert by["2026-08"]["state"] == "not_eligible"


def test_eligibility_confirmation_aparna(report_apr_sep):
    """Aparna A confirmed 2026-07-08 → not_eligible Apr-Jun, value/NC from Jul."""
    r = _find_row(report_apr_sep, "Aparna")
    if not r:
        pytest.skip("Aparna A not in dataset")
    by = {c["month"]: c for c in r["cells"]}
    for m in ["2026-04", "2026-05", "2026-06"]:
        assert by[m]["state"] == "not_eligible", f"{m} → {by[m]}"
    for m in ["2026-07", "2026-08"]:
        assert by[m]["state"] in ("value", "not_calculated"), f"{m} → {by[m]}"


def test_eligibility_gowtham(report_apr_sep):
    """Gowtham S inactive 2026-07-21 → not_eligible from Jul onwards."""
    r = _find_row(report_apr_sep, "Gowtham")
    if not r:
        pytest.skip("Gowtham S not in dataset")
    by = {c["month"]: c for c in r["cells"]}
    assert by["2026-07"]["state"] == "not_eligible"
    assert by["2026-08"]["state"] == "not_eligible"


# ---------------------- VALIDATION ----------------------
def test_from_after_to_400(admin_h):
    r = requests.get(f"{BASE_URL}/api/brsf/overall",
                     params={"from_month": "2026-09", "to_month": "2026-04"},
                     headers=admin_h, timeout=20)
    assert r.status_code == 400
    assert "From Month" in r.json().get("detail", "")


def test_range_over_36_months(admin_h):
    r = requests.get(f"{BASE_URL}/api/brsf/overall",
                     params={"from_month": "2020-01", "to_month": "2026-12"},
                     headers=admin_h, timeout=20)
    assert r.status_code == 400
    assert "36" in r.json().get("detail", "")


# ---------------------- FILTERS ----------------------
def test_employee_filter(report_apr_sep, admin_h):
    if not report_apr_sep["rows"]:
        pytest.skip("no rows")
    target = report_apr_sep["rows"][0]
    r = requests.get(f"{BASE_URL}/api/brsf/overall",
                     params={"from_month": "2026-04", "to_month": "2026-08",
                             "employee_id": target["id"]},
                     headers=admin_h, timeout=30)
    assert r.status_code == 200
    data = r.json()
    assert len(data["rows"]) == 1
    assert data["rows"][0]["id"] == target["id"]


def test_team_filter(report_apr_sep, admin_h):
    teams = report_apr_sep.get("teams") or []
    if "Framework - Parallelism" in teams:
        tname = "Framework - Parallelism"
        expected = 5
    else:
        tname = teams[0] if teams else None
        expected = None
    if not tname:
        pytest.skip("no teams")
    r = requests.get(f"{BASE_URL}/api/brsf/overall",
                     params={"from_month": "2026-04", "to_month": "2026-08", "team": tname},
                     headers=admin_h, timeout=30)
    assert r.status_code == 200
    rows = r.json()["rows"]
    for row in rows:
        assert row["team"] == tname
    if expected is not None:
        assert len(rows) == expected, f"expected {expected} for {tname}, got {len(rows)}"


# ---------------------- STARS MATCH EMPLOYEES TAB ----------------------
def test_stars_match_summary(admin_h):
    """Overall Star stars == Final Total Stars from /brsf/summary for the same month."""
    month = "2026-08"
    s = requests.get(f"{BASE_URL}/api/brsf/summary",
                     params={"month": month}, headers=admin_h, timeout=30)
    assert s.status_code == 200
    srows = s.json().get("rows") or []
    if not srows:
        pytest.skip("summary has no rows")
    target = None
    for row in srows:
        # summary uses "final_total" or similar — try common fields
        val = row.get("final_total")
        if val is None:
            val = (row.get("positive_total") or 0) + (row.get("negative_total") or 0)
        if val is not None:
            target = (row.get("employee_id") or row.get("id"), int(round(val)))
            break
    assert target, "no usable summary row"

    ov = requests.get(f"{BASE_URL}/api/brsf/overall",
                      params={"from_month": month, "to_month": month,
                              "employee_id": target[0]},
                      headers=admin_h, timeout=30).json()
    if not ov["rows"]:
        pytest.skip("employee not eligible in overall")
    cell = ov["rows"][0]["cells"][0]
    if cell["state"] != "value":
        pytest.skip(f"cell state {cell['state']}, cannot compare")
    assert cell["stars"] == target[1]


# ---------------------- RBAC ----------------------
def test_rbac_403_employee(emp_h):
    r = requests.get(f"{BASE_URL}/api/brsf/overall",
                     params={"from_month": "2026-04", "to_month": "2026-08"},
                     headers=emp_h, timeout=20)
    assert r.status_code == 403
    r2 = requests.get(f"{BASE_URL}/api/brsf/overall/export",
                      params={"from_month": "2026-04", "to_month": "2026-08"},
                      headers=emp_h, timeout=20)
    assert r2.status_code == 403


# ---------------------- EXPORTS ----------------------
def test_export_xlsx_structure(admin_h, report_apr_sep):
    r = requests.get(f"{BASE_URL}/api/brsf/overall/export",
                     params={"from_month": "2026-04", "to_month": "2026-09", "format": "xlsx"},
                     headers=admin_h, timeout=60)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[:4] == ["Employee Name", "Team Name", "Date Of Joining", "Date Of Employee Confirmation"]
    assert header[-1] == "Cash Rewards Total"
    # 5 completed months → 10 dynamic columns + 4 identity + 1 total = 15
    assert len(header) == 15
    # Freeze panes E2
    assert ws.freeze_panes == "E2"
    # Real Excel formula that sums ONLY cash columns
    n_rows = len(report_apr_sep["rows"])
    if n_rows == 0:
        pytest.skip("no rows to check formula")
    total_cell = ws.cell(row=2, column=15).value
    assert isinstance(total_cell, str) and total_cell.startswith("=SUM("), f"got {total_cell!r}"
    # Cash cols per convention: F, H, J, L, N (every 2nd starting F for 5 months)
    expected_cols = ["F2", "H2", "J2", "L2", "N2"]
    for col in expected_cols:
        assert col in total_cell, f"cash col {col} missing in {total_cell}"
    # Ensure star columns (E, G, I, K, M) are NOT in the formula
    for star_col in ["E2", "G2", "I2", "K2", "M2"]:
        assert star_col not in total_cell, f"star col {star_col} should not be in {total_cell}"


def test_export_xlsx_preserves_dash_and_nc(admin_h, report_apr_sep):
    r = requests.get(f"{BASE_URL}/api/brsf/overall/export",
                     params={"from_month": "2026-04", "to_month": "2026-09", "format": "xlsx"},
                     headers=admin_h, timeout=60)
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    found_dash = False
    found_nc = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        for v in row[4:-1]:
            if v == "-":
                found_dash = True
            if v == "NC":
                found_nc = True
    # At least dash likely exists (some employees inactive in Aug/confirmed later)
    assert found_dash or found_nc, "expected at least one of '-' / 'NC' as text"


def test_export_csv_headers(admin_h):
    r = requests.get(f"{BASE_URL}/api/brsf/overall/export",
                     params={"from_month": "2026-04", "to_month": "2026-09", "format": "csv"},
                     headers=admin_h, timeout=60)
    assert r.status_code == 200
    text = r.content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    hdr = rows[0]
    assert hdr[:4] == ["employee_name", "team_name", "joining_date", "confirmation_date"]
    assert hdr[-1] == "cash_reward_total"
    for m in ["2026-04", "2026-05", "2026-06", "2026-07", "2026-08"]:
        assert f"{m}_stars" in hdr and f"{m}_cash_reward" in hdr


def test_export_honors_team_filter(admin_h, report_apr_sep):
    teams = report_apr_sep.get("teams") or []
    if not teams:
        pytest.skip("no teams")
    tname = teams[0]
    r = requests.get(f"{BASE_URL}/api/brsf/overall/export",
                     params={"from_month": "2026-04", "to_month": "2026-08",
                             "team": tname, "format": "csv"},
                     headers=admin_h, timeout=60)
    assert r.status_code == 200
    text = r.content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    for data_row in rows[1:]:
        assert data_row[1] == tname


# ---------------------- READ-ONLY / NO SIDE EFFECTS ----------------------
def test_no_side_effects_on_star_lines(admin_h):
    """Loading the report must NOT create/modify any brsf_star_lines."""
    def count():
        # Use eligible-employees + summary is not a count; use a probe endpoint. Fall back to summary counts.
        s = requests.get(f"{BASE_URL}/api/brsf/summary",
                        params={"month": "2026-08"}, headers=admin_h, timeout=30).json()
        rows = s.get("rows") or []
        # signature: number of rows + sum of finals
        total = 0
        for row in rows:
            v = row.get("final_total")
            if v is None:
                v = (row.get("positive_total") or 0) + (row.get("negative_total") or 0)
            total += float(v or 0)
        return (len(rows), round(total, 3))

    before = count()
    for _ in range(3):
        requests.get(f"{BASE_URL}/api/brsf/overall",
                     params={"from_month": "2026-04", "to_month": "2026-09"},
                     headers=admin_h, timeout=60)
    after = count()
    assert before == after, f"star lines changed after overall load: {before} → {after}"
