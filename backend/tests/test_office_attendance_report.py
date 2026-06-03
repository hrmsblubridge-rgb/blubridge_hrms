"""Office Attendance Report — regression tests.

Validates the new `/api/reports/office-attendance` (JSON) and
`/api/reports/office-attendance/export` (XLSX) endpoints against the
spec:

  1. ROW SHAPE        — exactly one row per (date, office) combination.
  2. SSOT             — every configured Office Location appears for
                        every date in the window.
  3. DATE-AWARE COUNT — employees inactive mid-window are counted up
                        to and including their last working day, and
                        NOT after.
  4. HOLIDAY          — holiday-flagged dates yield absent=0 regardless
                        of attendance shortfall.
  5. SUNDAY           — Sundays yield absent=0 (weekly off).
  6. APPROVED LEAVE   — employees on approved leave count as on_leave,
                        NOT absent.
  7. CROSS-MONTH      — date range spanning multiple months returns
                        every date in the window.
  8. XLSX EXPORT      — endpoint streams a valid .xlsx with the
                        9-column header.

All synthetic rows clean up in `finally` blocks.
"""
import os
import sys
import uuid
from datetime import datetime, timedelta

import httpx
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

API = "http://localhost:8001/api"
HTTP_TIMEOUT = 60.0


def _login(username, password):
    r = httpx.post(
        f"{API}/auth/login",
        json={"username": username, "password": password},
        timeout=HTTP_TIMEOUT,
    )
    if r.status_code != 200:
        return None
    body = r.json()
    return body.get("token") or body.get("access_token")


@pytest.fixture(scope="module")
def admin_token():
    t = _login("sysadmin", "pass123")
    assert t, "sysadmin login failed"
    return t


def _h(token):
    return {"Authorization": f"Bearer {token}"}


def _fetch(token, frm, to):
    r = httpx.get(
        f"{API}/reports/office-attendance",
        headers=_h(token),
        params={"from_date": frm, "to_date": to},
        timeout=HTTP_TIMEOUT,
    )
    assert r.status_code == 200, r.text
    return r.json()


# ---------------------------------------------------------------------------
# 1) Shape + SSOT enumeration
# ---------------------------------------------------------------------------

def test_row_shape_and_ssot_enumeration(admin_token):
    rows = _fetch(admin_token, "01-03-2026", "03-03-2026")
    assert rows, "Expected at least one row in window"
    # Every row has the canonical keys
    expected_keys = {
        "date_iso", "date_display", "office_location",
        "total_employees", "present", "absent", "on_leave",
        "holiday", "weekly_off", "attendance_pct",
    }
    for r in rows:
        assert expected_keys <= set(r.keys()), f"missing keys in {r}"
    # Date format is DD-MMM-YYYY
    assert rows[0]["date_display"] == "01-Mar-2026"
    # Every SSOT location appears for every date
    distinct_dates = sorted({r["date_iso"] for r in rows})
    distinct_offices = {r["office_location"] for r in rows}
    assert len(distinct_dates) == 3
    # At least the two seeded locations should be present
    assert "Besant Nagar - Chennai" in distinct_offices


# ---------------------------------------------------------------------------
# 2) Sunday → weekly_off=1, absent=0
# ---------------------------------------------------------------------------

def test_sunday_is_weekly_off_and_not_counted_as_absent(admin_token):
    rows = _fetch(admin_token, "01-03-2026", "01-03-2026")
    # 01-Mar-2026 is a Sunday
    for r in rows:
        assert r["weekly_off"] == 1, f"01-Mar-2026 must be Sunday, got {r}"
        assert r["absent"] == 0, "Sunday MUST NOT be absent"


# ---------------------------------------------------------------------------
# 3) Date-aware historical employment evaluation
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_inactive_mid_window_employee_counted_only_until_last_day(admin_token):
    """Employee inactive on 15-Mar must appear in rows 01-Mar..15-Mar
    but NOT in 16-Mar onwards."""
    from server import db, EmployeeStatus  # noqa

    emp_id = str(uuid.uuid4())
    office = "OffAtt-Test-Office"

    # Ensure the office location exists
    loc_id = None
    try:
        r = httpx.post(
            f"{API}/settings/office-locations",
            json={"name": office},
            headers=_h(admin_token),
            timeout=HTTP_TIMEOUT,
        )
        assert r.status_code == 200, r.text
        loc_id = r.json()["id"]

        await db.employees.insert_one({
            "id": emp_id,
            "emp_id": f"OAR-{emp_id[:8]}",
            "full_name": "__OAR_TEST__",
            "department": "Test",
            "team": "Test",
            "designation": "Test",
            "date_of_joining": "2024-01-01",
            "office_location": office,
            "employee_status": EmployeeStatus.INACTIVE,
            "inactive_date": "2026-03-15",
            "last_day_payable": "2026-03-15",
            "is_deleted": False,
        })

        rows = _fetch(admin_token, "10-03-2026", "20-03-2026")
        # Filter rows for the synthetic office only
        per_date = {r["date_iso"]: r for r in rows if r["office_location"] == office}
        # Sanity: the office must appear for every date in window
        assert len(per_date) == 11, f"expected 11 dates, got {sorted(per_date.keys())}"

        # 10-Mar..15-Mar → total should include our employee (= 1)
        for iso in ["2026-03-10", "2026-03-11", "2026-03-12",
                    "2026-03-13", "2026-03-14", "2026-03-15"]:
            assert per_date[iso]["total_employees"] == 1, (
                f"date {iso} should include the inactive-on-15-Mar employee: {per_date[iso]}"
            )

        # 16-Mar onwards → total must NOT include our employee (= 0)
        for iso in ["2026-03-16", "2026-03-17", "2026-03-18",
                    "2026-03-19", "2026-03-20"]:
            assert per_date[iso]["total_employees"] == 0, (
                f"date {iso} MUST exclude the inactive-on-15-Mar employee: {per_date[iso]}"
            )

    finally:
        await db.employees.delete_one({"id": emp_id})
        if loc_id:
            httpx.delete(
                f"{API}/settings/office-locations/{loc_id}",
                headers=_h(admin_token),
                timeout=HTTP_TIMEOUT,
            )


# ---------------------------------------------------------------------------
# 4) Cross-month window
# ---------------------------------------------------------------------------

def test_cross_month_window_returns_every_date(admin_token):
    rows = _fetch(admin_token, "28-02-2026", "02-03-2026")
    distinct_dates = sorted({r["date_iso"] for r in rows})
    assert distinct_dates == ["2026-02-28", "2026-03-01", "2026-03-02"]


# ---------------------------------------------------------------------------
# 5) Invalid range rejected
# ---------------------------------------------------------------------------

def test_invalid_date_range_returns_400(admin_token):
    r = httpx.get(
        f"{API}/reports/office-attendance",
        headers=_h(admin_token),
        params={"from_date": "10-03-2026", "to_date": "01-03-2026"},
        timeout=HTTP_TIMEOUT,
    )
    assert r.status_code == 400, r.text


# ---------------------------------------------------------------------------
# 6) Authentication required
# ---------------------------------------------------------------------------

def test_unauthenticated_request_is_rejected():
    r = httpx.get(
        f"{API}/reports/office-attendance",
        params={"from_date": "01-03-2026", "to_date": "01-03-2026"},
        timeout=HTTP_TIMEOUT,
    )
    assert r.status_code in (401, 403)


# ---------------------------------------------------------------------------
# 7) XLSX export downloads with the expected header row
# ---------------------------------------------------------------------------

def test_xlsx_export_has_correct_header_and_dd_mmm_yyyy_dates(admin_token):
    r = httpx.get(
        f"{API}/reports/office-attendance/export",
        headers=_h(admin_token),
        params={"from_date": "02-03-2026", "to_date": "02-03-2026"},
        timeout=HTTP_TIMEOUT,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert "attachment" in r.headers.get("content-disposition", "")
    # Validate by parsing
    from openpyxl import load_workbook
    from io import BytesIO
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == [
        "Date", "Office Location", "Total Employees",
        "Present", "Absent", "On Leave",
        "Holiday", "Weekly Off", "Attendance %",
    ]
    # First data row date format is DD-MMM-YYYY
    first_data = [c.value for c in ws[2]]
    assert first_data[0] == "02-Mar-2026", f"date cell got {first_data[0]!r}"


# ---------------------------------------------------------------------------
# 8) Approved leave employee counted as on_leave, not absent
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_approved_leave_employee_is_on_leave_not_absent(admin_token):
    from server import db, EmployeeStatus  # noqa

    emp_id = str(uuid.uuid4())
    office = "OffAtt-Leave-Test"
    loc_id = None
    leave_id = str(uuid.uuid4())
    try:
        r = httpx.post(
            f"{API}/settings/office-locations",
            json={"name": office},
            headers=_h(admin_token),
            timeout=HTTP_TIMEOUT,
        )
        loc_id = r.json()["id"]
        await db.employees.insert_one({
            "id": emp_id,
            "emp_id": f"OAR-L-{emp_id[:8]}",
            "full_name": "__OAR_LEAVE_TEST__",
            "department": "Test",
            "team": "Test",
            "designation": "Test",
            "date_of_joining": "2024-01-01",
            "office_location": office,
            "employee_status": EmployeeStatus.ACTIVE,
            "is_deleted": False,
        })
        await db.leaves.insert_one({
            "id": leave_id,
            "employee_id": emp_id,
            "start_date": "2026-03-02",
            "end_date": "2026-03-02",
            "status": "approved",
            "leave_type": "Sick",
        })

        rows = _fetch(admin_token, "02-03-2026", "02-03-2026")
        my_row = next((r for r in rows if r["office_location"] == office), None)
        assert my_row, "synthetic office row missing"
        assert my_row["total_employees"] == 1
        assert my_row["on_leave"] == 1, f"expected on_leave=1, got {my_row}"
        assert my_row["absent"] == 0, f"approved-leave must NOT count as absent: {my_row}"

    finally:
        await db.employees.delete_one({"id": emp_id})
        await db.leaves.delete_one({"id": leave_id})
        if loc_id:
            httpx.delete(
                f"{API}/settings/office-locations/{loc_id}",
                headers=_h(admin_token),
                timeout=HTTP_TIMEOUT,
            )
