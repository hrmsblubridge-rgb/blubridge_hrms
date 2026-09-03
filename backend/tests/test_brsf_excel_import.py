"""BRSF Excel/CSV export + import — regression tests.

Covers the areas the main-agent probe could not fully verify:
  - non-HR 403 on export/import endpoints
  - audit entries created by a confirmed import
  - imported override survives an Auto Calculate run
  - CSV import applying an actual change
  - wrong-month guard, re-confirm idempotency, xlsx structural checks

Runs against localhost:8001 to bypass the Cloudflare bot filter on the preview URL.
"""
import io
import os
import time
import uuid

import pytest
import requests
from openpyxl import load_workbook

API = os.environ.get("BRSF_API", "http://localhost:8001")
MONTH = os.environ.get("BRSF_MONTH", "2026-09")
YEAR, MON = int(MONTH[:4]), int(MONTH[5:7])


# ---------- helpers ----------
def _login(username, password):
    r = requests.post(f"{API}/api/auth/login",
                      json={"username": username, "password": password}, timeout=90)
    if r.status_code != 200:
        pytest.skip(f"Login failed for {username}: {r.status_code} {r.text[:120]}")
    return {"Authorization": f"Bearer {r.json()['token']}"}


@pytest.fixture(scope="module")
def hr_headers():
    return _login("admin", "HrAdmin786$")


@pytest.fixture(scope="module")
def emp_headers():
    return _login("user", "pass123")


def _export(headers, fmt="xlsx", month=MONTH):
    r = requests.get(f"{API}/api/brsf/export",
                     params={"month": month, "format": fmt},
                     headers=headers, timeout=90)
    return r


def _preview(headers, buf, filename, month=MONTH):
    return requests.post(f"{API}/api/brsf/import/preview",
                         data={"month": month},
                         files={"file": (filename, buf)},
                         headers=headers, timeout=90)


def _first_eligible_row(headers):
    """Return (row_index, employee_id, employee_name) for the first eligible employee."""
    r = _export(headers)
    assert r.status_code == 200, r.text[:200]
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb[wb.sheetnames[0]]
    meta = wb["_BRSF_Metadata"]
    row = 3
    name = ws[f"A{row}"].value
    # meta row 2 == excel row 3
    emp_id = meta.cell(row=2, column=2).value
    return row, emp_id, name, wb, ws


# ================================================================
# 1. XLSX structural checks — merged bands, formulas, hidden meta
# ================================================================
class TestExportStructure:
    def test_xlsx_headers_and_formulas(self, hr_headers):
        r = _export(hr_headers)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb[wb.sheetnames[0]]

        assert ws["E1"].value == "Positive Stars"
        assert ws["K1"].value == "Negative Stars"
        merged = {str(m) for m in ws.merged_cells.ranges}
        assert "E1:J1" in merged and "K1:R1" in merged
        for col in ("A", "B", "C", "D", "S", "T", "U"):
            assert f"{col}1:{col}2" in merged, f"missing vertical merge on {col}"

        # freeze panes
        assert ws.freeze_panes == "A3"

        # every data row has real formulas in S/T/U
        assert ws.max_row >= 3
        for row in range(3, ws.max_row + 1):
            assert ws[f"S{row}"].value == f"=SUM(E{row}:J{row})"
            assert ws[f"T{row}"].value == f"=SUM(K{row}:R{row})"
            assert ws[f"U{row}"].value == f"=S{row}+T{row}"

        # hidden metadata sheet
        assert "_BRSF_Metadata" in wb.sheetnames
        meta = wb["_BRSF_Metadata"]
        assert meta.sheet_state == "hidden"
        header = [c.value for c in meta[1]]
        assert header[:5] == ["Sheet Row", "Employee ID", "Employee Email", "BRSF Month", "Export Batch ID"]
        assert meta.cell(row=2, column=4).value == MONTH

    def test_csv_headers(self, hr_headers):
        r = _export(hr_headers, "csv")
        assert r.status_code == 200
        first = r.content.decode("utf-8-sig").splitlines()[0]
        assert first.startswith("Employee Name,Team Name,Date Of Joining,Date Of Employee Confirmation")
        assert first.endswith("Total Positive Stars,Total Negative Stars,Total Stars")

    def test_wrong_month_scope(self, hr_headers):
        # August export should differ from September export in metadata month
        r = _export(hr_headers, month="2026-08")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        assert wb["_BRSF_Metadata"].cell(row=2, column=4).value == "2026-08"


# ================================================================
# 2. RBAC — employee token gets 403 on export/import endpoints
# ================================================================
class TestRBAC:
    def test_employee_cannot_export_xlsx(self, emp_headers):
        r = _export(emp_headers)
        assert r.status_code == 403

    def test_employee_cannot_export_csv(self, emp_headers):
        r = _export(emp_headers, "csv")
        assert r.status_code == 403

    def test_employee_cannot_preview_import(self, emp_headers):
        buf = io.BytesIO(b"dummy")
        r = _preview(emp_headers, buf, "x.csv")
        assert r.status_code == 403

    def test_employee_cannot_confirm_import(self, emp_headers):
        r = requests.post(f"{API}/api/brsf/import/confirm",
                          json={"batch_id": str(uuid.uuid4())},
                          headers=emp_headers, timeout=30)
        assert r.status_code == 403


# ================================================================
# 3. Wrong-month guard on preview
# ================================================================
class TestWrongMonthGuard:
    def test_september_file_previewed_as_august_rejected(self, hr_headers):
        r = _export(hr_headers)
        assert r.status_code == 200
        r2 = _preview(hr_headers, io.BytesIO(r.content), "sept.xlsx", month="2026-08")
        assert r2.status_code == 400
        assert "September" in r2.json().get("detail", "")


# ================================================================
# 4. Full flow: preview -> confirm -> audit -> re-export -> Auto Calc survives
# ================================================================
@pytest.fixture(scope="module")
def round_trip(hr_headers):
    """Perform a Performance=4 edit, confirm, return state for downstream tests."""
    r = _export(hr_headers)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb[wb.sheetnames[0]]
    meta = wb["_BRSF_Metadata"]

    # find an eligible row (row 3 exists per structure test)
    excel_row = 3
    emp_id = meta.cell(row=2, column=2).value
    emp_name = ws[f"A{excel_row}"].value

    # capture existing Performance final
    existing_G = ws[f"G{excel_row}"].value or 0

    # target value: pick 4 unless that equals existing, else 3
    target = 3 if existing_G == 4 else 4

    ws[f"G{excel_row}"] = target
    # also test blank-cell rule
    ws[f"E{excel_row}"] = None
    out = io.BytesIO()
    wb.save(out); out.seek(0)

    pr = _preview(hr_headers, out, "rt.xlsx")
    assert pr.status_code == 200, pr.text[:200]
    pdata = pr.json()

    # find our change
    our_change = next((c for c in pdata["changes"]
                       if c["employee_id"] == emp_id and c["code"] == "P02"), None)
    assert our_change, f"no P02 change found; changes={pdata['changes'][:3]}"
    assert our_change["existing"] == int(round(existing_G))
    assert our_change["imported"] == target
    assert our_change["target"] == "manual"

    # confirm
    cf = requests.post(f"{API}/api/brsf/import/confirm",
                       json={"batch_id": pdata["batch_id"]},
                       headers=hr_headers, timeout=90)
    assert cf.status_code == 200, cf.text[:200]
    cdata = cf.json()
    assert cdata["applied"] >= 1

    # re-confirm must be idempotent-fail (400)
    cf2 = requests.post(f"{API}/api/brsf/import/confirm",
                        json={"batch_id": pdata["batch_id"]},
                        headers=hr_headers, timeout=30)
    assert cf2.status_code == 400

    return {"emp_id": emp_id, "emp_name": emp_name, "target": target,
            "prev_value": int(round(existing_G)), "batch_id": pdata["batch_id"]}


class TestConfirmedImport:
    def test_confirm_applied_manual_value(self, hr_headers, round_trip):
        # GET /api/brsf/stars and check P02 line
        r = requests.get(f"{API}/api/brsf/stars",
                         params={"employee_id": round_trip["emp_id"], "month": MONTH},
                         headers=hr_headers, timeout=60)
        assert r.status_code == 200
        lines = {l["code"]: l for l in r.json()["lines"]}
        p02 = lines["P02"]
        assert p02["manual_value"] == round_trip["target"]
        assert p02["entry_mode"] == "monthly"
        assert int(round(p02["final_value"])) == round_trip["target"]

    def test_audit_entry_created(self, hr_headers, round_trip):
        r = requests.get(f"{API}/api/brsf/audit",
                         params={"employee_id": round_trip["emp_id"], "month": MONTH},
                         headers=hr_headers, timeout=30)
        assert r.status_code == 200
        entries = r.json().get("audit") or []
        matching = [e for e in entries
                    if e.get("import_batch_id") == round_trip["batch_id"]
                    and e.get("code") == "P02"]
        assert matching, f"no audit entry for import; sample={entries[:2]}"
        e = matching[0]
        assert e.get("action") in ("Excel Import", "CSV Import")
        assert e.get("new_value") == round_trip["target"]
        assert "rt.xlsx" in (e.get("reason") or "")

    def test_re_export_shows_imported_value(self, hr_headers, round_trip):
        r = _export(hr_headers)
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb[wb.sheetnames[0]]
        # find employee row via meta
        meta = wb["_BRSF_Metadata"]
        row = None
        for m_row in meta.iter_rows(min_row=2, values_only=True):
            if m_row[1] == round_trip["emp_id"]:
                row = m_row[0]; break
        assert row is not None
        assert ws[f"G{row}"].value == round_trip["target"]
        assert ws[f"S{row}"].value == f"=SUM(E{row}:J{row})"

    def test_override_survives_auto_calculate(self, hr_headers, round_trip):
        # Auto Calculate for the month
        r = requests.post(f"{API}/api/brsf/recalculate",
                          json={"month": MONTH, "employee_id": round_trip["emp_id"]},
                          headers=hr_headers, timeout=180)
        assert r.status_code == 200
        # Re-fetch
        r = requests.get(f"{API}/api/brsf/stars",
                         params={"employee_id": round_trip["emp_id"], "month": MONTH},
                         headers=hr_headers, timeout=60)
        assert r.status_code == 200
        p02 = {l["code"]: l for l in r.json()["lines"]}["P02"]
        # manual monthly value must persist across auto recompute
        assert p02["manual_value"] == round_trip["target"], (
            f"manual value lost after Auto Calculate: {p02}")
        assert int(round(p02["final_value"])) == round_trip["target"]


# ================================================================
# 5. CSV import applying an actual change
# ================================================================
class TestCsvImport:
    def test_csv_change_and_confirm(self, hr_headers, round_trip):
        # export csv
        r = _export(hr_headers, "csv")
        text = r.content.decode("utf-8-sig")
        rows = text.splitlines()
        header = rows[0].split(",")
        # column indexes (0-based): 0..3 A-D employee cols, then 4..17 SHEET_COLUMNS
        # G = Performance = SHEET_COLUMNS[2] -> csv col index 4+2 = 6
        # Change Performance for round_trip employee (find their row by name)
        target_name = round_trip["emp_name"]
        new_value = round_trip["target"]  # already same; pick a different valid value
        alt = 2 if new_value != 2 else 1
        found = False
        for i in range(1, len(rows)):
            cols = rows[i].split(",")
            if cols[0].strip() == (target_name or "").strip():
                cols[6] = str(alt)  # Performance column
                rows[i] = ",".join(cols)
                found = True
                break
        assert found, f"could not find row for {target_name} in csv"

        new_csv = "\n".join(rows).encode("utf-8")
        pr = _preview(hr_headers, io.BytesIO(new_csv), "edited.csv")
        assert pr.status_code == 200, pr.text[:200]
        data = pr.json()
        assert data["source"] == "CSV Import"
        our = [c for c in data["changes"] if c["code"] == "P02" and c["employee"] == target_name]
        assert our, f"no P02 CSV change; changes={data['changes'][:3]}"
        assert our[0]["imported"] == alt

        cf = requests.post(f"{API}/api/brsf/import/confirm",
                           json={"batch_id": data["batch_id"]},
                           headers=hr_headers, timeout=90)
        assert cf.status_code == 200
        assert cf.json()["applied"] >= 1

        # verify
        r = requests.get(f"{API}/api/brsf/stars",
                         params={"employee_id": round_trip["emp_id"], "month": MONTH},
                         headers=hr_headers, timeout=60)
        p02 = {l["code"]: l for l in r.json()["lines"]}["P02"]
        assert p02["manual_value"] == alt


# ================================================================
# 6. Cleanup — revert Performance back to 0 for the touched employee
# ================================================================
def teardown_module(_module):
    try:
        h = _login("admin", "HrAdmin786$")
    except Exception:
        return
    # Fetch summary; not strictly needed. We restore P02 manual_value to 0 for row3 emp.
    r = _export(h)
    if r.status_code != 200:
        return
    wb = load_workbook(io.BytesIO(r.content))
    meta = wb["_BRSF_Metadata"]
    emp_id = meta.cell(row=2, column=2).value
    # Use direct entry endpoint (mimic UI). If unavailable, best-effort skip.
    try:
        # find P02 line
        rr = requests.get(f"{API}/api/brsf/stars",
                          params={"employee_id": emp_id, "month": MONTH},
                          headers=h, timeout=60).json()
        p02 = {l["code"]: l for l in rr["lines"]}.get("P02")
        if p02 and (p02.get("manual_value") or 0) != 0:
            # write manual monthly 0 via the standard endpoint
            requests.put(f"{API}/api/brsf/stars/{p02['id']}/manual",
                         json={"entry_mode": "monthly", "monthly_value": 0,
                               "reason": "revert after brsf_excel test"},
                         headers=h, timeout=30)
    except Exception as e:
        print("teardown revert skipped:", e)
